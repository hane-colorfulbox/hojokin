"""チェック可能 Excel 生成 — 営業ヒアリングシート(05)のオンライン商談・入力版。

使い方:
    python make_check_xlsx.py <05_面談ヒアリングシート.md>
入力 md と同じ場所に `<basename>_オンライン入力用.xlsx` を出力する（md は出力名の決定だけに使う）。
設問データは hearing_form.FORM を参照する。

設計（Google スプレッドシート互換）:
- チェックは「未/済」のインライン list データ検証（ドロップダウン）。
  Excel のフォームコントロール／新セルチェックボックスは openpyxl で作れず Google でも消えるため使わない。
  インライン list の入力規則は Google スプレッドシートへインポートしてもドロップダウンとして残る
  （既存 `ツール/ヒアリングシート2026_*.xlsx` と同方式）。
- 値は環境非依存の「未／済」（☐☑✓ は環境依存で豆腐化するため値に使わない）。
  「済」セルは条件付き書式で緑塗りにして視認性を担保する。
- 1設問を「設問(A・縦結合) / 選択肢(B) / チェック(C=未済) / 自由記述・メモ(D)」の縦型チェックリストにする。

⚠️ 同期注意：設問・選択肢は hearing_form.FORM が単一の真実。05.md / FORM を変えたら本スクリプトを
   再実行して xlsx を作り直すこと（make_fillable_pdf.py / make_check_md.py も同様）。
"""
import sys
import pathlib

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

from hearing_form import FORM, split_check_row, freetext_items  # 設問データと共有の解釈ヘルパー

# ---- レイアウト定数（マジックナンバーは上部に集約） ----
FONT_NAME = "游ゴシック"
TITLE_FONT = Font(name=FONT_NAME, size=15, bold=True)
INSTR_FONT = Font(name=FONT_NAME, size=10, color="555555")
SECTION_FONT = Font(name=FONT_NAME, size=11.5, bold=True, color="1F3A52")
NOTE_FONT = Font(name=FONT_NAME, size=9, color="666666")
PROMPT_FONT = Font(name=FONT_NAME, size=10.5, bold=True)
OPTION_FONT = Font(name=FONT_NAME, size=10.5)
SUBHEAD_FONT = Font(name=FONT_NAME, size=10.5, bold=True, color="1F3A52")

SECTION_FILL = PatternFill("solid", fgColor="DCE6F1")
INPUT_FILL = PatternFill("solid", fgColor="FFFDF2")
DONE_FILL = PatternFill("solid", fgColor="C6EFCE")  # 「済」セルの緑塗り（条件付き書式）

_thin = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(bottom=_thin)
INPUT_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

TOP = Alignment(vertical="top", wrap_text=True)
TOPLEFT = Alignment(vertical="top", horizontal="left", wrap_text=True)
CENTER = Alignment(vertical="center", horizontal="center", wrap_text=True)
VCENTER = Alignment(vertical="center", wrap_text=True)

COL_PROMPT, COL_OPTION, COL_CHECK, COL_MEMO = 1, 2, 3, 4
COL_WIDTHS = {COL_PROMPT: 24, COL_OPTION: 32, COL_CHECK: 9, COL_MEMO: 46}
LAST_COL = COL_MEMO

CHECK_VALUES = ("未", "済")
CHECK_DEFAULT = "未"

H_TITLE, H_SECTION = 28, 22  # 行高さは原則 Excel/Sheets の自動調整に任せ、この2つだけ明示

TITLE_TEXT = "面談ヒアリングシート（オンライン商談・入力用）"
INSTR_TEXT = (
    "各項目の「チェック」列で 未／済 を選んでください（あてはまるものを『済』に。複数選択可）。"
    "選んだ『済』は緑色になります。自由記述・メモ欄は直接入力できます。"
)


class Sheet:
    def __init__(self):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.active
        self.ws.title = "ヒアリング"
        self.r = 1
        self.dv = DataValidation(
            type="list", formula1=f'"{",".join(CHECK_VALUES)}"', allow_blank=True
        )
        self.dv.error = "未 か 済 を選んでください"
        self.ws.add_data_validation(self.dv)
        self._check_first = None
        for col, w in COL_WIDTHS.items():
            self.ws.column_dimensions[chr(ord("A") + col - 1)].width = w

    def _merge(self, r0, c0, r1, c1):
        self.ws.merge_cells(start_row=r0, start_column=c0, end_row=r1, end_column=c1)

    def _set(self, r, c, value, font, align, fill=None, border=None):
        cell = self.ws.cell(row=r, column=c, value=value)
        cell.font = font
        cell.alignment = align
        if fill is not None:
            cell.fill = fill
        if border is not None:
            cell.border = border
        return cell

    def title(self, text):
        self._set(self.r, COL_PROMPT, text, TITLE_FONT, TOPLEFT)
        self._merge(self.r, COL_PROMPT, self.r, LAST_COL)
        self.ws.row_dimensions[self.r].height = H_TITLE
        self.r += 1

    def instruction(self, text):
        self._set(self.r, COL_PROMPT, text, INSTR_FONT, TOPLEFT)
        self._merge(self.r, COL_PROMPT, self.r, LAST_COL)
        self.r += 1

    def section(self, text):
        for c in range(COL_PROMPT, LAST_COL + 1):
            self._set(self.r, c, text if c == COL_PROMPT else None, SECTION_FONT, VCENTER, SECTION_FILL)
        self._merge(self.r, COL_PROMPT, self.r, LAST_COL)
        self.ws.row_dimensions[self.r].height = H_SECTION
        self.r += 1

    def note(self, text):
        self._set(self.r, COL_PROMPT, text, NOTE_FONT, TOPLEFT)
        self._merge(self.r, COL_PROMPT, self.r, LAST_COL)
        self.r += 1

    def label_line(self, text):
        self._set(self.r, COL_PROMPT, text, SUBHEAD_FONT, TOPLEFT)
        self._merge(self.r, COL_PROMPT, self.r, LAST_COL)
        self.r += 1

    def _check_cell(self, r):
        cell = self._set(r, COL_CHECK, CHECK_DEFAULT, OPTION_FONT, CENTER, border=CELL_BORDER)
        self.dv.add(cell)
        if self._check_first is None:
            self._check_first = r
        return cell

    def _input_cell(self, r, c0, c1):
        self._set(r, c0, None, OPTION_FONT, TOPLEFT, INPUT_FILL, INPUT_BORDER)
        if c1 > c0:
            self._merge(r, c0, r, c1)

    def check_group(self, prompt, options):
        r0 = self.r
        for kind, label in options:
            self._set(self.r, COL_OPTION, label, OPTION_FONT, VCENTER, border=CELL_BORDER)
            if kind == "check":
                self._check_cell(self.r)
                self._set(self.r, COL_MEMO, None, OPTION_FONT, TOP, border=CELL_BORDER)
            elif kind == "write":
                if "その他" in label:  # その他＝選択もできる＋内容も書ける
                    self._check_cell(self.r)
                else:
                    self._set(self.r, COL_CHECK, None, OPTION_FONT, CENTER, border=CELL_BORDER)
                self._input_cell(self.r, COL_MEMO, COL_MEMO)
            self.r += 1
        if prompt:
            self._set(r0, COL_PROMPT, prompt, PROMPT_FONT, TOP, border=CELL_BORDER)
            if self.r - 1 > r0:
                self._merge(r0, COL_PROMPT, self.r - 1, COL_PROMPT)

    def freetext_row(self, items):
        # 候補日（① 日付 / 時間 / 〜）は1行にまとめ、1つの記入欄で受ける
        if items and items[0][0][:1] in "①②③":
            label = " ".join(lbl for lbl, _ in items if lbl)
            self._set(self.r, COL_PROMPT, label, PROMPT_FONT, TOP, border=CELL_BORDER)
            self._input_cell(self.r, COL_OPTION, COL_MEMO)
            self.r += 1
            return
        for label, has_input in items:
            if has_input:
                self._set(self.r, COL_PROMPT, label, PROMPT_FONT, TOP, border=CELL_BORDER)
                self._input_cell(self.r, COL_OPTION, COL_MEMO)
            else:
                self._set(self.r, COL_PROMPT, label, NOTE_FONT, TOPLEFT)
                self._merge(self.r, COL_PROMPT, self.r, LAST_COL)
            self.r += 1

    def build(self):
        self.title(TITLE_TEXT)
        self.instruction(INSTR_TEXT)
        for block in FORM:
            kind = block[0]
            if kind == "title":
                continue  # 既定タイトルは上で出力済み
            if kind == "note":
                self.note(block[1])
            elif kind == "section":
                self.section(block[1])
            elif kind == "space":
                continue
            elif kind == "row":
                segments = block[1]
                has_c = any(s[0] == "c" for s in segments)
                has_f = any(s[0] == "f" for s in segments)
                if has_c:
                    prompt, options = split_check_row(segments)
                    self.check_group(prompt, options)
                elif has_f:
                    self.freetext_row(freetext_items(segments))
                else:
                    self.label_line(segments[0][1])
        # 「済」セルを緑塗りにする条件付き書式
        if self._check_first is not None:
            rng = f"{chr(ord('A') + COL_CHECK - 1)}{self._check_first}:{chr(ord('A') + COL_CHECK - 1)}{self.r - 1}"
            self.ws.conditional_formatting.add(
                rng, CellIsRule(operator="equal", formula=['"済"'], fill=DONE_FILL)
            )
        self.ws.freeze_panes = "A3"  # タイトル＋操作説明を固定
        self.ws.sheet_view.showGridLines = False

    def save(self, path):
        self.wb.save(str(path))


def main():
    if len(sys.argv) < 2:
        print("input .md を指定してください（出力名の決定に使います）")
        sys.exit(1)
    md = pathlib.Path(sys.argv[1])
    out = md.with_name(md.stem + "_オンライン入力用.xlsx")
    sheet = Sheet()
    sheet.build()
    sheet.save(out)
    print(f"OK: {out.name}  ({out.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
