# -*- coding: utf-8 -*-
"""
京のお肉処弘 - 給与支給総額計算ファイル作成
賃金状況報告シート + 損益計算書データ → 給与支給総額計算Excel
"""
import sys, openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path
from copy import copy

sys.stdout.reconfigure(encoding='utf-8')

# === 定数 ===
BASE = Path('c:/Users/user/projects/カラフルボックス/補助金/京の食事処資料')
TEMPLATE_DIR = Path('c:/Users/user/projects/カラフルボックス/補助金')

# P/L データ（第2期: R6.4.1～R7.3.31）
PL = {
    '売上高': 1_476_107_192,
    '売上原価': 852_862_100,
    '売上総利益': 623_244_955,
    '販管費合計': 626_819_960,
    '営業利益': -3_575_005,
    '経常利益': -2_034_572,
    '当期純利益': -2_336_572,
    # 販管費内訳
    '給料手当': 102_890_664,
    '雑給': 79_487_112,
    '賞与': 11_501_000,
    '法定福利費': 21_340_513,
    '福利厚生費': 1_241_702,
    '減価償却費': 374_218,
    '旅費交通費': 6_533_919,
}

# 従業員分類（賃金状況報告シート + PDF照合結果）
# 正社員: No.1-19, パート・アルバイト: No.20-42
SEISHAIN_COUNT = 19
PART_COUNT = 23
YAKUIN_COUNT = 1  # 役員は従業員に含まず
YAKUIN_HOSHU_3M = 1_383_333  # 役員報酬3ヶ月合計（賃金状況報告シートより）
YAKUIN_HOSHU_ANNUAL = YAKUIN_HOSHU_3M * 4  # 年間概算

# 年間平均労働時間（正社員基準: 月176h × 12 = 2,112h 概算）
STANDARD_ANNUAL_HOURS = 2_080  # 標準年間労働時間


def find_file(directory, keyword, exclude_keywords=None):
    """ディレクトリからキーワードでファイル検索"""
    for p in directory.iterdir():
        if keyword in p.name and not p.name.startswith('~$'):
            if exclude_keywords and any(k in p.name for k in exclude_keywords):
                continue
            return p
    return None


def read_wage_report():
    """既存の賃金状況報告シートから従業員データ読取"""
    teishutsu = BASE / '提出資料'
    f = None
    for p in teishutsu.iterdir():
        if '賃金状況報告シート' in p.name and '再修正' in p.name and 'コピー' not in p.name:
            f = p
            break

    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb['【必須】賃金状況報告シート']

    employees = []
    for row in ws.iter_rows(min_row=19, max_row=60):
        no = row[1].value
        name = row[2].value
        if name is None:
            continue

        area = row[3].value
        min_wage = row[4].value
        # 3ヶ月分の基本給と時給
        m1_base = row[5].value or 0
        m1_hr = row[6].value or 0
        m2_base = row[8].value or 0
        m2_hr = row[9].value or 0
        m3_base = row[11].value or 0
        m3_hr = row[12].value or 0
        judge = row[14].value if len(row) > 14 else None

        # 雇用形態判定（No.1-19: 正社員, No.20以降: パート/アルバイト）
        emp_type = '正社員' if no <= SEISHAIN_COUNT else 'パート・アルバイト'

        # 月間労働時間推定（基本給 ÷ 時給）
        hours = []
        for base, hr in [(m1_base, m1_hr), (m2_base, m2_hr), (m3_base, m3_hr)]:
            if hr > 0 and base > 0:
                hours.append(base / hr)
            else:
                hours.append(0)

        avg_monthly_hours = sum(hours) / len([h for h in hours if h > 0]) if any(h > 0 for h in hours) else 0

        employees.append({
            'no': no,
            'name': name.strip(),
            'area': area,
            'type': emp_type,
            'min_wage': min_wage,
            'm1_base': m1_base,
            'm2_base': m2_base,
            'm3_base': m3_base,
            'm1_hr': m1_hr,
            'm2_hr': m2_hr,
            'm3_hr': m3_hr,
            'avg_monthly_hours': avg_monthly_hours,
            'judge': judge,
        })

    wb.close()
    return employees


def create_wage_calc_excel(employees):
    """給与支給総額計算Excelファイル作成"""
    wb = openpyxl.Workbook()

    # スタイル定義
    title_font = Font(name='游ゴシック', size=14, bold=True)
    header_font = Font(name='游ゴシック', size=10, bold=True)
    normal_font = Font(name='游ゴシック', size=10)
    small_font = Font(name='游ゴシック', size=9)
    number_fmt = '#,##0'
    pct_fmt = '0.00%'
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(name='游ゴシック', size=10, bold=True, color='FFFFFF')
    light_blue = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    light_yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    light_green = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')

    # ========================================
    # Sheet 1: 給与支給総額計算（メイン）
    # ========================================
    ws1 = wb.active
    ws1.title = '給与支給総額計算'
    ws1.sheet_properties.tabColor = '4472C4'

    # タイトル
    ws1['B2'] = '給与支給総額計算書'
    ws1['B2'].font = title_font
    ws1['B3'] = '株式会社 京のお肉処弘'
    ws1['B3'].font = Font(name='游ゴシック', size=12)
    ws1['B4'] = '事業年度: 令和6年4月1日～令和7年3月31日（第2期）'
    ws1['B4'].font = normal_font

    # --- セクション1: 損益計算書データ ---
    r = 6
    ws1.cell(r, 2, '【損益計算書データ（販管費）】').font = header_font

    headers_pl = [
        ('科目', '金額（円）', '備考'),
    ]
    items_pl = [
        ('給料手当', PL['給料手当'], '正社員給与（役員報酬含む可能性あり）'),
        ('雑給', PL['雑給'], 'パート・アルバイト給与'),
        ('賞与', PL['賞与'], ''),
        ('法定福利費', PL['法定福利費'], '※給与支給総額から除外'),
        ('福利厚生費', PL['福利厚生費'], '※給与支給総額から除外'),
    ]

    r += 1
    for i, h in enumerate(['科目', '金額（円）', '備考']):
        c = ws1.cell(r, 2 + i, h)
        c.font = header_font_white
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    for item_name, amount, note in items_pl:
        r += 1
        ws1.cell(r, 2, item_name).font = normal_font
        ws1.cell(r, 2).border = thin_border
        c = ws1.cell(r, 3, amount)
        c.font = normal_font
        c.number_format = number_fmt
        c.border = thin_border
        ws1.cell(r, 4, note).font = small_font
        ws1.cell(r, 4).border = thin_border
        if '除外' in note:
            for col in range(2, 5):
                ws1.cell(r, col).fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 小計行
    r += 1
    ws1.cell(r, 2, '給与関連合計（A）').font = Font(name='游ゴシック', size=10, bold=True)
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 2).fill = light_blue
    total_wage_pl = PL['給料手当'] + PL['雑給'] + PL['賞与']
    c = ws1.cell(r, 3, total_wage_pl)
    c.font = Font(name='游ゴシック', size=10, bold=True)
    c.number_format = number_fmt
    c.border = thin_border
    c.fill = light_blue
    ws1.cell(r, 4, '給料手当 + 雑給 + 賞与').font = small_font
    ws1.cell(r, 4).border = thin_border
    ws1.cell(r, 4).fill = light_blue
    row_total_a = r

    # --- セクション2: 役員報酬 ---
    r += 2
    ws1.cell(r, 2, '【役員報酬の控除】').font = header_font
    r += 1
    ws1.cell(r, 2, '役員報酬（3ヶ月合計）').font = normal_font
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 3, YAKUIN_HOSHU_3M).font = normal_font
    ws1.cell(r, 3).number_format = number_fmt
    ws1.cell(r, 3).border = thin_border
    ws1.cell(r, 4, '賃金状況報告シートより').font = small_font
    ws1.cell(r, 4).border = thin_border

    r += 1
    ws1.cell(r, 2, '役員報酬（年間概算）（B）').font = Font(name='游ゴシック', size=10, bold=True)
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 2).fill = light_yellow
    c = ws1.cell(r, 3, YAKUIN_HOSHU_ANNUAL)
    c.font = Font(name='游ゴシック', size=10, bold=True)
    c.number_format = number_fmt
    c.border = thin_border
    c.fill = light_yellow
    ws1.cell(r, 4, '3ヶ月合計 × 4').font = small_font
    ws1.cell(r, 4).border = thin_border
    ws1.cell(r, 4).fill = light_yellow
    row_yakuin = r

    # --- セクション3: 給与支給総額 ---
    r += 2
    ws1.cell(r, 2, '【給与支給総額の算定】').font = header_font
    r += 1
    ws1.cell(r, 2, '給与支給総額（役員報酬込）').font = normal_font
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 3, total_wage_pl).font = normal_font
    ws1.cell(r, 3).number_format = number_fmt
    ws1.cell(r, 3).border = thin_border
    ws1.cell(r, 4, '(A) ※テンプレートE13相当').font = small_font
    ws1.cell(r, 4).border = thin_border

    r += 1
    wage_excl_yakuin = total_wage_pl - YAKUIN_HOSHU_ANNUAL
    ws1.cell(r, 2, '給与支給総額（役員報酬除外）').font = normal_font
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 3, wage_excl_yakuin).font = normal_font
    ws1.cell(r, 3).number_format = number_fmt
    ws1.cell(r, 3).border = thin_border
    ws1.cell(r, 4, '(A) - (B) ※賃上げ計算用').font = small_font
    ws1.cell(r, 4).border = thin_border

    # --- セクション4: 従業員数・1人当たり ---
    r += 2
    ws1.cell(r, 2, '【従業員数と1人当たり給与支給総額】').font = header_font

    r += 1
    for i, h in enumerate(['項目', '人数/金額', '備考']):
        c = ws1.cell(r, 2 + i, h)
        c.font = header_font_white
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    emp_items = [
        ('正規雇用従業員', f'{SEISHAIN_COUNT}人', ''),
        ('契約社員', '0人', ''),
        ('パート・アルバイト', f'{PART_COUNT}人', ''),
        ('役員', f'{YAKUIN_COUNT}人', '※従業員数に含まず'),
    ]
    for item_name, val, note in emp_items:
        r += 1
        ws1.cell(r, 2, item_name).font = normal_font
        ws1.cell(r, 2).border = thin_border
        ws1.cell(r, 3, val).font = normal_font
        ws1.cell(r, 3).border = thin_border
        ws1.cell(r, 4, note).font = small_font
        ws1.cell(r, 4).border = thin_border

    total_emp = SEISHAIN_COUNT + PART_COUNT
    r += 1
    ws1.cell(r, 2, '従業員合計（C）').font = Font(name='游ゴシック', size=10, bold=True)
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 2).fill = light_blue
    ws1.cell(r, 3, f'{total_emp}人').font = Font(name='游ゴシック', size=10, bold=True)
    ws1.cell(r, 3).border = thin_border
    ws1.cell(r, 3).fill = light_blue
    ws1.cell(r, 4, '').border = thin_border
    ws1.cell(r, 4).fill = light_blue

    # パートFTE換算
    r += 2
    ws1.cell(r, 2, '【パート・アルバイトFTE換算（参考）】').font = header_font
    r += 1

    part_employees = [e for e in employees if e['type'] == 'パート・アルバイト']
    total_part_hours = 0
    for e in part_employees:
        total_part_hours += e['avg_monthly_hours'] * 12

    standard_monthly = STANDARD_ANNUAL_HOURS / 12  # ~173.3h
    fte_total = 0
    for e in part_employees:
        if e['avg_monthly_hours'] > 0:
            fte = e['avg_monthly_hours'] / standard_monthly
            fte_total += fte

    ws1.cell(r, 2, '標準年間労働時間').font = normal_font
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 3, f'{STANDARD_ANNUAL_HOURS}時間').font = normal_font
    ws1.cell(r, 3).border = thin_border
    ws1.cell(r, 4, '40h/週 × 52週').font = small_font
    ws1.cell(r, 4).border = thin_border

    r += 1
    ws1.cell(r, 2, 'パートFTE換算合計').font = normal_font
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 3, round(fte_total, 2)).font = normal_font
    ws1.cell(r, 3).number_format = '0.00'
    ws1.cell(r, 3).border = thin_border
    ws1.cell(r, 4, '各パートの月間時間÷標準月間時間').font = small_font
    ws1.cell(r, 4).border = thin_border

    fte_adjusted = SEISHAIN_COUNT + fte_total
    r += 1
    ws1.cell(r, 2, 'FTE換算後従業員数（D）').font = Font(name='游ゴシック', size=10, bold=True)
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 2).fill = light_green
    ws1.cell(r, 3, round(fte_adjusted, 2)).font = Font(name='游ゴシック', size=10, bold=True)
    ws1.cell(r, 3).number_format = '0.00'
    ws1.cell(r, 3).border = thin_border
    ws1.cell(r, 3).fill = light_green
    ws1.cell(r, 4, f'正社員{SEISHAIN_COUNT} + パートFTE{round(fte_total, 2)}').font = small_font
    ws1.cell(r, 4).border = thin_border
    ws1.cell(r, 4).fill = light_green

    # --- セクション5: 最終計算結果 ---
    r += 2
    ws1.cell(r, 2, '【1人当たり給与支給総額】').font = Font(name='游ゴシック', size=12, bold=True)

    r += 1
    for i, h in enumerate(['算出方法', '金額', '']):
        c = ws1.cell(r, 2 + i, h)
        c.font = header_font_white
        c.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
        c.border = thin_border
        c.alignment = Alignment(horizontal='center')

    # 方法1: 全従業員頭数割り
    per_capita_head = total_wage_pl / total_emp
    r += 1
    ws1.cell(r, 2, '(A)÷(C) 頭数割り').font = normal_font
    ws1.cell(r, 2).border = thin_border
    c = ws1.cell(r, 3, round(per_capita_head))
    c.font = Font(name='游ゴシック', size=11, bold=True)
    c.number_format = number_fmt
    c.border = thin_border
    ws1.cell(r, 4, f'{total_wage_pl:,} ÷ {total_emp}').font = small_font
    ws1.cell(r, 4).border = thin_border

    # 方法2: 役員除外・頭数割り
    per_capita_excl = wage_excl_yakuin / total_emp
    r += 1
    ws1.cell(r, 2, '(A-B)÷(C) 役員報酬除外・頭数').font = normal_font
    ws1.cell(r, 2).border = thin_border
    c = ws1.cell(r, 3, round(per_capita_excl))
    c.font = normal_font
    c.number_format = number_fmt
    c.border = thin_border
    ws1.cell(r, 4, f'{wage_excl_yakuin:,} ÷ {total_emp}').font = small_font
    ws1.cell(r, 4).border = thin_border

    # 方法3: FTE換算
    per_capita_fte = total_wage_pl / fte_adjusted
    r += 1
    ws1.cell(r, 2, '(A)÷(D) FTE換算').font = normal_font
    ws1.cell(r, 2).border = thin_border
    c = ws1.cell(r, 3, round(per_capita_fte))
    c.font = normal_font
    c.number_format = number_fmt
    c.border = thin_border
    ws1.cell(r, 4, f'{total_wage_pl:,} ÷ {round(fte_adjusted, 2)}').font = small_font
    ws1.cell(r, 4).border = thin_border

    # 方法4: 役員除外・FTE換算（推奨）
    per_capita_best = wage_excl_yakuin / fte_adjusted
    r += 1
    ws1.cell(r, 2, '(A-B)÷(D) 役員除外・FTE（推奨）').font = Font(name='游ゴシック', size=10, bold=True)
    ws1.cell(r, 2).border = thin_border
    ws1.cell(r, 2).fill = light_green
    c = ws1.cell(r, 3, round(per_capita_best))
    c.font = Font(name='游ゴシック', size=12, bold=True, color='C00000')
    c.number_format = number_fmt
    c.border = thin_border
    c.fill = light_green
    ws1.cell(r, 4, f'{wage_excl_yakuin:,} ÷ {round(fte_adjusted, 2)}').font = small_font
    ws1.cell(r, 4).border = thin_border
    ws1.cell(r, 4).fill = light_green

    # --- テンプレート転記用サマリ ---
    r += 2
    ws1.cell(r, 2, '【2026テンプレート転記用】').font = header_font
    r += 1
    template_items = [
        ('給料手当（販管費E5）', PL['給料手当']),
        ('雑給（販管費E6）', PL['雑給']),
        ('賞与手当（販管費E7）', PL['賞与']),
        ('売上高（B10）', PL['売上高']),
        ('粗利益（B11）', PL['売上総利益']),
        ('営業利益（B12）', PL['営業利益']),
        ('経常利益（B13）', PL['経常利益']),
        ('減価償却費（B14）', PL['減価償却費']),
    ]
    for item_name, val in template_items:
        r += 1
        ws1.cell(r, 2, item_name).font = normal_font
        ws1.cell(r, 2).border = thin_border
        c = ws1.cell(r, 3, val)
        c.font = normal_font
        c.number_format = number_fmt
        c.border = thin_border

    # 列幅調整
    ws1.column_dimensions['A'].width = 2
    ws1.column_dimensions['B'].width = 38
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 40

    # ========================================
    # Sheet 2: 従業員別明細
    # ========================================
    ws2 = wb.create_sheet('従業員別明細')
    ws2.sheet_properties.tabColor = '70AD47'

    ws2['B2'] = '従業員別給与明細（直近3ヶ月）'
    ws2['B2'].font = title_font
    ws2['B3'] = '出典: 賃金状況報告シート（再修正分）'
    ws2['B3'].font = small_font

    headers = ['No', '氏名', '雇用形態', '1月基本給', '2月基本給', '3月基本給',
               '3ヶ月平均', '時給', '月間平均時間', 'FTE', '最低賃金判定']
    r = 5
    for i, h in enumerate(headers):
        c = ws2.cell(r, 2 + i, h)
        c.font = header_font_white
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    for e in employees:
        r += 1
        avg3 = (e['m1_base'] + e['m2_base'] + e['m3_base']) / 3
        avg_hr = (e['m1_hr'] + e['m2_hr'] + e['m3_hr']) / 3
        fte = e['avg_monthly_hours'] / (STANDARD_ANNUAL_HOURS / 12) if e['type'] != '正社員' else 1.0

        values = [
            e['no'], e['name'], e['type'],
            e['m1_base'], e['m2_base'], e['m3_base'],
            round(avg3), round(avg_hr), round(e['avg_monthly_hours'], 1),
            round(fte, 2) if e['type'] != '正社員' else 1.00,
            e['judge']
        ]

        for i, v in enumerate(values):
            c = ws2.cell(r, 2 + i, v)
            c.font = normal_font
            c.border = thin_border
            if i in (3, 4, 5, 6):
                c.number_format = number_fmt
            elif i == 9:
                c.number_format = '0.00'
            if e['type'] == 'パート・アルバイト':
                c.fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # 合計行
    r += 1
    ws2.cell(r, 2, '').border = thin_border
    ws2.cell(r, 3, '合計/平均').font = Font(name='游ゴシック', size=10, bold=True)
    ws2.cell(r, 3).border = thin_border

    seishain_list = [e for e in employees if e['type'] == '正社員']
    part_list = [e for e in employees if e['type'] == 'パート・アルバイト']

    for col_idx, month_key in [(5, 'm1_base'), (6, 'm2_base'), (7, 'm3_base')]:
        total = sum(e[month_key] for e in employees)
        c = ws2.cell(r, col_idx, total)
        c.font = Font(name='游ゴシック', size=10, bold=True)
        c.number_format = number_fmt
        c.border = thin_border

    # FTE合計
    fte_sum = SEISHAIN_COUNT + sum(
        e['avg_monthly_hours'] / (STANDARD_ANNUAL_HOURS / 12)
        for e in part_list if e['avg_monthly_hours'] > 0
    )
    c = ws2.cell(r, 11, round(fte_sum, 2))
    c.font = Font(name='游ゴシック', size=10, bold=True)
    c.number_format = '0.00'
    c.border = thin_border

    # 列幅
    col_widths = [4, 5, 14, 12, 12, 12, 12, 12, 8, 13, 8, 12]
    for i, w in enumerate(col_widths):
        ws2.column_dimensions[get_column_letter(i + 1)].width = w

    # ========================================
    # Sheet 3: 賃上げ計画シミュレーション
    # ========================================
    ws3 = wb.create_sheet('賃上げ計画')
    ws3.sheet_properties.tabColor = 'ED7D31'

    ws3['B2'] = '賃上げ計画シミュレーション'
    ws3['B2'].font = title_font
    ws3['B3'] = '※テンプレート行47相当の計算'
    ws3['B3'].font = small_font

    # 賃上げ率シミュレーション
    r = 5
    headers3 = ['', '直近決算期\n(実績値)', '1年目計画', '2年目計画', '3年目計画']
    for i, h in enumerate(headers3):
        c = ws3.cell(r, 2 + i, h)
        c.font = header_font_white
        c.fill = header_fill
        c.border = thin_border
        c.alignment = Alignment(horizontal='center', wrap_text=True)

    # 給与支給総額の推移（年率1.5%増）
    growth_rate = 0.015  # 1.5%/年
    base_wage = total_wage_pl
    projections = [base_wage]
    for _ in range(3):
        projections.append(round(projections[-1] * (1 + growth_rate)))

    r += 1
    labels = ['給与支給総額']
    ws3.cell(r, 2, '給与支給総額').font = Font(name='游ゴシック', size=10, bold=True)
    ws3.cell(r, 2).border = thin_border
    for i, p in enumerate(projections):
        c = ws3.cell(r, 3 + i, p)
        c.font = normal_font
        c.number_format = number_fmt
        c.border = thin_border

    r += 1
    ws3.cell(r, 2, '増加率（対基準年）').font = normal_font
    ws3.cell(r, 2).border = thin_border
    ws3.cell(r, 3, '-').font = normal_font
    ws3.cell(r, 3).border = thin_border
    for i in range(1, 4):
        rate = (projections[i] - projections[0]) / projections[0]
        c = ws3.cell(r, 3 + i, rate)
        c.font = normal_font
        c.number_format = pct_fmt
        c.border = thin_border

    r += 1
    ws3.cell(r, 2, '年率').font = normal_font
    ws3.cell(r, 2).border = thin_border
    ws3.cell(r, 3, '-').border = thin_border
    for i in range(1, 4):
        c = ws3.cell(r, 3 + i, growth_rate)
        c.font = normal_font
        c.number_format = pct_fmt
        c.border = thin_border

    # 注記
    r += 2
    notes = [
        '【算定ルール（IT導入補助金2025/2026）】',
        '・給与支給総額 = 給料 + 賃金 + 賞与 + 各種手当（決算書の販管費より）',
        '・除外項目: 福利厚生費、法定福利費、退職金',
        '・役員報酬の取扱い: 申請枠・要件により異なる（要確認）',
        '・通年で給与を支給していない従業員は対象外',
        '・パート・アルバイトはFTE換算（実労働時間÷標準労働時間）',
        '',
        '【賃上げ要件（通常枠）】',
        '・事業計画期間において、給与支給総額を年率平均1.5%以上増加',
        '・事業計画終了時点で、給与支給総額が基準値以上',
    ]
    for note in notes:
        ws3.cell(r, 2, note).font = small_font if not note.startswith('【') else header_font
        r += 1

    # 列幅
    ws3.column_dimensions['A'].width = 2
    ws3.column_dimensions['B'].width = 28
    for col in ['C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 18

    # === 保存 ===
    out_dir = BASE
    out_path = out_dir / '京のお肉処弘_給与支給総額計算.xlsx'
    # pathlib日本語パス問題回避: 親ディレクトリから書き込み
    import tempfile, shutil
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name
    wb.save(tmp_path)
    wb.close()

    # 一時ファイルから目的地にコピー
    shutil.copy2(tmp_path, str(out_path))
    Path(tmp_path).unlink()

    print(f'出力: {out_path.name}')
    return out_path


def main():
    print('=== 京のお肉処弘 給与支給総額計算 ===\n')

    # 従業員データ読込
    print('1. 賃金状況報告シートから従業員データ読込...')
    employees = read_wage_report()
    print(f'   従業員数: {len(employees)}人')
    seishain = [e for e in employees if e['type'] == '正社員']
    part = [e for e in employees if e['type'] == 'パート・アルバイト']
    print(f'   正社員: {len(seishain)}人, パート: {len(part)}人')

    # P/Lデータ表示
    print(f'\n2. 損益計算書データ（第2期）')
    print(f'   給料手当: {PL["給料手当"]:>15,}円')
    print(f'   雑給:     {PL["雑給"]:>15,}円')
    print(f'   賞与:     {PL["賞与"]:>15,}円')
    total = PL['給料手当'] + PL['雑給'] + PL['賞与']
    print(f'   合計:     {total:>15,}円')

    # FTE計算
    standard_monthly = STANDARD_ANNUAL_HOURS / 12
    fte_total = sum(
        e['avg_monthly_hours'] / standard_monthly
        for e in part if e['avg_monthly_hours'] > 0
    )
    fte_adjusted = SEISHAIN_COUNT + fte_total
    print(f'\n3. FTE換算')
    print(f'   正社員: {SEISHAIN_COUNT}人')
    print(f'   パートFTE: {fte_total:.2f}人')
    print(f'   合計FTE: {fte_adjusted:.2f}人')

    # 1人当たり計算
    per_capita = total / fte_adjusted
    per_capita_excl = (total - YAKUIN_HOSHU_ANNUAL) / fte_adjusted
    print(f'\n4. 1人当たり給与支給総額')
    print(f'   役員報酬込: {per_capita:>12,.0f}円')
    print(f'   役員報酬除: {per_capita_excl:>12,.0f}円')

    # Excel出力
    print(f'\n5. Excelファイル作成...')
    out = create_wage_calc_excel(employees)
    print(f'\n完了: {out}')


if __name__ == '__main__':
    main()
