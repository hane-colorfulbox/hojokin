# -*- coding: utf-8 -*-
"""
賃金台帳Excel読み取り + 加点措置判定

対応方針:
  ヘッダー別名辞書で正規化マッチ → 月の並び方を値ベースで自動判定する
  柔軟パーサー(_read_flexible) を採用。複数の実フォーマット差異を1本で吸収する。
  配置パターン:
    (a) 1月〜12月が列見出し   … 集計表型（1行1人）
    (b) 対象年月/給与年月列あり … 月別行型（1人×N行、月は明示列）
    (c) 先頭列に YYYYMM 値     … YYYYMM月次型（給与ソフト出力・1ファイル1人のケース）
  個人台帳型（行=項目、列=月、月度給与ブロック）は別ルートで温存。

加点措置の判定ロジック:
  ①用（インボイス枠/セキュリティ枠）:
    R6年10月～R7年9月の間で、地域別最低賃金以上かつR7年度改定後未満で
    雇用していた従業員が全従業員の30%以上いる月が3か月以上あるか
  ②用（共通）:
    交付申請の直近月における事業場内最低賃金が、
    R7年7月の事業場内最低賃金+63円以上の水準か
"""
from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
import pandas as pd

from .config import MIN_WAGE_MAP

logger = logging.getLogger(__name__)

# R6年度の最低賃金（加点措置①の下限判定に使用）
MIN_WAGE_R6 = {
    '北海道': 1010, '青森県': 953, '岩手県': 952, '宮城県': 973,
    '秋田県': 951, '山形県': 955, '福島県': 955, '東京都': 1163,
    '茨城県': 1005, '栃木県': 1004, '群馬県': 985, '埼玉県': 1078,
    '千葉県': 1076, '神奈川県': 1162, '新潟県': 985, '富山県': 998,
    '石川県': 984, '福井県': 984, '山梨県': 988, '長野県': 998,
    '岐阜県': 1001, '静岡県': 1034, '愛知県': 1077, '京都府': 1058,
    '大阪府': 1114, '三重県': 1023, '滋賀県': 1017, '兵庫県': 1052,
    '奈良県': 986, '和歌山県': 980, '鳥取県': 957, '島根県': 962,
    '岡山県': 982, '広島県': 1020, '山口県': 979, '徳島県': 980,
    '香川県': 970, '愛媛県': 956, '高知県': 952, '福岡県': 992,
    '佐賀県': 956, '長崎県': 953, '大分県': 954, '熊本県': 952,
    '宮崎県': 952, '鹿児島県': 953, '沖縄県': 952,
}

BONUS_THRESHOLD_YEN = 63

MONTH_NAMES = ['1月', '2月', '3月', '4月', '5月', '6月',
               '7月', '8月', '9月', '10月', '11月', '12月']

# ── 加点判定の対象暦月（デジタル化・AI導入補助金2026 公募要領で固定）──
# 加点措置①（補助率1/2→2/3 トリガー兼用／加点項目14）の対象期間。
# 令和6年10月〜令和7年9月の暦月12か月。Oct-Dec は令和6年・Jan-Sep は令和7年と
# 年をまたぐため、暦月12スロット配列では表現できない。(西暦年, 月) タプルで固定保持する。
BONUS1_WINDOW: list[tuple[int, int]] = [
    (2024, 10), (2024, 11), (2024, 12),
    (2025, 1), (2025, 2), (2025, 3), (2025, 4), (2025, 5), (2025, 6),
    (2025, 7), (2025, 8), (2025, 9),
]
# 加点措置②（加点項目15）の基準月＝令和7年7月。比較対象は交付申請の直近月。
BONUS2_BASE_YM: tuple[int, int] = (2025, 7)


def prev_month(ym: tuple[int, int]) -> tuple[int, int]:
    """(年, 月) の前月を返す（公式②シートの N14=EDATE(申請月, -1) と同義）。"""
    year, month = ym
    return (year - 1, 12) if month == 1 else (year, month - 1)


def ym_label(ym: tuple[int, int]) -> str:
    """(年, 月) → 和暦ラベル '令和7年7月'。"""
    return wareki_label(ym[0], ym[1])


# ── 加点判定用賃金台帳テンプレート（ツール/加点判定用賃金台帳テンプレート.xlsx）の固定レイアウト ──
# リーダー(read_bonus_wage_ledger)とライター(bonus_wage_ledger_writer)で共有する単一の真実。
BWL_SHEET_NAME = '加点判定用明細'
BWL_HEADER_ROW = 6
BWL_DATA_START_ROW = 7
BWL_COL_NO = 2          # B
BWL_COL_NAME = 3        # C
BWL_COL_EMPTYPE = 4     # D
BWL_COL_HOURS = 5       # E 月間所定労働時間
BWL_COL_WINDOW_START = 6  # F〜Q: BONUS1_WINDOW（令和6年10月〜令和7年9月）の基本給12列
BWL_COL_LATEST = 18       # R: 交付申請直近月の基本給
BWL_PREF_CELL = (2, 3)    # C2: 事業場所在地（都道府県）
BWL_APPYM_CELL = (3, 3)   # C3: 交付申請月（yyyy/mm）


# ============================================================
# 事業年度ウィンドウ / 年月ヘッダー ユーティリティ
# ============================================================
# 賃金台帳の月列は「直近事業年度の12ヶ月」を決算月起点で時系列に並べ、列ヘッダーを
# 和暦年月でスタンプする（暦年 1〜12月 との取り違え＝R216 期間ズレを構造的に防ぐ）。
# 内部表現 monthly_wages[12] は **暦月Index固定**（Index 0=1月）のまま維持する。
# 事業年度の12ヶ月は各暦月に重複なく1対1対応するため、暦月Indexで可逆に保持できる。
# 表示順（writer のヘッダー並び）と賞与の年度判定だけが事業年度ウィンドウを使う。

_REIWA_EPOCH = 2018   # 令和N年 = 2018 + N（令和元年 = 2019）
_HEISEI_EPOCH = 1988  # 平成N年 = 1988 + N（平成元年 = 1989）


def resolve_fiscal_window(fiscal_period_hint: str | None) -> list[tuple[int, int]] | None:
    """fiscal_period_hint（例 '2024-08〜2025-07'）から事業年度12ヶ月の
    (西暦年, 月) を決算期首→期末の時系列順で返す。

    形式不明・None の場合は None を返す（呼び出し側で暦年フォールバック）。
    """
    if not fiscal_period_hint:
        return None
    m = re.search(r'(\d{4})-(\d{1,2})', fiscal_period_hint)
    if not m:
        return None
    start_year, start_month = int(m.group(1)), int(m.group(2))
    if not 1 <= start_month <= 12:
        return None
    window: list[tuple[int, int]] = []
    for i in range(12):
        total = (start_month - 1) + i
        window.append((start_year + total // 12, total % 12 + 1))
    return window


def wareki_label(year: int, month: int) -> str:
    """西暦 (year, month) → 和暦ラベル '令和6年8月'。令和未満は西暦表記にフォールバック。"""
    if year >= 2019:
        return f'令和{year - _REIWA_EPOCH}年{month}月'
    if year >= 1989:
        return f'平成{year - _HEISEI_EPOCH}年{month}月'
    return f'{year}年{month}月'


def parse_ym_header(text) -> tuple[int | None, int] | None:
    """列ヘッダー文字列から (西暦年 or None, 月) を抽出する。

    対応: '令和6年8月' / 'R6.8' / 'R6/8' / '2024年8月' / '2024/08' /
          '2024-08' / '202408' / 素の '8月'。
    年が判別できない素の '8月' は (None, 8)。月が取れなければ None。
    """
    if text is None:
        return None
    s = unicodedata.normalize('NFKC', str(text))
    s = re.sub(r'[\s　]+', '', s)
    if not s:
        return None
    su = s.upper()
    m = re.search(r'令和(\d+)年(\d{1,2})月', s)
    if m:
        return _REIWA_EPOCH + int(m.group(1)), int(m.group(2))
    m = re.search(r'平成(\d+)年(\d{1,2})月', s)
    if m:
        return _HEISEI_EPOCH + int(m.group(1)), int(m.group(2))
    m = re.match(r'R(\d+)[./-](\d{1,2})月?$', su)
    if m:
        return _REIWA_EPOCH + int(m.group(1)), int(m.group(2))
    m = re.match(r'H(\d+)[./-](\d{1,2})月?$', su)
    if m:
        return _HEISEI_EPOCH + int(m.group(1)), int(m.group(2))
    m = re.search(r'(\d{4})[年/\-.](\d{1,2})月?', s)
    if m and 1 <= int(m.group(2)) <= 12:
        return int(m.group(1)), int(m.group(2))
    m = re.fullmatch(r'(\d{4})(\d{2})', s)
    if m and 1 <= int(m.group(2)) <= 12:
        return int(m.group(1)), int(m.group(2))
    # 先頭一致（fullmatch ではなく match）で末尾の注記を許容する。
    # 例: スキルがヘッダーを '1月\n(R7.1)' のように注記付きで出しても、空白除去後
    # '1月(R7.1)' の先頭 '1月' を月として拾える（fullmatch だと注記で None になり、
    # ツールが月列を検出できず台帳を0名と誤読する事故が起きた）。和暦/西暦/RN.M は
    # 上で先に評価済みなので、ここに来るのは素の 'N月[...]' 系のみ。
    m = re.match(r'(\d{1,2})月', s)
    if m and 1 <= int(m.group(1)) <= 12:
        return None, int(m.group(1))
    return None


# ============================================================
# データ構造
# ============================================================

@dataclass
class WageEmployee:
    """賃金台帳から読み取った従業員"""
    no: int
    name: str
    employment_type: str  # 正社員 / パート・アルバイト
    monthly_avg_hours: float
    hourly_rate: float  # 代表的な時給（フォーマット1用、他は月別から算出）
    monthly_wages: list[float | None]  # 12か月分の支給合計
    monthly_hourly_rates: list[float | None] = field(
        default_factory=lambda: [None] * 12
    )
    monthly_hours: list[float | None] = field(
        default_factory=lambda: [None] * 12
    )
    # データソース（抽出根拠の元ファイル名）。複数ファイル統合時は '統合(Nファイル)'
    source_file: str = ''
    # 年間通勤手当（非課税分のみ、円）。AI が抽出できた場合のみ > 0。
    # R216 算定対象から除外するため、ツール側で在籍月数で均等割して各月から減算する。
    # 既存呼出は 0.0（未設定）で動作、賃金台帳の作成タスクで S列に書き出す用途に使う。
    annual_transport_allowance: float = 0.0
    # 年間賞与（課税賞与の合計、円）。R216（給与支給総額）に算入するが、月次セル
    # （monthly_wages）には混ぜない。賞与を月次に混ぜると最低賃金判定（加点①）や
    # 直近3ヶ月の月額が歪むため、専用フィールド＝テンプレ T列「年間賞与」で隔離する。
    # 賞与の支給月は R216 では不問（年間合計に算入されれば月は問わない）。年度の
    # 帰属判定（対象事業年度内か）だけが効くため、ウィンドウ外賞与は算入しない。
    annual_bonus: float = 0.0
    # 事業年度ウィンドウ（[(西暦年, 月)] 12件、決算期首→期末順）。AI抽出時に決算月から
    # 決定論で確定する。writer が和暦年月ヘッダーを事業年度順に並べるために使う。
    # None の場合（決算月不明・暦年運用）は暦年 1〜12月でフォールバックする。
    fiscal_window: list[tuple[int, int]] | None = None
    # 事業年度ウィンドウ選択時の注意書き（年の自動補正・低カバレッジ等）。空なら問題なし。
    # per-ledger の情報を全従業員に同値で載せ、賃金台帳作成タスクが変換メモ／ステータスに
    # surfacing する用途に使う（R216 期間ズレのサイレント過少計上を可視化するため）。
    fiscal_window_note: str = ''
    # 支給日（支給年月）が読めず、事業年度ウィンドウで絞り込めないまま年間賞与に算入した
    # 賞与額の合計（円）。年間集計表など支給日のない台帳で > 0 になる。非暦年決算では
    # 暦年集計とのズレで R216 がズレ得るため、賃金台帳作成タスクが警告 surfacing に使う。
    bonus_undated_total: float = 0.0
    # 対象事業年度ウィンドウ外として除外した「支給年月あり」賞与額の合計（円）。
    # AI が誤った年月を付けた賞与が無言で R216 から落ちるのを変換メモで可視化するための情報。
    # 多年度台帳での正常な除外でも > 0 になり得る（アラートではなく情報表示用）。
    bonus_dropped_total: float = 0.0

    @property
    def is_full_year(self) -> bool:
        return all(w is not None for w in self.monthly_wages)

    def months_with_data(self) -> list[int]:
        return [i for i, w in enumerate(self.monthly_wages) if w is not None]

    def get_hourly_for_month(self, month_idx: int) -> float | None:
        """指定月の時給を取得（月別データ優先、なければ代表時給）"""
        if self.monthly_hourly_rates[month_idx] is not None:
            return self.monthly_hourly_rates[month_idx]
        return self.hourly_rate if self.hourly_rate > 0 else None


@dataclass
class BonusWageEmployee:
    """加点判定用賃金台帳から読み取った従業員（(年,月)キーの基本給ベース）。

    R215/R216 用の WageEmployee（暦月12スロット）とは別物。加点判定は暦月固定で
    令和6年10月〜令和7年9月＋申請直近月をまたいで見るため、(西暦年,月)→基本給 の
    辞書で保持する。時間換算給与 = 基本給 ÷ 所定労働時間（残業・通勤・各種手当を除く）。
    """
    no: int
    name: str
    employment_type: str = ''
    # 月間所定労働時間（最賃比較の分母。月給÷1か月平均所定労働時間 ＝厚労省の換算式）。
    scheduled_hours: float | None = None
    # (西暦年, 月) → 基本給（所定内賃金、円）
    monthly_base: dict[tuple[int, int], float] = field(default_factory=dict)
    # (西暦年, 月) → 所定労働時間の月次上書き（任意。無ければ scheduled_hours を使う）
    monthly_hours_override: dict[tuple[int, int], float] = field(default_factory=dict)
    source_file: str = ''

    @property
    def is_officer(self) -> bool:
        return '役員' in (self.employment_type or '')

    def hours_for(self, ym: tuple[int, int]) -> float | None:
        h = self.monthly_hours_override.get(ym)
        if h is not None and h > 0:
            return h
        return self.scheduled_hours if (self.scheduled_hours or 0) > 0 else None

    def base_for(self, ym: tuple[int, int]) -> float | None:
        return self.monthly_base.get(ym)

    def hourly_for(self, ym: tuple[int, int]) -> float | None:
        """その月の時間換算給与 = 基本給 ÷ 所定労働時間。算出不能なら None。"""
        base = self.monthly_base.get(ym)
        hours = self.hours_for(ym)
        if base is None or base <= 0 or hours is None or hours <= 0:
            return None
        return base / hours


@dataclass
class BonusWageLedger:
    """加点判定用賃金台帳の全体（会社単位の都道府県・交付申請月 + 従業員）。"""
    prefecture: str = ''
    application_ym: tuple[int, int] | None = None  # 交付申請月
    employees: list[BonusWageEmployee] = field(default_factory=list)
    # 読み取り時の注意書き（対象月欠落・所在地未設定など）。呼出側が surfacing する。
    notes: list[str] = field(default_factory=list)

    @property
    def latest_ym(self) -> tuple[int, int] | None:
        """加点措置②の比較対象＝交付申請の直近月（申請月の前月）。"""
        return prev_month(self.application_ym) if self.application_ym else None


@dataclass
class BonusPointResult:
    """加点措置の判定結果（(年,月)ベース）。"""
    bonus1_eligible: bool = False
    # 30%以上を満たした対象月（(年,月) のリスト）
    bonus1_months_met: list[tuple[int, int]] = field(default_factory=list)
    # BONUS1_WINDOW の各月の詳細。要素: {'ym','label','total','under_r7','ratio',
    #   'meets_30pct','has_data','employees':[{'name','base','hourly','is_target'}]}
    bonus1_details: list[dict] = field(default_factory=list)

    bonus2_eligible: bool = False
    bonus2_min_wage_july: float = 0.0
    bonus2_min_wage_latest: float = 0.0
    bonus2_diff: float = 0.0
    # 加点措置②の各期間詳細。{'ym','label','min_wage','employees':[{'name','base','hourly'}]}
    bonus2_july_detail: dict = field(default_factory=dict)
    bonus2_latest_detail: dict = field(default_factory=dict)

    prefecture: str = ''
    min_wage_r6: int = 0
    min_wage_r7: int = 0
    application_ym: tuple[int, int] | None = None
    latest_ym: tuple[int, int] | None = None  # 加点②の比較対象月（申請月の前月）
    # 判定上の注意（対象月欠落・所在地未設定・最賃マスタ欠落など）
    notes: list[str] = field(default_factory=list)


# ============================================================
# 柔軟パーサー（集計表型 / 月別行型 / YYYYMM月次型を統一処理）
# ============================================================

# ヘッダー別名辞書（正規化後に部分一致 or 完全一致で判定）
_HEADER_ALIASES = {
    'name':       ['氏名', '従業員氏名', '社員氏名', '名前'],
    'emp_id':     ['従業員番号', '従業員コード', '社員番号', 'no', 'ＮＯ', 'Ｎｏ'],
    'emp_type':   ['雇用形態', '区分', '従業員区分'],
    'base_wage':  ['基本給'],
    'hourly_wage':['基本給(時給)', '時給', '時間給'],
    'hours':      ['所定労働時間', '労働時間', '月間平均時間', '平均時間'],
    # 非課税額（通勤費の非課税分など）。total(gross) 採用時に差し引いて課税額へ補正。
    # 公募要領 p.10「給与所得として課税対象」+ 国税庁 No.2585（通勤手当は限度額まで非課税）。
    # ★宣言順を total_taxable / total より前に置く理由★: _match_alias は部分一致を許すため
    #   「非課税支給合計」が「課税支給合計」(total_taxable) や「支給合計」(total) の別名に
    #   部分一致してしまう。nontax を先に走査して『非課税◯◯』を確実に nontax へ確定させ、
    #   課税列・支給合計列への誤割当（R216 過小計上）を防ぐ。
    'nontax':     ['非課税額', '非課税合計', '非課税支給合計', '非課税分'],
    # 課税支給合計（R216 が求める「給与所得として課税対象」の合計）。最優先で採用。
    # 給与ソフトにより「課税分合計」「課税給与合計」等の表記揺れがあるため広めに拾う。
    # 注: 「課税対象額」は社保控除後の所得税課税ベースを指すソフトもあり ambiguous なので含めない。
    'total_taxable': ['課税支給合計', '課税分合計', '課税給与合計', '課税支給額'],
    # 支給合計（非課税通勤費等を含む可能性あり）。total_taxable が無い時のフォールバック。
    # この値を採用する場合は nontax 列があれば差し引いて課税額に補正する。
    # 注: 「差引支給合計」は社保・税控除後の手取りであり給与支給総額ではない（R216 過小計上に
    #     なる）ため、ここには含めない。
    'total':      ['支給合計額', '支給合計', '総支給額', '総支給'],
    'paid_date':  ['支給日', '支払日'],
    'month_col':  ['対象年月', '給与年月', '支給年月', '年月'],
    # 年間通勤手当（非課税分）。集計表型のみ有効。在籍月数で均等割して
    # monthly_wages から減算する（R216 公募要領 p.10 の課税給与定義に揃える）
    'transport_annual': ['年間通勤手当', '年間通勤費',
                         '通勤手当(年間)', '通勤費(年間)',
                         '非課税通勤手当(年間)'],
    # 年間賞与（課税賞与の年間合計）。集計表型のみ有効。月次セルには混ぜず R216 に加算する
    # （賞与を月次に入れると最低賃金判定や直近3ヶ月が歪むため専用列で隔離）。
    'bonus_annual': ['年間賞与', '賞与(年間)', '年間賞与額', '賞与年間合計', '年間賞与合計'],
}


def _norm(val) -> str:
    """文字列を正規化（NFKC・空白除去・小文字化）"""
    if val is None:
        return ''
    s = unicodedata.normalize('NFKC', str(val))
    s = s.replace('\u3000', '').replace(' ', '').strip()
    return s.lower()


def _match_alias(val: str, aliases: list[str]) -> bool:
    v = _norm(val)
    if not v:
        return False
    for a in aliases:
        na = _norm(a)
        if v == na or (len(na) >= 2 and na in v):
            return True
    return False


def _detect_field_map(ws, header_row: int) -> dict[str, int]:
    """指定行をヘッダーと見なし、各フィールドの列番号を割り出す"""
    fmap: dict[str, int] = {}
    month_cols: dict[int, int] = {}
    for c in range(1, min(ws.max_column + 1, 80)):
        val = ws.cell(header_row, c).value
        s = _norm(val)
        if not s:
            continue
        # 月列 → 集計表型。暦月「1月〜12月」に加え、和暦/西暦の年月スタンプ
        # （'令和6年8月' / '2024/08' / 'R6.8' 等）も暦月Index に割り付ける。
        # 年月は parse_ym_header が一手に解釈する（暦月Index 0=1月へ写像）。
        ym = parse_ym_header(val)
        if ym is not None:
            idx = ym[1] - 1
            if 0 <= idx <= 11:
                month_cols[idx] = c
                continue
        for key, aliases in _HEADER_ALIASES.items():
            if key in fmap:
                continue
            if _match_alias(val, aliases):
                # 部分一致の落とし穴対策（R216 母数の取り違え防止）:
                #   - '差引支給合計' は「支給合計」に部分一致するが手取り(控除後)なので総額系に割り当てない
                #   - '非課税支給合計' は「課税支給合計」「支給合計」に部分一致するが非課税分なので同様
                # nontax を先に走査しているので非課税列は通常そちらで確定するが、多重防御として弾く。
                nval = _norm(val)
                if key in ('total', 'total_taxable') and ('差引' in nval or '非課税' in nval):
                    continue
                fmap[key] = c
                break
    if month_cols:
        fmap['_month_cols'] = month_cols  # type: ignore[assignment]
    return fmap


def _find_header_rows(ws) -> list[tuple[int, dict]]:
    """シート内のヘッダー行を全て発見（給与/賞与セクション両方を取るため）"""
    rows = []
    for r in range(1, min(ws.max_row + 1, 40)):
        fmap = _detect_field_map(ws, r)
        has_name = 'name' in fmap
        # 課税支給合計（total_taxable）も合計列として認める。これが無いと、
        # 課税支給合計のみ（支給合計列なし・月列なし）の行型台帳でヘッダーが
        # 検出されず当該従業員の給与が丸ごと欠落し R216=0 になる。
        has_total = ('total' in fmap) or ('total_taxable' in fmap)
        has_month_cols = '_month_cols' in fmap
        if has_name and (has_total or has_month_cols):
            rows.append((r, fmap))
    return rows


def _parse_month(val, paid_date_val=None) -> int | None:
    """セル値から月インデックス(0-11)を抽出。YYYYMM数値/'〇年〇月'/支給日まで対応。

    給与ソフトの特殊コード対応:
      YYYYMM の月部分が 21 → 7月（夏季賞与） Index 6
      YYYYMM の月部分が 22 → 12月（冬季賞与） Index 11
      21,22 は一部医療機関向け給与ソフト出力で観測された賞与識別子。

    重要: 21/22 マッピングは給与ソフト依存のため、paid_date_val（支給日）が
    与えられている場合はそちらを優先する。これにより「21=春賞与/22=夏賞与」のように
    異なる慣習を持つソフトでも実支給月で正しく月マッピングできる。
    """
    def _normal_month_to_idx(m: int) -> int | None:
        """通常の月（1〜12）のみ受け付ける。21/22 は除外"""
        if 1 <= m <= 12:
            return m - 1
        return None

    def _bonus_code_to_idx(m: int) -> int | None:
        """賞与識別コード（21/22）のフォールバック。paid_date が無いときのみ使う"""
        if m == 21:
            return 6   # 夏季賞与 → 7月に加算（給与ソフト慣習。実 paid_date があれば paid_date 優先）
        if m == 22:
            return 11  # 冬季賞与 → 12月に加算
        return None

    # ── ステップ1: 通常の月（1〜12）の判定 ──
    bonus_code: int | None = None  # 21/22 を後回しにするため記録
    if val is not None:
        # YYYYMM 数値
        if isinstance(val, (int, float)):
            n = int(val)
            if 100000 <= n <= 999999:
                month = n % 100
                idx = _normal_month_to_idx(month)
                if idx is not None:
                    return idx
                if month in (21, 22):
                    bonus_code = month
        s = str(val)
        # '2025年3月' 等
        m = re.search(r'(\d{4})[年/\-](\d{1,2})', s)
        if m:
            idx = _normal_month_to_idx(int(m.group(2)))
            if idx is not None:
                return idx
        # '3月' 単独
        m = re.search(r'(\d{1,2})月', s)
        if m:
            idx = _normal_month_to_idx(int(m.group(1)))
            if idx is not None:
                return idx
        # 純粋なYYYYMM文字列
        m = re.fullmatch(r'\d{6}', s.strip())
        if m:
            month = int(s.strip()) % 100
            idx = _normal_month_to_idx(month)
            if idx is not None:
                return idx
            if month in (21, 22) and bonus_code is None:
                bonus_code = month
    # ── ステップ2: 支給日（paid_date）から月を取る（21/22 ボーナスコードより優先） ──
    # paid_date があれば、それが「実際に支給された月」なので最も信頼できる
    if paid_date_val is not None:
        s = str(paid_date_val)
        m = re.search(r'\d{4}[/\-年](\d{1,2})', s)
        if m:
            month = int(m.group(1))
            if 1 <= month <= 12:
                return month - 1

    # ── ステップ3: 21/22 賞与コードのフォールバック ──
    # paid_date が無い場合のみ、給与ソフト慣習の固定マッピングを使う
    if bonus_code is not None:
        idx = _bonus_code_to_idx(bonus_code)
        if idx is not None:
            return idx
    return None


def _to_float(val) -> float | None:
    if val is None:
        return None
    try:
        f = float(val)
        return f
    except (ValueError, TypeError):
        return None


def _new_emp_record(name: str, emp_type: str = '') -> dict:
    return {
        'name': name,
        'employment_type': emp_type,
        'monthly_wages': [None] * 12,
        'monthly_hourly_rates': [None] * 12,
        'monthly_hours': [None] * 12,
        'hourly_rate_flat': 0.0,
        'avg_hours_flat': 0.0,
        'annual_bonus': 0.0,
    }


def _parse_section_rowwise(ws, header_row: int, end_row: int,
                           fmap: dict, emp_data: dict) -> None:
    """月別行型 or YYYYMM月次型 のデータ行を処理（月=行方向）"""
    col_name = fmap['name']
    col_total_taxable = fmap.get('total_taxable')
    col_total = fmap.get('total')
    col_nontax = fmap.get('nontax')
    col_type = fmap.get('emp_type')
    col_month = fmap.get('month_col')
    col_base = fmap.get('base_wage')
    col_hours = fmap.get('hours')
    col_paid = fmap.get('paid_date')

    for r in range(header_row + 1, end_row):
        name_val = ws.cell(r, col_name).value
        if not name_val:
            continue
        name = str(name_val).replace('\u3000', ' ').strip()
        if not name:
            continue

        # 月を特定: 各 _parse_month 呼出に paid_date_val を渡す。
        # _parse_month は内部で「通常月(1-12) > paid_date > 21/22 賞与コード」の優先順序で
        # 月を決定するため、paid_date を毎回渡しておけば 21/22 のような ambiguous な
        # ボーナスコードに対しても、実支給月（paid_date）が優先される。
        paid_val = ws.cell(r, col_paid).value if col_paid else None
        month_idx = None
        if col_month:
            month_idx = _parse_month(ws.cell(r, col_month).value, paid_val)
        if month_idx is None:
            # 先頭列がYYYYMM(例: 202503)
            month_idx = _parse_month(ws.cell(r, 1).value, paid_val)
        if month_idx is None:
            continue

        if name not in emp_data:
            et = ''
            if col_type:
                et = str(ws.cell(r, col_type).value or '')
            emp_data[name] = _new_emp_record(name, et)

        rec = emp_data[name]

        # 月の給与額（課税支給合計）を決定。優先順（公募要領 p.10「課税対象となる経費」準拠）:
        #   1) 課税支給合計/課税分合計 列をそのまま採用（最も確実）
        #   2) 支給合計 − 非課税額（非課税通勤費等を差し引いて課税額へ補正）
        #   3) 支給合計 のみ（非課税分を含む可能性。雇用形態欠落警告で人手確認を促す）
        t = None
        if col_total_taxable:
            t = _to_float(ws.cell(r, col_total_taxable).value)
        elif col_total:
            gross = _to_float(ws.cell(r, col_total).value)
            if gross is not None:
                nontax = _to_float(ws.cell(r, col_nontax).value) if col_nontax else None
                t = gross - (nontax or 0)
                # 非課税額 > 支給合計（入力ミス・列誤マッピング・部分月行など）で
                # 負値になった場合は、負の給与を R216 に混入させず gross にフォールバック。
                if t < 0:
                    logger.warning(
                        f'非課税額が支給合計を超過（{name}）: 支給合計{gross:,.0f} '
                        f'< 非課税{(nontax or 0):,.0f} → 支給合計を採用し要確認'
                    )
                    t = gross
        if t is not None:
            # 給与＋賞与セクション両方が来たら加算（同月の別セクション）
            existing = rec['monthly_wages'][month_idx]
            rec['monthly_wages'][month_idx] = (existing or 0) + t

        if col_base and col_hours:
            base = _to_float(ws.cell(r, col_base).value)
            hours = _to_float(ws.cell(r, col_hours).value)
            if base is not None and hours is not None and hours > 0:
                rec['monthly_hours'][month_idx] = hours
                rec['monthly_hourly_rates'][month_idx] = base / hours


def _parse_section_summary(ws, header_row: int, fmap: dict,
                           emp_data: dict) -> None:
    """集計表型（列=月）のデータ行を処理"""
    col_name = fmap['name']
    col_type = fmap.get('emp_type')
    col_hours = fmap.get('hours')
    col_hourly = fmap.get('hourly_wage')
    col_transport = fmap.get('transport_annual')
    col_bonus = fmap.get('bonus_annual')
    month_cols: dict[int, int] = fmap['_month_cols']  # type: ignore[assignment]

    for r in range(header_row + 1, ws.max_row + 1):
        name_val = ws.cell(r, col_name).value
        if not name_val:
            continue
        name = str(name_val).replace('\u3000', ' ').strip()
        if not name:
            continue

        if name not in emp_data:
            et = ''
            if col_type:
                et = str(ws.cell(r, col_type).value or '')
            emp_data[name] = _new_emp_record(name, et)
        rec = emp_data[name]

        if col_hours:
            h = _to_float(ws.cell(r, col_hours).value)
            if h is not None:
                rec['avg_hours_flat'] = h
        if col_hourly:
            hr = _to_float(ws.cell(r, col_hourly).value)
            if hr is not None:
                rec['hourly_rate_flat'] = hr

        for midx, c in month_cols.items():
            v = _to_float(ws.cell(r, c).value)
            if v is not None:
                existing = rec['monthly_wages'][midx]
                rec['monthly_wages'][midx] = (existing or 0) + v
                if rec['hourly_rate_flat'] > 0:
                    rec['monthly_hourly_rates'][midx] = rec['hourly_rate_flat']

        # \u5e74\u9593\u901a\u52e4\u624b\u5f53\uff08\u975e\u8ab2\u7a0e\u5206\uff09\u3092\u5728\u7c4d\u6708\u6570\u3067\u5747\u7b49\u5272\u3057\u3066 monthly_wages \u304b\u3089\u6e1b\u7b97\u3002
        # R216 \u7b97\u5165\u984d\u3092\u300c\u7d66\u4e0e\u6240\u5f97\u3068\u3057\u3066\u8ab2\u7a0e\u5bfe\u8c61\u300d\u3060\u3051\u306b\u63c3\u3048\u308b\u305f\u3081\u306e\u88dc\u6b63\u3002
        # \u901a\u52e4\u624b\u5f53\u306f\u56fd\u7a0e\u5e81 No.2585 \u3067\u670815\u4e07\u5186\u307e\u3067\u975e\u8ab2\u7a0e\u306e\u305f\u3081\u7d66\u4e0e\u6240\u5f97\u306b\u542b\u307e\u308c\u306a\u3044\u3002
        # \uff08\u516c\u52df\u8981\u9818 \u7b2c6\u56de p.10 / docs/\u88dc\u52a9\u91d1_\u5b9f\u52d9\u77e5\u8b58\u30d9\u30fc\u30b9.md \u53c2\u7167\uff09
        if col_transport:
            ta = _to_float(ws.cell(r, col_transport).value)
            if ta is not None and ta > 0:
                non_null = [
                    m for m in range(12)
                    if rec['monthly_wages'][m] is not None
                ]
                if non_null:
                    per_month = ta / len(non_null)
                    for m in non_null:
                        rec['monthly_wages'][m] = max(
                            0.0, rec['monthly_wages'][m] - per_month
                        )
                    logger.info(
                        f'\u901a\u52e4\u624b\u5f53\u6e1b\u7b97: {name} \u5e74\u9593{ta:,.0f}\u5186 \u00f7 '
                        f'\u5728\u7c4d{len(non_null)}\u30f6\u6708 = \u6708{per_month:,.0f}\u5186\u6e1b'
                    )

        # \u5e74\u9593\u8cde\u4e0e\uff08T\u5217\uff09: \u6708\u6b21\u30bb\u30eb\u306b\u306f\u6df7\u305c\u305a\u5c02\u7528\u30d5\u30a3\u30fc\u30eb\u30c9\u306b\u4fdd\u6301 \u2192 R216 \u306b\u52a0\u7b97\uff08\u516c\u52df\u8981\u9818 p.10\uff09\u3002
        # \u8cde\u4e0e\u3092\u6708\u6b21\u306b\u6df7\u305c\u308b\u3068\u6700\u4f4e\u8cc3\u91d1\u5224\u5b9a\u30fb\u76f4\u8fd13\u30f6\u6708\u304c\u6b6a\u3080\u305f\u3081\u9694\u96e2\u3059\u308b\u3002
        if col_bonus:
            ba = _to_float(ws.cell(r, col_bonus).value)
            if ba is not None and ba > 0:
                rec['annual_bonus'] = (rec.get('annual_bonus') or 0.0) + ba


def _read_csv(path: Path, emp_data: dict | None = None) -> None:
    """CSV ファイルの簡易フォールバック読込み。

    通常は AI 経路（read_wage_ledgers_with_ai → _csv_to_tsv）で処理されるため
    この関数が呼ばれるのは USE_AI_WAGE_EXTRACTION=false または AI 抽出失敗時のみ。
    対応範囲：氏名列＋年間給与列がある単純な集計表型 CSV のみ。
    医療機関の月別行型（YYYYMM 1行=1月）には対応しない（AI 経路で処理）。
    """
    if emp_data is None:
        emp_data = {}

    # gzip CSV (magic 0x1F 0x8B) なら自動解凍
    raw_bytes: bytes | None = None
    try:
        with open(path, 'rb') as f:
            head = f.read(2)
        if head == b'\x1f\x8b':
            import gzip
            with gzip.open(path, 'rb') as gz:
                raw_bytes = gz.read()
    except Exception:
        head = b''

    last_error: Exception | None = None
    df = None
    for encoding in ('utf-8-sig', 'utf-8', 'cp932', 'shift_jis'):
        try:
            if raw_bytes is not None:
                import io as _io
                df = pd.read_csv(_io.BytesIO(raw_bytes), header=0,
                                 na_values=['', 'N/A', 'null'], encoding=encoding)
            else:
                df = pd.read_csv(path, header=0, na_values=['', 'N/A', 'null'],
                                 encoding=encoding)
            break
        except (UnicodeDecodeError, UnicodeError) as e:
            last_error = e
            continue
        except Exception as e:
            # encoding 以外の理由（パースエラー等）は早期返却
            level = _csv_decode_warning_level(path)
            msg = f'CSV 読込失敗: {path.name} ({e})'
            if level == 'info':
                logger.info(msg + ' → 同名 xlsx/xlsm を優先使用します')
            else:
                logger.warning(msg)
            return

    if df is None:
        # 全エンコーディング失敗 → 同名 xlsx/xlsm があれば info 降格
        level = _csv_decode_warning_level(path)
        msg = (
            f'CSV読込失敗（対応外形式の可能性、先頭バイト=0x{head.hex() if head else ""}）: '
            f'{path.name} ({last_error})'
        )
        if level == 'info':
            logger.info(msg + ' → 同名 xlsx/xlsm を優先使用します')
        else:
            logger.warning(msg)
        return

    # 列の正規化（日本語テキストの NFD→NFC 変換）
    df.columns = [unicodedata.normalize('NFC', str(col)) for col in df.columns]

    # 氏名列を探す（複数パターン対応）
    name_cols = [c for c in df.columns if '氏名' in c or '名前' in c or 'name' in c.lower()]
    if not name_cols:
        logger.warning(f'CSV に氏名列が見つかりません: {path.name}')
        return
    name_col = name_cols[0]

    # 給与列を探す（複数パターン対応）
    wage_col_patterns = ['支給', '給与', '合計', '給', 'wage', 'total']
    wage_cols = [c for c in df.columns if any(p in c for p in wage_col_patterns)]
    if not wage_cols:
        logger.warning(f'CSV に給与列が見つかりません: {path.name}')
        return

    # 各行を処理（従業員追加）
    for idx, row in df.iterrows():
        name_val = row[name_col]
        if pd.isna(name_val) or not str(name_val).strip():
            continue

        name = str(name_val).replace('　', ' ').strip()
        if not name or name in emp_data:
            continue

        # 初期化（_new_emp_record と同じ構造）
        emp_type = ''
        if '雇用形態' in df.columns:
            et = row['雇用形態']
            if not pd.isna(et):
                emp_type = str(et)
        emp_data[name] = {
            'name': name,
            'employment_type': emp_type,
            'monthly_wages': [None] * 12,
            'monthly_hours': [None] * 12,
            'hourly_rate_flat': 0.0,
            'avg_hours_flat': 0.0,
            'monthly_hourly_rates': [None] * 12,
        }

        # 給与合計を取得（最初の給与列を使用）
        if wage_cols:
            total_wage = row[wage_cols[0]]
            if not pd.isna(total_wage):
                try:
                    annual_total = float(total_wage)
                    # 年間総額を12で除算して月別平均を算出
                    monthly_avg = annual_total / 12
                    emp_data[name]['monthly_wages'] = [monthly_avg] * 12
                except (ValueError, TypeError):
                    pass

        logger.debug(f'CSV から従業員を追加: {name}')


def _visible_worksheets(wb: openpyxl.Workbook) -> list:
    """人間に見える（表示中の）シートだけ返す。非表示シートは読まない。

    人手修正で旧版シートが「×」付きのまま非表示で残ることがあり、全シート走査だと
    名寄せで二重計上になる（hidden/veryHidden を除外し、表示中シートのみを正本扱い）。
    """
    return [ws for ws in wb.worksheets if ws.sheet_state == 'visible']


def _read_flexible(wb: openpyxl.Workbook,
                   emp_data: dict | None = None) -> dict:
    """柔軟パーサー本体（emp_dataに蓄積）"""
    if emp_data is None:
        emp_data = {}

    for ws in _visible_worksheets(wb):
        header_rows = _find_header_rows(ws)
        if not header_rows:
            continue

        for i, (hr, fmap) in enumerate(header_rows):
            end = (header_rows[i + 1][0]
                   if i + 1 < len(header_rows) else ws.max_row + 1)
            if '_month_cols' in fmap:
                _parse_section_summary(ws, hr, fmap, emp_data)
            else:
                _parse_section_rowwise(ws, hr, end, fmap, emp_data)

    return emp_data


# ============================================================
# フォーマット3: 個人台帳型（給与ソフト出力）
# ============================================================

def _parse_hours_str(val) -> float:
    """時間を数値に変換: 248, '248:00', '168:30' → float"""
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    m = re.match(r'(\d+):(\d+)', s)
    if m:
        return int(m.group(1)) + int(m.group(2)) / 60
    try:
        return float(s)
    except ValueError:
        return 0.0


def _extract_name_from_cell(text: str) -> str:
    """'007\\n嘉口澪\\xa0(女)' → '嘉口澪'"""
    # 改行で分割して最後の部分（名前部分）を取得
    parts = str(text).split('\n')
    name_part = parts[-1] if len(parts) > 1 else parts[0]
    # 先頭の番号を除去
    name_part = re.sub(r'^\d+\s*', '', name_part)
    # 性別マーカーを除去: (女) (男) （女） （男）
    name_part = re.sub(r'\s*[\(（][男女][\)）]\s*$', '', name_part)
    # 不要な空白を整理
    name_part = name_part.replace('\xa0', ' ').replace('\u3000', ' ').strip()
    return name_part


def _parse_month_from_header(text: str) -> int | None:
    """'令和 7年\\n1月度給与' → 0 (1月=index0)"""
    m = re.search(r'(\d+)月度給与', str(text))
    if m:
        month = int(m.group(1))
        if 1 <= month <= 12:
            return month - 1
    return None


def _read_individual_ledger(wb: openpyxl.Workbook) -> list[WageEmployee]:
    """フォーマット3: 行=項目、列=月、1人1ブロック"""
    employees = []

    for ws in _visible_worksheets(wb):
        # シート内のブロックを探す（「賃金台帳」を含むセルが区切り）
        blocks = _find_individual_blocks(ws)

        for block_start, block_end in blocks:
            emp = _parse_individual_block(ws, block_start, block_end)
            if emp:
                emp.no = len(employees) + 1
                employees.append(emp)

    return employees


def _find_individual_blocks(ws) -> list[tuple[int, int]]:
    """個人台帳のブロック開始・終了行を特定"""
    blocks = []
    block_start = None

    for r in range(1, ws.max_row + 1):
        val = str(ws.cell(r, 1).value or '')
        # 「賃金台帳」または「頁」を含む行がブロック開始
        if '賃金台帳' in val or '頁' in val:
            if block_start is not None:
                blocks.append((block_start, r - 1))
            block_start = r

    # 最後のブロック
    if block_start is not None:
        blocks.append((block_start, ws.max_row))

    # ブロックが見つからなかった場合、シート全体を1ブロックとする
    if not blocks:
        blocks = [(1, ws.max_row)]

    return blocks


def _parse_individual_block(ws, start_row: int, end_row: int) -> WageEmployee | None:
    """個人台帳の1ブロックを解析"""
    # 名前を探す（開始行付近のA列）
    name = ''
    for r in range(start_row, min(start_row + 5, end_row + 1)):
        val = str(ws.cell(r, 1).value or '')
        # 番号+改行+名前のパターン、または名前を含む行
        if '\n' in val and re.search(r'\d+\n', val):
            name = _extract_name_from_cell(val)
            break

    if not name:
        return None

    # 月列のマッピングを構築（ヘッダー行から）
    month_cols: dict[int, int] = {}  # month_index → column
    for r in range(start_row, min(start_row + 5, end_row + 1)):
        for c in range(2, ws.max_column + 1):
            val = str(ws.cell(r, c).value or '')
            m_idx = _parse_month_from_header(val)
            if m_idx is not None:
                month_cols[m_idx] = c

    if not month_cols:
        return None

    # 行ラベルのマッピングを構築
    row_labels: dict[str, int] = {}
    for r in range(start_row, end_row + 1):
        label = str(ws.cell(r, 1).value or '').strip()
        if label:
            row_labels[label] = r

    # 基本給の行を特定
    base_wage_row = row_labels.get('基本給')
    # 所定労働時間の行を特定
    hours_row = row_labels.get('所定労働時間')
    # 支給合計の行を特定（候補順）。R216 は課税支給合計を最優先（公募要領 p.10）。
    total_row = (
        row_labels.get('課税支給合計')
        or row_labels.get('課税分合計')
        or row_labels.get('課税給与合計')
        or row_labels.get('支給合計')
        or row_labels.get('差引支給合計')
    )
    # 基本給(時給)があれば時給ベースの判別に使える
    hourly_base_row = row_labels.get('基本給(時給)')

    # 月別データを抽出
    monthly_wages = [None] * 12
    monthly_hourly = [None] * 12
    monthly_hours_list: list[float | None] = [None] * 12

    for m_idx, col in month_cols.items():
        # 支給合計
        if total_row:
            val = ws.cell(total_row, col).value
            if val is not None:
                try:
                    monthly_wages[m_idx] = float(val)
                except (ValueError, TypeError):
                    pass

        # 月別の労働時間
        if hours_row:
            hours_val = ws.cell(hours_row, col).value
            if hours_val is not None:
                h = _parse_hours_str(hours_val)
                if h > 0:
                    monthly_hours_list[m_idx] = h

        # 時給計算
        if base_wage_row and hours_row:
            base = ws.cell(base_wage_row, col).value
            hours_val = ws.cell(hours_row, col).value
            if base is not None and hours_val is not None:
                try:
                    base_f = float(base)
                    hours_f = _parse_hours_str(hours_val)
                    if hours_f > 0:
                        monthly_hourly[m_idx] = base_f / hours_f
                except (ValueError, TypeError):
                    pass

    # 代表時給を算出
    valid_hourly = [h for h in monthly_hourly if h is not None]
    avg_hourly = sum(valid_hourly) / len(valid_hourly) if valid_hourly else 0

    # 平均労働時間（月別データがあればそこから算出）
    valid_hours = [h for h in monthly_hours_list if h is not None and h > 0]
    avg_hours = sum(valid_hours) / len(valid_hours) if valid_hours else 0

    # 雇用形態の推定（基本給(時給)行にデータがあればパート系）
    emp_type = ''
    if hourly_base_row:
        hourly_vals = [
            ws.cell(hourly_base_row, col).value
            for col in month_cols.values()
        ]
        has_hourly = any(v and float(v) > 0 for v in hourly_vals
                        if v is not None)
        if has_hourly:
            emp_type = 'パート・アルバイト'

    return WageEmployee(
        no=0,
        name=name,
        employment_type=emp_type,
        monthly_avg_hours=round(avg_hours, 1),
        hourly_rate=round(avg_hourly, 1),
        monthly_wages=monthly_wages,
        monthly_hourly_rates=monthly_hourly,
        monthly_hours=monthly_hours_list,
    )


# ============================================================
# メイン読み取り関数
# ============================================================

def _is_individual_ledger(wb: openpyxl.Workbook) -> bool:
    """個人台帳型（月度給与ブロック）かどうか判定"""
    for ws in _visible_worksheets(wb):
        for r in range(1, min(ws.max_row + 1, 30)):
            for c in range(1, min(ws.max_column + 1, 30)):
                val = str(ws.cell(r, c).value or '')
                if '月度給与' in val:
                    return True
    return False


def _emp_dict_to_list(emp_data: dict) -> list[WageEmployee]:
    """内部dict表現 → WageEmployeeリスト変換"""
    employees = []
    for i, (name, data) in enumerate(emp_data.items()):
        hourly_rates = [h for h in data['monthly_hourly_rates'] if h is not None]
        hours_list = [h for h in data['monthly_hours'] if h is not None]
        if hourly_rates:
            avg_hourly = sum(hourly_rates) / len(hourly_rates)
        else:
            avg_hourly = data.get('hourly_rate_flat', 0.0)
        if hours_list:
            avg_hours = sum(hours_list) / len(hours_list)
        else:
            avg_hours = data.get('avg_hours_flat', 0.0)

        employees.append(WageEmployee(
            no=i + 1,
            name=data['name'],
            employment_type=data['employment_type'],
            monthly_avg_hours=round(avg_hours, 1),
            hourly_rate=round(avg_hourly, 1),
            monthly_wages=data['monthly_wages'],
            monthly_hourly_rates=data['monthly_hourly_rates'],
            monthly_hours=data['monthly_hours'],
            annual_bonus=float(data.get('annual_bonus') or 0.0),
        ))
    return employees


def read_wage_ledger(file_path: Path) -> list[WageEmployee]:
    """
    単一の賃金台帳Excelを読み取る。
    個人台帳型（月度給与ブロック）は専用パーサー、それ以外は柔軟パーサーで統一処理。
    """
    wb = openpyxl.load_workbook(str(file_path), data_only=True)

    if _is_individual_ledger(wb):
        employees = _read_individual_ledger(wb)
        fmt = 'individual'
    else:
        emp_data = _read_flexible(wb)
        employees = _emp_dict_to_list(emp_data)
        fmt = 'flexible'

    wb.close()
    logger.info(f'賃金台帳読み取り完了: {file_path.name} → {len(employees)}名 ({fmt})')
    return employees


def _workbook_to_tsv(wb: openpyxl.Workbook, file_label: str) -> str:
    """ワークブック全シートをTSV文字列に変換（AI入力用）。"""
    parts: list[str] = [f'### ファイル: {file_label} ###']
    for ws in _visible_worksheets(wb):
        parts.append(f'\n--- シート: {ws.title} ---')
        for row in ws.iter_rows(values_only=True):
            # 末尾の None だけのセルは無視して圧縮
            cells = list(row)
            while cells and cells[-1] is None:
                cells.pop()
            if not cells:
                continue
            line = '\t'.join('' if v is None else str(v) for v in cells)
            parts.append(line)
    return '\n'.join(parts)


def _pdf_to_tsv(path: Path) -> str:
    """PDFからテキストを抽出してTSV風文字列に変換（AI入力用）。
    テキスト層が薄い場合は RuntimeError を送出 → 呼出し側でバイナリ送信にフォールバックする。
    """
    import fitz  # PyMuPDF
    MIN_CHARS_PER_PAGE = 50   # これ未満なら画像PDF扱いとしてフォールバック
    MAX_TOTAL_CHARS = 500_000  # 超大量テキストへのガード（約125Kトークン相当）

    doc = fitz.open(str(path))
    page_count = len(doc)  # close前に保持
    parts: list[str] = [f'### ファイル: {path.name} ###']
    total_chars = 0
    try:
        for page_num in range(page_count):
            text = doc[page_num].get_text('text')
            if text.strip():
                parts.append(f'\n--- ページ {page_num + 1} ---')
                parts.append(text)
                total_chars += len(text)
                if total_chars > MAX_TOTAL_CHARS:
                    logger.warning(
                        f'PDF テキスト量上限超過({total_chars:,}文字)でページ{page_num+1}以降を切り捨て: {path.name}'
                    )
                    break
    finally:
        doc.close()

    avg_chars = total_chars / max(page_count, 1)
    if avg_chars < MIN_CHARS_PER_PAGE:
        raise RuntimeError(
            f'PDF テキスト層が薄すぎます ({avg_chars:.0f}文字/ページ): {path.name}'
        )
    return '\n'.join(parts)


def _csv_decode_warning_level(path: Path) -> str:
    """CSV decode 失敗時に warning と info のどちらで出力するかを判定する。

    同じディレクトリに同名の .xlsx / .xlsm が存在し正常そうなら、
    そちら経由で情報を拾えるので info 降格。それ以外は warning 維持。
    """
    stem = path.stem
    for ext in ('.xlsx', '.xlsm'):
        sibling = path.parent / f'{stem}{ext}'
        if sibling.exists() and sibling.stat().st_size > 0:
            return 'info'
    return 'warning'


def _csv_to_tsv(path: Path) -> str:
    """CSVファイルをTSV文字列に変換（AI入力用）。
    ヘッダー位置・列構成は AI に解釈させるため、全行を文字列のまま渡す。
    複数エンコーディング（UTF-8 / CP932）を順次試す。
    gzip圧縮 CSV (magic 0x1F 0x8B) の場合は自動解凍してから decode。
    """
    # gzip magic を見て解凍を試みる
    try:
        with open(path, 'rb') as f:
            head = f.read(2)
    except Exception:
        head = b''

    raw_bytes: bytes | None = None
    if head == b'\x1f\x8b':
        # gzip CSV — 解凍してから渡す
        try:
            import gzip
            with gzip.open(path, 'rb') as gz:
                raw_bytes = gz.read()
            logger.info(f'CSV(gzip): {path.name} を解凍しました ({len(raw_bytes):,}バイト)')
        except Exception as e:
            logger.warning(f'CSV(gzip) 解凍失敗: {path.name} ({e})')

    last_error: Exception | None = None
    for encoding in ('utf-8-sig', 'utf-8', 'cp932', 'shift_jis'):
        try:
            if raw_bytes is not None:
                import io as _io
                df = pd.read_csv(
                    _io.BytesIO(raw_bytes),
                    header=None,
                    dtype=str,
                    keep_default_na=False,
                    encoding=encoding,
                )
            else:
                df = pd.read_csv(
                    path,
                    header=None,
                    dtype=str,
                    keep_default_na=False,
                    encoding=encoding,
                )
            break
        except (UnicodeDecodeError, UnicodeError) as e:
            last_error = e
            continue
    else:
        # decode 全敗 — 同名 xlsx があれば info 降格、なければ warning
        level = _csv_decode_warning_level(path)
        msg = (
            f'CSV読込失敗（対応外形式の可能性、先頭バイト=0x{head.hex() if head else ""}）: '
            f'{path.name}'
        )
        if level == 'info':
            msg += ' → 同名 xlsx/xlsm を優先使用します'
            logger.info(msg)
        else:
            logger.warning(msg)
        raise last_error or RuntimeError(f'CSV decode failed: {path.name}')

    parts: list[str] = [f'### ファイル: {path.name} ###']
    for row in df.itertuples(index=False, name=None):
        cells = [str(c) if c is not None else '' for c in row]
        while cells and cells[-1] == '':
            cells.pop()
        if not cells:
            continue
        parts.append('\t'.join(cells))
    return '\n'.join(parts)


# ============================================================
# AI抽出データ → 事業年度ウィンドウ適用（決定論）
# ============================================================
# AI（賃金台帳の作成タスク）は各従業員の月次を「年月付きの生データ」で全件返す。
# 暦年と事業年度の取り違えを AI 判断に委ねず、決算月から確定した事業年度12ヶ月を
# **ここで決定論的に選択** する。賞与も支給年月で年度フィルタしてから合算する。

def _ai_entry_ym(entry) -> tuple[int | None, int] | None:
    if not isinstance(entry, dict):
        return None
    return parse_ym_header(entry.get('ym'))


def _build_windowed_arrays(
    monthly_entries: list, fiscal_window: list[tuple[int, int]],
) -> tuple[list, list, list]:
    """年月付き月次エントリ群を、事業年度12ヶ月（暦月Index 0=1月）に割り付ける。

    同一 (年,月) が重複した場合は課税額の大きい方を採用（部分入力・欠損救済）。
    年が判別できないエントリは、その月が窓内で一意のときだけ best-effort 採用。
    """
    by_ym: dict[tuple[int, int], tuple] = {}
    by_month_noyear: dict[int, list[tuple]] = {}
    for e in monthly_entries or []:
        ym = _ai_entry_ym(e)
        if ym is None:
            continue
        y, mo = ym
        rec = (_to_float(e.get('taxable')), _to_float(e.get('hours')),
               _to_float(e.get('work_days')))
        if y is None:
            by_month_noyear.setdefault(mo, []).append(rec)
        else:
            prev = by_ym.get((y, mo))
            if prev is None or (rec[0] or -1) > (prev[0] or -1):
                by_ym[(y, mo)] = rec
    wages: list = [None] * 12
    hours: list = [None] * 12
    days: list = [None] * 12
    for (y, mo) in fiscal_window:
        rec = by_ym.get((y, mo))
        if rec is None:
            cand = by_month_noyear.get(mo)
            if cand and len(cand) == 1:
                rec = cand[0]
        if rec is None:
            continue
        idx = mo - 1
        wages[idx], hours[idx], days[idx] = rec
    return wages, hours, days


def _build_calendar_arrays_recent(monthly_entries: list) -> tuple[list, list, list]:
    """決算月不明時のフォールバック: 暦月Indexに「最新年の値」を割り付ける（旧 直近12ヶ月相当）。"""
    by_month: dict[int, tuple[int, tuple]] = {}
    for e in monthly_entries or []:
        ym = _ai_entry_ym(e)
        if ym is None:
            continue
        y, mo = ym
        yk = y if y is not None else -1
        rec = (_to_float(e.get('taxable')), _to_float(e.get('hours')),
               _to_float(e.get('work_days')))
        prev = by_month.get(mo)
        if prev is None or yk >= prev[0]:
            by_month[mo] = (yk, rec)
    wages: list = [None] * 12
    hours: list = [None] * 12
    days: list = [None] * 12
    for mo, (_yk, rec) in by_month.items():
        idx = mo - 1
        wages[idx], hours[idx], days[idx] = rec
    return wages, hours, days


def _collect_present_yms(ai_data: list) -> set[tuple[int, int]]:
    """全従業員の monthly エントリから (西暦年, 月) 集合を返す（年不明は除外）。"""
    present: set[tuple[int, int]] = set()
    for emp in ai_data or []:
        if not isinstance(emp, dict):
            continue
        for e in emp.get('monthly') or []:
            ym = _ai_entry_ym(e)
            if ym is None:
                continue
            y, mo = ym
            if y is not None:
                present.add((y, mo))
    return present


def _window_ending(end_year: int, end_month: int) -> list[tuple[int, int]]:
    """期末 (end_year, end_month) から事業年度12ヶ月を期首→期末順で返す。"""
    end_total = end_year * 12 + (end_month - 1)
    start_total = end_total - 11
    return [((start_total + i) // 12, (start_total + i) % 12 + 1) for i in range(12)]


def select_fiscal_window_from_data(
    ai_data: list, fiscal_period_hint: str | None,
) -> tuple[list[tuple[int, int]] | None, dict]:
    """決算月（hint由来＝信頼）と台帳の実在年から、カバレッジ最大の事業年度12ヶ月を選ぶ。

    fiscal_period_hint の **期末月** は決算月（ユーザー指定）として信頼するが、**期末年**は
    台帳に実在する年（monthly[].ym の YYYY）から決め直す。これにより「決算月＋今日基準」で
    組んだ窓が台帳の実データと1期ズレても（補助金申請が集中する期末〜申告期限の谷間で頻発）、
    台帳側に追従して正しい12ヶ月を選べる。年が判別できない台帳ではヒント窓のまま返す。

    タイブレーク: カバレッジ降順 → 最新年（直近事業年度）優先。
    進行中の未完了FYは月が揃わずカバレッジが低いため自然に選ばれない。

    Returns:
        (fiscal_window or None, info)
        info = {'end_month','hint_year','chosen_year','coverage',
                'hint_coverage','has_year_data','shifted'}
    """
    base = resolve_fiscal_window(fiscal_period_hint)
    if base is None:
        return None, {}
    hint_year, end_month = base[-1]  # 期末 = 決算月（信頼）/ 年は今日基準の推定
    present = _collect_present_yms(ai_data)
    info = {
        'end_month': end_month, 'hint_year': hint_year, 'chosen_year': hint_year,
        'coverage': 0, 'hint_coverage': 0, 'has_year_data': bool(present),
        'shifted': False,
    }
    if not present:
        # 台帳に西暦年が無い（ym=null 等）→ ヒント窓のまま（by_month_noyear で月割当）
        return base, info
    years = {y for (y, _m) in present}
    # FY末年候補: 実在年とその +1（非12月決算は2暦年に跨るため）+ ヒント年
    candidates = sorted(years | {y + 1 for y in years} | {hint_year})
    best_key: tuple[int, int] | None = None
    best: tuple[int, list[tuple[int, int]], int] | None = None
    for cy in candidates:
        win = _window_ending(cy, end_month)
        cov = sum(1 for ym in win if ym in present)
        if cy == hint_year:
            info['hint_coverage'] = cov
        key = (cov, cy)  # カバレッジ最大 → 同点なら最新年
        if best_key is None or key > best_key:
            best_key, best = key, (cy, win, cov)
    chosen_year, chosen_win, cov = best  # type: ignore[misc]
    info['chosen_year'] = chosen_year
    info['coverage'] = cov
    info['shifted'] = (chosen_year != hint_year)
    return chosen_win, info


def _format_window_note(info: dict) -> str:
    """select_fiscal_window_from_data の info を人間向け注意書きに整形（空＝問題なし）。"""
    if not info:
        return ''
    end_m = info.get('end_month')
    notes: list[str] = []
    if info.get('shifted'):
        cy, hy = info['chosen_year'], info['hint_year']
        win = _window_ending(cy, end_m)
        period = f'{wareki_label(*win[0])}〜{wareki_label(*win[-1])}'
        notes.append(
            f'⚠ 対象年を自動補正: 決算月{end_m}月＋今日基準の推定では{hy}年期末でしたが、'
            f'台帳の実データは{cy}年期末の事業年度（{period}）でした。'
            f'台帳に合わせて{cy}年期末を採用しました。'
            f'意図と異なる場合は決算月の指定と対象台帳を確認してください。'
        )
    cov = info.get('coverage', 0)
    if info.get('has_year_data') and cov < 12:
        win = _window_ending(info['chosen_year'], end_m)
        period = f'{wareki_label(*win[0])}〜{wareki_label(*win[-1])}'
        notes.append(
            f'⚠ 対象事業年度（{period}）の12ヶ月のうち、台帳にデータが存在するのは{cov}ヶ月のみ。'
            f'期間ズレ・月欠損の可能性があるため、PDF原本で対象期間を確認してください。'
        )
    return ' / '.join(notes)


def _sum_window_bonuses(
    bonus_entries: list, fiscal_window: list[tuple[int, int]] | None,
    *, undated_paid_months: list[tuple[int | None, int]] | None = None,
) -> tuple[float, list[float], list[float]]:
    """賞与を事業年度ウィンドウでフィルタして合算。

    (合計, 年度判定不能で算入した額, 対象年度外として除外した支給年月あり賞与額) を返す。
    第3要素 dropped は「支給年月が読めたが対象事業年度ウィンドウ外だった」賞与額の一覧。
    合計(total)の計算には一切影響しない（窓外は元から非算入）。AI の年月誤読による
    R216 サイレント過少を変換メモで可視化するための情報として返す（多年度台帳の正常な
    除外でも入る＝アラートではなく情報）。

    - (年,月) が窓内 → 算入
    - 年不明だが月が窓内 → 算入（best-effort, undated に記録して警告対象に）
    - 支給日完全不明（ym パース不可）→ 後述「支給日不明バケット」で処理
    - 窓外（翌期・前期）→ 算入しない（R216 は対象事業年度の賞与のみ）
    重複（分割PDFの重なり）は (ym, 金額) で de-dup。

    undated_paid_months（ユーザーが「賃金台帳の作成」タスクで入力した賞与支給月リスト）:
      支給日完全不明の賞与に対し、年間集計表など「年間合計しか無い」資料では金額の
      月按分が原理的に不能。よってバケット全体に対して保守的に判定する:
        - 指定支給月が全て窓内 → 全額算入（解決済み・undated に積まない＝警告不要）
        - 指定支給月が全て窓外 → 当年度に含めない（暦年賞与を非暦年決算に誤算入する事故を防ぐ）
        - 窓内外が混在 → 按分不能のため全額算入＋undated 記録（要人手確認）
      未指定（None/空）のときは従来どおり全額算入＋undated 記録（警告対象）。
    """
    total = 0.0
    undated: list[float] = []
    dropped: list[float] = []  # 支給年月ありだが対象年度ウィンドウ外で除外した賞与額（情報用）
    undated_amounts: list[float] = []  # ym 完全不明の賞与額（後でまとめて判定）
    seen: set = set()
    window_set = set(fiscal_window) if fiscal_window else None
    window_months = {mo for (_y, mo) in fiscal_window} if fiscal_window else None
    for b in bonus_entries or []:
        if not isinstance(b, dict):
            continue
        amt = _to_float(b.get('amount'))
        if amt is None or amt <= 0:
            continue
        ymraw = b.get('ym')
        dedup_key = (str(ymraw), round(amt, 2))
        if dedup_key in seen:
            continue
        seen.add(dedup_key)
        ym = parse_ym_header(ymraw)
        if window_set is None:
            total += amt
            continue
        if ym is None:
            undated_amounts.append(amt)
            continue
        y, mo = ym
        if y is None:
            if mo in window_months:
                total += amt
                undated.append(amt)
            else:
                dropped.append(amt)  # 月のみ判明だが窓の月に該当せず除外
            continue
        if (y, mo) in window_set:
            total += amt
        else:
            dropped.append(amt)  # 年月判明だが対象事業年度外 → 除外（情報として記録）

    # 支給日完全不明バケットの処理
    if undated_amounts:
        if window_set is not None and undated_paid_months:
            def _in_window(m: tuple[int | None, int]) -> bool:
                y, mo = m
                return (mo in window_months) if y is None else ((y, mo) in window_set)
            in_win = [m for m in undated_paid_months if _in_window(m)]
            if not in_win:
                pass  # 全て窓外 → 当年度に含めない（total に積まない）
            elif len(in_win) == len(undated_paid_months):
                total += sum(undated_amounts)  # 全て窓内 → 全額算入（解決済み）
            else:
                total += sum(undated_amounts)  # 混在 → 按分不能・全額算入＋要確認
                undated.extend(undated_amounts)
        else:
            total += sum(undated_amounts)  # 支給月未指定 → 従来動作
            undated.extend(undated_amounts)
    return total, undated, dropped


def _normalize_ai_employee_dict(
    emp: dict, fiscal_window: list[tuple[int, int]] | None,
    *, undated_paid_months: list[tuple[int | None, int]] | None = None,
) -> dict:
    """新スキーマ（monthly/bonuses）を旧スキーマ（monthly_wages[12] 等）に正規化する。

    旧スキーマ（monthly_wages を持つ）はそのまま返す（後方互換）。
    新スキーマは決算月ウィンドウで12ヶ月を確定し、賞与を年間合計に集約する。
    """
    if 'monthly' not in emp:
        return emp
    monthly_entries = emp.get('monthly') or []
    bonus_entries = emp.get('bonuses') or []
    if fiscal_window:
        wages, hours, days = _build_windowed_arrays(monthly_entries, fiscal_window)
    else:
        wages, hours, days = _build_calendar_arrays_recent(monthly_entries)
    bonus_total, undated, dropped = _sum_window_bonuses(
        bonus_entries, fiscal_window, undated_paid_months=undated_paid_months,
    )
    out = dict(emp)
    out['monthly_wages'] = wages
    out['monthly_hours'] = hours
    out['monthly_work_days'] = days
    out['annual_bonus'] = max(0.0, bonus_total)
    if undated:
        out['_bonus_undated'] = undated
    if dropped:
        out['_bonus_dropped'] = dropped
    return out


def _validate_ai_employee(emp: dict) -> tuple[bool, str]:
    """AI抽出した1従業員データの妥当性チェック。(OK?, エラー理由)"""
    name = emp.get('name')
    if not name or not isinstance(name, str):
        return False, 'name が空または文字列でない'
    monthly_wages = emp.get('monthly_wages')
    monthly_hours = emp.get('monthly_hours')
    if not isinstance(monthly_wages, list) or len(monthly_wages) != 12:
        return False, f'monthly_wages が12要素のリストでない (len={len(monthly_wages) if isinstance(monthly_wages, list) else "N/A"})'
    if not isinstance(monthly_hours, list) or len(monthly_hours) != 12:
        return False, f'monthly_hours が12要素のリストでない'
    # 金額の現実的範囲チェック (0〜1000万円/月)
    for i, w in enumerate(monthly_wages):
        if w is None:
            continue
        if not isinstance(w, (int, float)) or w < 0 or w > 10_000_000:
            return False, f'{i+1}月の給与額が異常: {w}'
    # 労働時間の現実的範囲チェック (0〜400時間/月)
    for i, h in enumerate(monthly_hours):
        if h is None:
            continue
        if not isinstance(h, (int, float)) or h < 0 or h > 400:
            return False, f'{i+1}月の労働時間が異常: {h}'
    # monthly_work_days は任意フィールド（無くても可）。あれば妥当性チェック
    monthly_work_days = emp.get('monthly_work_days')
    if monthly_work_days is not None:
        if not isinstance(monthly_work_days, list) or len(monthly_work_days) != 12:
            return False, 'monthly_work_days が12要素のリストでない'
        for i, d in enumerate(monthly_work_days):
            if d is None:
                continue
            if not isinstance(d, (int, float)) or d < 0 or d > 31:
                return False, f'{i+1}月の労働日数が異常: {d}'
    return True, ''


def _ai_data_to_wage_employees(
    ai_data: list[dict], fiscal_period_hint: str | None = None,
    *, derive_year_from_data: bool = False,
    undated_paid_months: list[tuple[int | None, int]] | None = None,
) -> list[WageEmployee]:
    """AI抽出データを WageEmployee リストに変換（バリデーション付き）。

    新スキーマ（monthly[{ym,taxable,hours,work_days}] / bonuses[{ym,amount}]）は、
    fiscal_period_hint から確定した事業年度12ヶ月を決定論で選択し、賞与を年間合計に
    集約してから旧スキーマ（monthly_wages[12] 等）に正規化する。
    旧スキーマ（monthly_wages を直接持つ）はそのまま処理（後方互換）。

    derive_year_from_data=True（賃金台帳作成タスク）の場合、期末「年」を台帳の実在年から
    決め直す（決算月は hint を信頼）。決算書を読まず今日基準で年を推定する作成タスクで、
    台帳が1期ズレてもサイレント過少計上しないための補正。年補正・低カバレッジは
    fiscal_window_note に載せて呼出側が surfacing する。
    （申請書/給与計算タスクは決算書 fiscal_year_end が期末年の正なので False のまま据置。）

    労働時間が「ない / 異常に少ない（残業時間と誤認の疑い）」場合は、
    労働日数×8時間で補完する。役員は労働時間補完の対象外。
    """
    HOURS_PER_DAY = 8.0
    SUSPICIOUS_AVG_HOURS = 50.0  # 役員/パート以外で月平均がこれ未満なら誤認の疑い

    window_note = ''
    if derive_year_from_data and fiscal_period_hint:
        fiscal_window, _winfo = select_fiscal_window_from_data(ai_data, fiscal_period_hint)
        window_note = _format_window_note(_winfo)
        if window_note:
            logger.warning(f'事業年度ウィンドウ選択: {window_note}')
    else:
        fiscal_window = resolve_fiscal_window(fiscal_period_hint)
    employees: list[WageEmployee] = []
    for i, emp in enumerate(ai_data):
        if not isinstance(emp, dict):
            logger.warning(f'AI抽出: index={i} が辞書でないためスキップ: {type(emp).__name__}')
            continue
        # 新スキーマ → 旧スキーマ正規化（事業年度ウィンドウ適用・賞与集約）
        emp = _normalize_ai_employee_dict(
            emp, fiscal_window, undated_paid_months=undated_paid_months,
        )
        if emp.get('_bonus_undated'):
            logger.warning(
                f'AI抽出: {emp.get("name", "?")} の賞与に支給日不明分があり年間賞与へ算入しました '
                f'（{emp["_bonus_undated"]} 円）。対象事業年度内かは PDF 原本で要確認'
            )
        ok, reason = _validate_ai_employee(emp)
        if not ok:
            logger.warning(f'AI抽出: index={i} ({emp.get("name", "?")}) バリデーション失敗: {reason}')
            continue

        emp_name = str(emp['name']).strip()
        emp_type = str(emp.get('employment_type', '') or '').strip()
        is_officer = '役員' in emp_type
        is_part = 'パート' in emp_type or 'アルバイト' in emp_type

        monthly_wages = [
            float(w) if w is not None else None for w in emp['monthly_wages']
        ]
        monthly_hours: list[float | None] = [
            float(h) if h is not None else None for h in emp['monthly_hours']
        ]
        monthly_work_days = emp.get('monthly_work_days') or [None] * 12

        valid_hours = [h for h in monthly_hours if h is not None and h > 0]
        avg_hours = sum(valid_hours) / len(valid_hours) if valid_hours else 0.0

        # 月給制（正社員 = 役員でもパートでもない）で時間データが部分的（12ヶ月のうち
        # PARTIAL_HOURS_THRESHOLD 未満）の場合は、月給制と判断して時間情報を全 null にする。
        # → 賃金台帳に「平日普通4時間、休日普通15時間」等の勤怠内訳が部分記録された月だけ
        #    AI が拾ってしまい、後段で時給を誤計算する問題を回避。
        PARTIAL_HOURS_THRESHOLD = 3  # 3ヶ月未満なら「月給制で部分記録」と判定
        is_monthly_paid_full_time = not is_officer and not is_part
        if is_monthly_paid_full_time and 0 < len(valid_hours) < PARTIAL_HOURS_THRESHOLD:
            logger.info(
                f'AI抽出補正: {emp_name} (正社員) は時間データが {len(valid_hours)}ヶ月のみで '
                f'月給制と判断 → monthly_hours を全 null に補正'
            )
            monthly_hours = [None] * 12
            valid_hours = []
            avg_hours = 0.0

        # 労働時間が無い、または役員/パート以外で異常に少ない場合は労働日数×8hで補完
        needs_fallback = not valid_hours or (
            not is_officer
            and not is_part
            and avg_hours < SUSPICIOUS_AVG_HOURS
        )
        if needs_fallback and not is_officer:
            valid_days = [
                d for d in monthly_work_days
                if d is not None and isinstance(d, (int, float)) and d > 0
            ]
            # 正社員で労働日数も部分的（PARTIAL_HOURS_THRESHOLD 未満）の場合は補完しない
            # → 中途入社月給制で、入社月だけ勤怠記録が残るケース対策
            if is_monthly_paid_full_time and 0 < len(valid_days) < PARTIAL_HOURS_THRESHOLD:
                logger.info(
                    f'AI抽出補正: {emp_name} (正社員) は労働日数も {len(valid_days)}ヶ月のみで '
                    '月給制と判断 → 労働日数×8h補完をスキップ'
                )
            elif valid_days:
                old_avg = avg_hours
                monthly_hours = [
                    float(d) * HOURS_PER_DAY if (d is not None and isinstance(d, (int, float)) and d > 0) else None
                    for d in monthly_work_days
                ]
                valid_hours = [h for h in monthly_hours if h is not None and h > 0]
                avg_hours = sum(valid_hours) / len(valid_hours) if valid_hours else 0.0
                logger.info(
                    f'AI抽出補完: {emp_name} の労働時間を労働日数×{HOURS_PER_DAY}hで再計算 '
                    f'(月平均 {old_avg:.1f}h → {avg_hours:.1f}h)'
                )

        # 時給算出: 月別 monthly_wages / monthly_hours から計算
        # 加点判定 (judge_bonus_points) と給与計算シートの時給表示で使用
        # 役員は時給を出さない（月給制で意味がないため）
        monthly_hourly_rates: list[float | None] = []
        for w, h in zip(monthly_wages, monthly_hours):
            if not is_officer and w is not None and h is not None and h > 0:
                monthly_hourly_rates.append(w / h)
            else:
                monthly_hourly_rates.append(None)
        valid_hourly = [h for h in monthly_hourly_rates if h is not None and h > 0]
        avg_hourly = sum(valid_hourly) / len(valid_hourly) if valid_hourly else 0.0

        # 年間通勤手当（AI が抽出できた場合のみ取得、なければ 0）
        atransport = emp.get('annual_transport_allowance', 0)
        try:
            atransport_val = float(atransport) if atransport is not None else 0.0
        except (TypeError, ValueError):
            atransport_val = 0.0
        if atransport_val < 0:
            atransport_val = 0.0  # 負値は無視

        # 非課税通勤手当が monthly_wages（支給合計＝通勤込み）に含まれている場合の課税額補正。
        # AI プロンプトは「monthly_wages に課税支給合計を入れた場合は通勤費除外済み→atransport=0」
        # を指示しているが、課税列が無い台帳で AI が通勤込みを入れ atransport>0 を返すことがある。
        # その場合は在籍月数で均等割して monthly_wages から減算し、課税額（R216 母数）に揃える
        # （決定論パーサー _parse_section_summary の S列減算と同一ロジック）。
        # 減算後は atransport_val=0 とする（消費済み）。これにより賃金台帳作成タスクで S列に
        # 二重計上→再読込時に二重減算、という不整合を防ぐ。不変条件「monthly_wages＝課税額」を担保。
        if atransport_val > 0:
            _non_null = [m for m in range(12) if monthly_wages[m] is not None]
            if _non_null:
                _per_month = atransport_val / len(_non_null)
                monthly_wages = [
                    (max(0.0, w - _per_month) if w is not None else None)
                    for w in monthly_wages
                ]
                logger.info(
                    f'AI抽出: {emp_name} の年間通勤手当{atransport_val:,.0f}円を'
                    f'在籍{len(_non_null)}ヶ月で均等割し monthly_wages から減算（課税額補正）'
                )
            atransport_val = 0.0

        employees.append(WageEmployee(
            no=i + 1,
            name=emp_name,
            employment_type=emp_type,
            monthly_avg_hours=round(avg_hours, 1),
            hourly_rate=round(avg_hourly, 1),
            monthly_wages=monthly_wages,
            monthly_hourly_rates=monthly_hourly_rates,
            monthly_hours=monthly_hours,
            annual_transport_allowance=atransport_val,
            annual_bonus=float(emp.get('annual_bonus') or 0.0),
            fiscal_window=fiscal_window,
            fiscal_window_note=window_note,
            bonus_undated_total=float(sum(emp.get('_bonus_undated') or [])),
            bonus_dropped_total=float(sum(emp.get('_bonus_dropped') or [])),
        ))
    return employees


def _load_wage_ai_payload(
    file_paths: list[Path],
) -> tuple[list[str], list[tuple[str, bytes]]]:
    """賃金台帳ファイル群を AI 抽出用の (TSVブロック list, PDFバイナリ list) に変換する。

    PDF はテキスト層があれば TSV 化、薄ければバイナリ添付にフォールバック。
    CSV/Excel は TSV 化。read_wage_ledgers_with_ai と加点用抽出で共有する。
    """
    tsv_blocks: list[str] = []
    pdf_files: list[tuple[str, bytes]] = []
    for path in file_paths:
        ext = path.suffix.lower()
        if ext == '.pdf':
            try:
                tsv_blocks.append(_pdf_to_tsv(path))
                logger.info(f'賃金台帳PDF→テキスト変換: {path.name}')
            except RuntimeError as e:
                logger.warning(f'{e} → バイナリ送信にフォールバック')
                try:
                    pdf_files.append((path.name, path.read_bytes()))
                except Exception as e2:
                    logger.warning(f'賃金台帳PDF読込失敗(AI経路): {path.name} ({e2})')
            except Exception as e:
                logger.warning(f'賃金台帳PDF読込失敗(AI経路): {path.name} ({e})')
            continue
        if ext == '.csv':
            try:
                tsv_blocks.append(_csv_to_tsv(path))
            except Exception as e:
                level = _csv_decode_warning_level(path)
                msg = f'賃金台帳CSV読込失敗(AI経路): {path.name} ({e})'
                logger.info(msg) if level == 'info' else logger.warning(msg)
            continue
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True)
        except Exception as e:
            logger.warning(f'賃金台帳読込失敗(AI経路): {path.name} ({e})')
            continue
        tsv_blocks.append(_workbook_to_tsv(wb, path.name))
        wb.close()
    return tsv_blocks, pdf_files


def _ai_data_to_bonus_employees(ai_data: list[dict]) -> list[BonusWageEmployee]:
    """AI抽出データ（ymキーの monthly[{ym,base,hours,...}]）を BonusWageEmployee に変換。

    12スロットには畳まず (年,月)→基本給 を保持する（加点①は R6/10〜R7/9 をまたぐため）。
    所定労働時間は月次 hours の中央値を代表値に置く（無ければ None＝後で人手記入）。
    """
    out: list[BonusWageEmployee] = []
    for i, emp in enumerate(ai_data):
        if not isinstance(emp, dict):
            continue
        name = str(emp.get('name') or '').strip()
        if not name:
            continue
        emp_type = str(emp.get('employment_type', '') or '').strip()
        base_by_ym: dict[tuple[int, int], float] = {}
        hours_by_ym: dict[tuple[int, int], float] = {}
        hours_samples: list[float] = []
        for m in (emp.get('monthly') or []):
            if not isinstance(m, dict):
                continue
            parsed = parse_ym_header(m.get('ym'))
            if not parsed or parsed[0] is None:
                continue
            ym = (parsed[0], parsed[1])
            base = _to_float(m.get('base'))
            if base is not None and base > 0:
                base_by_ym[ym] = base
            h = _to_float(m.get('hours'))
            if h is not None and h > 0:
                hours_by_ym[ym] = h
                hours_samples.append(h)
        sched = sorted(hours_samples)[len(hours_samples) // 2] if hours_samples else None
        out.append(BonusWageEmployee(
            no=i + 1, name=name, employment_type=emp_type,
            scheduled_hours=sched, monthly_base=base_by_ym,
            monthly_hours_override=hours_by_ym,
        ))
    return out


def read_bonus_source_employees(
    file_paths: list[Path],
    extractor,
    *,
    disable_image_fallback: bool = False,
) -> list[BonusWageEmployee]:
    """生の賃金台帳/給与明細から加点判定用の従業員（(年,月)→基本給）を AI 抽出する。

    read_wage_ledgers_with_ai と同じ抽出経路を使うが、事業年度12スロットに畳まず
    ymキーの基本給を保持する（加点①の R6/10〜R7/9 は年をまたぐため）。
    """
    if not file_paths:
        return []
    tsv_blocks, pdf_files = _load_wage_ai_payload(file_paths)
    if not tsv_blocks and not pdf_files:
        return []
    combined_tsv = '\n\n'.join(tsv_blocks) if tsv_blocks else ''
    logger.info(
        f'加点用AI抽出開始: TSV{len(tsv_blocks)}ブロック({len(combined_tsv):,}文字)'
        + (f' + PDFバイナリ{len(pdf_files)}件' if pdf_files else '')
    )
    try:
        ai_data = extractor.extract_wage_ledger(
            combined_tsv, None,
            pdf_files=pdf_files if pdf_files else None,
            disable_image_fallback=disable_image_fallback,
        )
    except Exception as e:
        from .ai_extractor import APICreditExhaustedError, ImageFallbackBlockedError
        if isinstance(e, (APICreditExhaustedError, ImageFallbackBlockedError)):
            raise
        logger.error(f'加点用AI抽出例外: {e}', exc_info=True)
        return []
    employees = _ai_data_to_bonus_employees(ai_data)
    logger.info(f'加点用AI抽出結果: 入力{len(ai_data)}名 → {len(employees)}名')
    return employees


def read_wage_ledgers_with_ai(
    file_paths: list[Path],
    extractor,
    fiscal_period_hint: str | None = None,
    *,
    disable_image_fallback: bool = False,
    derive_year_from_data: bool = False,
    undated_paid_months: list[tuple[int | None, int]] | None = None,
) -> list[WageEmployee]:
    """
    AI による賃金台帳読み取り。
    Excel(.xlsx/.xlsm)・CSV は TSV に変換、PDF はそのまま添付して1回の API 呼出しで抽出する。
    各フォーマット混在も可。バリデーション失敗時は空リストを返す（呼出し側で fallback 判断）。

    Args:
        disable_image_fallback: True の場合、Document AI 失敗時に Sonnet 画像経路へ
            落ちる前に ImageFallbackBlockedError が送出される（伝播してくる）。
            「賃金台帳の作成」タスク向け。
    """
    if not file_paths:
        return []

    tsv_blocks, pdf_files = _load_wage_ai_payload(file_paths)
    if not tsv_blocks and not pdf_files:
        return []

    combined_tsv = '\n\n'.join(tsv_blocks) if tsv_blocks else ''
    pdf_count_binary = len(pdf_files)
    pdf_total_mb = sum(len(p[1]) for p in pdf_files) / 1_000_000
    logger.info(
        f'AI抽出開始: TSV{len(tsv_blocks)}ブロック({len(combined_tsv):,}文字)'
        + (f' + PDFバイナリ{pdf_count_binary}件({pdf_total_mb:.2f}MB)' if pdf_files else '')
        + (f' (前事業年度ヒント: {fiscal_period_hint})' if fiscal_period_hint else '')
    )

    try:
        ai_data = extractor.extract_wage_ledger(
            combined_tsv,
            fiscal_period_hint,
            pdf_files=pdf_files if pdf_files else None,
            disable_image_fallback=disable_image_fallback,
        )
    except Exception as e:
        # API残高切れ・画像フォールバック禁止例外は pipeline で全体停止する必要があるので再 raise
        from .ai_extractor import APICreditExhaustedError, ImageFallbackBlockedError
        if isinstance(e, (APICreditExhaustedError, ImageFallbackBlockedError)):
            raise
        logger.error(f'AI抽出例外: {e}', exc_info=True)
        return []

    employees = _ai_data_to_wage_employees(
        ai_data, fiscal_period_hint, derive_year_from_data=derive_year_from_data,
        undated_paid_months=undated_paid_months,
    )
    logger.info(
        f'AI抽出結果: 入力{len(ai_data)}名 → 妥当{len(employees)}名'
    )
    return employees


# OCR が混同しやすい異体字・類似字形の正規化辞書。
# 同一人物判定の取りこぼしを防ぐため、代表字に統一する（読みやすさより一貫性優先）。
# ここに足すかどうかの基準: 字形が似ていて OCR が両方候補に挙げる + 同一人物のはずが
# 別人扱いされる典型例を実観測している こと。
_OCR_VARIANT_MAP = {
    # 旧字 ⇔ 新字（実観測: 吉田 壽 ⇔ 吉田 靖）
    '壽': '寿',
    # 旧字柳 / 櫛 の混乱（実観測: 栁井 ⇔ 櫛井）
    '栁': '柳',
    '櫛': '柳',
    # 嶋 / 崎 / 島 の OCR 混乱（実観測: 大嶋 ⇔ 大崎、宮嶋 ⇔ 宮崎）
    '嶋': '島',
    '崎': '島',
    # 高 / 髙 の混乱（実観測: 髙橋 ⇔ 高橋）
    '髙': '高',
    # 斉 / 齊 / 斎 / 齋 の混乱（実観測: 斉藤 ⇔ 斎藤）
    '齊': '斉',
    '齋': '斉',
    '斎': '斉',
    # 渡辺 / 渡邉 / 渡邊 の混乱
    '邉': '辺',
    '邊': '辺',
    # 沢 / 澤 の混乱
    '澤': '沢',
    # 浜 / 濱 の混乱
    '濱': '浜',
}


def _apply_variant_normalization(s: str) -> str:
    """OCR 異体字を代表字に置換"""
    return ''.join(_OCR_VARIANT_MAP.get(c, c) for c in s)


def _normalize_name_key(name: str) -> str:
    """同一人物判定用の正規化キー。

    - NFKC正規化（全角英数→半角、半角カナ→全角等）
    - 全角・半角の空白をすべて除去
    - OCR 異体字を代表字に置換（壽→寿、栁→柳、嶋→島 等）
    - 大文字小文字統一はしない（漢字氏名の運用なので影響小）

    例: '長塚 典子' / '長塚　典子' / '長塚典子' は同じキー '長塚典子' になる。
    例: '吉田 壽' と '吉田 靖' は字形違いで別人扱いされやすいが、
        '壽' は '寿' に正規化される。'靖' は別字なので統合されないが、
        '吉田寿' に近い候補として後段（_merge_similar_names）で統合判定される。
    """
    if not name:
        return ''
    n = unicodedata.normalize('NFKC', str(name))
    n = re.sub(r'[\s　]+', '', n)
    n = _apply_variant_normalization(n)
    return n


def _merge_two_employees(a: WageEmployee, b: WageEmployee) -> WageEmployee:
    """同一人物と判定された2件のWageEmployeeを統合する。

    同月衝突ポリシー:
        - 片方null → もう片方を採用（補完）
        - 両方値あり・差 < 1円(時間 < 0.1h) → そのまま採用
        - 両方値あり・差大 → WARNING ログ + **大きい方を採用**
          （理由: 部分入力・欠損で値が小さくなる典型的ミスを救う。
           給与は「完全データ ≥ 部分データ ≥ 0」なので max が最も完全な値を選ぶ）
    """
    new_wages: list[float | None] = []
    new_hours: list[float | None] = []
    for i in range(12):
        wa = a.monthly_wages[i] if i < len(a.monthly_wages) else None
        wb = b.monthly_wages[i] if i < len(b.monthly_wages) else None
        if wa is None:
            new_wages.append(wb)
        elif wb is None:
            new_wages.append(wa)
        elif abs(wa - wb) < 1.0:
            new_wages.append(wa)
        else:
            chosen = max(wa, wb)
            logger.warning(
                f'同名統合: "{a.name}" の{i+1}月給与に差分があります '
                f'({wa:,.0f} vs {wb:,.0f}) → 大きい方 {chosen:,.0f} を採用'
            )
            new_wages.append(chosen)

        ha = a.monthly_hours[i] if i < len(a.monthly_hours) else None
        hb = b.monthly_hours[i] if i < len(b.monthly_hours) else None
        if ha is None:
            new_hours.append(hb)
        elif hb is None:
            new_hours.append(ha)
        elif abs(ha - hb) < 0.1:
            new_hours.append(ha)
        else:
            # 時間も同じ理由で max（残業時間との誤判定で部分値が出るケースを救う）
            new_hours.append(max(ha, hb))

    # 月平均/時給を統合後の値で再計算
    valid_h = [h for h in new_hours if h is not None and h > 0]
    avg_h = sum(valid_h) / len(valid_h) if valid_h else 0.0

    is_officer = '役員' in (a.employment_type or '')
    new_rates: list[float | None] = []
    for w, h in zip(new_wages, new_hours):
        if not is_officer and w is not None and h is not None and h > 0:
            new_rates.append(w / h)
        else:
            new_rates.append(None)
    valid_r = [r for r in new_rates if r is not None and r > 0]
    avg_r = sum(valid_r) / len(valid_r) if valid_r else 0.0

    # source_file の統合: 両方あれば連結（重複排除）、片方なら採用
    sa = (a.source_file or '').strip()
    sb = (b.source_file or '').strip()
    if sa and sb:
        if sa == sb:
            merged_source = sa
        else:
            merged_source = f'{sa} + {sb}'
    else:
        merged_source = sa or sb

    return WageEmployee(
        no=a.no,
        name=a.name,
        employment_type=a.employment_type or b.employment_type,
        monthly_avg_hours=round(avg_h, 1),
        hourly_rate=round(avg_r, 1),
        monthly_wages=new_wages,
        monthly_hourly_rates=new_rates,
        monthly_hours=new_hours,
        source_file=merged_source,
        # 年間通勤手当・年間賞与は同一人物の重複（名寄せ揺れ）で二重計上しないよう max を採る
        annual_transport_allowance=max(
            a.annual_transport_allowance or 0.0, b.annual_transport_allowance or 0.0
        ),
        annual_bonus=max(a.annual_bonus or 0.0, b.annual_bonus or 0.0),
        fiscal_window=a.fiscal_window or b.fiscal_window,
        fiscal_window_note=a.fiscal_window_note or b.fiscal_window_note,
        bonus_undated_total=max(
            a.bonus_undated_total or 0.0, b.bonus_undated_total or 0.0
        ),
        bonus_dropped_total=max(
            a.bonus_dropped_total or 0.0, b.bonus_dropped_total or 0.0
        ),
    )


def _dedupe_employees_by_normalized_name(
    employees: list[WageEmployee],
) -> list[WageEmployee]:
    """正規化キー（NFKC + 全空白除去）で同一人物の重複を統合する。

    例: '長塚典子' (xlsx由来) と '長塚 典子' (ファイル名由来) を同一として統合。
    """
    if not employees:
        return employees

    merged: dict[str, WageEmployee] = {}
    order: list[str] = []
    for emp in employees:
        key = _normalize_name_key(emp.name)
        if not key:
            continue
        if key not in merged:
            merged[key] = emp
            order.append(key)
        else:
            existing = merged[key]
            if existing.name != emp.name:
                logger.info(
                    f'同名統合: "{existing.name}" と "{emp.name}" を同一人物として統合'
                )
            merged[key] = _merge_two_employees(existing, emp)

    result = [merged[k] for k in order]
    for i, emp in enumerate(result):
        emp.no = i + 1
    return result


def _reconcile_midyear_positions_with_deterministic(
    ai_employees: list[WageEmployee],
    file_paths: list[Path],
    fiscal_period_hint: str | None,
) -> list[WageEmployee]:
    """AI 抽出された中途者の monthly_wages 位置を決定論パーサーで補正。

    背景:
        AI 抽出（Claude）は短期在籍者（在籍 1〜数ヶ月）の monthly_wages を
        誤った位置（暦月とズレた index）に格納する事例が観測されている。
        値そのものは正しいが index がズレるので、出力備考の月ラベルが
        実体と食い違う（例: 賃金台帳の暦4月の値が AI 出力では index 5＝6月扱い）。
        さらに同一ファイルでも実行ごとに変動するため不安定。

    対処:
        - 「データ位置だけ違って値の集合は一致」しているケースに限定して補正
          → 月給合算など値が変わる処理を AI が施した行は触らない（安全側）
        - AI 側 or 決定論側どちらかが「12ヶ月未満」であれば候補（中途者扱い）
          → AI が誤って full_year=True を返した場合の取りこぼし対策
        - 名前は既存の _normalize_name_key で照合（OCR異体字・空白対応）
        - 同名 collision は曖昧として両方とも突合対象から除外（誤上書き防止）
        - 値の上書きは「中身が有効な月が1つ以上ある」場合のみ
          → 空リストで AI の有効データを潰さない

    決定論パーサーは API を呼ばないので追加コストは CPU のみ。
    PDF のみの賃金台帳（Excel/CSV なし）では決定論パーサーが失敗するため
    その場合は AI 結果をそのまま返す。

    Returns:
        補正後の ai_employees（破壊的に書き換え、戻り値は同一リスト）。
    """
    # 突合候補: AI 側で is_full_year=False の人だけでなく、AI が誤って
    # 全埋めしてきた可能性も考えるため、判定は後段で決定論側との比較で行う。
    # ここでは空集合の枝刈りだけ
    if not ai_employees:
        return ai_employees

    # 決定論パーサーで全件読む（CPU のみ、API ゼロ）
    try:
        det_employees = read_wage_ledgers(
            file_paths, extractor=None, fiscal_period_hint=fiscal_period_hint,
        )
    except Exception as e:
        logger.warning(
            f'位置突合用の決定論パーサー読込に失敗: {e} → AI 結果のままで継続'
        )
        return ai_employees

    if not det_employees:
        # PDF のみ等で決定論パーサーが読めないケース
        logger.info('決定論パーサーで0件 → AI 結果をそのまま採用（突合スキップ）')
        return ai_employees

    # 既存の名前正規化を使う（OCR異体字・空白除去を共有）
    # 同名 collision を検出して、その名前は両側とも除外（安全側）
    det_by_name: dict[str, WageEmployee] = {}
    ambiguous_names: set[str] = set()
    for de in det_employees:
        key = _normalize_name_key(de.name)
        if not key:
            continue
        if key in det_by_name:
            ambiguous_names.add(key)
            continue
        det_by_name[key] = de
    for k in ambiguous_names:
        det_by_name.pop(k, None)  # 曖昧キーは突合対象外
        logger.warning(
            f'同名 collision を検出: 正規化キー "{k}" が決定論側に複数あり、'
            '位置突合をスキップ（誤上書き回避）'
        )

    def _data_indices(monthly: list) -> set[int]:
        return {i for i, w in enumerate(monthly or []) if w is not None}

    def _data_multiset(monthly: list) -> tuple:
        """値の multiset。位置非依存で比較するための tuple（順序付け済み）。"""
        vals = [w for w in (monthly or []) if w is not None]
        return tuple(sorted(vals))

    fixed_count = 0
    for ai_emp in ai_employees:
        det = det_by_name.get(_normalize_name_key(ai_emp.name))
        if det is None:
            continue
        ai_months = _data_indices(ai_emp.monthly_wages)
        det_months = _data_indices(det.monthly_wages)
        # 突合候補は「どちらかが12ヶ月未満」かつ「インデックス集合が違う」
        if len(ai_months) == 12 and len(det_months) == 12:
            continue
        if ai_months == det_months:
            continue
        # 値の集合（multiset）が一致するか確認
        # → 一致なら「位置だけ違う、値は同じ」ことが確定。安全に上書き
        # → 不一致なら値そのものが違う可能性が高いので、警告だけ出して上書きしない
        ai_vals = _data_multiset(ai_emp.monthly_wages)
        det_vals = _data_multiset(det.monthly_wages)
        if ai_vals != det_vals:
            logger.warning(
                f'位置不一致だが値集合も不一致: {ai_emp.name} '
                f'AI 値集合={ai_vals} 決定論 値集合={det_vals} '
                f'→ 上書きせず AI 結果のまま（要目視確認）'
            )
            continue
        # 位置だけ違い、値は完全一致 → 決定論側で位置を上書き
        logger.warning(
            f'位置のみ不一致を検出: {ai_emp.name} '
            f'AI={sorted(ai_months)} 決定論={sorted(det_months)} '
            f'値集合は一致 → 決定論側で位置上書き'
        )
        ai_emp.monthly_wages = list(det.monthly_wages)
        # monthly_hours / monthly_hourly_rates は「有効な要素が1つ以上ある」場合のみ上書き
        # （[None]*12 のような空リストで AI の有効データを潰さない）
        if det.monthly_hours and any(h is not None for h in det.monthly_hours):
            ai_emp.monthly_hours = list(det.monthly_hours)
        if det.monthly_hourly_rates and any(
            h is not None for h in det.monthly_hourly_rates
        ):
            ai_emp.monthly_hourly_rates = list(det.monthly_hourly_rates)
        fixed_count += 1

    if fixed_count:
        logger.warning(
            f'AI抽出 vs 決定論パーサーの突合: {fixed_count}名の monthly_wages '
            f'位置を決定論側で上書き（中途者の備考月ラベルを正しく出すため）'
        )

    return ai_employees


def read_wage_ledgers(
    file_paths: list[Path],
    extractor=None,
    fiscal_period_hint: str | None = None,
    *,
    disable_image_fallback: bool = False,
    derive_year_from_data: bool = False,
    undated_paid_months: list[tuple[int | None, int]] | None = None,
) -> list[WageEmployee]:
    """
    複数の賃金台帳ファイルを読み、同名の従業員をマージして返す。
    1人1ファイル運用（給与ソフト出力）と、1ファイルに全員を入れる運用の両方に対応。

    extractor が渡され、かつ環境変数 USE_AI_WAGE_EXTRACTION が有効な場合は
    AI 抽出を優先し、結果が空なら決定論パーサーにフォールバックする。

    最終結果は NFKC 正規化キー（全空白除去）で同一人物の重複を統合する。
    'A' / 'A ' / 'A　' / 'Ａ' のような表記揺れを 1 人として扱う。

    Args:
        disable_image_fallback: True の場合、Document AI 失敗時に Sonnet 画像経路への
            フォールバックを禁止する（ImageFallbackBlockedError を送出）。
            「賃金台帳の作成」タスクで Document AI 一本に絞る用途。
    """
    if not file_paths:
        return []

    # AI 経路（extractor がある場合）
    if extractor is not None:
        from .config import USE_AI_WAGE_EXTRACTION
        if USE_AI_WAGE_EXTRACTION:
            ai_employees = read_wage_ledgers_with_ai(
                file_paths, extractor, fiscal_period_hint,
                disable_image_fallback=disable_image_fallback,
                derive_year_from_data=derive_year_from_data,
                undated_paid_months=undated_paid_months,
            )
            if ai_employees:
                ai_employees = _dedupe_employees_by_normalized_name(ai_employees)
                # employment_type 補完: provenance を残す形式「正社員(推定)」で埋める。
                # check_employment_type_missing は「(推定)」を含む値もカウントするため、
                # 補完しても警告は出る（人間チェック可能）。
                for emp in ai_employees:
                    if not (emp.employment_type or '').strip():
                        emp.employment_type = '正社員(推定)'
                # データソース（抽出根拠ファイル名）の補完。
                # AI 経路は複数 PDF をまとめて投入するため、ファイル単位の分離は困難。
                # ファイル名一覧として記録（全員同じ値になる）
                _assign_source_files(ai_employees, file_paths)
                # AI 抽出の中途者は monthly_wages の位置が暦月とズレる事例が観測
                # されている（実行ごとに +1〜+3ヶ月変動）。決定論パーサーは
                # Excel/CSV を直接読むので位置は確実。中途者だけ位置を突合補正する。
                ai_employees = _reconcile_midyear_positions_with_deterministic(
                    ai_employees, file_paths, fiscal_period_hint,
                )
                logger.info(f'賃金台帳合算結果(AI): {len(ai_employees)}名 ({len(file_paths)}ファイル)')
                return ai_employees
            logger.warning('AI抽出が0件を返したため、決定論パーサーにフォールバック')

    # 決定論パーサー経路（フォールバック or extractor なし）
    individual_ledger_paths = []
    merged_emp_data: dict = {}

    for path in file_paths:
        try:
            # CSV の場合は pandas で読込、以降の処理は Excel と同じ
            if path.suffix.lower() == '.csv':
                before = len(merged_emp_data)
                _read_csv(path, merged_emp_data)
                logger.info(
                    f'賃金台帳読み取り(CSV): {path.name} '
                    f'→ 追加/更新 {len(merged_emp_data) - before}名 (累計{len(merged_emp_data)}名)'
                )
                continue

            # Excel の場合は既存処理
            wb = openpyxl.load_workbook(str(path), data_only=True)
        except Exception as e:
            logger.warning(f'賃金台帳読込失敗: {path.name} ({e})')
            continue

        if _is_individual_ledger(wb):
            individual_ledger_paths.append(path)
            wb.close()
            continue

        before = len(merged_emp_data)
        _read_flexible(wb, merged_emp_data)
        wb.close()
        logger.info(
            f'賃金台帳読み取り: {path.name} '
            f'→ 追加/更新 {len(merged_emp_data) - before}名 (累計{len(merged_emp_data)}名)'
        )

    employees = _emp_dict_to_list(merged_emp_data)

    # 個人台帳型ファイルは別途パース（統合が複雑なためファイル単位で結合）
    for path in individual_ledger_paths:
        wb = openpyxl.load_workbook(str(path), data_only=True)
        extra = _read_individual_ledger(wb)
        wb.close()
        logger.info(f'賃金台帳読み取り(個人台帳型): {path.name} → {len(extra)}名')
        for e in extra:
            e.no = len(employees) + 1
            # 個人台帳型は 1ファイル=1〜数名なので、source_file を直接記録
            if not e.source_file:
                e.source_file = path.name
            employees.append(e)

    employees = _dedupe_employees_by_normalized_name(employees)

    # employment_type の補完: provenance を残す形式「正社員(推定)」で埋める。
    # 給与計算で「正社員」「パート」のどちらにもカウントされないと人数 0 になるため、
    # 補完は必須だが、推定値であることを明示することで wage_validator や UI が
    # 「人間チェック対象」として扱える。
    for emp in employees:
        if not (emp.employment_type or '').strip():
            emp.employment_type = '正社員(推定)'

    # データソース（抽出根拠ファイル名）の補完
    # 統合読み込み（_read_flexible / _read_csv）由来の従業員は path 単位で分離できないので
    # 「統合」表記。個人台帳型は _read_individual_ledger 内で既に設定済み。
    _assign_source_files(employees, file_paths)

    logger.info(f'賃金台帳合算結果(決定論): {len(employees)}名 ({len(file_paths)}ファイル)')
    return employees


def _assign_source_files(
    employees: list[WageEmployee],
    file_paths: list[Path],
) -> None:
    """source_file 未設定の従業員に既定値を割り当てる。

    - 1ファイルのみ: そのファイル名
    - 複数ファイル: '統合(Nファイル)' とファイル名先頭3つを併記
    """
    if not employees or not file_paths:
        return
    if len(file_paths) == 1:
        default = file_paths[0].name
    else:
        # 多すぎる場合は先頭3つだけ列挙（セル幅対策）
        names = [p.name for p in file_paths[:3]]
        suffix = f' 他{len(file_paths) - 3}件' if len(file_paths) > 3 else ''
        default = f'統合({len(file_paths)}ファイル): ' + ', '.join(names) + suffix
    for emp in employees:
        if not emp.source_file:
            emp.source_file = default


# ============================================================
# 賃金台帳一覧Excel出力（チェック用）
# ============================================================

def _is_excluded_from_wage_total(emp: WageEmployee) -> bool:
    """給与支給総額（R216）の集計から除外される従業員か判定。

    除外条件（公募要領準拠）:
      - 役員（employment_type に「役員」を含む）
      - 基準年度に全月分の給与支給を受けていない（中途入退社等）
    """
    if '役員' in (emp.employment_type or ''):
        return True
    if not emp.is_full_year:
        return True
    return False


def export_wage_ledger_summary(
    employees: list[WageEmployee],
    output_path: Path,
    company_name: str = '',
    extraction_method: str = '',
) -> Path:
    """
    賃金台帳から読み取ったデータを一覧Excelに出力（チェック用）

    出力内容:
      シート1「賃金台帳一覧」:
        左ブロック  : 月別課税対象額（12か月）+ 年間合計賃金
        右ブロック  : 月別労働時間（12か月）+ 年間合計時間 + 月平均労働時間
        右端       : データソース（抽出根拠の元ファイル名）
      シート2「算定根拠」:
        - 採用列・含む経費・除外経費・役員除外ルール
        - 抽出経路（AI / 決定論）
        - IT導入補助金 2026 通常枠 公募要領 p.10 原文引用
        - データソースとなった賃金台帳ファイル一覧（重複排除）

    集計対象外（役員 or 非全月在籍）の行は薄いグレーで塗り、目視で除外行を判別可能にする。

    引数:
      extraction_method: 抽出経路の説明（'AI抽出（Claude Sonnet 4.6）' / '決定論パーサー' 等）。
                        空文字なら算定根拠シートでは「未指定」表示。
    """
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '賃金台帳一覧'

    # スタイル定義
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    group_fill = PatternFill(start_color='8FAADC', end_color='8FAADC', fill_type='solid')
    excluded_fill = PatternFill(  # 集計対象外行（役員 / 非全月在籍）のグレー塗り
        start_color='E7E6E6', end_color='E7E6E6', fill_type='solid',
    )
    header_font_white = Font(bold=True, size=10, color='FFFFFF')
    number_fmt = '#,##0'
    hours_fmt = '#,##0.0'
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # タイトル行
    title = '賃金台帳 読取データ一覧'
    if company_name:
        title = f'{company_name} — {title}'
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)

    # 抽出経路の明示（AI抽出かどうかを最上部に大きく表示）
    is_ai_extraction = 'AI' in (extraction_method or '')
    extraction_label = extraction_method or '未指定'
    if is_ai_extraction:
        warn_msg = (
            f'⚠ 抽出経路: {extraction_label} '
            '— AI（Claude）による読み取りのため誤読の可能性があります。'
            '賃金台帳原本と必ず照合してください。'
        )
    else:
        warn_msg = f'抽出経路: {extraction_label}'
    cell_warn = ws.cell(row=2, column=1, value=warn_msg)
    cell_warn.font = Font(bold=True, size=10, color='C00000' if is_ai_extraction else '333333')
    if is_ai_extraction:
        # 視認性のため薄オレンジで強調（給与支給総額計算.xlsx の AI 色と統一）
        cell_warn.fill = PatternFill(
            start_color='FCE4D6', end_color='FCE4D6', fill_type='solid',
        )

    ws.cell(
        row=3, column=1,
        value=(
            '※採用列: 賃金台帳の「課税支給合計」（給与所得として課税対象となる経費。'
            '非課税通勤手当・社保等控除前）。'
            '出典: IT導入補助金 2026 通常枠 公募要領 p.10。'
            '抽出経路と公募要領原文の引用は「算定根拠」シートを参照。'
        ),
    )
    ws.cell(row=3, column=1).font = Font(size=9, color='666666')
    ws.cell(
        row=4, column=1,
        value='※グレー行は給与支給総額（R216）の集計対象外 — 役員 or 非全月在籍（中途入退社）',
    )
    ws.cell(row=4, column=1).font = Font(size=9, color='666666')

    # 列レイアウト
    # 1: No, 2: 従業員名, 3: 雇用形態,
    # 4-15: 1月〜12月 賃金, 16: 年間合計賃金,
    # 17-28: 1月〜12月 時間, 29: 年間合計時間, 30: 月平均労働時間,
    # 31: データソース
    wage_start = 4
    wage_total_col = wage_start + 12  # 16
    hours_start = wage_total_col + 1  # 17
    hours_total_col = hours_start + 12  # 29
    avg_hours_col = hours_total_col + 1  # 30
    source_col = avg_hours_col + 1  # 31

    # グループヘッダー（5行目）— 抽出経路メッセージを 2行目に追加したため1行ずれる
    group_row = 5
    ws.cell(row=group_row, column=wage_start, value='月別課税対象額（円）')
    ws.merge_cells(start_row=group_row, start_column=wage_start,
                   end_row=group_row, end_column=wage_total_col)
    ws.cell(row=group_row, column=hours_start, value='月別労働時間')
    ws.merge_cells(start_row=group_row, start_column=hours_start,
                   end_row=group_row, end_column=avg_hours_col)
    for c in (wage_start, hours_start):
        cell = ws.cell(row=group_row, column=c)
        cell.font = header_font_white
        cell.fill = group_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # 列ヘッダー（6行目）
    header_row = 6
    headers = (
        ['No', '従業員名', '雇用形態']
        + MONTH_NAMES + ['年間合計']
        + MONTH_NAMES + ['年間合計', '月平均']
        + ['データソース']
    )
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=c, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # データ行
    for i, emp in enumerate(employees):
        r = header_row + 1 + i
        is_excluded = _is_excluded_from_wage_total(emp)

        ws.cell(row=r, column=1, value=emp.no).border = thin_border
        ws.cell(row=r, column=2, value=emp.name).border = thin_border
        ws.cell(row=r, column=3, value=emp.employment_type).border = thin_border

        annual_wage = 0.0
        for m in range(12):
            cell = ws.cell(row=r, column=wage_start + m)
            cell.border = thin_border
            val = emp.monthly_wages[m]
            if val is not None:
                cell.value = val
                cell.number_format = number_fmt
                annual_wage += val
        wage_total_cell = ws.cell(row=r, column=wage_total_col, value=annual_wage)
        wage_total_cell.number_format = number_fmt
        wage_total_cell.font = Font(bold=True)
        wage_total_cell.border = thin_border

        annual_hours = 0.0
        has_any_hours = False
        for m in range(12):
            cell = ws.cell(row=r, column=hours_start + m)
            cell.border = thin_border
            val = emp.monthly_hours[m] if m < len(emp.monthly_hours) else None
            if val is not None and val > 0:
                cell.value = val
                cell.number_format = hours_fmt
                annual_hours += val
                has_any_hours = True

        # 年間合計時間（月別データが無ければ 月平均×月数 で代用）
        hours_total_cell = ws.cell(row=r, column=hours_total_col)
        hours_total_cell.border = thin_border
        hours_total_cell.number_format = hours_fmt
        hours_total_cell.font = Font(bold=True)
        if has_any_hours:
            hours_total_cell.value = round(annual_hours, 1)
        elif emp.monthly_avg_hours > 0:
            months_with_wage = sum(
                1 for w in emp.monthly_wages if w is not None
            )
            hours_total_cell.value = round(
                emp.monthly_avg_hours * months_with_wage, 1
            )

        avg_hours_cell = ws.cell(row=r, column=avg_hours_col,
                                 value=emp.monthly_avg_hours)
        avg_hours_cell.number_format = hours_fmt
        avg_hours_cell.border = thin_border

        # データソース（抽出根拠の元ファイル名）
        source_cell = ws.cell(row=r, column=source_col, value=emp.source_file)
        source_cell.border = thin_border
        source_cell.font = Font(size=9, color='666666')
        source_cell.alignment = Alignment(horizontal='left', vertical='center')

        # 集計対象外行（役員 or 非全月在籍）はグレーに塗る
        if is_excluded:
            for c in range(1, source_col + 1):
                ws.cell(row=r, column=c).fill = excluded_fill

    # ── 合計行（全員 / 集計対象のみの2段）────────────────────────────
    # 後段の給与支給総額計算や申請書 R216 との突合用。SUM 式で記述しておくと
    # 行追加・編集後も自動再計算されるので、人間チェックの再利用性が上がる。
    if employees:
        from openpyxl.styles import Font as _Font  # 上で import 済みだが明示
        subtotal_fill_all = PatternFill(
            start_color='B4C7E7', end_color='B4C7E7', fill_type='solid',
        )
        subtotal_fill_target = PatternFill(
            start_color='C6E0B4', end_color='C6E0B4', fill_type='solid',
        )
        first_data_row = header_row + 1
        last_data_row = header_row + len(employees)
        total_all_row = last_data_row + 1
        total_target_row = last_data_row + 2

        ws.cell(row=total_all_row, column=2, value='合計（全員）').font = Font(bold=True, size=10)
        ws.cell(row=total_target_row, column=2,
                value='合計（集計対象のみ）').font = Font(bold=True, size=10)

        # 集計対象行のインデックス（行番号は first_data_row 基準）
        target_row_nums = [
            first_data_row + i for i, emp in enumerate(employees)
            if not _is_excluded_from_wage_total(emp)
        ]

        def _set_total_cell(row: int, col: int, formula: str, fmt: str, fill):
            c = ws.cell(row=row, column=col, value=formula)
            c.font = Font(bold=True)
            c.number_format = fmt
            c.fill = fill
            c.border = thin_border

        # 賃金: 1月〜12月 + 年間合計
        for col in list(range(wage_start, wage_start + 12)) + [wage_total_col]:
            col_letter = get_column_letter(col)
            # 全員
            _set_total_cell(
                total_all_row, col,
                f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})',
                number_fmt, subtotal_fill_all,
            )
            # 集計対象のみ（個別セルの加算式）
            if target_row_nums:
                parts = '+'.join(f'{col_letter}{r}' for r in target_row_nums)
                target_formula = f'={parts}'
            else:
                target_formula = 0
            _set_total_cell(
                total_target_row, col, target_formula,
                number_fmt, subtotal_fill_target,
            )

        # 労働時間: 1月〜12月 + 年間合計
        for col in list(range(hours_start, hours_start + 12)) + [hours_total_col]:
            col_letter = get_column_letter(col)
            _set_total_cell(
                total_all_row, col,
                f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})',
                hours_fmt, subtotal_fill_all,
            )
            if target_row_nums:
                parts = '+'.join(f'{col_letter}{r}' for r in target_row_nums)
                target_formula = f'={parts}'
            else:
                target_formula = 0
            _set_total_cell(
                total_target_row, col, target_formula,
                hours_fmt, subtotal_fill_target,
            )

        # 備考: 「集計対象のみ」の年間合計＝R216 母数。ただし R列(年間合計)は月次課税給与のみで、
        # R216（給与支給総額）は「月次年計＋年間賞与」。賞与は月次セルに混ぜない設計（年間賞与は
        # WageEmployee.annual_bonus に隔離）のため、R列だけを R216 母数と書くと申請書 R216
        # （賞与込み）とダブルチェックで賞与額ぶん食い違う。集計対象の年間賞与額と R216 実値を
        # 備考に明示し、申請書 R216 と突合できるようにする。
        wage_total_letter = get_column_letter(wage_total_col)
        included_emps = [e for e in employees if not _is_excluded_from_wage_total(e)]
        target_wage_sum = sum(
            sum(w for w in e.monthly_wages if w is not None) for e in included_emps
        )
        target_bonus_sum = sum(
            (getattr(e, 'annual_bonus', 0.0) or 0.0) for e in included_emps
        )
        if target_bonus_sum > 0:
            note_value = (
                f'※「合計（集計対象のみ）」年間合計（{wage_total_letter}列）'
                f'＝月次課税給与の年計 {target_wage_sum:,.0f}円。'
                f'R216 給与支給総額＝これ＋年間賞与 {target_bonus_sum:,.0f}円'
                f'＝ {target_wage_sum + target_bonus_sum:,.0f}円'
                f'（賞与は月次セルに含めない設計のため別途加算）'
            )
        else:
            note_value = (
                f'※「合計（集計対象のみ）」の年間合計（{wage_total_letter}列）'
                f'＝ R216 給与支給総額の母数 {target_wage_sum:,.0f}円（賞与なし）'
            )
        note_cell = ws.cell(
            row=total_target_row, column=source_col, value=note_value,
        )
        note_cell.font = Font(size=9, color='666666')
        note_cell.fill = subtotal_fill_target
        note_cell.border = thin_border
        note_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 全員行の備考: グレー行（役員・中途）も含む合計である旨を明示
        note_all_cell = ws.cell(
            row=total_all_row, column=source_col,
            value='※役員・中途入退社を含む全行合計（R216 母数には未調整）',
        )
        note_all_cell.font = Font(size=9, color='666666')
        note_all_cell.fill = subtotal_fill_all
        note_all_cell.border = thin_border
        note_all_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # 列幅調整
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 14
    for c in range(wage_start, avg_hours_col + 1):
        ws.column_dimensions[get_column_letter(c)].width = 11
    ws.column_dimensions[get_column_letter(wage_total_col)].width = 13
    ws.column_dimensions[get_column_letter(hours_total_col)].width = 13
    ws.column_dimensions[get_column_letter(avg_hours_col)].width = 11
    ws.column_dimensions[get_column_letter(source_col)].width = 40

    # 算定根拠シート（採用列・除外ルール・公募要領原文・データソース一覧）
    _write_calculation_basis_sheet(
        wb,
        employees=employees,
        extraction_method=extraction_method,
        thin_border=thin_border,
    )

    wb.save(str(output_path))
    wb.close()
    logger.info(f'賃金台帳一覧出力: {output_path} ({len(employees)}名)')
    return output_path


def _write_calculation_basis_sheet(
    wb,
    employees: list[WageEmployee],
    extraction_method: str,
    thin_border,
) -> None:
    """『算定根拠』シートを追加。R216 算定の根拠条文・採用列・データソースを記録。"""
    from openpyxl.styles import Font, Alignment, PatternFill

    ws = wb.create_sheet(title='算定根拠')

    title_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=10, color='FFFFFF')
    section_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    label_font = Font(bold=True, size=10)
    quote_font = Font(size=10, italic=True, color='333333')
    note_font = Font(size=9, color='666666')

    ws.cell(row=1, column=1, value='給与支給総額（R216）の算定根拠').font = title_font
    ws.cell(row=2, column=1, value='IT導入補助金 2026 通常枠 公募要領 準拠').font = note_font

    rows: list[tuple[str, str]] = [
        ('対象補助金', 'IT導入補助金 2026（デジタル化・AI導入補助金2026、通常枠／インボイス対応類型）'),
        ('採用列', '賃金台帳の「課税支給合計」（給与所得として課税対象となる経費）'),
        (
            '含む経費',
            '給料・賃金・賞与・各種手当（残業手当／休日出勤手当／'
            '職務手当／地域手当／家族(扶養)手当／住宅手当）等',
        ),
        (
            '含まない経費',
            '福利厚生費・法定福利費・退職金・'
            '非課税通勤手当（限度額内分、国税庁 No.2585 により給与所得に含まれない）',
        ),
        ('役員の扱い', '集計対象外（IT導入補助金 2026 通常枠 公募要領 p.10「役員報酬…は除く」）。ただし従業員0名の法人のみ役員で読み替え可'),
        (
            '中途入退社者の扱い',
            '集計対象外（基準年度に全月分の給与等の支給を受けていない従業員）',
        ),
        ('賞与の合算', '月別の課税給与に合算（賞与シートが別ファイルの場合は対応月に加算）'),
        ('抽出経路', extraction_method or '未指定'),
        (
            '決定論パーサーの列優先順',
            '① 課税支給合計 → ② 支給合計 → ③ 差引支給合計（手取り）',
        ),
    ]

    start_row = 4
    ws.cell(row=start_row, column=1, value='項目').font = section_font
    ws.cell(row=start_row, column=1).fill = section_fill
    ws.cell(row=start_row, column=2, value='値・説明').font = section_font
    ws.cell(row=start_row, column=2).fill = section_fill
    ws.cell(row=start_row, column=1).border = thin_border
    ws.cell(row=start_row, column=2).border = thin_border

    for i, (label, value) in enumerate(rows, 1):
        r = start_row + i
        c1 = ws.cell(row=r, column=1, value=label)
        c1.font = label_font
        c1.alignment = Alignment(vertical='top', wrap_text=True)
        c1.border = thin_border
        c2 = ws.cell(row=r, column=2, value=value)
        c2.alignment = Alignment(vertical='top', wrap_text=True)
        c2.border = thin_border

    # IT導入補助金 2026 通常枠 公募要領 p.10 原文
    quote_row = start_row + len(rows) + 3
    ws.cell(row=quote_row, column=1, value='IT導入補助金 2026 通常枠 公募要領 p.10 原文').font = label_font
    ws.merge_cells(start_row=quote_row + 1, start_column=1,
                   end_row=quote_row + 1, end_column=2)
    quote_cell = ws.cell(
        row=quote_row + 1, column=1,
        value=(
            '算定対象となる給与等は、給料、賃金、賞与、各種手当'
            '（残業手当、休日出勤手当、職務手当、地域手当、家族（扶養）手当、住宅手当）等、'
            '給与所得として課税対象となる経費を指す。'
            '役員報酬、福利厚生費、法定福利費や退職金は除く。'
        ),
    )
    quote_cell.font = quote_font
    quote_cell.alignment = Alignment(vertical='top', wrap_text=True)
    ws.row_dimensions[quote_row + 1].height = 60

    # 公式 URL
    url_row = quote_row + 3
    ws.cell(row=url_row, column=1, value='IT導入補助金 2026 通常枠 公募要領 PDF').font = label_font
    ws.cell(
        row=url_row, column=2,
        value='https://it-shien.smrj.go.jp/pdf/it2026_koubo_tsujyo.pdf',
    )
    ws.cell(row=url_row + 1, column=1, value='IT導入補助金 2026 交付申請マニュアル PDF').font = label_font
    ws.cell(
        row=url_row + 1, column=2,
        value='https://it-shien.smrj.go.jp/pdf/it2026_manual_kofu.pdf',
    )
    ws.cell(row=url_row + 2, column=1, value='通勤手当 非課税限度額').font = label_font
    ws.cell(
        row=url_row + 2, column=2,
        value='https://www.nta.go.jp/taxes/shiraberu/taxanswer/gensen/2585.htm（国税庁 No.2585）',
    )

    # データソース一覧（emp.source_file から重複排除）
    source_files: list[str] = []
    seen: set[str] = set()
    for emp in employees:
        src = (emp.source_file or '').strip()
        if src and src not in seen:
            seen.add(src)
            source_files.append(src)

    src_row = url_row + 4
    ws.cell(
        row=src_row, column=1,
        value=f'参照した賃金台帳ファイル（{len(source_files)}件）',
    ).font = label_font
    if not source_files:
        ws.cell(row=src_row + 1, column=2, value='（ファイル情報なし）').font = note_font
    else:
        for i, src in enumerate(source_files):
            ws.cell(row=src_row + 1 + i, column=2, value=src).font = note_font

    # 列幅
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 90


# ============================================================
# 加点措置判定
# ============================================================

def _parse_app_month(v) -> tuple[int, int] | None:
    """交付申請月セルの値を (西暦年, 月) に正規化する。datetime / 'yyyy/mm' / 和暦に対応。"""
    import datetime as _dt
    if isinstance(v, (_dt.datetime, _dt.date)):
        return (v.year, v.month)
    parsed = parse_ym_header(v)
    if parsed and parsed[0] is not None:
        return (parsed[0], parsed[1])
    return None


def read_bonus_wage_ledger(path: Path) -> BonusWageLedger:
    """加点判定用賃金台帳（専用テンプレ）を決定論で直読みして BonusWageLedger を返す。

    AI 再抽出はしない。月列は暦月固定（F〜Q=令和6年10月〜令和7年9月、R=交付申請直近月）。
    時間換算給与は持たず、基本給と所定労働時間から judge_bonus_points が算出する。
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb[BWL_SHEET_NAME] if BWL_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    ledger = BonusWageLedger()
    pref = ws.cell(*BWL_PREF_CELL).value
    ledger.prefecture = str(pref).strip() if pref else ''
    ledger.application_ym = _parse_app_month(ws.cell(*BWL_APPYM_CELL).value)
    latest_ym = ledger.latest_ym

    for r in range(BWL_DATA_START_ROW, ws.max_row + 1):
        name_val = ws.cell(r, BWL_COL_NAME).value
        if name_val is None or not str(name_val).strip():
            continue
        name = str(name_val).replace('　', ' ').strip()
        no_val = ws.cell(r, BWL_COL_NO).value
        try:
            no = int(no_val) if no_val is not None else (r - BWL_DATA_START_ROW + 1)
        except (TypeError, ValueError):
            no = r - BWL_DATA_START_ROW + 1
        emp_type = str(ws.cell(r, BWL_COL_EMPTYPE).value or '').strip()
        emp = BonusWageEmployee(
            no=no, name=name, employment_type=emp_type,
            scheduled_hours=_to_float(ws.cell(r, BWL_COL_HOURS).value),
            source_file=path.name,
        )
        for i, ym in enumerate(BONUS1_WINDOW):
            v = _to_float(ws.cell(r, BWL_COL_WINDOW_START + i).value)
            if v is not None and v > 0:
                emp.monthly_base[ym] = v
        if latest_ym is not None:
            v = _to_float(ws.cell(r, BWL_COL_LATEST).value)
            if v is not None and v > 0:
                emp.monthly_base[latest_ym] = v
        ledger.employees.append(emp)
    wb.close()

    if not ledger.prefecture:
        ledger.notes.append('台帳に事業場所在地（都道府県）が入力されていません（C2セル）。')
    if ledger.application_ym is None:
        ledger.notes.append(
            '台帳に交付申請月（yyyy/mm）が入力されていません（C3セル）。加点措置②は判定できません。'
        )
    if not ledger.employees:
        ledger.notes.append('台帳から従業員データを読み取れませんでした。氏名列（C列）を確認してください。')
    return ledger


def is_bonus_wage_ledger(path: Path) -> bool:
    """ファイルが加点判定用賃金台帳（専用シート『加点判定用明細』を持つ）か判定する。"""
    if path.suffix.lower() not in ('.xlsx', '.xlsm'):
        return False
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True)
        ok = BWL_SHEET_NAME in wb.sheetnames
        wb.close()
        return ok
    except Exception:
        return False


def judge_bonus_points(ledger: BonusWageLedger) -> BonusPointResult:
    """加点措置①②を (年,月) 固定の暦月で判定する。

    加点①（補助率1/2→2/3 トリガー兼用／加点項目14）:
      令和6年10月〜令和7年9月の各暦月で「時間換算給与 < R7改定後地域別最賃」の
      従業員が（役員を除く）全従業員の30%以上である月が3か月以上 → 対象。
    加点②（加点項目15）:
      令和7年7月 と 交付申請直近月（申請月の前月）の事業場内最低賃金（時間換算給与の
      最小値）を比較し、差が63円以上 → 対象。
    時間換算給与 = 基本給 ÷ 所定労働時間（残業・通勤・各種手当を除く）。役員は母数から除外。
    公式シート①の判定式 `IF(AND(時間換算給与 < R7改定後, 時間換算給与 > 0),"対象")` に準拠
    （下限は >0 のみ。R6改定前は判定に用いない）。
    """
    prefecture = ledger.prefecture
    result = BonusPointResult(
        prefecture=prefecture,
        min_wage_r6=MIN_WAGE_R6.get(prefecture, 0),
        min_wage_r7=MIN_WAGE_MAP.get(prefecture, 0),
        application_ym=ledger.application_ym,
        latest_ym=ledger.latest_ym,
        notes=list(ledger.notes),
    )

    if not result.min_wage_r7:
        result.notes.append(
            f'最低賃金マスタに「{prefecture}」が見つかりません。都道府県を確認してください。'
        )
        logger.warning(f'最低賃金が見つかりません: {prefecture}')
        return result

    mw_r7 = result.min_wage_r7
    # 判定母数は労働者のみ（役員報酬は最賃規制の対象外。R215/R216 と同じ役員定義）。
    workers = [e for e in ledger.employees if not e.is_officer]

    # ── 加点措置① ──
    months_met: list[tuple[int, int]] = []
    for ym in BONUS1_WINDOW:
        total = 0
        under = 0
        emps_detail: list[dict] = []
        for emp in workers:
            hourly = emp.hourly_for(ym)
            if hourly is None or hourly <= 0:
                continue
            # 公式シートの時間換算給与列は整数（円）。事務局の判定式 IF(I<G) はその整数で
            # 再計算されるため、画面判定とExcelを一致させるべく丸めた整数で比較する。
            hourly = round(hourly)
            total += 1
            is_target = hourly < mw_r7  # >0 は上の continue で担保
            if is_target:
                under += 1
            emps_detail.append({
                'name': emp.name,
                'base': emp.base_for(ym),
                'hourly': hourly,
                'is_target': is_target,
            })
        ratio = (under / total) if total > 0 else 0.0
        meets = total > 0 and ratio >= 0.30
        if meets:
            months_met.append(ym)
        result.bonus1_details.append({
            'ym': ym,
            'label': ym_label(ym),
            'total': total,
            'under_r7': under,
            'ratio': ratio,
            'meets_30pct': meets,
            'has_data': total > 0,
            'employees': emps_detail,
        })
        if total == 0:
            result.notes.append(
                f'加点①対象月 {ym_label(ym)} の時間換算給与データがありません'
                '（基本給/所定労働時間の欠落）。台帳の対象月入力を確認してください。'
            )

    result.bonus1_months_met = months_met
    result.bonus1_eligible = len(months_met) >= 3
    logger.info(
        f'加点措置①: {len(months_met)}か月が30%以上 '
        f'→ {"対象" if result.bonus1_eligible else "対象外"}'
    )

    # ── 加点措置② ──
    def _period_detail(ym: tuple[int, int]) -> tuple[dict, list[float]]:
        emps: list[dict] = []
        hourlies: list[float] = []
        for emp in workers:
            hourly = emp.hourly_for(ym)
            if hourly is None or hourly <= 0:
                continue
            hourly = round(hourly)  # 公式②シートの MIN は整数の時間換算給与で再計算される
            hourlies.append(hourly)
            emps.append({
                'name': emp.name,
                'base': emp.base_for(ym),
                'hourly': hourly,
            })
        min_wage = min(hourlies) if hourlies else 0.0
        return ({'ym': ym, 'label': ym_label(ym),
                 'min_wage': min_wage, 'employees': emps}, hourlies)

    july_detail, july_h = _period_detail(BONUS2_BASE_YM)
    result.bonus2_july_detail = july_detail
    result.bonus2_min_wage_july = july_detail['min_wage']

    latest_ym = ledger.latest_ym
    latest_h: list[float] = []
    if latest_ym is None:
        result.notes.append('交付申請月が未入力のため、加点措置②の直近月を確定できません。')
    else:
        latest_detail, latest_h = _period_detail(latest_ym)
        result.bonus2_latest_detail = latest_detail
        result.bonus2_min_wage_latest = latest_detail['min_wage']

    if july_h and latest_h:
        result.bonus2_diff = result.bonus2_min_wage_latest - result.bonus2_min_wage_july
        result.bonus2_eligible = result.bonus2_diff >= BONUS_THRESHOLD_YEN
    else:
        if not july_h:
            result.notes.append(
                f'加点②基準月 {ym_label(BONUS2_BASE_YM)} の時間換算給与データがありません。'
            )
        if latest_ym is not None and not latest_h:
            result.notes.append(
                f'加点②直近月 {ym_label(latest_ym)} の時間換算給与データがありません。'
            )

    logger.info(
        f'加点措置②: {ym_label(BONUS2_BASE_YM)}={result.bonus2_min_wage_july:.0f}円 → '
        f'直近={result.bonus2_min_wage_latest:.0f}円 '
        f'(差額{result.bonus2_diff:.0f}円) '
        f'→ {"対象" if result.bonus2_eligible else "対象外"}'
    )
    return result


# ============================================================
# 加点措置シートへの自動入力
# ============================================================

def _bw_set(ws, row: int, col: int, value) -> None:
    """結合セルでも安全に値を書く（結合範囲の左上に書き込む）。数式セルは呼び出さない前提。"""
    from openpyxl.cell.cell import MergedCell
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
                ws.cell(row=rng.min_row, column=rng.min_col, value=value)
                return
        return
    cell.value = value


def _detect_bonus_data_start(ws, header_col: int = 2, default: int = 18) -> int:
    """データ開始行を動的検出する。

    公式シートは「No」見出し行 → 「例」サンプル行 → データ行 の順。
    ①用=見出し16/データ18、補助率引上げ①用=見出し17/データ19、②用=見出し15/データ17。
    いずれも『見出し行 + 2』がデータ開始行（事務局の COUNTIF/MIN もその行から集計）。
    """
    for r in range(1, 40):
        v = ws.cell(row=r, column=header_col).value
        if isinstance(v, str) and v.strip() == 'No':
            return r + 2
    return default


def _select_bonus1_periods(result: BonusPointResult, n: int = 3) -> list[dict]:
    """加点①で公式シートに出す賃金計算期間（最大n=3か月）を選ぶ。

    30%以上を達成した月を優先し、比率が高い順。3か月に満たない場合は、データのある
    月のうち比率が高い順で補う（事務局が「最も要件に近い月」を確認できるようにする）。
    """
    candidates = [d for d in result.bonus1_details if d['has_data']]
    candidates.sort(key=lambda d: (d['meets_30pct'], d['ratio']), reverse=True)
    return candidates[:n]


def fill_bonus_sheet_1(
    template_path: Path,
    output_path: Path,
    result: BonusPointResult,
    selected_periods: list[dict] | None = None,
) -> Path:
    """加点措置①用（または補助率引上げ・加点措置①用）シートに従業員データを入力。

    シート構成: 3つの賃金計算期間を横並び（期間① B-K / ② M-U / ③ W-AE）。
    各期間に H/R/AB=基本給、I/S/AC=時間換算給与 を書く。判定式・VLOOKUP・COUNTIF は温存。
    データ開始行はテンプレ別に動的検出（①用=18 / 補助率引上げ①用=19）。
    """
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb[wb.sheetnames[0]]

    if selected_periods is None:
        selected_periods = _select_bonus1_periods(result, 3)

    # base=基本給(H/R/AB列), hourly=時間換算給与(I/S/AC列)
    period_cols = [
        {'no': 2, 'name': 3, 'pref': 4, 'base': 8, 'hourly': 9},
        {'no': 13, 'name': 14, 'pref': 15, 'base': 18, 'hourly': 19},
        {'no': 23, 'name': 24, 'pref': 25, 'base': 28, 'hourly': 29},
    ]
    data_start = _detect_bonus_data_start(ws, header_col=2, default=18)

    for period_idx, detail in enumerate(selected_periods[:3]):
        cols = period_cols[period_idx]
        for i, emp in enumerate(detail.get('employees', [])):
            row = data_start + i
            _bw_set(ws, row, cols['no'], i + 1)
            _bw_set(ws, row, cols['name'], emp['name'])
            _bw_set(ws, row, cols['pref'], result.prefecture)
            if emp.get('base') is not None:
                _bw_set(ws, row, cols['base'], round(emp['base']))
            _bw_set(ws, row, cols['hourly'], round(emp['hourly']))

    wb.save(str(output_path))
    wb.close()
    logger.info(f'加点措置①シート保存: {output_path}')
    return output_path


def fill_bonus_sheet_2(
    template_path: Path,
    output_path: Path,
    result: BonusPointResult,
) -> Path:
    """加点措置②用シートに従業員データを入力。

    シート構成: 期間①（令和7年7月・F14固定）と期間②（交付申請直近月・N14=EDATE(D5,-1)）を
    横並び。F/N=基本給, G/O=時間換算給与 を書く。事業場内最低賃金 D7/D8 は時間換算給与列の
    MIN 配列式・D10 の判定式で自動算出されるため温存。
    交付申請月を D5 に日付で書き込み、N14=EDATE(D5,-1) の直近月ラベルを駆動する。
    """
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb[wb.sheetnames[0]]

    # base=基本給(F/N列), hourly=時間換算給与(G/O列)
    period_cols = [
        {'no': 2, 'name': 3, 'pref': 4, 'base': 6, 'hourly': 7},
        {'no': 10, 'name': 11, 'pref': 12, 'base': 14, 'hourly': 15},
    ]
    data_start = _detect_bonus_data_start(ws, header_col=2, default=17)

    # D5（申請月）を日付で設定 → N14=EDATE(D5,-1) が直近月を表示・D7/D8 の集計対象を駆動
    if result.application_ym:
        import datetime as _dt
        _bw_set(ws, 5, 4, _dt.datetime(result.application_ym[0], result.application_ym[1], 1))

    details = [result.bonus2_july_detail, result.bonus2_latest_detail]
    for period_idx, detail in enumerate(details):
        if not detail:
            continue
        cols = period_cols[period_idx]
        for i, emp in enumerate(detail.get('employees', [])):
            row = data_start + i
            _bw_set(ws, row, cols['no'], i + 1)
            _bw_set(ws, row, cols['name'], emp['name'])
            _bw_set(ws, row, cols['pref'], result.prefecture)
            if emp.get('base') is not None:
                _bw_set(ws, row, cols['base'], round(emp['base']))
            _bw_set(ws, row, cols['hourly'], round(emp['hourly']))

    wb.save(str(output_path))
    wb.close()
    logger.info(f'加点措置②シート保存: {output_path}')
    return output_path
