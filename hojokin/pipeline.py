# -*- coding: utf-8 -*-
"""処理パイプライン: ファイル検出 → 抽出 → 転記 → 出力"""
from __future__ import annotations

import logging
import re
import tempfile
import unicodedata
from pathlib import Path

from .config import get_mapping, CLAUDE_API_KEY, BIZ_DESC
from .models import ExtractionResult, ProcessingStatus, CompanyInfo
from .ai_extractor import create_extractor, BaseExtractor, StubExtractor
from .hearing_reader import read_hearing_sheet
from .template_filler import fill_template
from .wage_calculator import (
    create_wage_calculation,
    PayrollEmployee,
    calculate_per_capita_wage,
    is_full_time_employment,
    is_zero_wage_detail,
    wage_employees_to_payroll,
)
from .wage_reader import read_wage_ledger, read_wage_ledgers, export_wage_ledger_summary
from .pdf_reader import pdf_to_images

logger = logging.getLogger(__name__)


# ───────────────────────── ファイル名年月パース ─────────────────────────
# 決算書ファイル名から「期末年月 (year, month)」を取り出すユーティリティ。
# 例:
#   令和7年3月決算書        → (2025, 3)
#   令和6年3月決算書        → (2024, 3)
#   R7.3決算書              → (2025, 3)
#   2025年3月期決算書        → (2025, 3)
#   2025.03_決算書          → (2025, 3)
#   平成31年4月決算書        → (2019, 4)
# 取れなければ None。年月の妥当性チェック付き（年 1900-2100、月 1-12）。

_REIWA_RE = re.compile(r'令和\s*(\d{1,2})\s*年\s*(\d{1,2})\s*月')
_REIWA_GANNEN_RE = re.compile(r'令和\s*元\s*年\s*(\d{1,2})\s*月')
_HEISEI_RE = re.compile(r'平成\s*(\d{1,2})\s*年\s*(\d{1,2})\s*月')
_HEISEI_GANNEN_RE = re.compile(r'平成\s*元\s*年\s*(\d{1,2})\s*月')
_RY_DOT_RE = re.compile(r'(?<![A-Za-z0-9])R\s*(\d{1,2})\s*[\.\-_／/]\s*(\d{1,2})(?!\d)')
_YYYY_KANJI_RE = re.compile(r'(\d{4})\s*年\s*(\d{1,2})\s*月')
_YYYY_SEP_RE = re.compile(r'(?<!\d)(\d{4})[\.\-_／/](\d{1,2})(?!\d)')


# 月情報なしファイル名フォールバック用パターン（fiscal_month_override 必須時のみ動作）
# 「令和6年決算書」「R6年決算書」「2024年決算書」のように年号のみ含まれるケースを救済
_REIWA_GANNEN_NOMONTH_RE = re.compile(r'令和\s*元\s*年(?!\s*\d+\s*月)')
_REIWA_NOMONTH_RE = re.compile(r'令和\s*(\d{1,2})\s*年(?!\s*\d+\s*月)')
_HEISEI_NOMONTH_RE = re.compile(r'平成\s*(\d{1,2})\s*年(?!\s*\d+\s*月)')
_RY_NOMONTH_RE = re.compile(r'(?<![A-Za-z0-9])R\s*(\d{1,2})\s*年(?!\s*\d+\s*月)')
_YYYY_NOMONTH_RE = re.compile(r'(?<!\d)(\d{4})\s*年(?!\s*\d+\s*月)')


def _parse_fiscal_end_from_filename(
    name: str,
    fiscal_month_override: int | None = None,
) -> tuple[int, int] | None:
    """ファイル名から期末年月 (year, month) を取り出す。失敗時 None。

    NFC 正規化してから複数のパターンを試行し、最初に成立したものを返す。

    fiscal_month_override が指定されている場合、月情報の無いファイル名
    （「R6年決算書」「令和6年決算書」「2024年決算書」など）も年単独で救済する。
    その際の月は fiscal_month_override を組み合わせて返す（同一会社内の決算書を
    年号で並べ替える比較目的の擬似値）。期首年・期末年どちらで命名するかは会社
    依存だが、同一会社内の相対順序（より新しい決算書）の判定には影響しない。
    """
    s = unicodedata.normalize('NFC', name)

    def _valid(y: int, m: int) -> tuple[int, int] | None:
        if 1900 <= y <= 2100 and 1 <= m <= 12:
            return (y, m)
        return None

    # 期間レンジ「RX.M〜RY.N」「YYYY/M〜YYYY/N」形式は末尾(=期末)を採る。
    # 単一日付パターンは先頭の日付にマッチして期首を返してしまうため、
    # レンジ判定を最優先する（例:「決算書37期（R6.9〜R7.8）」→ 期末2025-08）。
    period = _parse_wage_ledger_period(s)
    if period is not None:
        return _valid(*period[1])

    # 令和元年 (=2019)
    m = _REIWA_GANNEN_RE.search(s)
    if m:
        return _valid(2019, int(m.group(1)))
    # 令和N年M月 (令和N = 2018 + N)
    m = _REIWA_RE.search(s)
    if m:
        return _valid(2018 + int(m.group(1)), int(m.group(2)))
    # 平成元年 (=1989)
    m = _HEISEI_GANNEN_RE.search(s)
    if m:
        return _valid(1989, int(m.group(1)))
    # 平成N年M月 (平成N = 1988 + N)
    m = _HEISEI_RE.search(s)
    if m:
        return _valid(1988 + int(m.group(1)), int(m.group(2)))
    # RN.M / RN-M / RN_M (略式の令和)
    m = _RY_DOT_RE.search(s)
    if m:
        return _valid(2018 + int(m.group(1)), int(m.group(2)))
    # YYYY年M月
    m = _YYYY_KANJI_RE.search(s)
    if m:
        return _valid(int(m.group(1)), int(m.group(2)))
    # YYYY-MM / YYYY.MM / YYYY_MM
    m = _YYYY_SEP_RE.search(s)
    if m:
        return _valid(int(m.group(1)), int(m.group(2)))

    # ---- フォールバック: 月情報なしファイル名 (fiscal_month_override 必須) ----
    if fiscal_month_override is not None:
        m = _REIWA_GANNEN_NOMONTH_RE.search(s)
        if m:
            return _valid(2019, fiscal_month_override)
        m = _REIWA_NOMONTH_RE.search(s)
        if m:
            return _valid(2018 + int(m.group(1)), fiscal_month_override)
        m = _HEISEI_NOMONTH_RE.search(s)
        if m:
            return _valid(1988 + int(m.group(1)), fiscal_month_override)
        m = _RY_NOMONTH_RE.search(s)
        if m:
            return _valid(2018 + int(m.group(1)), fiscal_month_override)
        m = _YYYY_NOMONTH_RE.search(s)
        if m:
            return _valid(int(m.group(1)), fiscal_month_override)

    return None


def _parse_fiscal_year_from_filename(name: str) -> int | None:
    """ファイル名から決算期の「西暦年」だけを取り出す（順位付け専用）。失敗時 None。

    `_parse_fiscal_end_from_filename` は月情報が無いと（fiscal_month_override 未指定時に）
    None を返すが、同一会社内で「どちらがより新しい期か」を決めるには年だけで足りる。
    「令和N年度」「RN年」「YYYY年度」のような月なしファイル名でも年を抽出し、
    決算書ファイルの自動選択・UI 表示が前年度を誤選択しないようにする。
    """
    s = unicodedata.normalize('NFC', name)

    def _valid_year(y: int) -> int | None:
        return y if 1900 <= y <= 2100 else None

    # 月ありで取れるならその年（期末年）を採用
    full = _parse_fiscal_end_from_filename(s)
    if full is not None:
        return full[0]

    # 月なしフォールバック（fiscal_month_override 不要・年のみ）
    if _REIWA_GANNEN_NOMONTH_RE.search(s):
        return _valid_year(2019)
    m = _REIWA_NOMONTH_RE.search(s)
    if m:
        return _valid_year(2018 + int(m.group(1)))
    m = _HEISEI_NOMONTH_RE.search(s)
    if m:
        return _valid_year(1988 + int(m.group(1)))
    m = _RY_NOMONTH_RE.search(s)
    if m:
        return _valid_year(2018 + int(m.group(1)))
    m = _YYYY_NOMONTH_RE.search(s)
    if m:
        return _valid_year(int(m.group(1)))
    return None


_WAGE_PERIOD_REIWA_RE = re.compile(
    r'R\s*(\d{1,2})\s*[\.\-_／/]\s*(\d{1,2})\s*[-〜~～ー]\s*R\s*(\d{1,2})\s*[\.\-_／/]\s*(\d{1,2})'
)
_WAGE_PERIOD_YYYY_RE = re.compile(
    r'(\d{4})\s*[\.\-_／/年]\s*(\d{1,2})\s*月?\s*[-〜~～ー]\s*(\d{4})\s*[\.\-_／/年]\s*(\d{1,2})'
)


def _parse_wage_ledger_period(name: str) -> tuple[tuple[int, int], tuple[int, int]] | None:
    """賃金台帳のファイル名から (期首年月, 期末年月) を取り出す。失敗時 None。

    例:
      R6.4-R7.3賃金台帳    → ((2024, 4), (2025, 3))
      2024-04～2025-03賃金 → ((2024, 4), (2025, 3))
    """
    s = unicodedata.normalize('NFC', name)

    def _valid(y: int, mo: int) -> bool:
        return 1900 <= y <= 2100 and 1 <= mo <= 12

    m = _WAGE_PERIOD_REIWA_RE.search(s)
    if m:
        sy = 2018 + int(m.group(1))
        sm = int(m.group(2))
        ey = 2018 + int(m.group(3))
        em = int(m.group(4))
        if _valid(sy, sm) and _valid(ey, em):
            return ((sy, sm), (ey, em))
    m = _WAGE_PERIOD_YYYY_RE.search(s)
    if m:
        sy = int(m.group(1))
        sm = int(m.group(2))
        ey = int(m.group(3))
        em = int(m.group(4))
        if _valid(sy, sm) and _valid(ey, em):
            return ((sy, sm), (ey, em))
    return None


def _parse_year_month_from_iso(s: str) -> tuple[int, int] | None:
    """'2025-03' / '2025-03-31' / '2025/03' → (2025, 3)。失敗時 None。"""
    if not s:
        return None
    s = s.strip()
    m = re.match(r'^(\d{4})[\-\./](\d{1,2})', s)
    if not m:
        return None
    y, mo = int(m.group(1)), int(m.group(2))
    if 1900 <= y <= 2100 and 1 <= mo <= 12:
        return (y, mo)
    return None


def _record_pl_selection(
    status,
    detector: 'FileDetector',
    pl_path: Path | None,
    fiscal_month_override: int | None,
) -> None:
    """選定された決算書ファイル名・推定期末年月・選定警告を ProcessingStatus に記録。

    UI の処理開始ボタン直前 / 処理結果カードで「📄 直近年度として『...』を使用」
    と明示するためのデータソース。ファイル名から年月が抽出できない場合は
    pl_selected_end は空のままにし、UI 側で「期末: 不明（AI推定値で補完）」のように
    表示するか、財務AI抽出後に上書きする。
    """
    if pl_path is None:
        return
    status.pl_selected_filename = pl_path.name
    ym = _parse_fiscal_end_from_filename(
        pl_path.name, fiscal_month_override=fiscal_month_override
    )
    if ym is not None:
        status.pl_selected_end = f'{ym[0]:04d}-{ym[1]:02d}'
    if detector.pl_selection_warnings:
        status.pl_selection_warnings = list(detector.pl_selection_warnings)


def _check_pl_wage_period_consistency(
    detector: 'FileDetector',
    financial,  # FinancialData | None
    fiscal_month_override: int | None,
) -> tuple[object, str]:
    """賃金台帳ファイル名の期間と PL 期末年月の整合性をチェック。

    ズレを検出したら financial の数値情報を空に戻して「決算書由来の値を出力に
    乗せない」+ 強警告メッセージを返す。給与支給総額の本算出は賃金台帳ベースで
    続行されるが、テンプレ転記用の財務値はスキップされる。

    Returns:
        (financial_or_reset, warning_message)
        - financial_or_reset: 整合OK時はそのまま。NG時は revenue=0 にリセット
        - warning_message: 空 or 強警告（status.message 末尾に追加する想定）
    """
    if financial is None:
        return financial, ''

    pl_end = _parse_year_month_from_iso(getattr(financial, 'fiscal_year_end', '') or '')

    # 賃金台帳ファイル名から期末を集める
    ledger_paths = detector.get_all('wage_ledger')
    wage_ends: list[tuple[int, int]] = []
    for p in ledger_paths:
        period = _parse_wage_ledger_period(p.name)
        if period is not None:
            wage_ends.append(period[1])

    msgs: list[str] = []

    # ── 賃金台帳期末 vs PL期末 ──
    if wage_ends and pl_end is not None:
        # 賃金台帳が複数あっても、通常は同一期末年月のはず。バラついていたら最頻値
        # を採用（端末ばらつき疑い）。ここでは最も新しいものを採用して比較
        latest_wage_end = max(wage_ends)
        if latest_wage_end != pl_end:
            msgs.append(
                f' ⛔ 致命的不整合: 賃金台帳期末={latest_wage_end[0]:04d}-{latest_wage_end[1]:02d} ／ '
                f'決算書期末={pl_end[0]:04d}-{pl_end[1]:02d}。'
                f'別の期の決算書が読み込まれた可能性が高いため、決算書由来の'
                f'財務値（売上・粗利・営業利益等）の転記をスキップしました。'
                f'資料フォルダから前期の決算書を退避してから再実行してください。'
            )

    # ── 賃金台帳期末 vs ユーザー指定の決算月 ──
    if wage_ends and fiscal_month_override is not None:
        latest_wage_end = max(wage_ends)
        if latest_wage_end[1] != fiscal_month_override:
            msgs.append(
                f' ⛔ 致命的不整合: 賃金台帳期末月={latest_wage_end[1]}月 ／ '
                f'ユーザー指定決算月={fiscal_month_override}月。'
                f'賃金台帳の対象期間がユーザー指定と矛盾しています。'
                f'決算月の設定または賃金台帳の差し替えを確認してください。'
            )

    # ── PL期末 vs ユーザー指定の決算月 ──
    if pl_end is not None and fiscal_month_override is not None and not wage_ends:
        # 賃金台帳ファイル名から期間が取れなかった時のフォールバック
        if pl_end[1] != fiscal_month_override:
            msgs.append(
                f' ⚠ 警告: 決算書期末月={pl_end[1]}月 ／ '
                f'ユーザー指定決算月={fiscal_month_override}月で不一致。'
                f'別の期の決算書が読み込まれた可能性があります。'
            )

    warning = ''.join(msgs)
    if warning and '致命的不整合' in warning:
        # 致命的不整合: financial を「revenue=0」状態にリセットして、テンプレ転記の
        # 起点である「PL値の出力」を全部止める。給与支給総額は賃金台帳から再算出されるが
        # 売上・粗利・営業利益・経常利益・減価償却費・役員報酬の決算書由来値は出ない。
        from .models import FinancialData
        reset = FinancialData()
        # 賃金台帳から取れる情報（人件費周辺）は維持したいが、PL値はリセット
        logger.warning(f'PL/賃金台帳整合性チェック失敗: {warning}')
        return reset, warning

    if warning:
        logger.warning(f'PL/賃金台帳整合性チェック警告: {warning}')
    return financial, warning


def _pick_full_version(candidates: list[Path]) -> Path:
    """同一期PDF候補から「フル版」を選ぶ。

    PDFsam 等の部分抜粋を除外してサイズ最大を採用。
    全候補が抜粋マーカー付きなら、消去法でサイズ最大を採用。
    """
    if len(candidates) == 1:
        return candidates[0]
    EXCLUDE_MARKERS = ('_PDFsam_', 'PDFsam', '部分抜粋', '抜粋')
    full_versions = [
        p for p in candidates
        if not any(m in p.name for m in EXCLUDE_MARKERS)
    ]
    pool = full_versions or candidates
    return max(pool, key=lambda p: p.stat().st_size)


_PL_PERIOD_RE = re.compile(r'第(\d+)期')


def _rank_pl_names(names, fiscal_month_override=None) -> dict:
    """ファイル名のみで PL 候補を順位付けし、最上位タイ候補と根拠を返す（純関数・副作用なし）。

    get_pl_latest（実選択）と app._auto_selected_for_display（UI 表示）が共有する
    単一の真実。サイズ/mtime に依存する最終同点崩しはここでは行わず、呼び出し側に委ねる。
    これにより「UI が具体ファイル名を出すなら、それは必ず実選択と一致する」が構造的に成立する。

    Returns:
        {'top': list[str], 'basis': 'period'|'yearmonth'|'year'|'none',
         'month_fallback': bool}
        - top: 名前ベースで最上位のタイ候補（複数なら size/mtime で最終決定が必要）
        - month_fallback: 決算月指定ありだが一致ファイルが無く年月最新へ代替した
    """
    names = list(names)

    # ステップ1: 第N期（最大N。1件でも第N期があれば最優先＝get_pl_latest と同順）
    def _period_num(n: str) -> int:
        m = _PL_PERIOD_RE.search(n)
        return int(m.group(1)) if m else -1

    nums = [(n, _period_num(n)) for n in names]
    max_num = max((v for _, v in nums), default=-1)
    if max_num >= 0:
        return {
            'top': [n for n, v in nums if v == max_num],
            'basis': 'period', 'month_fallback': False,
        }

    # ステップ2: 期末年月
    date_pairs = [
        (n, _parse_fiscal_end_from_filename(
            n, fiscal_month_override=fiscal_month_override))
        for n in names
    ]
    with_date = [(n, ym) for n, ym in date_pairs if ym is not None]
    if with_date:
        month_fallback = False
        if fiscal_month_override is not None:
            mm = [(n, ym) for n, ym in with_date if ym[1] == fiscal_month_override]
            if mm:
                with_date = mm
            else:
                month_fallback = True
        max_ym = max(ym for _, ym in with_date)
        return {
            'top': [n for n, ym in with_date if ym == max_ym],
            'basis': 'yearmonth', 'month_fallback': month_fallback,
        }

    # ステップ2.5: 年のみ（令和N年度 等）
    year_pairs = [(n, _parse_fiscal_year_from_filename(n)) for n in names]
    with_year = [(n, y) for n, y in year_pairs if y is not None]
    if with_year:
        max_year = max(y for _, y in with_year)
        return {
            'top': [n for n, y in with_year if y == max_year],
            'basis': 'year', 'month_fallback': False,
        }

    # ステップ3: 名前からは判別不可（呼び出し側が mtime 最新で決める）
    return {'top': names, 'basis': 'none', 'month_fallback': False}


class FileDetector:
    """資料フォルダからファイルを自動分類"""

    PATTERNS = {
        'hearing': ['ヒアリング'],
        # 登記書類は通称が割れる（履歴事項全部証明書／登記簿謄本／登記事項証明書）。
        # app.py _FILE_CATEGORIES の registry と必ず同期すること。
        'registry': ['履歴事項', '登記簿', '登記事項'],
        'identity': ['運転免許証', '運転経歴証明書', '住民票', '本人確認'],
        'tax': ['納税証明'],
        'pl': ['損益計算書', '決算報告書', '決算書', '収支内訳書', '青色申告'],
        'cost_report': ['製造原価報告書', '原価報告書'],
        'estimate': ['見積', 'お見積'],
        'wage_report': ['賃金状況報告'],
        # 賃金台帳: 「給与台帳」「賃金台帳」両方を検出（給与ソフト出力で「給与台帳」表記が多い）
        'wage_ledger': ['賃金台帳', '給与台帳'],
        'wage_data': ['支給控除一覧', '給与データ'],
    }

    # 自己参照ループ防止: 過去の出力ファイル（前回ツールで生成された Excel）を入力として誤検出しないよう、
    # ファイル名にこれらのパターンを含むファイルは検出対象から除外する。
    # 出力ファイル命名規則 (app.py / pipeline.py を参照):
    #   - {会社名}_{枠}_AI版.xlsx
    #   - {会社名}_給与支給総額計算.xlsx
    #   - {会社名}_賃金台帳_AI集計.xlsx
    #   - {会社名}_加点①_結果.xlsx / 加点②_結果.xlsx
    # 旧名 `_賃金台帳一覧` はユーザーが手動で作る入力ファイル名と衝突するため
    # マーカーから除外する（過去にツールが同名で出力していた経緯あり）。
    OUTPUT_FILE_MARKERS = (
        '_AI版', '_給与支給総額計算', '_一人当たり給与支給総額',
        '_賃金台帳_AI集計',
        '_加点①', '_加点②',
    )

    # カテゴリ別の許可拡張子（小文字で比較）
    # openpyxl系のカテゴリに .csv が混入すると読み込み時に例外が出て pipeline が全滅するため、
    # 検出段階で弾く。拡張子は後段の読み取り処理に合わせて絞っている。
    ALLOWED_EXTS = {
        'hearing':     {'.xlsx', '.xlsm'},
        'registry':    {'.pdf'},
        'identity':    {'.pdf'},
        'tax':         {'.pdf'},
        'pl':          {'.pdf'},
        'cost_report': {'.pdf'},
        'estimate':    {'.xlsx', '.xlsm', '.pdf'},
        'wage_report': {'.xlsx', '.xlsm'},
        # 2026-05 方針変更: 賃金台帳の回収は Excel/CSV に集約。PDF は受け付けない
        # （PDFで届いた場合はローカルで Excel/CSV に変換してから投入する運用）
        'wage_ledger': {'.xlsx', '.xlsm', '.csv'},
        'wage_data':   {'.pdf'},
    }

    def __init__(
        self,
        folder: Path,
        selection_override: dict[str, list[Path]] | None = None,
        extra_allowed_exts: dict[str, set[str]] | None = None,
    ):
        """フォルダをスキャンしてファイルを自動分類する。

        Args:
            folder: スキャン対象のフォルダ
            selection_override: ユーザーが UI で明示指定したファイル群（カテゴリ別）。
                指定されたカテゴリは自動検出結果を上書きし、`get_pl_latest` の
                自動選定ロジックもバイパスする（手動指定 = ユーザーの責任で確定）。
                値が None または欠落のカテゴリは自動検出を維持。
                値が `[]` のカテゴリは「対象外」として明示的に空にする。
            extra_allowed_exts: カテゴリ別の許可拡張子を一時的に追加する。
                例: 「賃金台帳の作成」タスクで `{'wage_ledger': {'.pdf'}}` を渡せば
                通常は除外される賃金台帳PDFも検出対象になる。
                クラス属性 ALLOWED_EXTS の合集合で評価。
        """
        self.folder = folder
        self.files: dict[str, list[Path]] = {k: [] for k in self.PATTERNS}
        self.skipped: list[tuple[str, str, str]] = []  # (category, filename, reason)
        # get_pl_latest が積む UI 表示用警告（毎回の呼び出しで上書き）
        self.pl_selection_warnings: list[str] = []
        # ユーザーが UI で明示指定したカテゴリ集合（get_pl_latest のバイパス判定で使う）
        self._manual_categories: set[str] = set()
        # インスタンス別の許可拡張子テーブル（クラス属性の複製 + 追加分のマージ）
        self.allowed_exts: dict[str, set[str]] = {
            k: set(v) for k, v in self.ALLOWED_EXTS.items()
        }
        if extra_allowed_exts:
            for cat, exts in extra_allowed_exts.items():
                if cat in self.allowed_exts:
                    self.allowed_exts[cat] |= set(exts)
                else:
                    self.allowed_exts[cat] = set(exts)
        self._scan()
        if selection_override:
            self._apply_override(selection_override)

    def _apply_override(self, override: dict[str, list[Path]]) -> None:
        """ユーザー指定のファイル選択を適用する。

        指定されたカテゴリの `self.files[cat]` を override の値で上書きし、
        `_manual_categories` に追加する。未知のカテゴリは無視。
        値が None のカテゴリは自動検出を維持（上書きしない）。
        """
        for cat, paths in override.items():
            if cat not in self.files:
                continue
            if paths is None:
                continue
            self.files[cat] = list(paths)
            self._manual_categories.add(cat)
            logger.info(
                f'手動選択: [{cat}] {[p.name for p in paths] if paths else "（対象外）"}'
            )

    def is_manual(self, category: str) -> bool:
        """ユーザーが UI で明示指定したカテゴリか。

        テスト用サブクラスなど `__init__` をスキップする呼び出し元に備えて
        getattr フォールバック付き（未初期化なら自動検出扱い = False）。
        """
        return category in getattr(self, '_manual_categories', set())

    def _scan(self):
        """フォルダを再帰的にスキャンしてファイル分類"""
        for p in self._iter_files(self.folder):
            if p.name.startswith('~$'):
                continue
            # Google Drive等で作成されたファイル名はNFD分解形式(例: グ=ク+濁点)で
            # 保存されていることがあり、NFCのキーワードと素直にin比較すると一致しない。
            # 正規化してから判定する。
            name_nfc = unicodedata.normalize('NFC', p.name)
            # 自己参照ループ防止: 過去の出力ファイルは入力として扱わない
            if any(marker in name_nfc for marker in self.OUTPUT_FILE_MARKERS):
                self.skipped.append(('output_file', p.name, '過去の出力ファイル（処理対象外）'))
                logger.info(f'除外: [output_file] {p.name} (過去の出力ファイル)')
                continue
            for category, keywords in self.PATTERNS.items():
                if any(kw in name_nfc for kw in keywords):
                    allowed = self.allowed_exts.get(category)
                    if allowed is not None and p.suffix.lower() not in allowed:
                        self.skipped.append((category, p.name, f'拡張子{p.suffix}は{category}では非対応'))
                        logger.info(f'除外: [{category}] {p.name} (許可拡張子: {sorted(allowed)})')
                        break
                    self.files[category].append(p)
                    logger.debug(f'検出: [{category}] {p.name}')
                    break

    def _iter_files(self, directory: Path):
        """日本語パス対応の再帰ファイル探索"""
        try:
            for p in directory.iterdir():
                if p.is_dir() and not p.name.startswith(('.', '_')):
                    yield from self._iter_files(p)
                elif p.is_file():
                    yield p
        except PermissionError:
            logger.warning(f'アクセス拒否: {directory}')

    def get(self, category: str) -> Path | None:
        """カテゴリの最初のファイルを返す"""
        files = self.files.get(category, [])
        return files[0] if files else None

    def get_all(self, category: str) -> list[Path]:
        """カテゴリの全ファイルを返す"""
        return self.files.get(category, [])

    def get_pl_latest(self, fiscal_month_override: int | None = None) -> Path | None:
        """損益計算書の直近期を返す。

        判定優先順（fiscal_month_override 指定時は決算月一致を最優先）:
          1. ファイル名に「第N期」を含む → N が最大のものを採用（事業年数の進んだ会社対応）
          2. ファイル名から期末年月を抽出（令和X年Y月 / RY.M / YYYY年M月 / YYYY-MM 等）
             → fiscal_month_override が指定されていれば、月情報なしファイル名
               （「R6年決算書」「2024年決算書」等）も年単独で救済する
             → 月一致する候補のみで比較し、期末年月が最新のものを採用
          3. 同期が複数あれば、PDFsam等の部分抜粋を除外し、フル版（サイズ最大）を採用
          4. 上記いずれも該当なければ更新日時最新を採用（強警告：実務上ほぼ起こらないはず）

        副作用:
          self.pl_selection_warnings に UI 表示用の警告を積む（mtime フォールバック等）

        旧実装の問題: ステップ1・2が無いと「令和7年3月決算書」「令和6年3月決算書」のような
        ファイル名（第N期表記なし）が並んだ際、mtime ガチャで前期決算書が選ばれて
        賃金台帳の期間と整合しない財務値が転記される誤動作があった。
        """
        # 選定過程の警告を UI に伝えるため、毎回リセットしてから積み直す
        self.pl_selection_warnings: list[str] = []

        pls = self.files.get('pl', [])
        if not pls:
            return None

        # ---- 手動選択バイパス: ユーザーが UI で明示指定した場合 ----
        # ファイル名から年月が取れないケースでも、ユーザーの選択を尊重する。
        # 複数指定された場合は最初の1件を採用（通常は1件のはず）。
        if self.is_manual('pl'):
            chosen = pls[0]
            logger.info(f'手動選択 PL: {chosen.name}（自動選定ロジックをスキップ）')
            return chosen

        # ---- ステップ1/2/2.5: 名前ベースの順位付け（_rank_pl_names と共有）----
        # 表示用 app._auto_selected_for_display と同一の純関数で順位付けし、
        # 「UI 表示＝実選択」を構造的に保証する。サイズ/mtime に依存する最終決定だけ
        # ここで行う（UI は再現不能なので、その場合 UI 側は「手動選択を促す」に倒す）。
        ranked = _rank_pl_names([p.name for p in pls], fiscal_month_override)
        if ranked['month_fallback']:
            logger.warning(
                f'決算月{fiscal_month_override}月と一致するファイル名が無く、年月最新で選びます'
            )
            self.pl_selection_warnings.append(
                f'決算月{fiscal_month_override}月と一致する決算書ファイル名が'
                f'見つからず、年月最新で代替選択しました'
            )
        if ranked['basis'] != 'none':
            top = set(ranked['top'])
            latest = [p for p in pls if p.name in top]
            if len(latest) > 1:
                logger.info(
                    f'同一順位のPL候補が{len(latest)}件あり（{ranked["basis"]}）'
                    f' → フル版優先で選択'
                )
            return _pick_full_version(latest)

        # ---- ステップ3: フォールバック（mtime 最新）— 強警告 ----
        logger.warning(
            f'PL候補のファイル名から期番号も期末年月も抽出できないため、'
            f'更新日時最新でフォールバックします: {[p.name for p in pls]}'
        )
        self.pl_selection_warnings.append(
            '⚠️ 決算書ファイル名から年度が判別できず、Drive更新日時で代替選択しました。'
            '誤った期の決算書が選ばれている可能性があります。'
            'ファイル名に「令和N年」「RN年」「YYYY年」など年情報を含めて再アップロードを推奨します。'
            f'候補: {[p.name for p in pls]}'
        )
        return max(pls, key=lambda p: p.stat().st_mtime)

    def summary(self) -> str:
        """検出結果のサマリ"""
        lines = ['検出されたファイル:']
        for cat, files in self.files.items():
            if files:
                names = [f.name for f in files]
                lines.append(f'  {cat}: {", ".join(names)}')
            else:
                lines.append(f'  {cat}: なし')
        if self.skipped:
            lines.append('')
            lines.append('除外されたファイル（拡張子不一致）:')
            for cat, name, reason in self.skipped:
                lines.append(f'  [{cat}] {name} — {reason}')
        return '\n'.join(lines)


def _adopt_proprietor_as_representative(extraction, mapping) -> bool:
    """個人事業主: 決算書の申告者氏名・フリガナを代表者として採用する。

    個人には履歴事項全部証明書が無く、代表者氏名の取得元が
    青色申告決算書/収支内訳書（PL抽出の proprietor_name/kana）しかない。
    履歴事項由来の代表者氏名が既にあればそちらを優先して何もしない。

    Returns:
        True = 採用した / False = 条件を満たさず何もしなかった
    """
    if not mapping.is_kojin:
        return False
    if extraction.company.representative_name:
        return False
    if not extraction.financial.proprietor_name:
        return False
    extraction.company.representative_name = extraction.financial.proprietor_name
    extraction.company.representative_kana = extraction.financial.proprietor_kana
    return True


def run_application_transfer(
    resource_folder: Path,
    template_path: Path,
    template_type: str,
    output_path: Path,
    extractor: BaseExtractor | None = None,
    fiscal_month_override: int | None = None,
    has_cost_report_hint: bool = False,
    selection_override: dict[str, list[Path]] | None = None,
) -> ProcessingStatus:
    """
    タスク1: 申請書転記の実行

    Args:
        resource_folder: 資料フォルダ
        template_path: テンプレートExcelパス
        template_type: '通常枠_2026' or 'インボイス枠_2026'
        output_path: 出力ファイルパス
        extractor: AI抽出器（省略時は自動選択）
        fiscal_month_override: ユーザー指定の決算月（1〜12）。指定時はAI推定より優先
        has_cost_report_hint: ユーザーが「製造原価報告書あり」とチェック済みなら True。
            自動検出されなかった場合の警告強化に使う
        selection_override: ユーザーが UI で明示指定したファイル群（カテゴリ別）。
            FileDetector に渡して自動検出を上書きする。
    """
    status = ProcessingStatus(
        company_name=resource_folder.name,
        template_type=template_type,
        status='処理中',
    )

    try:
        from .ai_extractor import APICreditExhaustedError
        mapping = get_mapping(template_type)

        if extractor is None:
            extractor = create_extractor(CLAUDE_API_KEY)

        # ファイル検出
        detector = FileDetector(resource_folder, selection_override=selection_override)
        logger.info(detector.summary())

        extraction = ExtractionResult()

        # API残高切れ等で Phase 2 が部分実行になった場合の理由
        api_skipped_reason: str = ''

        # ===== Phase 1: API不要な処理（確実に動く・コスト¥0）=====
        # ヒアリングシート読取（Excel直接読取）
        hearing_path = detector.get('hearing')
        hearing_data = {}
        if hearing_path:
            hearing_data = read_hearing_sheet(hearing_path)
            logger.info(f'ヒアリングシート: {len(hearing_data)}行読込')
        else:
            logger.warning('ヒアリングシートが見つかりません')

        # 見積書: Excel なら API 不要なのでここで処理（PDF は Phase 2）
        estimate_path = detector.get('estimate')
        estimate_pdf_path = None
        if estimate_path:
            if estimate_path.suffix == '.xlsx':
                # Excel の見積書は直接読取（API不要）
                import openpyxl
                wb_est = openpyxl.load_workbook(estimate_path, data_only=True)
                ws = wb_est[wb_est.sheetnames[0]]
                tool_name_keywords = ['件名', '品名', 'ツール名', '商品名', 'サービス名']
                found = False
                for row in ws.iter_rows(min_row=1, max_row=30):
                    for i, cell in enumerate(row):
                        v = cell.value
                        if v and isinstance(v, str):
                            if any(kw in v for kw in tool_name_keywords):
                                if i + 1 < len(row) and row[i + 1].value:
                                    extraction.estimate.tool_name = str(row[i + 1].value)
                                    found = True
                                    break
                    if found:
                        break
                if not found:
                    import re
                    name = estimate_path.stem
                    for remove in ['お見積り', 'お見積', '見積り', '見積', '御見積', '_', '様']:
                        name = name.replace(remove, '')
                    name = re.sub(r'\d{8}', '', name)
                    name = re.sub(r'\d{4}[-/]\d{2}[-/]\d{2}', '', name)
                    name = name.strip()
                    if len(name) > 2:
                        extraction.estimate.tool_name = name
                wb_est.close()
            else:
                # PDF は Phase 2 で処理
                estimate_pdf_path = estimate_path

        # ===== Phase 2: API使用（残高切れなら部分スキップ。Phase 1 結果は維持）=====
        # 製造原価報告書の自動検出フラグは Phase 2 内で例外発生しても参照されるため
        # try ブロック前に初期化しておく（UnboundLocalError 回避）
        cost_report_detected: bool = False
        # データソースシート用に検出済みファイルパスを保持（API 残高切れで Phase 2 が
        # 中断しても、データソースシートには「ファイルは検出されたが AI 抽出失敗」と
        # 出せるよう、try ブロック外で参照可能にする）
        registry_path: Path | None = None
        pl_path: Path | None = None
        cost_report_path: Path | None = None
        tax_path: Path | None = None
        try:
            # 履歴事項PDF → CompanyInfo
            registry_path = detector.get('registry')
            if registry_path:
                images = pdf_to_images(registry_path)
                extraction.company = extractor.extract_registry(images)
                logger.info(f'履歴事項: {extraction.company.name}')

            # 損益計算書PDF → FinancialData
            # 決算月指定があれば、ファイル名年月と突合して直近期を確定（誤読防止）
            pl_path = detector.get_pl_latest(fiscal_month_override=fiscal_month_override)
            _record_pl_selection(status, detector, pl_path, fiscal_month_override)
            pl_period_warning = ''
            cost_report_path = detector.get('cost_report')
            if pl_path:
                logger.info(f'直近期決算書として採用: {pl_path.name}')
                images = pdf_to_images(pl_path)
                if cost_report_path:
                    images += pdf_to_images(cost_report_path)
                    logger.info(f'製造原価報告書も読取: {cost_report_path.name}')
                    cost_report_detected = True
                extraction.financial = extractor.extract_pl(images)
                logger.info(f'損益計算書: 売上{extraction.financial.revenue:,}')

                # 個人事業主: 決算書の申告者氏名を代表者として採用
                # （財務値リセットの整合性チェックより前に確定させる）
                if _adopt_proprietor_as_representative(extraction, mapping):
                    logger.info(
                        f'個人事業主の代表者氏名を決算書から採用: '
                        f'{extraction.company.representative_name}'
                    )

                # 賃金台帳期間 vs PL期末 / ユーザー指定決算月の整合性チェック
                # ズレを検出したら財務値転記をスキップ + 強警告
                extraction.financial, pl_period_warning = (
                    _check_pl_wage_period_consistency(
                        detector, extraction.financial, fiscal_month_override,
                    )
                )

            # 納税証明書PDF
            tax_path = detector.get('tax')
            if tax_path:
                images = pdf_to_images(tax_path)
                extraction.tax = extractor.extract_tax(images)

            # 見積書PDF (Excelは Phase 1 で処理済)
            if estimate_pdf_path:
                images = pdf_to_images(estimate_pdf_path)
                extraction.estimate = extractor.extract_estimate(images)

            # AI判断（ヒアリングデータも渡してIT投資状況等の矛盾を防ぐ）
            extraction.ai_judgment = extractor.generate_ai_judgment(
                extraction.company,
                extraction.financial,
                extraction.estimate.tool_name,
                hearing_data=hearing_data,
            )
        except APICreditExhaustedError as e:
            # API残高切れ → Phase 1 の結果（ヒアリング・賃金台帳Excel等）で申請書を出力
            # 確認キューにAI由来項目が「未取得」として一覧される
            api_skipped_reason = str(e)
            logger.warning(
                f'[Phase 2] API残高切れで以降の API 呼出をスキップ: {e}'
            )

        # 賃金台帳 → 1人当たり給与支給総額の計画値 + 一覧Excel出力
        # 申請書作成タスクでは賃金台帳を AI で再抽出しない（決定論パーサー一本）。
        # 運用上、賃金台帳は事前に「賃金台帳の作成」タスクで標準テンプレ形式に整え、
        # 「一人当たり給与支給総額」タスクで人間が数値を承認している前提で動く。
        # AI を使わないことで:
        #   - per_employee_wage と完全に同じ結果になる（再現性・突合可能）
        #   - API コストゼロ・処理時間短縮
        #   - API 残高切れの影響を受けない
        # 非標準フォーマットの Excel が渡されて決定論で読めない場合は
        # wage_status='no_data' で警告し、R215/R216 等は空欄で続行する。
        wage_extraction_method = '決定論パーサー（賃金台帳Excel直読）'
        try:
            wage_plan, ledger_employees, wage_status = _calc_wage_plan_from_ledger(
                detector, extraction.financial, extractor=None,
                fiscal_month_override=fiscal_month_override,
            )
        except Exception as e:
            logger.warning(f'賃金台帳処理エラー（申請書作成は続行）: {e}', exc_info=True)
            wage_plan, ledger_employees, wage_status = None, [], 'error'

        # ユーザー指定の決算月 vs AI 推定の照合（警告のみ。処理は続行）
        _, fiscal_month_warning = _resolve_fiscal_period(
            extraction.financial, fiscal_month_override,
        )

        # ユーザーが「製造原価報告書あり」とチェックしたのに自動検出されなかった場合の警告
        # （ファイル名キーワード未一致や、損益計算書PDFに統合されているケースなど）
        cost_report_warning = ''
        if has_cost_report_hint and not cost_report_detected:
            cost_report_warning = (
                ' ⚠ 「製造原価報告書あり」とチェックされていますが、製造原価報告書のPDFを'
                '検出できませんでした。ファイル名に「製造原価報告書」を含めるか、'
                '損益計算書PDFに統合されている場合は原本を目視で確認し、'
                '原価部の人件費（労務費・賞与等）が抽出値に含まれているかご確認ください。'
            )

        # AI 生成の事業内容の文字数チェック（240〜255文字が望ましい）
        # フロー: AI候補採用 / 機械削除フォールバック / 失敗（原文残し）の3パターンで警告分岐。
        biz_desc_warning = ''
        biz_desc = (extraction.ai_judgment.business_description or '').strip()
        if biz_desc:
            n_orig = len(biz_desc)
            if n_orig > BIZ_DESC.HARD_MAX:
                result = None
                if extractor is not None:
                    try:
                        result = extractor.shorten_business_description(
                            biz_desc, max_len=BIZ_DESC.SHORTEN_MAX_DEFAULT,
                        )
                    except Exception as e:
                        logger.warning(f'事業内容の自動短縮に失敗（警告にフォールバック）: {e}', exc_info=True)
                        result = None

                # 最終案を採用 → その案に対して1回だけ文字数判定して警告を分岐
                if result and result.text and result.source == 'ai':
                    extraction.ai_judgment.business_description = result.text
                    if result.length < BIZ_DESC.TARGET_MIN:
                        # 短縮できたが推奨下限(240)を割った（全候補が短い等）。
                        # 黙って採用せず⚠で人へ厚み追加を促す（silent な過少を防ぐ）。
                        biz_desc_warning = (
                            f' ⚠ 事業内容が文字数制限超過（原文{n_orig}文字）のため AI 再生成で'
                            f'短縮しましたが、結果が{result.length}文字と推奨下限（240文字）を下回りました。'
                            f'4要素（現状・課題・解決策・期待効果）が書き切れているか、'
                            f'ヒアリング情報を追記して厚みを出してください。'
                        )
                        logger.warning(
                            f'事業内容を自動短縮(AI採用)したが下限未達: {n_orig}文字 → {result.length}文字'
                        )
                    else:
                        biz_desc_warning = (
                            f' ℹ 事業内容が文字数制限超過（原文{n_orig}文字）だったため、'
                            f'AI で再生成した3案から最適なもの（{result.length}文字）を採用しました。'
                            f'提出前に内容を必ず目視確認してください。'
                        )
                        logger.warning(f'事業内容を自動短縮(AI採用): {n_orig}文字 → {result.length}文字')
                elif result and result.text and result.source == 'mechanical':
                    extraction.ai_judgment.business_description = result.text
                    biz_desc_warning = (
                        f' ⚠ 事業内容が文字数制限超過（原文{n_orig}文字）。AI再生成3案がすべて'
                        f'255文字超だったため、機械的に末尾の文を削って{result.length}文字に収めました。'
                        f'4要素のうち「期待効果」が削られている可能性があるため、'
                        f'必ず原文と照合して目視確認してください。'
                    )
                    logger.warning(f'事業内容を自動短縮(機械削除): {n_orig}文字 → {result.length}文字')
                else:
                    biz_desc_warning = (
                        f' ⚠ 事業内容が文字数制限超過（{n_orig}文字 / 上限255文字）。'
                        f'自動短縮（AI再生成 + 機械削除）に失敗しました。'
                        f'原稿を手動で短縮してください。'
                    )
                    logger.warning(f'事業内容の自動短縮に失敗: {n_orig}文字のまま残置')
            elif n_orig < BIZ_DESC.TARGET_MIN:
                # 240字未満は Opus(writing_model) で 4要素を厚くして 240〜252 字に増量する。
                # 成否で >255 短縮と同じ3パターン警告構造に揃える。
                result = None
                if extractor is not None:
                    try:
                        result = extractor.expand_business_description(biz_desc)
                    except Exception as e:
                        logger.warning(f'事業内容の自動増量に失敗（警告にフォールバック）: {e}', exc_info=True)
                        result = None

                if result and result.text and result.source == 'ai':
                    extraction.ai_judgment.business_description = result.text
                    biz_desc_warning = (
                        f' ℹ 事業内容が短め（原文{n_orig}文字）だったため、AI で再生成し'
                        f'4要素を厚くした案（{result.length}文字）を採用しました。'
                        f'提出前に内容を必ず目視確認してください。'
                    )
                    logger.warning(f'事業内容を自動増量(AI採用): {n_orig}文字 → {result.length}文字')
                elif result and result.text and result.source == 'mechanical':
                    extraction.ai_judgment.business_description = result.text
                    biz_desc_warning = (
                        f' ℹ 事業内容が短め（原文{n_orig}文字）だったため増量し、'
                        f'{result.length}文字に調整しました。提出前に内容を必ず目視確認してください。'
                    )
                    logger.warning(f'事業内容を自動増量(機械調整): {n_orig}文字 → {result.length}文字')
                else:
                    biz_desc_warning = (
                        f' ⚠ 事業内容が短すぎます（{n_orig}文字 / 推奨240〜255文字）。'
                        f'4要素（現状・課題・解決策・期待効果）が十分書き切れているか、'
                        f'ヒアリング情報を追記して厚みを出してください。'
                    )

        # 賃金台帳に PDF がアップロードされた場合の警告
        # 2026-05 方針: 賃金台帳は Excel/CSV のみ受け付ける。
        # PDF は detector でスキップされるが、ユーザーには明示警告して再投入を促す
        wage_pdf_warning = ''
        wage_pdf_files = [
            name for cat, name, _ in detector.skipped
            if cat == 'wage_ledger' and name.lower().endswith('.pdf')
        ]
        if wage_pdf_files:
            wage_pdf_warning = (
                f' ⚠ 賃金台帳PDFは受け付け対象外です（{len(wage_pdf_files)}件: '
                f'{", ".join(wage_pdf_files[:3])}'
                f'{"…" if len(wage_pdf_files) > 3 else ""}）。'
                f'ローカルでExcel/CSV形式に変換してから再投入してください。'
            )

        # テンプレート転記（Phase 1 + Phase 2成功分のみ。残高切れ項目は confidence='low' で空欄）
        empty_cells = fill_template(
            template_path=template_path,
            output_path=output_path,
            mapping=mapping,
            hearing_data=hearing_data,
            extraction=extraction,
            wage_plan=wage_plan,
        )

        status.status = '完了' if not api_skipped_reason else '部分完了'
        status.output_files = [output_path.name]
        status.empty_cells = empty_cells
        # 後続タスク（給与計算/加点判定）で再利用するため抽出結果を status に保持
        # （financial=PL の AI 抽出結果、ledger_employees=賃金台帳の決定論パーサー結果）
        status.financial = extraction.financial
        status.ledger_employees = ledger_employees or []
        # Phase 4: 低信頼項目を「確認キュー」として集約
        status.confidence_warnings = _build_confidence_warnings(extraction.financial)
        # 賃金台帳の読み取り状況に応じて完了メッセージに警告を追記（処理は続行）
        wage_warning = ''
        if wage_status == 'no_data':
            wage_warning = (
                ' ⚠ 賃金台帳が読み取れませんでした（給与支給総額は空欄）。'
                '推奨フロー: ①「賃金台帳の作成」タスクで標準テンプレ形式に変換 → '
                '②「一人当たり給与支給総額」タスクで数値を確認 → ③ 申請書作成'
            )
        elif wage_status == 'zero_total':
            wage_warning = (
                ' ⚠ 賃金台帳の給与支給総額が0でした（R215/R216は空欄）。'
                '賃金台帳の内容・対象期間をご確認ください'
            )
        elif wage_status == 'error':
            wage_warning = (
                ' ⚠ 賃金台帳処理中にエラーが発生しました。'
                '「賃金台帳の作成」タスクで標準テンプレ形式に変換してから再実行してください'
            )
        elif wage_status == 'officer_only':
            wage_warning = (
                ' ⛔ 強警告: 賃金台帳に非役員の従業員が0名でした（役員のみの法人）。'
                'IT導入補助金 公募要領 p.10 では「従業員を雇用していない法人は、従業員を'
                '役員と読み替え、役員報酬と役員数に応じて算出」と定めています。'
                '申請書 R215（従業員数）・R216（給与支給総額）・R217〜R219（賃上げ計画）は '
                '空欄のままです。役員報酬・役員数ベースで手動入力してください'
            )
        elif wage_status == 'no_full_year_target':
            wage_warning = (
                ' ⛔ 強警告: 賃金台帳の全月在籍者が0名でした（従業員はいるが全員が直近事業年度の'
                '全月分給与を受けていない）。考えられる原因は2つです。'
                '【原因①】賃金台帳の対象期間が直近決算期12ヶ月とズレている（支給日ベース等）'
                '→ まず提出期間が直近決算期12ヶ月を含むか確認し、ズレていれば正しい期間で再取得。'
                '【原因②】設立間もない等で全員が在籍12ヶ月未満 → 役員報酬ベースでの算定が'
                '必要か公募要領で確認（事務局確認推奨）。'
                'いずれもR215（従業員数）・R216（給与支給総額）・R217〜R219は空欄のままです。'
                '原因を確定のうえ手動で値を入力してください'
            )
        elif wage_status == 'low_full_year_ratio':
            wage_warning = (
                ' ⚠ 注意: 賃金台帳の全月在籍者が会社規模に対して少なめです'
                '（中途入退社が多い案件）。公募要領通り「全月分の給与支給を受けた従業員」'
                'のみで R215/R216 を自動算出・転記しました。「一人当たり給与支給総額」'
                'タスクの出力と数値が一致しているかと、賃金台帳期間が直近決算期と'
                '揃っているかを念のためご確認ください'
            )

        # パート時間欠落で FTE=1.0 サイレント昇格があれば追加警告
        # IT導入補助金 公募要領 p.10「パートタイム従業員については正社員の就業時間に換算」を満たすため、
        # 時間データなしのパートが含まれると R215 が過大計上になる
        if wage_plan and wage_plan.get('part_fte_fallback_count', 0) > 0:
            _fb_n = wage_plan['part_fte_fallback_count']
            _fb_names = wage_plan.get('part_fte_fallback_names', []) or []
            _name_preview = ', '.join(_fb_names[:3])
            if len(_fb_names) > 3:
                _name_preview += f' ほか{len(_fb_names) - 3}名'
            wage_warning += (
                f' ⚠ パート・アルバイト {_fb_n}名で労働時間データなし'
                f'({_name_preview}) → R215（FTE換算従業員数）が過大計上の可能性。'
                f'IT導入補助金 公募要領 p.10 は「正社員の就業時間に換算」を要求。'
                f'賃金台帳テンプレ E列「月間平均時間」を顧客確認のうえ手入力してください'
            )

        # 給与支給0円の従業員を算定対象外にした場合の通知（黙って R215/R216 が
        # 小さくなるとユーザーが原因を追えないため、除外人数と対象者を明示する）
        if wage_plan and wage_plan.get('excluded_zero_count', 0) > 0:
            _z_names = wage_plan.get('excluded_zero_names', []) or []
            _z_preview = ', '.join(_z_names[:3])
            if len(_z_names) > 3:
                _z_preview += f' ほか{len(_z_names) - 3}名'
            wage_warning += (
                f' ℹ 賃金台帳で給与支給0円の従業員 {wage_plan["excluded_zero_count"]}名'
                f'（{_z_preview}）を R215/R216 の算定対象外にしました'
                f'（公募要領「全月分の給与等の支給を受けた従業員」に非該当）。'
                f'退職済み行・無給の家族従業者等でないか賃金台帳をご確認ください'
            )

        # 0円支給月と支給月が混在する従業員（中途入社・退職・無給休職等）を
        # 算定対象外にした場合の通知（v0.2.69。同じく黙って減らさない）
        if wage_plan and wage_plan.get('excluded_partial_zero_count', 0) > 0:
            _p_names = wage_plan.get('excluded_partial_zero_names', []) or []
            _p_preview = ', '.join(_p_names[:3])
            if len(_p_names) > 3:
                _p_preview += f' ほか{len(_p_names) - 3}名'
            wage_warning += (
                f' ℹ 賃金台帳で0円支給月がある従業員 '
                f'{wage_plan["excluded_partial_zero_count"]}名（{_p_preview}）を '
                f'R215/R216 の算定対象外にしました'
                f'（公募要領「全月分の給与等の支給を受けた従業員」に非該当）。'
                f'中途入社・退職なら正しい挙動です。産休・育休等の休職で算入したい場合'
                f'（除外は任意）や、0円月に賞与支給があり全月支給に該当する場合は、'
                f'賃金台帳の該当月を実支給額に修正して再実行してください'
            )
        # 整合性チェック: 賃金台帳合計と損益計算書の人件費の差が大きいと AI 抽出ミスの疑い
        consistency_warning = _check_wage_pl_consistency(wage_plan, extraction.financial)
        # 会計式整合（売上 − 原価 = 粗利）— ()書きマイナス誤読の自動検出
        pl_accounting_warning = _check_pl_accounting_consistency(extraction.financial)
        # 業種コードのフォーマット検証（旧3桁体系・自己流コード検出）
        industry_code_warning = _check_industry_code_format(extraction.ai_judgment)
        # 役員名の誤入力検出（従業員代表者・事業所内最低賃金者の欄に役員名が入っていたら警告）
        officer_alert_warning = _check_representative_against_officers(
            hearing_data, extraction.company,
        )
        # 賃金台帳抽出結果の自動品質検証（人数妥当性・月別カバレッジ・値分布・賞与未参照）
        from .wage_validator import run_all_validations
        validation_warnings = ''.join(
            run_all_validations(hearing_data, ledger_employees, extraction.financial)
        )
        # API残高切れで Phase 2 がスキップされた場合は冒頭にお知らせを追加
        api_skip_msg = ''
        if api_skipped_reason:
            api_skip_msg = (
                f'⛔ APIエラーでAI抽出部分がスキップされました ({api_skipped_reason})。'
                f'ヒアリングシート・賃金台帳など API不要部分のみ転記しました。'
                f'残高チャージ後に再実行すると AI 部分も埋まります。 '
            )
        status.message = (
            api_skip_msg
            + f'完了。空欄{len(empty_cells)}件{wage_warning}{consistency_warning}'
            + validation_warnings
            + fiscal_month_warning
            + pl_period_warning
            + cost_report_warning
            + biz_desc_warning
            + wage_pdf_warning
            + pl_accounting_warning
            + industry_code_warning
            + officer_alert_warning
        )
        logger.info(f'申請書作成完了: {output_path.name} (空欄{len(empty_cells)}件{wage_warning})')

        # 賃金台帳集計Excel出力（チェック用）— 決定論パーサーで読み取った賃金台帳を集計。
        # 「一人当たり給与支給総額」タスクの出力と数値が完全一致する想定なので、
        # 申請書作成時のクロスチェックに使う。
        # NOTE: ファイル名の `_AI集計` 表記は過去 AI 抽出経路だった名残（既存ファイル名
        # 互換のため維持）。中身は決定論パーサー由来で、誤読リスクは無い。
        if ledger_employees:
            company = output_path.stem.split('_')[0]
            ledger_output = output_path.parent / f'{company}_賃金台帳_AI集計.xlsx'
            export_wage_ledger_summary(
                ledger_employees, ledger_output, company,
                extraction_method=wage_extraction_method,
            )
            status.output_files.append(ledger_output.name)

        # データソースシート追加（申請書出力 Excel の末尾に追記）。
        # 1次振り返り MTG（2026-05-14）の要望:
        #   「抽出した数値の根拠となるデータソースを記録し、出力資料に含める」
        # PDF はテキスト層が取れればページ番号を、Excel/CSV はファイル名のみを記録。
        # 失敗してもログ警告のみで申請書本体は維持（補助情報のため）。
        from .template_filler import add_data_source_sheet
        # PL値→ページ逆引きは PL.pdf + 製造原価報告書.pdf 両方を探索する。
        # extract_pl が両方の画像をまとめて AI に渡しているため、
        # 製造原価部由来の値（労務費等）も逆引き可能にしないと出所が出ない。
        pl_value_pages = _compute_pl_value_pages(
            pl_path, extraction.financial, cost_report_path=cost_report_path,
        )
        wage_ledger_paths_list = detector.get_all('wage_ledger') or []
        add_data_source_sheet(
            output_path,
            hearing_path=hearing_path,
            registry_path=registry_path,
            pl_path=pl_path,
            cost_report_path=cost_report_path,
            tax_path=tax_path,
            # estimate_path は Phase 1 で Excel/PDF 両対応に解決済（None 含む）
            estimate_path=estimate_path,
            wage_ledger_paths=wage_ledger_paths_list,
            extraction=extraction,
            pl_value_pages=pl_value_pages,
            wage_plan=wage_plan,
            wage_extraction_method=wage_extraction_method,
        )

        # 一人当たり給与支給総額シートを AI 版に統合（API 追加コスト 0）。
        # 申請書作成タスクで既に取得済みの financial / ledger_employees / company を
        # cached_* として渡して per_employee_only 経路を実行 → 内部で API 呼出スキップ。
        # 賃金台帳が読めなかったケース（ledger_employees 空）はそもそも給与計算できないので
        # スキップ。失敗してもログ警告のみで申請書本体は維持する（補助情報のため）。
        if ledger_employees:
            try:
                # 会社名は出力ファイル名の先頭セグメントを使う。
                # resource_folder.name は Streamlit 実行時に一時フォルダ名
                # （tmpXXXX）になり、給与計算シートの社名表示が壊れるため不可。
                _attach_per_employee_wage_sheet(
                    output_path=output_path,
                    resource_folder=resource_folder,
                    company_name=output_path.stem.split('_')[0],
                    extractor=extractor,
                    cached_financial=extraction.financial,
                    cached_ledger_employees=ledger_employees,
                    cached_company=extraction.company,
                    fiscal_month_override=fiscal_month_override,
                    selection_override=selection_override,
                )
            except Exception as e:
                logger.warning(
                    f'一人当たり給与支給総額シートの統合に失敗（申請書本体は維持）: {e}',
                    exc_info=True,
                )

    except Exception as e:
        # 通常の例外（ファイル不在等）: ステータスをエラーに
        # API残高切れは Phase 2 内で個別ハンドル済みなのでここには到達しない
        status.status = 'エラー'
        status.message = str(e)
        logger.error(f'エラー: {e}', exc_info=True)

    return status


def _attach_per_employee_wage_sheet(
    output_path: Path,
    resource_folder: Path,
    company_name: str,
    extractor: BaseExtractor | None,
    cached_financial: 'FinancialData | None',
    cached_ledger_employees: list,
    cached_company: 'CompanyInfo | None',
    fiscal_month_override: int | None,
    selection_override: dict[str, list[Path]] | None,
) -> None:
    """申請書作成タスクの出力 Excel に「一人当たり給与支給総額」シートを統合する。

    動作概要:
      1. 一時ファイルパスで run_wage_calculation(per_employee_only=True) を実行
         （cached_* を渡すことで履歴事項・PL・賃金台帳の AI 抽出を全てスキップ → API 追加コスト 0）
      2. 一時ファイルから「給与支給総額計算」「従業員別明細」シートを copy_sheet_to_workbook で
         AI 版.xlsx 末尾にコピー
      3. 一時ファイル削除

    失敗時は呼び出し側で warning ログを出して申請書本体は維持する設計。
    """
    import tempfile
    from .template_filler import copy_sheet_to_workbook

    with tempfile.TemporaryDirectory(prefix='hojokin_wage_') as tmpdir:
        tmp_path = Path(tmpdir) / f'{company_name}_一人当たり給与支給総額.xlsx'
        wage_status = run_wage_calculation(
            resource_folder=resource_folder,
            company_name=company_name,
            output_path=tmp_path,
            extractor=extractor,
            cached_financial=cached_financial,
            cached_ledger_employees=cached_ledger_employees,
            cached_company=cached_company,
            fiscal_month_override=fiscal_month_override,
            selection_override=selection_override,
            per_employee_only=True,
        )

        if wage_status.status not in ('完了', '部分完了') or not tmp_path.exists():
            logger.warning(
                f'一人当たり給与支給総額シート: 生成失敗 (status={wage_status.status}) — '
                f'統合をスキップ'
            )
            return

        # 一時 Excel を開いてシートを抽出 → AI版.xlsx に追記
        import openpyxl
        src_wb = openpyxl.load_workbook(tmp_path)
        dest_wb = openpyxl.load_workbook(output_path)
        try:
            copied = 0
            # 給与計算 Excel のシート構成は create_wage_calculation 仕様:
            #   1. 給与支給総額計算（メイン）
            #   2. 従業員別明細（任意。賃金台帳に12ヶ月明細がある場合のみ生成）
            for sheet_name in ('給与支給総額計算', '従業員別明細'):
                if sheet_name in src_wb.sheetnames:
                    copy_sheet_to_workbook(src_wb[sheet_name], dest_wb)
                    copied += 1
            dest_wb.save(output_path)
            logger.info(
                f'一人当たり給与支給総額シート: {copied}シートを AI 版に統合（API追加 0回）'
            )
        finally:
            src_wb.close()
            dest_wb.close()


def run_wage_calculation(
    resource_folder: Path,
    company_name: str,
    output_path: Path,
    extractor: BaseExtractor | None = None,
    cached_financial: 'FinancialData | None' = None,
    cached_ledger_employees: list | None = None,
    fiscal_month_override: int | None = None,
    cached_company: 'CompanyInfo | None' = None,
    selection_override: dict[str, list[Path]] | None = None,
    per_employee_only: bool = False,
) -> ProcessingStatus:
    """
    タスク2: 給与支給総額計算の実行

    cached_financial / cached_ledger_employees が渡された場合は API 呼出を省略する
    （申請書作成タスクの結果を再利用してコスト2重化を防ぐ）。

    fiscal_month_override (1〜12) が指定された場合、ユーザー指定の決算月で
    賃金台帳の対象期間（直近12ヶ月）を確定する。AI 推定とズレていれば警告。

    per_employee_only=True のとき「一人当たり給与支給総額」タスクとして動作:
        - 決算書PDF（損益計算書）は一切読まない
        - 出力Excelから決算書由来セクション（給料手当〜減価償却費）を削除
        - 賃金台帳のみが情報源
    """
    status = ProcessingStatus(
        company_name=company_name,
        template_type='一人当たり給与支給総額' if per_employee_only else '給与計算',
        status='処理中',
    )

    try:
        if extractor is None:
            extractor = create_extractor(CLAUDE_API_KEY)

        detector = FileDetector(resource_folder, selection_override=selection_override)
        logger.info(detector.summary())

        # 損益計算書（任意: あれば精度向上）— キャッシュがあれば再利用
        # per_employee_only モードでは PL を一切参照しない
        financial = cached_financial
        pl_period_warning = ''
        if per_employee_only:
            from .models import FinancialData
            financial = FinancialData()
            logger.info('per_employee_only: 決算書PDFは参照せず賃金台帳のみで計算')
        elif financial is None:
            pl_path = detector.get_pl_latest(fiscal_month_override=fiscal_month_override)
            _record_pl_selection(status, detector, pl_path, fiscal_month_override)
            if pl_path:
                logger.info(f'直近期決算書として採用: {pl_path.name}')
                images = pdf_to_images(pl_path)
                financial = extractor.extract_pl(images)
                # 賃金台帳期間 vs PL期末 / ユーザー指定決算月の整合性チェック
                financial, pl_period_warning = _check_pl_wage_period_consistency(
                    detector, financial, fiscal_month_override,
                )
        else:
            logger.info('PL: 申請書作成タスクの結果を再利用（API呼出スキップ）')
            # 申請書作成タスクで既にチェック済みだが、キャッシュ経路でも再確認
            financial, pl_period_warning = _check_pl_wage_period_consistency(
                detector, financial, fiscal_month_override,
            )

        if financial is None or financial.revenue == 0:
            from .models import FinancialData
            if financial is None:
                financial = FinancialData()
            if not per_employee_only:
                logger.info('損益計算書なし → 賃金台帳ベースで計算')

        # 賃金状況報告シートから従業員データ読取（あれば）
        employees_detail = None
        seishain_count = 0
        part_count = 0
        yakuin_hoshu_3m = 0

        # 役員数の取得（履歴事項証明書から動的に決定。固定値ハードコードを廃止）
        # 申請書作成タスクから cached_company が渡ってきていればそれを優先。
        # 無ければ resource_folder 内の履歴事項証明書を AI 抽出する（API 1回）。
        # 全く取れない場合のフォールバックは1人（代表取締役を最低限想定）。
        #
        # 注意: AI プロンプト（ai_extractor.py:287）で「代表者は officers に含めない」
        # と明示しており、CompanyInfo.officers は代表取締役を除いたリストになる。
        # よって役員総数 = 1（代表者）+ len(officers)。template_filler.py:121 と同じ規約。
        from .ai_extractor import APICreditExhaustedError
        yakuin_count = 1
        company = cached_company
        if company is None:
            registry_path = detector.get('registry')
            if registry_path:
                try:
                    images = pdf_to_images(registry_path)
                    company = extractor.extract_registry(images)
                    logger.info(
                        f'履歴事項: {company.name}（officers={len(company.officers)}名）'
                    )
                except APICreditExhaustedError:
                    # 残高切れは他の API 呼び出しも全部失敗するので即停止
                    raise
                except Exception as e:
                    logger.warning(f'履歴事項証明書の AI 抽出に失敗: {e}')
                    company = None
        if company is not None:
            # 代表者(+1) + 取締役・監査役等(officers リスト)
            yakuin_count = 1 + len(company.officers)
            logger.info(
                f'役員数: {yakuin_count}名（代表者1 + その他役員{len(company.officers)}名）'
            )
        else:
            logger.info('役員数: 1名（履歴事項証明書なし or 取得失敗のためフォールバック）')

        wage_report_path = detector.get('wage_report')
        if wage_report_path:
            employees_detail, seishain_count, part_count, yakuin_hoshu_3m = \
                _read_wage_report(wage_report_path)
            logger.info(f'賃金状況報告シート: 正社員{seishain_count}, パート{part_count}')

        # 賃金状況報告シート未取込 + PL に役員報酬がある → PL の役員報酬 ÷ 4 で機械計算
        # （正確な3ヶ月実績ではないが、年間役員報酬を均等割した推定値として埋めておく）
        if yakuin_hoshu_3m == 0 and financial.officer_compensation > 0:
            yakuin_hoshu_3m = int(financial.officer_compensation / 4)
            logger.info(
                f'役員報酬3ヶ月合計を PL から推定: {yakuin_hoshu_3m:,}円 '
                f'(年額{financial.officer_compensation:,}円 ÷ 4)'
            )

        # フォールバック: 賃金状況報告シートで人数が取れなかった場合、賃金台帳から補完
        if seishain_count + part_count == 0:
            # キャッシュがあれば再利用（API呼出スキップ）
            ledger_emps = cached_ledger_employees
            if ledger_emps:
                logger.info(f'賃金台帳: 申請書作成タスクの結果を再利用（API呼出スキップ、{len(ledger_emps)}名）')
            else:
                ledger_paths = detector.get_all('wage_ledger')
                if ledger_paths:
                    fiscal_hint, _ = _resolve_fiscal_period(financial, fiscal_month_override)
                    # 賃金台帳は決定論パーサーで直読する（extractor=None）。
                    # 申請書作成タスク（run_application_transfer → _calc_wage_plan_from_ledger）が
                    # extractor=None で読むのと完全に揃え、②「一人当たり給与支給総額」と
                    # ③「申請書作成」で R215/R216 が一致することを担保する。
                    # AI 抽出経路は中途者の monthly_wages 位置が実行ごとにズレ／1ヶ月丸ごと
                    # 欠落する事例があり（wage_reader._reconcile_midyear_positions_with_deterministic
                    # の補正は値集合一致時のみ＝丸ごと欠落は救えない）、②だけ特定月が
                    # 除外される不整合の原因になっていた。運用上、賃金台帳は事前に
                    # 「賃金台帳の作成」タスクで標準テンプレ形式に整っている前提なので
                    # 決定論パーサーで確実に読める。
                    ledger_emps = read_wage_ledgers(
                        ledger_paths,
                        extractor=None,
                        fiscal_period_hint=fiscal_hint,
                    )
            if ledger_emps:
                # fiscal_hint を渡して時系列順で直近3ヶ月を抽出（非1月始まり対応）
                _fiscal_hint_for_detail, _ = _resolve_fiscal_period(financial, fiscal_month_override)
                employees_detail = _build_employees_detail_from_ledger(
                    ledger_emps, fiscal_period_hint=_fiscal_hint_for_detail,
                )
                # 「契約社員」は正規雇用相当として seishain_count にカウント
                # （wage_calculator.is_full_time_employment と整合）。
                # 給与支給0円の人は R215 算定対象外のため人数にも数えない。
                seishain_count = sum(
                    1 for e in employees_detail
                    if is_full_time_employment(e['type']) and not is_zero_wage_detail(e)
                )
                part_count = sum(
                    1 for e in employees_detail
                    if e['type'] in ('パート・アルバイト',) and not is_zero_wage_detail(e)
                )
                logger.info(
                    f'賃金台帳フォールバック: 正社員{seishain_count}, '
                    f'パート・契約{part_count}'
                )

        # 支給控除一覧/給与データの PDF が唯一の賃金ソースのケース。
        # PDF からの直接読取は未対応（正規ルートは「賃金台帳の作成」タスクで
        # 標準テンプレ形式の Excel に変換してから投入する）。ここで AI 抽出を
        # 呼んでも結果を employees_detail に組み込む経路が無く、API 課金だけが
        # 無駄に走って無言で R215/R216 が空になる。呼び出さず明示警告で変換を促す。
        wage_data_pdf_warning = ''
        wage_data_pdfs = detector.get_all('wage_data')
        if wage_data_pdfs and not employees_detail:
            _names = [p.name for p in sorted(wage_data_pdfs)]
            wage_data_pdf_warning = (
                f' ⚠ 支給控除一覧/給与データのPDF（{len(_names)}件: '
                f'{", ".join(_names[:3])}{"…" if len(_names) > 3 else ""}）は'
                f'直接読み取れません。「賃金台帳の作成」タスクで標準テンプレ形式の'
                f'Excelに変換してから投入してください。'
            )
            logger.warning(
                f'wage_data PDF を検出したが直接読取は未対応のためスキップ（API未呼出）: {_names}'
            )

        # 表示用の期間ラベル。ユーザーが決算月を指定 + AI 推定と不一致なら
        # override 反映後の期間を表示（賃金計算と帳票表示の整合を取る）
        _resolved_period, _ = _resolve_fiscal_period(financial, fiscal_month_override)
        if fiscal_month_override is not None and _resolved_period and '〜' in _resolved_period:
            fiscal_label = _resolved_period
        else:
            fiscal_label = f'{financial.fiscal_year_start} ～ {financial.fiscal_year_end}'

        # データソース（ファイル名）— 人間チェックの突合用に各セクションに表示する
        # per_employee_only モードでは PL 関連の取得・検証も一切行わない
        if per_employee_only:
            _pl_path = None
        else:
            _pl_path = detector.get_pl_latest(fiscal_month_override=fiscal_month_override)
            # cached_financial 経路で line 865 をスキップした場合に備えて、ここでも記録
            if not status.pl_selected_filename:
                _record_pl_selection(status, detector, _pl_path, fiscal_month_override)
        _ledger_paths = detector.get_all('wage_ledger')
        _wage_report = detector.get('wage_report')
        _registry = detector.get('registry')

        # PL の各値が決算書PDFのどのページに記載されているかを機械的に逆引き
        # （AI抽出結果の検証 + 人間チェック時のページ番号案内に兼用）
        # 加えて breakdown 各内訳（販管費側/原価部側）について、AI判定がPDFの
        # セクション見出しと整合しているかも機械的に検証する。
        # PDFテキスト層が無い画像PDFの場合は空辞書になる（フォールバック側で扱う）
        pl_value_pages: dict[str, list[int]] = {}
        pl_breakdown_verification: dict[str, dict] = {}
        if _pl_path and financial is not None and not per_employee_only:
            try:
                from .pdf_text_extractor import (
                    get_pdf_pages_text, find_value_pages,
                    detect_pl_sections, classify_value_by_section,
                )
                _pages = get_pdf_pages_text(_pl_path)
                if _pages and any(p.strip() for p in _pages):
                    for key in ('salary', 'misc_wages', 'bonus', 'legal_welfare',
                                'welfare', 'officer_compensation', 'revenue',
                                'gross_profit', 'operating_profit', 'ordinary_profit',
                                'depreciation'):
                        val = getattr(financial, key, 0)
                        # 0 円の科目は決算書に記載がないことが多いので検証スキップ
                        # （AI誤読ではなく「該当する経費が無い」が正しいケースが大半）
                        if not val:
                            continue
                        pl_value_pages[key] = find_value_pages(_pages, val)
                    logger.info(
                        f'PL 値→ページ逆引き完了: '
                        f'{ {k: v for k, v in pl_value_pages.items() if v} }'
                    )

                    # breakdown 各内訳（pl_section / cost_section）の機械検証
                    # PDFの「販売費及び一般管理費」「製造原価報告書」等の見出しを検出し、
                    # AIが pl_section に分類した値が本当に販管費表に載っているか照合する。
                    pl_sections = detect_pl_sections(_pages)
                    breakdown = getattr(financial, 'breakdown', {}) or {}
                    for key, bd in breakdown.items():
                        if not isinstance(bd, dict):
                            continue
                        pl_v = int(bd.get('pl_section') or 0)
                        cost_v = int(bd.get('cost_section') or 0)
                        pl_breakdown_verification[key] = {
                            'pl_section_value': pl_v,
                            'cost_section_value': cost_v,
                            'pl_section_class': (
                                classify_value_by_section(_pages, pl_sections, pl_v)
                                if pl_v > 0 else 'none'
                            ),
                            'cost_section_class': (
                                classify_value_by_section(_pages, pl_sections, cost_v)
                                if cost_v > 0 else 'none'
                            ),
                            'pl_section_pages': (
                                find_value_pages(_pages, pl_v) if pl_v > 0 else []
                            ),
                            'cost_section_pages': (
                                find_value_pages(_pages, cost_v) if cost_v > 0 else []
                            ),
                        }
                    if pl_breakdown_verification:
                        logger.info(
                            f'PL内訳の機械検証完了: {len(pl_breakdown_verification)}件 '
                            f'(セクション検出ページ数: {len(pl_sections)})'
                        )
                else:
                    logger.info(
                        '決算書PDFのテキスト層が取得できなかったため、ページ番号特定はスキップ'
                    )
            except Exception as e:
                logger.warning(f'PL ページ番号逆引きに失敗: {e}')

        source_files = {
            'pl': _pl_path.name if _pl_path else '',
            'wage_ledger': (
                _ledger_paths[0].name if len(_ledger_paths) == 1
                else f'{_ledger_paths[0].name} 他 {len(_ledger_paths) - 1} 件'
                if _ledger_paths else ''
            ),
            'wage_report': _wage_report.name if _wage_report else '',
            'registry': _registry.name if _registry else '',
            'pl_value_pages': pl_value_pages,
            'pl_breakdown_verification': pl_breakdown_verification,
        }

        create_wage_calculation(
            output_path=output_path,
            company_name=company_name,
            fiscal_year_label=fiscal_label,
            financial=financial,
            seishain_count=seishain_count,
            part_count=part_count,
            yakuin_count=yakuin_count,
            yakuin_hoshu_3m=yakuin_hoshu_3m,
            employees_detail=employees_detail,
            source_files=source_files,
        )

        # per_employee_only: 出力Excelから決算書PDF由来セクションを機械削除
        if per_employee_only:
            from .template_filler import strip_pl_section_from_wage_sheet
            removed = strip_pl_section_from_wage_sheet(output_path)
            logger.info(f'per_employee_only: 決算書由来セクション {removed}行を削除')

        # ユーザー指定の決算月 vs AI 推定の照合（警告のみ）
        _, fiscal_month_warning = _resolve_fiscal_period(financial, fiscal_month_override)

        status.status = '完了'
        status.output_files = [output_path.name]
        task_label = '一人当たり給与支給総額計算' if per_employee_only else '給与支給総額計算'
        status.message = (
            f'{task_label} 完了'
            + fiscal_month_warning + pl_period_warning + wage_data_pdf_warning
        )
        logger.info(f'{task_label} 完了: {output_path.name}')

    except Exception as e:
        from .ai_extractor import APICreditExhaustedError
        if isinstance(e, APICreditExhaustedError):
            status.status = 'エラー'
            status.message = (
                f'⛔ {e}\n'
                f'処理を中断しました。チャージ後にもう一度実行してください。'
            )
            logger.error(f'API残高切れ: 給与計算中断 ({e})')
        else:
            status.status = 'エラー'
            status.message = str(e)
            logger.error(f'エラー: {e}', exc_info=True)

    return status


def _format_fiscal_period(financial: 'FinancialData') -> str | None:
    """FinancialData の fiscal_year_start/end から AI 用ヒント文字列を組み立てる。

    例: '2024-05-01' / '2025-04-30' → '2024-05〜2025-04'
    """
    start = (financial.fiscal_year_start or '').strip()
    end = (financial.fiscal_year_end or '').strip()
    if not start and not end:
        return None
    # YYYY-MM-DD → YYYY-MM
    def _ym(s: str) -> str:
        if len(s) >= 7 and s[4] == '-':
            return s[:7]
        return s
    s_ym = _ym(start)
    e_ym = _ym(end)
    if s_ym and e_ym:
        return f'{s_ym}〜{e_ym}'
    return s_ym or e_ym


def _guess_recent_fiscal_end_year(fiscal_month: int) -> int:
    """指定された決算月から、今日基準で「直近の確定済み決算期末年」を推定する。

    判定は安全側に倒す: 決算月そのもの の月内は **まだ確定していない**とみなし前年扱い。
    （実際の月末日は 28/30/31 の差があり、決算実務でも申告までは「直近期＝先期」と
    扱うのが普通）

    例:
      今日=2026-05-14, fiscal_month=3 → 2026（2026-03 はすでに過ぎている）
      今日=2026-05-14, fiscal_month=5 → 2025（2026-05 はまだ進行中）
      今日=2026-05-14, fiscal_month=6 → 2025（2026-06 はまだ来てない）
    """
    from datetime import date
    today = date.today()
    # 今日の月が決算月より大きい場合のみ「直近期末は今年」と判定。
    # 同月の場合は決算月内＝進行中なので、前年扱いに倒す。
    if today.month > fiscal_month:
        return today.year
    return today.year - 1


def _resolve_fiscal_period(
    financial: 'FinancialData',
    fiscal_month_override: int | None = None,
) -> tuple[str | None, str]:
    """fiscal_period_hint を解決する。

    fiscal_month_override（1〜12）が指定されている場合：
      - その月を期末月としてヒント文字列を再構築（ユーザー指定優先）
      - 期末年は financial.fiscal_year_end が取れていればその年、なければ今日から推定
      - AI 推定の期末月と override がズレていれば警告メッセージを返す

    fiscal_month_override が None の場合：
      - 従来通り AI 抽出の fiscal_year_start/end から組み立てる

    Returns:
        (fiscal_period_hint, warning_message)
          - fiscal_period_hint: '2024-05〜2025-04' 形式 or None
          - warning_message: 不一致警告。一致時 or 未指定時は ''
    """
    warning = ''
    if not fiscal_month_override:
        return _format_fiscal_period(financial), warning

    # 期末年の決定:
    #   - AI 推定の月が override と一致 → AI の year を採用（AI が信頼できる）
    #   - AI 推定の月が override と不一致 → AI の year も誤読の可能性が高いので
    #     今日基準で推定し直す（例: AI=2026-01 のはずが実は 2025-12 だった場合、
    #     AI year 2026 をそのまま流用すると未来の 2026-12 を生成してしまう）
    #   - AI 推定なし → 今日基準で推定
    end_str = (financial.fiscal_year_end or '').strip()
    ai_month: int | None = None
    ai_year: int | None = None
    if end_str and len(end_str) >= 7 and '-' in end_str:
        try:
            ai_year = int(end_str.split('-')[0])
            ai_month = int(end_str.split('-')[1])
        except (ValueError, IndexError):
            ai_year = None
            ai_month = None

    if ai_month is not None and ai_month == fiscal_month_override and ai_year is not None:
        # 月が一致するなら AI year を信用
        end_year = ai_year
    else:
        # 月が不一致 or AI 推定なし → 今日基準で推定
        end_year = _guess_recent_fiscal_end_year(fiscal_month_override)

    # AI 推定月と override の照合
    if ai_month is not None and ai_month != fiscal_month_override:
        warning = (
            f' ⚠ 決算月の不一致: ユーザー指定={fiscal_month_override}月 / '
            f'AI推定={ai_month}月。決算書PDFを目視確認してください'
            f'（ユーザー指定値を優先しました）。'
        )

    # 期首 = 期末の翌月 - 12ヶ月前
    end_ym = f'{end_year:04d}-{fiscal_month_override:02d}'
    if fiscal_month_override == 12:
        start_year = end_year
        start_month = 1
    else:
        start_year = end_year - 1
        start_month = fiscal_month_override + 1
    start_ym = f'{start_year:04d}-{start_month:02d}'

    return f'{start_ym}〜{end_ym}', warning


def _calc_wage_plan_from_ledger(
    detector: FileDetector,
    financial: 'FinancialData',
    extractor=None,
    fiscal_month_override: int | None = None,
) -> tuple[dict[str, float] | None, list, str]:
    """
    賃金台帳から給与支給総額を算出し、年3%成長の計画値を返す。

    extractor が渡されると AI 抽出を優先する（USE_AI_WAGE_EXTRACTION=true 時）。
    AI失敗時は決定論パーサーにフォールバック。

    fiscal_month_override (1〜12) が指定された場合、ユーザー指定の決算月を
    優先して fiscal_period_hint を構築する。AI 推定とズレていれば警告ログ。

    Returns:
        (plan_dict_or_None, employees_raw_list, status_message)
        status_message:
          - '': 正常
          - 'no_ledger': 賃金台帳なし
          - 'no_data': 賃金台帳はあるがデータ抽出失敗
          - 'low_full_year_ratio': 全月在籍者が非役員数の50%未満（自動転記はするが要確認）
          - 'officer_only': 非役員従業員0名＝役員のみの法人（公募要領特例で役員報酬ベース算定が必要）
          - 'no_full_year_target': 非役員はいるが全月在籍者0名（①台帳期間ズレ ②全員12ヶ月未満の可能性）
          - 'zero_total': 上記以外で給与支給総額が0以下
          - 'error': 例外発生
    """
    from .wage_reader import read_wage_ledgers

    ledger_paths = detector.get_all('wage_ledger')
    if not ledger_paths:
        logger.info('賃金台帳が見つかりません → 計画値転記をスキップ')
        return None, [], 'no_ledger'

    fiscal_hint, fiscal_warning = _resolve_fiscal_period(financial, fiscal_month_override)
    if fiscal_warning:
        logger.warning(f'決算月の不一致警告: {fiscal_warning.strip()}')

    try:
        employees_raw = read_wage_ledgers(
            ledger_paths,
            extractor=extractor,
            fiscal_period_hint=fiscal_hint,
        )
        if not employees_raw:
            logger.warning(
                f'賃金台帳からデータを読み取れませんでした '
                f'(ファイル: {[p.name for p in ledger_paths]}, '
                f'fiscal_hint={fiscal_hint})'
            )
            return None, [], 'no_data'

        logger.info(f'賃金台帳: {len(employees_raw)}名読取 ({len(ledger_paths)}ファイル)')

        # WageEmployee → PayrollEmployee に変換（変換ロジックは wage_calculator に集約。
        # 回帰テスト _debug/test_wage_regression.py が本番と同一の変換を検証できるようにする）。
        payroll_list, total_annual_hours, part_fte_fallback_names = (
            wage_employees_to_payroll(employees_raw)
        )
        part_fte_fallback_count = len(part_fte_fallback_names)

        result = calculate_per_capita_wage(payroll_list)

        # 決算書P/Lとのソフト突合（事業年度ウィンドウのズレ・賞与取りこぼし/二重計上の検知補助）。
        # ログのみ（ブロックしない）。賃金台帳R216は12ヶ月在籍者のみ（中途者除外）なので
        # 通常は「P/L人件費 ≥ R216」。R216 が P/L を大きく超えるのは窓ズレ/賞与二重計上の兆候。
        pl_emp_wage = (
            (financial.salary or 0) + (financial.misc_wages or 0) + (financial.bonus or 0)
        )
        if pl_emp_wage > 0 and result.total_salary > 0:
            ratio = result.total_salary / pl_emp_wage
            if ratio > 1.15:
                logger.warning(
                    f'R216×P/L突合: 賃金台帳R216={result.total_salary:,.0f}円 が '
                    f'決算書の給料手当+雑給+賞与={pl_emp_wage:,.0f}円 を大きく超過(×{ratio:.2f})。'
                    f'事業年度ウィンドウのズレ・賞与の二重計上がないか要確認'
                )
            elif ratio < 0.5:
                logger.info(
                    f'R216×P/L突合: 賃金台帳R216={result.total_salary:,.0f}円 が '
                    f'決算書人件費={pl_emp_wage:,.0f}円 の半分未満(×{ratio:.2f})。'
                    f'中途者除外が多い／台帳の月欠落の可能性（要確認）'
                )

        # ── 給与支給総額が0になる原因の切り分け ──────────────────────────
        # 役員除外・中途者除外の結果 total_salary=0 になるケースは、原因により
        # 担当者の取るべき対応（台帳期間の確認 / 役員報酬ベース算定）が異なるため、
        # 簡素な 'zero_total' で潰さず具体的に案内する。
        # 公募要領 p.10:「全月分の給与等の支給を受けた従業員」のみ算定対象。
        #   中途採用・退職等で全月分支給を受けていない者は「除く必要がある」。
        #   「従業員を雇用していない法人は…役員と読み替え、役員報酬と役員数に応じて算出」。
        #
        # 分岐:
        #   (a) non_officer_count == 0（役員のみの法人＝従業員0名）
        #       → 公募要領特例で役員報酬・役員数ベース算定が必要 ('officer_only')。
        #   (b) total_salary<=0 かつ included_count == 0（非役員はいるが全員が全月在籍でない）
        #       → ①台帳期間ズレ ②設立間もない等で全員12ヶ月未満（役員報酬ベース要否を確認）
        #         のいずれか。R215/R216は自動転記せず案内 ('no_full_year_target')。
        #   (c) 0 < included_count < non_officer_count * 0.5（中途入退社が多い案件）
        #       → 公募要領通り全月在籍者のみで算出は正しい。自動転記しつつ注意喚起
        #         ('low_full_year_ratio')。②タスクと結果を一致させる目的。
        non_officer_count = sum(1 for p in payroll_list if not p.is_officer)
        included_count = len(result.included)
        FISCAL_MISMATCH_RATIO = 0.5  # 全月在籍者が非役員数の50%未満は要確認

        if result.total_salary <= 0:
            if non_officer_count == 0:
                logger.warning(
                    '給与支給総額0：非役員の従業員が0名（役員のみの法人）。'
                    '公募要領p.10特例＝役員報酬・役員数ベースで算定が必要。'
                    'R215/R216 は自動転記せず手動入力（役員報酬ベース）が必要です'
                )
                return None, employees_raw, 'officer_only'
            if included_count == 0:
                zero_note = ''
                if result.excluded_zero_names:
                    zero_note = (
                        f'③賃金台帳で給与0円と明記された従業員'
                        f'{len(result.excluded_zero_names)}名は算定対象外、'
                    )
                if result.excluded_partial_zero_names:
                    zero_note += (
                        f'④0円支給月がある従業員'
                        f'{len(result.excluded_partial_zero_names)}名は算定対象外、'
                    )
                logger.warning(
                    f'給与支給総額0：非役員{non_officer_count}名いるが全月在籍者が0名。'
                    f'①賃金台帳の対象期間が直近事業年度12ヶ月とズレ（支給日ベース等）、'
                    f'②設立間もない等で全員が在籍12ヶ月未満（役員報酬ベース算定の要否を要確認）、'
                    f'{zero_note}'
                    f'のいずれかの可能性。R215/R216 は自動転記せず手動確認が必要です'
                )
                return None, employees_raw, 'no_full_year_target'
            logger.warning('給与支給総額が0以下 → 計画値転記をスキップ')
            return None, employees_raw, 'zero_total'

        low_full_year_ratio = (
            non_officer_count >= 2
            and included_count < non_officer_count * FISCAL_MISMATCH_RATIO
        )

        # 給与支給総額ベースで年3%成長の計画値を算出
        base = result.total_salary
        rate = result.GROWTH_RATE
        plan = {
            'employee_count_fte': result.employee_count_fte,
            'wage_total_base': base,
            'wage_total_y1': base * (1 + rate),
            'wage_total_y2': base * (1 + rate) ** 2,
            'wage_total_y3': base * (1 + rate) ** 3,
        }
        if total_annual_hours > 0:
            plan['total_annual_hours'] = round(total_annual_hours, 1)
        if part_fte_fallback_count > 0:
            plan['part_fte_fallback_count'] = part_fte_fallback_count
            plan['part_fte_fallback_names'] = part_fte_fallback_names
        if result.excluded_zero_names:
            plan['excluded_zero_count'] = len(result.excluded_zero_names)
            plan['excluded_zero_names'] = list(result.excluded_zero_names)
            logger.info(
                f'給与支給0円のため算定対象外: {len(result.excluded_zero_names)}名 '
                f'({", ".join(result.excluded_zero_names[:5])}'
                f'{"..." if len(result.excluded_zero_names) > 5 else ""})'
            )
        if result.excluded_partial_zero_names:
            plan['excluded_partial_zero_count'] = len(
                result.excluded_partial_zero_names
            )
            plan['excluded_partial_zero_names'] = list(
                result.excluded_partial_zero_names
            )
            logger.info(
                f'0円支給月があるため算定対象外: '
                f'{len(result.excluded_partial_zero_names)}名 '
                f'({", ".join(result.excluded_partial_zero_names[:5])}'
                f'{"..." if len(result.excluded_partial_zero_names) > 5 else ""})'
            )
        logger.info(
            f'給与支給総額: {base:,.0f}円 '
            f'(従業員FTE: {result.employee_count_fte:.1f}人, 年3%成長, '
            f'総労働時間: {total_annual_hours:,.0f}時間)'
        )
        if part_fte_fallback_count > 0:
            logger.warning(
                f'⚠ パート・アルバイト {part_fte_fallback_count}名で労働時間データなし '
                f'(対象: {", ".join(part_fte_fallback_names[:5])}{"..." if len(part_fte_fallback_names) > 5 else ""}) '
                f'→ FTE=1.0 にサイレント昇格、R215（FTE換算従業員数）が過大計上の可能性。'
                f'IT導入補助金 公募要領 p.10 / マニュアル p.86 では「正社員就業時間に換算した小数値」を要求。'
                f'賃金台帳テンプレ E列「月間平均時間」を顧客に確認のうえ手入力してください。'
            )
        if low_full_year_ratio:
            zero_n = len(result.excluded_zero_names)
            partial_n = len(result.excluded_partial_zero_names)
            # 除外集合は互いに素（partial は年計>0、zero は年計≤0）。引き忘れると
            # 中途入退社の人数が二重計上される
            midyear_n = non_officer_count - included_count - zero_n - partial_n
            zero_part = f'・給与支給0円{zero_n}名' if zero_n else ''
            partial_part = f'・0円支給月あり{partial_n}名' if partial_n else ''
            logger.warning(
                f'全月在籍者({included_count}名)が非役員数({non_officer_count}名)の'
                f'{FISCAL_MISMATCH_RATIO*100:.0f}%未満。中途入退社{midyear_n}名'
                f'{partial_part}{zero_part}は公募要領通り算出対象から除外。'
                f'R215/R216 は自動転記しますが、'
                f'対象者と賃金台帳期間を念のためご確認ください'
            )
            return plan, employees_raw, 'low_full_year_ratio'
        return plan, employees_raw, ''

    except Exception as e:
        # API残高切れは pipeline で全体停止する必要があるので再 raise
        from .ai_extractor import APICreditExhaustedError
        if isinstance(e, APICreditExhaustedError):
            raise
        logger.warning(f'賃金台帳処理エラー（申請書作成は続行）: {e}', exc_info=True)
        return None, [], 'error'


def _build_confidence_warnings(financial) -> list[dict]:
    """FinancialData.confidence から「確認キュー」用の警告リストを構築（Phase 4）。

    UI で「📋 確認キュー」セクションに表示する：項目・元値・根拠・警告理由。
    level='low' のフィールドだけを抽出（high/medium はスルー）。
    """
    if financial is None:
        return []
    conf = getattr(financial, 'confidence', None) or {}
    if not conf:
        return []
    # フィールド名 → 表示ラベル
    label_map = {
        'fiscal_year_start': '事業年度開始日',
        'fiscal_year_end': '事業年度終了日',
        'revenue': '売上高',
        'cost_of_sales': '売上原価',
        'gross_profit': '売上総利益',
        'operating_profit': '営業利益',
        'ordinary_profit': '経常利益',
        'net_profit': '当期純利益',
        'salary': '給料手当',
        'misc_wages': '雑給',
        'bonus': '賞与',
        'officer_compensation': '役員報酬',
        'legal_welfare': '法定福利費',
        'welfare': '福利厚生費',
        'depreciation': '減価償却費',
        'travel_expense': '旅費交通費',
    }
    # source_component → 表示ラベル
    source_map = {
        'basic': '損益計算書本表',
        'PL': '販管費明細',
        'cost': '原価報告書',
        'PL+cost': '販管費 + 原価報告書',
        'unknown': '不明',
    }
    warnings = []
    for field, c in conf.items():
        if not c or getattr(c, 'level', 'high') != 'low':
            continue
        value = getattr(financial, field, None)
        warnings.append({
            'field': field,
            'label': label_map.get(field, field),
            'value': str(value) if value is not None else '(空)',
            'source': source_map.get(getattr(c, 'source_component', ''), getattr(c, 'source_component', '')),
            'reason': getattr(c, 'reason', '') or '抽出失敗',
        })
    return warnings


def _compute_pl_value_pages(
    pl_path: Path | None,
    financial,
    cost_report_path: Path | None = None,
) -> dict[str, dict]:
    """損益計算書PDF（+ 製造原価報告書）のテキスト層を解析し、各 PL 値が
    現れるファイル / ページ番号を逆引きする。

    申請書作成タスクの「データソース」シートで「売上高=決算書PDF p.3」
    「給料手当=決算書PDF p.4 + 製造原価報告書 p.2」のような出所表示を
    出すために使う。

    extract_pl は pl_path + cost_report_path の画像を一括で AI に渡すため、
    両方の PDF を検索対象にしないと製造原価部由来の値（労務費等）が
    「PL.pdf に無い」扱いになって出所表示が空白になる。

    Returns:
        {key: {'pl': [pages], 'cost': [pages]}} 形式。
        ファイルが無い / テキスト層が無い場合は対応キーが空リスト。
    """
    pl_value_pages: dict[str, dict] = {}
    if financial is None:
        return pl_value_pages
    try:
        from .pdf_text_extractor import get_pdf_pages_text, find_value_pages

        # 各 PDF のテキスト層を取得（取れなければ空リスト）
        def _pages_for(path: Path | None) -> list[str]:
            if path is None:
                return []
            try:
                pages = get_pdf_pages_text(path)
                if pages and any(p.strip() for p in pages):
                    return pages
            except Exception as e:
                logger.warning(f'PDFテキスト抽出失敗 {path.name}: {e}')
            return []

        pl_pages_text = _pages_for(pl_path)
        cost_pages_text = _pages_for(cost_report_path)

        if not pl_pages_text and not cost_pages_text:
            # どちらも画像PDF or 失敗 → ページ特定不能
            return pl_value_pages

        for key in ('salary', 'misc_wages', 'bonus', 'legal_welfare',
                    'welfare', 'officer_compensation', 'revenue',
                    'gross_profit', 'operating_profit', 'ordinary_profit',
                    'depreciation', 'cost_of_sales', 'net_profit'):
            val = getattr(financial, key, 0)
            if not val:
                continue
            pl_value_pages[key] = {
                'pl':   find_value_pages(pl_pages_text, val) if pl_pages_text else [],
                'cost': find_value_pages(cost_pages_text, val) if cost_pages_text else [],
            }
    except Exception as e:
        logger.warning(f'PL値→ページ逆引きに失敗（データソースシート出力はスキップ可）: {e}')
    return pl_value_pages


def _check_representative_against_officers(hearing_data: dict, company) -> str:
    """従業員代表者・事業所内最低賃金者の欄に役員名が入っていたら警告文字列を返す。

    役員（代表取締役・取締役・監査役等）は「従業員代表者」「事業所内最低賃金者」の
    対象外。ヒアリングのこれらの欄に役員名が誤入力されているのを検知し、人手修正を促す。
    該当なし／役員不明（company is None・役員0名）なら '' を返す（誤検知防止）。

    氏名照合は wage_reader._normalize_name_key（NFKC＋空白除去＋OCR異体字）で正規化して
    完全一致。役員キー集合の作り方は wage_ledger_writer.match_officer_names_to_employees と同型。
    対象欄はヒアリングのラベル（'従業員代表'／'最低賃金者'）で枠非依存に拾う
    （金額欄 '事業所内最低賃金' は '者' を含まないため対象外）。
    """
    if not hearing_data or company is None:
        return ''
    from .wage_reader import _normalize_name_key

    officer_names: list[str] = []
    rep = getattr(company, 'representative_name', '') or ''
    if rep:
        officer_names.append(rep)
    for o in (getattr(company, 'officers', None) or []):
        n = (o.get('name') if isinstance(o, dict) else getattr(o, 'name', '')) or ''
        if n:
            officer_names.append(n)
    officer_keys = {_normalize_name_key(n) for n in officer_names if str(n).strip()}
    if not officer_keys:
        return ''

    TARGET_LABELS = ('従業員代表', '最低賃金者')
    hits: list[str] = []
    for _row, rec in (hearing_data or {}).items():
        if not isinstance(rec, dict):
            continue
        label = str(rec.get('label', '') or '')
        if not any(t in label for t in TARGET_LABELS):
            continue
        val = rec.get('value')
        if val is None or not str(val).strip():
            continue
        if _normalize_name_key(str(val)) in officer_keys:
            hits.append(f'{label}「{val}」')
    if not hits:
        return ''
    return (
        f' ⚠ 役員名の誤入力の可能性: {"、".join(hits)} が履歴事項証明書の役員と一致します。'
        '役員は従業員代表者・事業所内最低賃金者の対象外です。従業員の氏名に修正してください。'
    )


def _check_industry_code_format(ai_judgment) -> str:
    """AI 生成の業種コードが日本標準産業分類（令和5年6月改定）の細分類4桁形式かチェック。

    旧分類（3桁体系）や AI の自己流コードを検出して警告する。
    プロンプトを強化してもなお古いコードが返るケースを救う。
    AI が int を返してきたケースも考慮（str 化 + NFKC 正規化）。
    """
    if ai_judgment is None:
        return ''
    raw = getattr(ai_judgment, 'industry_code', '')
    if raw is None or raw == '':
        return ''
    # int で返ってきた場合や全角数字対策のため str 化 + NFKC 正規化
    code = unicodedata.normalize('NFKC', str(raw)).strip()
    if not code:
        return ''
    # 細分類は ASCII 半角の4桁数字（NFKC 後は全角数字 → 半角に揃う）
    if not (len(code) == 4 and code.isascii() and code.isdigit()):
        return (
            f' ⚠ 業種コード「{raw}」が日本標準産業分類（令和5年6月改定）の'
            f'細分類4桁形式と異なります。e-Statで再確認してください。'
        )
    return ''


def _check_pl_accounting_consistency(financial) -> str:
    """損益計算書の会計式整合（売上高 − 売上原価 = 売上総利益）をチェック。

    AI が決算書の `(1,234)` `△1,234` `▲1,234` をマイナスとして読み損ねたケースを
    機械的に検出する。整合 or 比較不能なら空文字列、不整合なら警告文字列を返す。

    判定方針:
      - 売上高・売上原価・売上総利益のいずれも 0/None なら判定不能（''）
      - 売上の 0.5% 以内の差は端数誤差として許容
      - それ以上の差は AI の符号読み違いを疑って警告
    """
    if financial is None:
        return ''
    revenue = financial.revenue or 0
    cost = financial.cost_of_sales or 0
    gross = financial.gross_profit or 0
    if revenue <= 0 and cost == 0 and gross == 0:
        return ''
    expected_gross = revenue - cost
    diff = abs(expected_gross - gross)
    diff_ratio = diff / revenue if revenue > 0 else (1.0 if diff > 0 else 0.0)
    if diff_ratio < 0.005:
        return ''
    return (
        f' ⚠ 決算書の会計式不整合: '
        f'売上高({revenue:,}) − 売上原価({cost:,}) = {expected_gross:,} のはずですが、'
        f'抽出された売上総利益は {gross:,}（差 {diff:,}）。'
        f'AIが括弧書きや△/▲記号をマイナスとして読み損ねている可能性があります。'
        f'決算書PDFを目視確認してください。'
    )


def _check_wage_pl_consistency(
    wage_plan: dict | None,
    financial: 'FinancialData',
    tolerance: float = 0.20,
    severe_threshold: float = 0.50,
) -> str:
    """賃金台帳合計と損益計算書の人件費を照合し、不整合があれば警告文字列を返す。

    Args:
        wage_plan: _calc_wage_plan_from_ledger の戻り値（基準年の給与支給総額を含む）
        financial: 損益計算書から抽出した財務データ
        tolerance: 通常警告の許容差（デフォルト ±20%）
        severe_threshold: 強警告に格上げする差分比（デフォルト 50%）

    Returns:
        警告文字列（不整合あり）または空文字列（整合 / 比較不能）

    判定方針:
        - wage_plan['wage_total_base'] は wage_calculator.calculate_per_capita_wage で
          **役員報酬を除外** して算出されている（[wage_calculator.py:49] 参照）。
          そのため PL 側からも officer_compensation を除いて比較する必要がある。
        - 比較対象 = PL の (給料手当 + 雑給 + 賞与) ※役員報酬除く
        - PL に給与系の計上が無い場合は判定不能 → '' を返す
        - 差が severe_threshold を超えたら強警告（⛔ + 「賃金台帳が著しく〜」）
        - 差が tolerance を超えたら通常警告（⚠ + 「〜の差があります」）
        - いずれも処理は続行する
    """
    if not wage_plan or 'wage_total_base' not in wage_plan:
        return ''
    # 役員報酬は除外（賃金台帳側も役員除外で集計しているため）
    pl_personnel = (
        (financial.salary or 0)
        + (financial.misc_wages or 0)
        + (financial.bonus or 0)
    )
    if pl_personnel <= 0:
        return ''  # PL データなし → 比較不能
    ledger_total = wage_plan['wage_total_base']
    if ledger_total <= 0:
        return ''
    diff_ratio = abs(ledger_total - pl_personnel) / pl_personnel
    if diff_ratio <= tolerance:
        return ''
    direction = '多い' if ledger_total > pl_personnel else '少ない'
    if diff_ratio >= severe_threshold:
        # 著しい乖離 — AI 抽出の打ち切り / 雇用区分誤認 / 中途退職者多数 等の可能性大
        return (
            f' ⛔ 強警告: 賃金台帳合計({ledger_total:,.0f}円)と損益計算書の人件費'
            f'({pl_personnel:,.0f}円, 役員報酬除く)に{diff_ratio*100:.0f}%の著しい差があります'
            f'（賃金台帳が著しく{direction}）。賃金台帳の抽出漏れがないか必ず確認してください'
        )
    return (
        f' ⚠ 賃金台帳合計({ledger_total:,.0f}円)と損益計算書の人件費'
        f'({pl_personnel:,.0f}円, 役員報酬除く)に{diff_ratio*100:.0f}%の差があります'
        f'（賃金台帳が{direction}）。AI抽出が月数や雇用区分を取り違えていないか確認してください'
    )


def _classify_emp_type(emp_type: str) -> str:
    """賃金台帳の雇用形態文字列を4分類に正規化"""
    t = (emp_type or '').strip()
    if '役員' in t or '取締役' in t:
        return '役員'
    if 'パート' in t or 'アルバイト' in t or '非常勤' in t:
        return 'パート・アルバイト'
    if '契約' in t:
        return '契約社員'
    # 雇用形態が空/不明な場合は正社員として集計（台帳に区分列が無い一般的なケース）
    return '正社員'


def _fiscal_month_order(fiscal_period_hint: str | None) -> list[int]:
    """fiscal_period_hint から「事業年度内の月の時系列順 Index リスト」を返す。

    例: '2024-05〜2025-04' → [4,5,6,7,8,9,10,11,0,1,2,3]（5月始まり、12月の次は1月）
    None や形式不明な場合は [0,1,...,11]（カレンダー順 = 1月始まり）にフォールバック。
    """
    if not fiscal_period_hint:
        return list(range(12))
    # 'YYYY-MM〜YYYY-MM' / 'YYYY-MM-DD〜...' / 'YYYY-MM' から開始月を抽出
    m = re.search(r'(\d{4})-(\d{1,2})', fiscal_period_hint)
    if not m:
        return list(range(12))
    start_month = int(m.group(2))  # 1〜12
    if not 1 <= start_month <= 12:
        return list(range(12))
    start_idx = start_month - 1  # 0〜11
    return [(start_idx + i) % 12 for i in range(12)]


def _build_employees_detail_from_ledger(
    employees,
    fiscal_period_hint: str | None = None,
) -> list[dict]:
    """賃金台帳の読取結果 → create_wage_calculation が期待する employees_detail 形式に変換。
    役員はカウント対象外として除外する。

    直近3ヶ月の判定方針:
        fiscal_period_hint があれば「事業年度内の月の時系列順」で末尾3つを採用。
        例: 5月開始の年度なら 5,6,7,...,2,3,4 の順序で並べた中の最後3つ。
        fiscal_period_hint がなければカレンダー順末尾3つ（フォールバック、従来動作）。
        この区別が無いと、12月始まりの台帳で「直近3 = 配列末尾の10/11/12月」になり、
        実際の時系列上の直近（9/10/11月）とズレる。
    """
    detail = []
    month_order = _fiscal_month_order(fiscal_period_hint)
    for emp in employees:
        classified = _classify_emp_type(emp.employment_type)
        if classified == '役員':
            continue

        # 事業年度内の時系列順で「支給>0 の月」を取り、末尾3つを直近として採用。
        # 0円明記月は「支給を受けていない月」（公募要領 p.10、v0.2.69）として
        # 在籍月に数えない（「賃金台帳の作成」産の台帳は不在月を0円で出力するため、
        # non-None 判定では中途入社者が full_year=True になり R215/R216 に誤算入）。
        # 支給>0 の月が0件の人（全月0円＝賞与のみ受給者等）は従来の non-None 判定に
        # フォールバックし、full_year=True を維持（賞与のみ算入仕様・v0.2.66）。
        # ※ wage_calculator.is_full_year_paid と同一ルール。変える場合は両方同期。
        paid_months_ordered = [
            idx for idx in month_order
            if idx < len(emp.monthly_wages) and (emp.monthly_wages[idx] or 0) > 0
        ]
        if paid_months_ordered:
            ordered_months_with_data = paid_months_ordered
        else:
            ordered_months_with_data = [
                idx for idx in month_order
                if idx < len(emp.monthly_wages) and emp.monthly_wages[idx] is not None
            ]
        tenure_months = len(ordered_months_with_data)
        full_year = tenure_months >= 12

        last_three = ordered_months_with_data[-3:]
        m_vals = [emp.monthly_wages[m] or 0 for m in last_three]
        # 中途者向け: 表示列に対応する暦月ラベル（実体が分かるように）
        last_three_labels = [f'{m + 1}月' for m in last_three]
        while len(m_vals) < 3:
            m_vals.append(0)
            last_three_labels.append('')

        # 12ヶ月分の生データ（事業年度内の時系列順で並べる）。
        # wage_calculator 側で12ヶ月明細表示と賃金台帳ベース給与支給総額の算定に使う。
        # データが無い月は値0 / マスクFalse として渡し、表示側でグレーアウト判定する。
        monthly_wages_full = [
            float(emp.monthly_wages[idx] or 0) for idx in month_order
        ]
        monthly_hours_full = [
            float(emp.monthly_hours[idx] or 0) for idx in month_order
        ]
        month_labels_full = [f'{idx + 1}月' for idx in month_order]
        month_data_mask = [
            emp.monthly_wages[idx] is not None for idx in month_order
        ]

        detail.append({
            'no': len(detail) + 1,
            'name': emp.name,
            'type': classified,
            'm1': m_vals[0],
            'm2': m_vals[1],
            'm3': m_vals[2],
            'hr': emp.hourly_rate,
            'monthly_hours': emp.monthly_avg_hours,
            'judge': '',
            # 中途入退社の扱いを正しくするための追加情報
            'tenure_months': tenure_months,
            'full_year': full_year,
            'last_three_labels': last_three_labels,
            # 12ヶ月明細・賃金台帳ベース集計用
            'monthly_wages_full': monthly_wages_full,
            'monthly_hours_full': monthly_hours_full,
            'month_labels_full': month_labels_full,
            'month_data_mask': month_data_mask,
            # 年間賞与（R216 に算入。月次セルには混ぜず Sheet2 の年間給与計に加算する）
            'annual_bonus': float(getattr(emp, 'annual_bonus', 0.0) or 0.0),
        })
    return detail


def _read_wage_report(path: Path) -> tuple[list[dict], int, int, int]:
    """
    賃金状況報告シートから従業員データを読取。
    Returns: (employees_detail, seishain_count, part_count, yakuin_hoshu_3m)
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)

    # シート名を探す
    ws = None
    for name in wb.sheetnames:
        if '賃金' in name and 'マスタ' not in name and '元データ' not in name:
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    def _to_num(v) -> float:
        # 賃金状況報告シートのセルは数値想定だが、ユーザー入力で
        # '300,000円' / '― ' / 'N/A' / '' のような文字列が混じることがあり
        # そのまま比較すると TypeError: '>' not supported between str and int で落ちる。
        if v is None:
            return 0.0
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).replace(',', '').replace(' ', '').replace('円', '').replace('時間', '').strip()
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0

    # 役員報酬（行13, D列）
    yakuin_hoshu_3m = _to_num(ws.cell(13, 4).value)

    employees = []
    for row in ws.iter_rows(min_row=19, max_row=200):
        no = row[1].value
        name = row[2].value
        if name is None or no is None:
            continue

        m1_base = _to_num(row[5].value)
        m1_hr = _to_num(row[6].value)
        m2_base = _to_num(row[8].value)
        m2_hr = _to_num(row[9].value)
        m3_base = _to_num(row[11].value)
        m3_hr = _to_num(row[12].value)
        judge = row[14].value if len(row) > 14 else ''

        # 時間推定
        hours = []
        for base, hr in [(m1_base, m1_hr), (m2_base, m2_hr), (m3_base, m3_hr)]:
            if hr > 0 and base > 0:
                hours.append(base / hr)
        avg_hours = sum(hours) / len(hours) if hours else 0

        # 正社員/パート判定: 時給1300円以上 and 月給18万以上 → 正社員の傾向
        avg_base = (m1_base + m2_base + m3_base) / 3
        avg_hr = (m1_hr + m2_hr + m3_hr) / 3
        emp_type = '正社員' if avg_base >= 180000 and avg_hr >= 1200 else 'パート・アルバイト'

        employees.append({
            'no': no,
            'name': str(name).strip(),
            'type': emp_type,
            'm1': m1_base,
            'm2': m2_base,
            'm3': m3_base,
            'hr': round(avg_hr),
            'monthly_hours': round(avg_hours, 1),
            'judge': judge or '',
            # 賃金状況報告シート由来は在籍中の社員のみが載る前提のため、
            # 全員 12ヶ月在籍扱いとする（賃金台帳の中途入退社検出とは別系統）
            'tenure_months': 12,
            'full_year': True,
            'last_three_labels': ['', '', ''],
        })

    wb.close()

    # 給与支給0円（3ヶ月とも0）の人は R215 算定対象外のため人数に数えない。
    # employees リストには残す（従業員別明細シートで人間が突合できるように）。
    countable = [e for e in employees if not is_zero_wage_detail(e)]
    seishain = [e for e in countable if e['type'] == '正社員']
    part = [e for e in countable if e['type'] != '正社員']
    return employees, len(seishain), len(part), yakuin_hoshu_3m


def run_bonus_wage_ledger_creation(
    resource_folder: Path,
    company_name: str,
    template_path: Path,
    output_path: Path,
    extractor: BaseExtractor | None = None,
    prefecture: str = '',
    application_ym: tuple[int, int] | None = None,
    selection_override: dict[str, list[Path]] | None = None,
) -> ProcessingStatus:
    """タスク「加点判定用賃金台帳の作成」: 生の賃金台帳/給与明細 → 加点判定用 Excel。

    基本給と所定労働時間を AI 抽出し、令和6年10月〜令和7年9月＋交付申請直近月の
    時間換算給与（＝基本給÷所定労働時間）ベースの専用台帳に変換する。
    Document AI 一本（Sonnet 画像経路フォールバックは無効）。
    抽出精度が加点判定（補助率2/3 トリガー）を左右するため、欠落を警告し人手確認を促す。
    """
    from .ai_extractor import APICreditExhaustedError, ImageFallbackBlockedError
    from .wage_reader import (
        read_bonus_source_employees, BONUS1_WINDOW, prev_month,
    )
    from .bonus_wage_ledger_writer import write_bonus_wage_ledger

    status = ProcessingStatus(
        company_name=company_name,
        template_type='加点判定用賃金台帳の作成',
        status='処理中',
    )
    try:
        if extractor is None:
            extractor = create_extractor(CLAUDE_API_KEY)

        detector = FileDetector(
            resource_folder,
            selection_override=selection_override,
            extra_allowed_exts={'wage_ledger': {'.pdf'}},
        )
        logger.info(detector.summary())
        wage_files = detector.get_all('wage_ledger')
        if not wage_files:
            status.status = 'エラー'
            status.message = (
                '賃金台帳/給与明細ファイルが見つかりません。'
                'PDF/Excel/CSV のいずれかをアップロードしてください'
                '（ファイル名に「賃金台帳」または「給与台帳」を含めてください）。'
            )
            return status

        try:
            employees = read_bonus_source_employees(
                wage_files, extractor=extractor, disable_image_fallback=True,
            )
        except ImageFallbackBlockedError as e:
            status.status = 'エラー'
            status.message = (
                f'⛔ Document AI で賃金台帳を抽出できませんでした（{str(e).split("。")[0]}）。'
                '手書き/画像PDF は手元の Claude Code（wagebook-convert）での変換をご検討ください。'
            )
            return status
        except APICreditExhaustedError as e:
            status.status = 'エラー'
            status.message = f'⛔ API残高切れで抽出を継続できませんでした（{e}）。'
            return status

        if not employees:
            status.status = 'エラー'
            status.message = (
                '⛔ 賃金台帳から従業員データを抽出できませんでした。'
                'PDF原本のレイアウト・手書きの有無を確認してください。'
            )
            return status

        write_bonus_wage_ledger(
            employees, template_path, output_path,
            prefecture=prefecture, application_ym=application_ym,
        )

        # カバレッジ警告（加点に必要な月の基本給・所定労働時間がどれだけ埋まったか）
        target_yms = list(BONUS1_WINDOW)
        latest_ym = prev_month(application_ym) if application_ym else None
        if latest_ym:
            target_yms.append(latest_ym)
        no_base = [
            e.name for e in employees
            if not any(ym in e.monthly_base for ym in target_yms)
        ]
        no_hours = [e.name for e in employees if e.scheduled_hours is None]

        msgs = [f'{len(employees)}名の加点判定用賃金台帳を作成しました。']
        if no_base:
            head = '、'.join(no_base[:5]) + ('…' if len(no_base) > 5 else '')
            msgs.append(
                f'⚠️ 加点対象月の基本給が取得できなかった従業員: {len(no_base)}名（{head}）。'
                '原本を確認し、各月の基本給（所定内賃金）を手入力してください。'
            )
        if no_hours:
            head = '、'.join(no_hours[:5]) + ('…' if len(no_hours) > 5 else '')
            msgs.append(
                f'⚠️ 月間所定労働時間が取得できなかった従業員: {len(no_hours)}名（{head}）。'
                'E列に会社の所定労働時間（例: 160）を手入力してください（最低賃金比較の分母）。'
            )
        if application_ym is None:
            msgs.append(
                '⚠️ 交付申請月が未指定のため、加点措置②の直近月列（R列）が空です。'
                'サイドバーで交付申請月を指定して再作成するか、台帳の C3 と R列を手入力してください。'
            )
        msgs.append(
            '※ 出力台帳は提出前に必ず人手で確認してください'
            '（基本給・所定労働時間の抽出精度が加点判定＝補助率2/3 を左右します）。'
        )

        status.status = '完了'
        status.message = '\n\n'.join(msgs)
        status.output_files = [output_path.name]
        return status

    except Exception as e:
        logger.error(f'加点判定用賃金台帳の作成エラー: {e}', exc_info=True)
        status.status = 'エラー'
        status.message = f'処理中にエラーが発生しました: {e}'
        return status


def _bonus_period_shift_warning(
    employees, fiscal_month_override: int | None,
) -> str | None:
    """支給日なし賞与 × 非暦年決算の賞与期間ズレ警告（③）。該当なければ None。

    支給日（支給年月）の無い賞与は事業年度ウィンドウで絞り込めず、暦年ベースの年間賞与が
    そのまま算入される（WageEmployee.bonus_undated_total > 0）。決算月が12月以外（非暦年）
    だと対象事業年度とズレ、R216（給与支給総額）の賞与額が不正確になり得る。
    「賃金台帳の作成」タスクで賞与の支給月を入力すれば対象年度に自動で絞り込む
    （_sum_window_bonuses の undated_paid_months）。入力が無い／窓内外が混在で按分不能の
    ときは bonus_undated_total が残るので、本警告で人手確認を促す。
    """
    undated = [
        getattr(e, 'name', '?') for e in (employees or [])
        if (getattr(e, 'bonus_undated_total', 0.0) or 0.0) > 0
    ]
    if not undated or fiscal_month_override == 12:
        return None
    month_label = (
        f'{fiscal_month_override}月決算'
        if fiscal_month_override else '決算月未指定'
    )
    return (
        f'⚠ 賞与の対象期間に注意（{month_label}・非暦年）: '
        f'{len(undated)}名の賞与に支給日（支給年月）が無く、'
        '対象事業年度の12ヶ月で絞り込めていません'
        '（暦年ベースの年間賞与をそのまま年間賞与として算入）。'
        'サイドバー「賞与の支給月」に支給年月を入力すると対象事業年度で自動補正します。'
        '入力できない場合は、対象事業年度内の賞与かを賞与明細・勘定科目内訳明細書・'
        f'PDF原本で確認してください。対象者: {undated}'
    )


def _bonus_dropped_info_note(employees) -> str | None:
    """対象事業年度ウィンドウ外として除外した賞与（支給年月あり）を情報表示（③の補足）。

    多年度分が載った台帳では正常な除外でも出るため、アラート（⚠）ではなく情報（ℹ）。
    AI の年月誤読で本来年度内の賞与が落ちていないかを人手確認する導線として surface する。
    WageEmployee.bonus_dropped_total（_sum_window_bonuses が窓外で除外した dated 賞与額）を集計。
    """
    rows = [
        (getattr(e, 'name', '?'), float(getattr(e, 'bonus_dropped_total', 0.0) or 0.0))
        for e in (employees or [])
        if (getattr(e, 'bonus_dropped_total', 0.0) or 0.0) > 0
    ]
    if not rows:
        return None
    total = sum(amt for _n, amt in rows)
    names = '、'.join(f'{n}({amt:,.0f}円)' for n, amt in rows)
    return (
        f'ℹ 対象事業年度外として除外した賞与: {len(rows)}名・計{total:,.0f}円。'
        '多年度分が載った台帳なら正常です。年月の誤読で本来年度内の賞与が落ちていないか、'
        f'賞与明細・PDF原本でご確認ください。対象: {names}'
    )


def _officer_autodetect_skipped_warning(registry_path) -> str | None:
    """履歴事項PDF未添付で役員の自動確定（氏名照合）が走らないときの注意（G）。

    registry_path が無いと officer_names=[] となり、役員の雇用形態上書き照合がスキップされる。
    役員が従業員に混在しているとR216（給与支給総額）が過大になり補助率判定に不利方向へ
    ズレ得るため、原本での雇用形態確認を促す。registry_path があれば None。
    """
    if registry_path:
        return None
    return (
        '⚠ 履歴事項全部証明書PDFが未添付のため、役員の自動確定（氏名照合）を実施していません。'
        '役員が従業員に混在しているとR216（給与支給総額）が過大になります。'
        '役員に該当する人の雇用形態を原本で確認してください。'
    )


def _recon_name_key(name: str) -> str:
    """氏名照合キー: NFKC正規化＋空白除去。"""
    return re.sub(r'\s+', '', unicodedata.normalize('NFKC', str(name or '')))


def _reconcile_monthly_with_layout(employees, pdf_layout) -> list[str]:
    """AI抽出の月配置を、PDFから決定論的に読んだ実列で補正する（落とし穴①の根治）。

    給与ソフトの「N月度（締め期間＝前月／末締め翌月払い）」レイアウトでは、AIが締め期間を
    勤務月と解釈して全体が1ヶ月ズレ、事業年度ウィンドウで境界月が脱落 → 令和N年12月が空 →
    全員「全月在籍でない」と誤判定され R216/R215 が 0 になる事故があった（実案件A）。
    決定論パーサー(`wage_pdf_layout_parser`)は「N月度」を月番号（度ラベル＝支給サイクル＝決算書
    と一致）で読むので、その月別課税支給合計を「月配置の正解」として AI 出力を上書き補正する。

    安全ゲート（誤補正＝別バグ混入の防止）:
      - 氏名一致する従業員のみ。
      - 決定論側の埋まり月数 >= AI側の埋まり月数（決定論が同等以上に揃っている）。
      - AIの非空値が決定論値の部分多重集合（＝同じ金額群＝純粋な配置ズレと確認できるときだけ上書き）。
        決定論が別年度・誤読で異なる値を持つ場合は上書きしない（AIのまま＋要確認）。

    Returns:
        補正・要確認のメモ文字列リスト（変換メモ／警告に載せる）。employees は in-place 補正。
    """
    if not employees or not pdf_layout:
        return []
    from collections import Counter

    by_name: dict[str, object] = {}
    for p in pdf_layout:
        key = _recon_name_key(getattr(p, 'name', ''))
        if key and key not in by_name:
            by_name[key] = p

    corrected: list[str] = []
    unverified: list[str] = []
    for e in employees:
        name = getattr(e, 'name', '') or ''
        ai = list(getattr(e, 'monthly_wages', []) or [])
        if len(ai) != 12:
            continue
        p = by_name.get(_recon_name_key(name))
        det_map = dict(getattr(p, 'monthly_taxable_totals', {}) or {}) if p else {}
        det = [det_map.get(m) for m in range(1, 13)]
        det_count = sum(1 for v in det if v)
        ai_count = sum(1 for v in ai if v)
        if det_count == 0:
            if ai_count:
                unverified.append(name)  # 決定論で読めず（極小ページ等）。AIのまま要確認。
            continue
        det_norm = [int(v) if v else None for v in det]
        ai_norm = [int(v) if v else None for v in ai]
        if det_norm == ai_norm:
            continue  # 既に一致。補正不要。
        subset = not (Counter(v for v in ai_norm if v) - Counter(v for v in det_norm if v))
        if det_count >= ai_count and subset:
            e.monthly_wages = [float(v) if v else None for v in det]
            corrected.append(name)
        else:
            unverified.append(name)  # 値が食い違う＝純粋なズレでない → 上書きせず要確認。

    notes: list[str] = []
    if corrected:
        notes.append(
            f'✅ 月配置をPDF実列で自動補正: {len(corrected)}名'
            '（AIの月ズレを決定論パーサーの月別課税支給合計で修正。'
            f'末締め翌月払い等の「N月度」レイアウト対策）: {corrected}'
        )
    if unverified:
        notes.append(
            f'⚠ 月配置の決定論照合ができなかった: {len(unverified)}名'
            '（極小ページ/値不一致）。PDF原本で月配置を確認してください: '
            f'{unverified}'
        )
    return notes


def run_wage_ledger_conversion(
    resource_folder: Path,
    company_name: str,
    template_path: Path,
    output_path: Path,
    extractor: BaseExtractor | None = None,
    fiscal_month_override: int | None = None,
    is_kojin: bool = False,
    selection_override: dict[str, list[Path]] | None = None,
    bonus_paid_months: list[tuple[int | None, int]] | None = None,
) -> ProcessingStatus:
    """タスク「賃金台帳の作成」: PDF/Excel/CSV → ツール規格 Excel 賃金台帳一覧。

    Document AI + Sonnet 4.6 一本で抽出する（Sonnet 画像経路フォールバックは無効）。
    Document AI が失敗した場合は明示エラーで停止し、
    ローカル（Claude Code）での手動変換を案内する。

    Args:
        resource_folder: 入力ファイルの置き場（FileDetector で wage_ledger を取得）
        company_name: 出力ファイル名・タイトルに使用
        template_path: 賃金台帳テンプレートExcel (`ツール/賃金台帳テンプレート.xlsx`)
        output_path: 出力 xlsx のパス（通常 `{会社名}_賃金台帳一覧.xlsx`）
        extractor: AI 抽出器（None なら新規作成）
        fiscal_month_override: 決算月（1〜12）。AI への事業年度フィルタ用
        is_kojin: 個人事業主テンプレ選択時に True（雇用形態正規化が変わる）
        selection_override: ファイル手動選択
        bonus_paid_months: 賞与の支給年月リスト [(年, 月), ...]（年は None 可）。
            支給日なし賞与（年間集計表等）を対象事業年度で絞り込むためのユーザー入力。
            全て窓内→全額算入 / 全て窓外→除外 / 混在→全額算入＋要確認。未指定なら従来動作。

    Returns:
        ProcessingStatus（status: '完了' / 'エラー'）
    """
    from .ai_extractor import APICreditExhaustedError, ImageFallbackBlockedError
    from .wage_ledger_writer import (
        detect_handwritten_pdf,
        write_wage_ledger_to_template,
    )
    from .models import FinancialData

    status = ProcessingStatus(
        company_name=company_name,
        template_type='賃金台帳の作成',
        status='処理中',
    )

    try:
        if extractor is None:
            extractor = create_extractor(CLAUDE_API_KEY)

        # 賃金台帳の作成タスクでは PDF も賃金台帳カテゴリで受け付ける（拡張子追加）
        # 履歴事項PDF（registry）も役員照合のために検出対象に含める
        detector = FileDetector(
            resource_folder,
            selection_override=selection_override,
            extra_allowed_exts={'wage_ledger': {'.pdf'}},
        )
        logger.info(detector.summary())

        wage_files = detector.get_all('wage_ledger')
        registry_path = detector.get('registry')
        if not wage_files:
            status.status = 'エラー'
            status.message = (
                '賃金台帳ファイルが見つかりません。'
                'PDF/Excel/CSV のいずれかをアップロードしてください'
                '（ファイル名に「賃金台帳」または「給与台帳」を含めてください）。'
            )
            return status

        # 決算月から事業年度ヒント文字列を組み立てる（決算書 PDF は読まない）
        fiscal_hint, _ = _resolve_fiscal_period(FinancialData(), fiscal_month_override)
        logger.info(
            f'賃金台帳作成: 対象期間ヒント={fiscal_hint or "(指定なし)"} / '
            f'対象ファイル={len(wage_files)}件'
        )

        # PDF ファイルだけ手書き判定（テキスト層が薄ければ警告対象に乗せる）
        handwritten_files: list[str] = []
        for f in wage_files:
            if f.suffix.lower() == '.pdf':
                is_handwritten, reason = detect_handwritten_pdf(f)
                if is_handwritten:
                    handwritten_files.append(f.name)
                    logger.warning(
                        f'手書きPDF判定: {f.name} ({reason}) — 抽出は継続、精度低下警告を出力'
                    )

        # ── 履歴事項PDFがあれば役員リストを取得（A 改善: 役員自動判定の最優先経路） ──
        officer_names: list[str] = []
        if registry_path:
            try:
                from .pdf_reader import pdf_to_images
                images = pdf_to_images(registry_path)
                company_info = extractor.extract_registry(images)
                if company_info:
                    if company_info.representative_name:
                        officer_names.append(company_info.representative_name)
                    for o in company_info.officers or []:
                        n = (o.get('name') if isinstance(o, dict) else getattr(o, 'name', '')) or ''
                        if n:
                            officer_names.append(n)
                    logger.info(
                        f'履歴事項PDF: {registry_path.name} → 役員{len(officer_names)}名 '
                        f'({officer_names})'
                    )
            except APICreditExhaustedError:
                # 残高切れは賃金台帳抽出にも失敗するので即停止せず、後段の挙動に任せる
                logger.warning('履歴事項PDF抽出で残高切れ — 役員自動判定はスキップして続行')
            except Exception as e:
                logger.warning(
                    f'履歴事項PDF抽出失敗: {e} — 役員自動判定はスキップして続行'
                )

        # AI抽出（Document AI 一本、Sonnet 画像フォールバック無効）
        try:
            employees = read_wage_ledgers(
                wage_files,
                extractor=extractor,
                fiscal_period_hint=fiscal_hint,
                disable_image_fallback=True,
                # 決算書を読まないこのタスクは今日基準で期末年を推定するため、台帳の実在年で
                # 12ヶ月窓を選び直す（1期ズレによる R216 サイレント過少計上を防ぐ）。
                derive_year_from_data=True,
                # 支給日なし賞与（年間集計表等）をユーザー入力の支給月で対象年度に絞る
                undated_paid_months=bonus_paid_months,
            )
        except ImageFallbackBlockedError as e:
            status.status = 'エラー'
            status.message = (
                f'⛔ Document AI で賃金台帳を抽出できませんでした（{str(e).split("。")[0]}）。'
                'PDF が画像品質的に Document AI で読めない可能性があります。'
                '手元の Claude Code に wagebook-convert Skill をインストールして手動変換してください — '
                'Streamlit アプリ上部「📘 賃金台帳の作成手順（CC向け Skill）」expander から ZIP を取得。'
            )
            logger.error(f'画像フォールバック禁止のため停止: {e}')
            return status
        except APICreditExhaustedError as e:
            status.status = 'エラー'
            status.message = (
                f'⛔ API残高切れで抽出を継続できませんでした（{e}）。'
                'API残高をチャージしてから再実行してください。'
            )
            return status

        if not employees:
            status.status = 'エラー'
            status.message = (
                '⛔ 賃金台帳から従業員データを抽出できませんでした。'
                'PDF原本を確認し、レイアウトが極端に崩れていないか、'
                '手書きでないかをチェックしてください。'
            )
            return status

        # 抽出経路ラベルをログから推定（直近の API 送信ログに記録される path=... を拾えないため、
        # ここではフラグ状態から推測する）
        from .config import (
            USE_DOCUMENT_AI_SONNET_EXTRACTION,
            USE_OCR_HAIKU_EXTRACTION,
            USE_DOCUMENT_AI_OCR,
        )
        if USE_DOCUMENT_AI_SONNET_EXTRACTION:
            extraction_path = 'C(DocAI+Sonnet)'
        elif USE_OCR_HAIKU_EXTRACTION:
            extraction_path = 'B(DocAI+Haiku)'
        elif USE_DOCUMENT_AI_OCR:
            extraction_path = 'C-implicit(DocAI+Sonnet)'
        else:
            extraction_path = 'TextOnly(Sonnet)'

        # データソースファイル名一覧（変換メモシート用）
        data_source_files = [f.name for f in wage_files]
        if registry_path:
            data_source_files.append(f'{registry_path.name}（履歴事項 — 役員照合用）')

        # ── セル単位整合性チェック ──
        # PDF テキストを別途取得して、AI 抽出結果と物理列構造を突合する。
        # 月給漏れ・賞与漏れ・月配置ズレを書き込み前に検知し、変換メモシートに警告を載せる。
        # PDF 以外（Excel/CSV）が混在する場合は、PDF ファイルだけに対して実施。
        #
        # 注意: pdf_text_extractor.extract_pdf_as_text_with_source は本番フラグ次第で
        # Document AI 経由のフラットなテキストを返す。本パーサーはテーブル形式
        # （pdfplumber/PyMuPDF のページ別テキスト）に依存しているため、
        # parse_wage_ledger_layout_from_pdf でテキスト取得経路を独自に確保する。
        cell_consistency_warnings: list[str] = []
        try:
            from .wage_pdf_layout_parser import (
                parse_wage_ledger_layout_from_pdf, summarize_layout,
            )
            from .wage_validator import check_cell_level_consistency

            pdf_layout_all = []
            for wf in wage_files:
                if wf.suffix.lower() != '.pdf':
                    continue
                try:
                    with open(wf, 'rb') as fp:
                        layout = parse_wage_ledger_layout_from_pdf(fp.read())
                    if layout:
                        pdf_layout_all.extend(layout)
                        logger.info(
                            f'PDFレイアウト解析: {wf.name} → {len(layout)}名'
                        )
                        logger.debug(summarize_layout(layout))
                    else:
                        logger.info(
                            f'PDFレイアウト解析: {wf.name} → 0名 '
                            f'(画像PDF or 未対応レイアウト、検証スキップ)'
                        )
                except Exception as inner:
                    logger.warning(
                        f'PDFレイアウト解析失敗（スキップ）: {wf.name} - {inner}'
                    )

            if pdf_layout_all:
                # 月配置の自己補正（決定論実列で AI の月ズレを直す）。employees を in-place 補正。
                # 「N月度＋締め前月」等のレイアウトで R216/R215 が0になる事故の根治（法人・個人共通）。
                correction_notes = _reconcile_monthly_with_layout(
                    employees, pdf_layout_all,
                )
                if correction_notes:
                    logger.info('月配置 自己補正:\n' + '\n'.join(correction_notes))

                # AI 抽出結果（補正後）を validator が読める dict 列に正規化して整合性チェック
                emp_dicts = [
                    {
                        'name': getattr(e, 'name', '') or '',
                        'employment_type': getattr(e, 'employment_type', '') or '',
                        'monthly_wages': list(getattr(e, 'monthly_wages', []) or []),
                    }
                    for e in employees
                ]
                cell_consistency_warnings = check_cell_level_consistency(
                    emp_dicts, pdf_layout_all,
                )
                # 補正メモを先頭に載せる（人手チェック用の変換メモへ）
                cell_consistency_warnings = correction_notes + cell_consistency_warnings
                if cell_consistency_warnings:
                    logger.warning(
                        f'セル単位整合性チェック: {len(cell_consistency_warnings)}件の警告\n'
                        + '\n'.join(cell_consistency_warnings)
                    )
                else:
                    logger.info('セル単位整合性チェック: 漏れ・誤配置の検知なし')
        except Exception as e:
            # 検証側のエラーで本体処理を止めない
            logger.warning(f'セル単位整合性チェックでエラー（処理続行）: {e}')

        # 事業年度ウィンドウ選択の注意書き（年の自動補正・低カバレッジ）を集約。
        # 全従業員に同値で載るため重複排除して surfacing する。
        window_notes: list[str] = []
        for e in employees:
            note = getattr(e, 'fiscal_window_note', '') or ''
            if note and note not in window_notes:
                window_notes.append(note)

        # 賞与の期間ズレ警告（③）。支給日なし賞与×非暦年決算で R216 がズレ得る点を surface。
        _bonus_note = _bonus_period_shift_warning(employees, fiscal_month_override)
        if _bonus_note:
            window_notes.append(_bonus_note)
        # 対象年度外として除外した「支給年月あり」賞与の情報表示（E1）。AI 年月誤読の見逃し防止。
        _dropped_note = _bonus_dropped_info_note(employees)
        if _dropped_note:
            window_notes.append(_dropped_note)
        # 履歴事項PDF未添付時は役員の自動確定がスキップされる旨を可視化（G）。
        # 個人事業主には役員概念も履歴事項も無いため、この警告は出さない。
        _officer_skip_note = (
            None if is_kojin
            else _officer_autodetect_skipped_warning(registry_path)
        )
        if _officer_skip_note:
            window_notes.append(_officer_skip_note)

        # テンプレートに書込
        write_result = write_wage_ledger_to_template(
            employees,
            template_path=template_path,
            output_path=output_path,
            company_name=company_name,
            fiscal_month=fiscal_month_override,
            is_kojin=is_kojin,
            extraction_path=extraction_path,
            handwritten_files=handwritten_files,
            additional_warnings=window_notes,
            officer_names=officer_names,
            data_source_files=data_source_files,
            cell_consistency_warnings=cell_consistency_warnings,
        )

        # ステータスメッセージ
        msg_parts = [f'完了。検出 {write_result.employee_count}名']
        if write_result.officer_count > 0:
            msg_parts.append(f'役員{write_result.officer_count}名')
        if write_result.officer_matches:
            msg_parts.append(f'履歴事項照合 {len(write_result.officer_matches)}名 → 役員上書き')
        if write_result.officer_suspects:
            msg_parts.append(
                f'⚠ 役員疑い {len(write_result.officer_suspects)}名 — 変換メモシートで確認'
            )
        if write_result.part_time_missing:
            msg_parts.append(
                f'⚠ パート時間欠落 {len(write_result.part_time_missing)}名 — '
                f'所定労働時間の手入力が必要'
            )
        if write_result.midyear_count > 0:
            msg_parts.append(f'中途入退社{write_result.midyear_count}名')
        if handwritten_files:
            msg_parts.append(
                f'⚠ 手書きPDF {len(handwritten_files)}件 — 精度低下の可能性あり、原本照合必須'
            )
        if cell_consistency_warnings:
            msg_parts.append(
                f'⚠ セル単位整合性 {len(cell_consistency_warnings)}件 — '
                f'変換メモシートで確認のうえ原本照合'
            )
        if window_notes:
            msg_parts.append(
                '⚠ 対象期間の自動補正/要確認あり — 変換メモシート「事業年度ウィンドウ」を確認'
            )

        status.status = '完了'
        status.message = ' / '.join(msg_parts)
        status.output_files = [output_path.name]
        status.ledger_employees = employees
        return status

    except Exception as e:
        logger.exception(f'賃金台帳作成タスクで予期しないエラー: {e}')
        status.status = 'エラー'
        status.message = f'予期しないエラーが発生しました: {e}'
        return status


def run_full_pipeline(
    resource_folder: Path,
    template_path: Path,
    template_type: str,
    company_name: str,
    fiscal_month_override: int | None = None,
    selection_override: dict[str, list[Path]] | None = None,
) -> list[ProcessingStatus]:
    """タスク1 + タスク2 を一括実行"""
    extractor = create_extractor(CLAUDE_API_KEY)
    results = []

    # タスク1: 申請書
    output_app = resource_folder / f'{company_name}_{template_type.replace("_", "_")}_AI版.xlsx'
    s1 = run_application_transfer(
        resource_folder, template_path, template_type, output_app, extractor,
        fiscal_month_override=fiscal_month_override,
        selection_override=selection_override,
    )
    results.append(s1)

    # タスク2: 給与計算
    output_wage = resource_folder / f'{company_name}_給与支給総額計算.xlsx'
    s2 = run_wage_calculation(
        resource_folder, company_name, output_wage, extractor,
        fiscal_month_override=fiscal_month_override,
        selection_override=selection_override,
    )
    results.append(s2)

    return results
