# -*- coding: utf-8 -*-
"""
給与支給総額計算Excel生成 + 1人当たり給与支給総額算出

2026年度要件:
- 指標: 1人当たり給与支給総額（非常勤を含む全従業員）
- 計算: 給与支給総額（役員報酬除く）÷ 従業員数（パートは正社員換算）
- 年平均成長率: 3%以上

対象給与: 給料、賃金、賞与、各種手当（残業手当、休日出勤手当、
         職務手当、地域手当、家族手当、住宅手当）等
除外: 役員報酬、福利厚生費、法定福利費、退職金

対象従業員: 全月分の給与を受けた従業員のみ（中途・退職者はその年度除外）
パート換算: 正社員の所定労働時間で換算
"""
from __future__ import annotations

import logging
from dataclasses import dataclass, field as dc_field
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .models import FinancialData, MonthlyWageData
from .config import STANDARD_ANNUAL_HOURS

logger = logging.getLogger(__name__)


# ── 1人当たり給与支給総額計算（2026年新要件）──

@dataclass
class PayrollEmployee:
    """賃金台帳から読み取った従業員1名のデータ"""
    name: str
    employment_type: str  # 正社員, 契約社員, パート, アルバイト, 役員
    monthly_salary: list[float] = dc_field(default_factory=list)  # 12ヶ月分の総支給額（賞与除く）
    monthly_hours: list[float] = dc_field(default_factory=list)   # 12ヶ月分の労働時間
    is_officer: bool = False
    is_excluded: bool = False      # 産休・育休等で除外
    full_year: bool = True         # 全月分の給与を受けたか
    annual_bonus: float = 0.0      # 年間賞与（R216 に算入。月次には混ぜない）
    # 支給>0 の月と 0円明記月が混在（v0.2.69 の新規除外対象）。monthly_salary は
    # None→0 変換済みで判別不能なため、wage_employees_to_payroll で焼き込む
    partial_zero: bool = False
    # 全月が0円明記（None なし）。賞与の有無と組み合わせて「賞与のみ受給で除外
    # (v0.2.70)」「給与支給0円で除外 (v0.2.66)」の警告を出し分ける。同上の理由で焼き込み
    zero_explicit: bool = False


@dataclass
class PerCapitaWageResult:
    """1人当たり給与支給総額の計算結果"""
    total_salary: float = 0.0              # 給与支給総額（役員報酬除く）
    employee_count_fte: float = 0.0        # 従業員数（正社員換算）
    per_person_salary: float = 0.0         # 1人当たり給与支給総額
    officer_compensation: float = 0.0      # 役員報酬合計
    regular_annual_hours: float = 0.0      # 正社員の年間所定労働時間
    included: list[PayrollEmployee] = dc_field(default_factory=list)
    excluded_names: list[str] = dc_field(default_factory=list)
    # excluded_names のうち「給与支給0円（年計≤0）」で除外した人（中途入退社とは区別して
    # 警告文に出すために分離保持。excluded_names にも重複して入る）
    excluded_zero_names: list[str] = dc_field(default_factory=list)
    # excluded_names のうち「0円明記月と支給月が混在」で除外した人（v0.2.69）。
    # None ベースの従来中途者と区別して警告文に出すために分離保持。
    # excluded_zero_names（年計≤0）とは互いに素（partial_zero は年計>0）。
    excluded_partial_zero_names: list[str] = dc_field(default_factory=list)
    # excluded_names のうち「全月0円明記＋賞与のみ受給」で除外した人（v0.2.70）。
    # 賞与が R216 から落ちることを黙らせないために分離保持。
    excluded_bonus_only_names: list[str] = dc_field(default_factory=list)

    GROWTH_RATE = 0.03  # 3%

    def plan_values(self) -> dict[str, float]:
        """3年分の計画数値（3%成長）"""
        b = self.per_person_salary
        r = self.GROWTH_RATE
        return {
            'year_0': b,
            'year_1': b * (1 + r),
            'year_2': b * (1 + r) ** 2,
            'year_3': b * (1 + r) ** 3,
        }


def is_full_time_employment(employment_type: str | None) -> bool:
    """雇用区分が「正規雇用相当（フルタイム）」か判定する共通 predicate。

    `_calc_fte` / 給与計算の人数集計 / 詳細表示の3箇所で同じ判定を使うことで、
    「契約社員」が場所によって正規雇用扱いになったりパート扱いになったりする
    不整合を避ける（Codex Round 4 指摘）。

    判定: '正社員' か '契約社員' を文字列に含む（provenance/修飾付きも許容）。
    例: '正社員', '正社員(推定)', '契約社員', '契約社員(臨時)' → True
        'パート', 'アルバイト', '日雇い', '役員' → False
    """
    et = employment_type or ''
    return '正社員' in et or '契約社員' in et


def is_zero_wage(monthly_wages, annual_bonus: float = 0.0) -> bool:
    """「年計（月次支給合計＋年間賞与）≤ 0」＝給与の支給を受けていない従業員か判定する。

    公募要領 p.10「全月分の給与等の支給を受けた従業員」に非該当のため、非役員であれば
    R215（FTE換算従業員数）/ R216（給与支給総額）/ 人数集計の算定対象から除外する
    （2026-06-10 補助金MTG決定: 賃金台帳に0円と明記された人を1人と数えて
    1人当たり給与が過小になる不具合の対処）。
    賞与のみ受給（月次0・賞与>0）は年計>0 のため本述語では False（除外判定は
    is_full_year_paid 側が担う。v0.2.70 から賞与のみ受給者も全月支給に非該当
    として除外＝2026-06-11 ルール変更。v0.2.66〜v0.2.69 は算入だった）。
    """
    total = sum(w or 0.0 for w in (monthly_wages or []))
    return total + (annual_bonus or 0.0) <= 0


def is_zero_wage_detail(e: dict) -> bool:
    """employees_detail の1人分 dict に対する is_zero_wage。

    12ヶ月モード（monthly_wages_full あり）は12ヶ月合計＋年間賞与、
    3ヶ月モード（賃金状況報告シート由来）は m1〜m3 合計で判定する。
    """
    wages = e.get('monthly_wages_full')
    if not wages:
        wages = [e.get('m1', 0), e.get('m2', 0), e.get('m3', 0)]
    return is_zero_wage(wages, e.get('annual_bonus', 0.0) or 0.0)


def is_full_year_paid(monthly_wages) -> bool:
    """「全月分の給与等の支給を受けた」（公募要領 p.10）か判定する単一の真実。

    公募要領 p.10:「中途採用や退職等で全月分の給与等の支給を受けていない従業員に
    ついては…算出の対象から除く必要がある」（除外は義務。産前産後・育児・介護休業
    等の休職は『除くことができる』＝任意だが、台帳からは中途と判別できないため
    一律除外し、警告で手動修正の経路を案内する）。

    ルール（v0.2.70、2026-06-11 確定）: 12スロット全てが 支給>0 のときのみ True。
    None（データ無し月）も 0円・負値の月も「支給を受けていない月」＝空欄と同等に扱う
    （「賃金台帳の作成」タスク産の台帳は不在月を 0円 で出力するため。v0.2.69）。

    賞与は月の支給判定に使わない: 標準テンプレ T列は賞与の年間合計のみで支給月の
    情報を持たないため、「0円月に賞与が支給されていた」ケースは機械判定できない。
    該当し得る人（賞与あり＋0円月）には警告を出し、算入する場合は該当月に賞与額を
    記入し T列から同額を差し引く手動修正で対応する。
    これに伴い全月0円＋賞与のみ受給者も除外（v0.2.66〜v0.2.69 の「算入する」仕様を
    2026-06-11 に変更。月給0円の月は支給なしと読む整理）。
    """
    wages = monthly_wages or []
    return all(w is not None and w > 0 for w in wages)


def has_partial_zero_months(monthly_wages) -> bool:
    """支給>0 の月と「明記された ≤0 の月」（None でない 0円・負値）が混在するか。

    is_full_year_paid=False のうち v0.2.69 で新たに除外対象になった人（0円明記月
    持ち）を、従来からの None ベース中途者と区別して警告に出すための述語。
    None のみで欠ける従来中途者には新しい警告を重複して出さない。
    """
    wages = monthly_wages or []
    if not any((w or 0) > 0 for w in wages):
        return False
    return any(w is not None and w <= 0 for w in wages)


def has_partial_zero_months_detail(e: dict) -> bool:
    """employees_detail の1人分 dict に対する has_partial_zero_months。

    12ヶ月モードは monthly_wages_full（None→0 変換済み）と month_data_mask
    （元データの non-None マスク）から復元して判定する。
    3ヶ月モード（賃金状況報告シート由来・monthly_wages_full なし）は全月判定が
    原理的に不能なため常に False（部分0円ルールの対象外）。
    """
    restored = _restore_monthly_from_detail(e)
    if restored is None:
        return False
    return has_partial_zero_months(restored)


def is_all_zero_explicit(monthly_wages) -> bool:
    """全月が「明記された0円以下」（None を含まない）か。

    is_full_year_paid=False のうち「全月0円明記」の人（賞与のみ受給者・退職済み行・
    無給の家族従業者など）を、None ベースの中途者と区別して警告・凡例に出し分ける
    ための述語。賞与の有無は見ない（呼出側が annual_bonus で出し分ける）。
    """
    wages = monthly_wages or []
    return bool(wages) and all(w is not None and w <= 0 for w in wages)


def is_all_zero_explicit_detail(e: dict) -> bool:
    """employees_detail の1人分 dict に対する is_all_zero_explicit。
    3ヶ月モード（monthly_wages_full なし）は常に False。
    """
    restored = _restore_monthly_from_detail(e)
    if restored is None:
        return False
    return is_all_zero_explicit(restored)


def _restore_monthly_from_detail(e: dict) -> list | None:
    """employees_detail から None 込みの月次配列を復元する（detail 系述語の共通部）。

    monthly_wages_full は None→0 変換済みのため、month_data_mask（元データの
    non-None マスク）で None を復元する。monthly_wages_full が無い（3ヶ月モード）
    場合は None を返す。
    """
    wages = e.get('monthly_wages_full')
    if not wages:
        return None
    mask = e.get('month_data_mask') or [True] * len(wages)
    return [
        w if (i < len(mask) and mask[i]) else None
        for i, w in enumerate(wages)
    ]


def _calc_fte(emp: PayrollEmployee, annual_hours: float) -> float:
    """パート・アルバイトを正社員換算（FTE換算）。フルタイム雇用は FTE 1.0。

    根拠: IT導入補助金 2026 通常枠 公募要領 p.10
        「パートタイム従業員については、正社員の就業時間に換算して人数を算出すること」
    交付申請マニュアル p.86 算定例
        「週20時間勤務（正規雇用は、週40時間労働）の場合 従業員数：20÷40＝0.5(人)」

    挙動:
      - 正社員・契約社員: 1.0
      - パート・アルバイト + monthly_hours 集計値あり: sum(monthly_hours) / annual_hours
      - パート・アルバイト + monthly_hours 集計値なし: **1.0 にサイレント昇格**
        （R215 過大計上のリスクあり。賃金台帳側で時間データを取り切る運用が望ましい。
         アプリ画面で「パート時間欠落で R215 過大計上の可能性」を警告すること）
    """
    if is_full_time_employment(emp.employment_type):
        return 1.0
    if not emp.monthly_hours:
        return 1.0
    return sum(emp.monthly_hours) / annual_hours


def calculate_per_capita_wage(
    employees: list[PayrollEmployee],
    regular_annual_hours: float = STANDARD_ANNUAL_HOURS,
) -> PerCapitaWageResult:
    """従業員リストから1人当たり給与支給総額を算出"""
    result = PerCapitaWageResult(regular_annual_hours=regular_annual_hours)

    for emp in employees:
        if emp.is_officer:
            result.officer_compensation += sum(emp.monthly_salary)
            continue
        if emp.is_excluded or not emp.full_year:
            if emp.partial_zero:
                # 0円明記月と支給月の混在＝「全月分の支給」非該当（公募要領 p.10、
                # v0.2.69）。None ベースの従来中途者と区別して警告に出す。
                if any((w or 0) < 0 for w in emp.monthly_salary):
                    logger.warning(
                        f'{emp.name}: 月次給与に負値が含まれるため算定対象外'
                        '（年末調整の返金処理や入力ミスの可能性。賃金台帳の値を要確認）'
                    )
                result.excluded_partial_zero_names.append(emp.name)
            elif emp.zero_explicit and (emp.annual_bonus or 0) > 0:
                # 全月0円明記＋賞与のみ受給（v0.2.70 から除外。賞与も R216 に入らない）
                result.excluded_bonus_only_names.append(emp.name)
            elif emp.zero_explicit:
                # 全月0円明記・賞与なし。v0.2.66 からの除外対象（従来は後段の
                # is_zero_wage で拾っていたが、full_year=False が先に立つため
                # ここで excluded_zero_names に入れて警告表示を維持する）
                result.excluded_zero_names.append(emp.name)
            result.excluded_names.append(emp.name)
            continue

        # 給与支給総額（R216）= 月次課税給与の年計 ＋ 年間賞与（公募要領 p.10: 賞与も対象）
        annual = sum(emp.monthly_salary) + (emp.annual_bonus or 0.0)
        # 年計0円の人は給与を受けていない（賃金台帳に0円明記・個人事業主本人・退職済み行など）。
        # 分子に0を足し分母（FTE）に+1すると1人当たりが不当に下がるため、計算対象から除外する。
        if is_zero_wage(emp.monthly_salary, emp.annual_bonus):
            if any((w or 0) < 0 for w in emp.monthly_salary):
                logger.warning(
                    f'{emp.name}: 月次給与に負値が含まれ年計0以下 → 算定対象外'
                    '（賃金台帳の値を要確認）'
                )
            result.excluded_names.append(emp.name)
            result.excluded_zero_names.append(emp.name)
            continue
        result.total_salary += annual
        fte = _calc_fte(emp, regular_annual_hours)
        result.employee_count_fte += fte
        result.included.append(emp)

    if result.employee_count_fte > 0:
        result.per_person_salary = result.total_salary / result.employee_count_fte

    logger.info(
        f'1人当たり計算: {result.total_salary:,.0f}円 / '
        f'{result.employee_count_fte:.1f}人 = {result.per_person_salary:,.0f}円'
    )
    return result


def wage_employees_to_payroll(
    employees,
) -> tuple[list[PayrollEmployee], float, list[str]]:
    """WageEmployee → PayrollEmployee へ変換する（pipeline と回帰テストで共有）。

    R215（FTE換算従業員数）/ R216（給与支給総額）の算定前段。元は pipeline 内に
    インライン実装されていたが、回帰テスト（_debug/test_wage_regression.py）が本番と
    **同一の変換ロジック**を検証できるよう共有関数に切り出した。WageEmployee を直接
    import せず属性アクセス（duck typing）で扱い、循環 import を避ける。

    Returns:
        (payroll_list, total_annual_hours, part_fte_fallback_names)
        - total_annual_hours: 役員を除く全従業員の年間総労働時間
        - part_fte_fallback_names: パート/アルバイトで労働時間データが無く
          _calc_fte で FTE=1.0 にサイレント昇格した従業員名（R215 過大計上の懸念対象）
    """
    payroll_list: list[PayrollEmployee] = []
    total_annual_hours = 0.0
    part_fte_fallback_names: list[str] = []

    for emp in employees:
        is_officer = '役員' in emp.employment_type
        emp_type = emp.employment_type if emp.employment_type else '正社員'

        # 全月分の給与を受けたか判定
        full_year = emp.is_full_year

        monthly_salary = [
            w if w is not None else 0.0 for w in emp.monthly_wages
        ]

        # 労働時間: 月別実績データがあればそれを優先。なければ月平均で補完
        has_monthly_hours = any(
            h is not None and h > 0 for h in emp.monthly_hours
        )
        if has_monthly_hours:
            monthly_hours = [
                h if (h is not None and h > 0) else 0.0
                for h in emp.monthly_hours
            ]
        elif emp.monthly_avg_hours > 0:
            # 月別データが取れないフォーマットは、在籍月数×月平均で概算
            months_with_wage = sum(
                1 for w in emp.monthly_wages if w is not None
            )
            months = months_with_wage if months_with_wage > 0 else 12
            monthly_hours = [emp.monthly_avg_hours] * months + [0.0] * (12 - months)
        else:
            monthly_hours = []

        annual_bonus = getattr(emp, 'annual_bonus', 0.0) or 0.0
        # 給与支給0円の非役員は R215/R216 算定対象外（calculate_per_capita_wage で除外）。
        # 総労働時間（生産性指標B40）・パートFTE警告にも入れない。
        zero_wage = not is_officer and is_zero_wage(monthly_salary, annual_bonus)

        payroll_list.append(PayrollEmployee(
            name=emp.name,
            employment_type=emp_type,
            monthly_salary=monthly_salary,
            monthly_hours=monthly_hours,
            is_officer=is_officer,
            full_year=full_year,
            annual_bonus=annual_bonus,
            # monthly_salary は None→0 変換済みのため、変換前の monthly_wages から判定
            partial_zero=has_partial_zero_months(emp.monthly_wages),
            zero_explicit=is_all_zero_explicit(emp.monthly_wages),
        ))

        # 役員を除く全従業員の年間総労働時間を集計
        if not is_officer and monthly_hours and not zero_wage:
            total_annual_hours += sum(monthly_hours)

        # パートで時間データが空 → _calc_fte で FTE=1.0 サイレント昇格になる人
        # （IT導入補助金は本来 FTE 換算が要件。R215 過大計上の警告対象）
        if (
            not is_officer
            and not zero_wage
            and not is_full_time_employment(emp_type)
            and not monthly_hours
            and full_year
        ):
            part_fte_fallback_names.append(emp.name)

    return payroll_list, total_annual_hours, part_fte_fallback_names

# ── スタイル定義 ──
TITLE_FONT = Font(name='游ゴシック', size=14, bold=True)
HEADER_FONT = Font(name='游ゴシック', size=10, bold=True)
NORMAL_FONT = Font(name='游ゴシック', size=10)
SMALL_FONT = Font(name='游ゴシック', size=9)
HEADER_FONT_WHITE = Font(name='游ゴシック', size=10, bold=True, color='FFFFFF')
BOLD_FONT = Font(name='游ゴシック', size=10, bold=True)
RESULT_FONT = Font(name='游ゴシック', size=12, bold=True, color='C00000')
NUMBER_FMT = '#,##0'
PCT_FMT = '0.00%'
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
FILL_HEADER = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
FILL_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
FILL_YELLOW = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
FILL_GREEN = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
FILL_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
FILL_HEADER_DARK = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
# AI抽出値（決算書PDFから Claude が読み取った値）の視認用。誤読の可能性があるため
# 人間チェックの対象として目立たせる。色は控えめなオレンジ系。
FILL_AI_EXTRACTED = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')


def _cell(ws, row, col, value, font=NORMAL_FONT, fmt=None, fill=None, border=THIN_BORDER):
    """セルに値とスタイルをまとめて設定"""
    c = ws.cell(row=row, column=col, value=value)
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if border:
        c.border = border
    return c


def _apply_readable_layout(ws, text_columns: tuple[int, ...], max_widths: dict[int, int]):
    """指定列のセルに wrap_text を一括適用し、長文行の行高を文字数から見積もって設定する。

    openpyxl は行高の自動計算をしないため、長文セルが入った行は明示設定しないと
    1行表示で切れたままになる。日本語は1文字≒全角幅なので、列幅 W に対して
    1行に入る文字数を W * 1.0（半角換算でないので 1.0）として行数を見積もる。

    Args:
        ws: 対象シート
        text_columns: 折り返し対象列のインデックス（1始まり、B=2, C=3, D=4 など）
        max_widths: 各列の表示幅（列インデックス → Excel 列幅）
    """
    line_height = 14  # 1行あたりの目安（Excel デフォルト 14.4pt 程度）
    max_height = 120  # 上限。これ以上は折り返さず横スクロールで対応
    for r in range(1, ws.max_row + 1):
        needed_lines = 1
        for col in text_columns:
            cell = ws.cell(r, col)
            if cell.value is None or not isinstance(cell.value, str):
                continue
            # 既存 alignment の horizontal を保持しつつ wrap_text と vertical='top' を適用
            existing = cell.alignment
            cell.alignment = Alignment(
                horizontal=existing.horizontal,
                vertical='top',
                wrap_text=True,
            )
            # 行数見積もり: 改行 + 列幅から計算
            text = cell.value
            chars_per_line = max(10, max_widths.get(col, 30))
            # 改行ごとに分割し、各セグメントが何行分か計算
            seg_lines = sum(
                max(1, (len(seg) + chars_per_line - 1) // chars_per_line)
                for seg in text.split('\n')
            )
            needed_lines = max(needed_lines, seg_lines)
        if needed_lines > 1:
            target_h = min(needed_lines * line_height + 4, max_height)
            current_h = ws.row_dimensions[r].height or 15
            if target_h > current_h:
                ws.row_dimensions[r].height = target_h


def create_wage_calculation(
    output_path: Path,
    company_name: str,
    fiscal_year_label: str,
    financial: FinancialData,
    seishain_count: int,
    part_count: int,
    yakuin_count: int,
    yakuin_hoshu_3m: int,
    employees_detail: list[dict] | None = None,
    source_files: dict[str, str] | None = None,
) -> Path:
    """
    給与支給総額計算Excelを作成。

    employees_detail: [{'no': 1, 'name': '氏名', 'type': '正社員',
                        'm1': 基本給, 'm2': 基本給, 'm3': 基本給,
                        'hr': 時給, 'monthly_hours': 月間時間, 'judge': '対象外'}, ...]
    source_files: 各データソースのファイル名（人間チェックの突合用）。
                  キー: 'pl' / 'wage_ledger' / 'wage_report' / 'registry'
                  値: ファイル名（パス除外）。AI抽出値の出所追跡用に各セクション
                  ヘッダー下に表示する。None または欠落キーは「未取得」表示。
    """
    sources = source_files or {}
    pl_source = sources.get('pl', '') or '（不明）'
    ledger_source = sources.get('wage_ledger', '') or '（不明）'
    wage_report_source = sources.get('wage_report', '')
    registry_source = sources.get('registry', '') or '（不明）'
    # 値→ページ番号の逆引き結果（pipeline 側で機械的に特定済み）
    pl_value_pages: dict[str, list[int]] = sources.get('pl_value_pages', {}) or {}
    # breakdown 各内訳（販管費/原価部）のセクション検証結果（pipeline 側で機械検証済み）
    # 値: {key: {'pl_section_class': 'pl'/'cost'/'both'/'absent'/'unknown'/'none',
    #            'cost_section_class': ...,
    #            'pl_section_pages': [...], 'cost_section_pages': [...]}}
    pl_breakdown_verification: dict[str, dict] = (
        sources.get('pl_breakdown_verification', {}) or {}
    )

    def _page_tag(key: str) -> str:
        """財務値が決算書PDFの何ページに見つかったかをタグ文字列化。

        例: [1] → ' (PDF p.1で確認済)'
            [3,5] → ' (PDF p.3,5で確認済)'
            [] → ' (⚠PDFに値なし／AI誤読の可能性)'
            キー未登録 → ''（ページ番号特定機能が動かなかった想定。何も表示しない）
        """
        if key not in pl_value_pages:
            return ''
        pages = pl_value_pages[key]
        if not pages:
            return ' (⚠PDFに値なし／AI誤読の可能性)'
        return f' (PDF p.{",".join(map(str, pages))}で確認済)'
    wb = openpyxl.Workbook()

    # 計算用定数
    total_wage_pl = financial.salary + financial.misc_wages + financial.bonus
    yakuin_annual = yakuin_hoshu_3m * 4
    wage_excl_yakuin = total_wage_pl - yakuin_annual
    total_emp = seishain_count + part_count
    standard_monthly = STANDARD_ANNUAL_HOURS / 12

    # 賃金台帳由来の12ヶ月明細が渡っているか（_build_employees_detail_from_ledger 経由）。
    # True なら Sheet 2 を12ヶ月レイアウトで出力し、Sheet 1 のメイン「給与支給総額」を
    # 賃金台帳合計（R216 公募要領定義に即した値）で算定する。
    # False なら賃金状況報告シート由来 or データなし → 従来の3ヶ月レイアウト＋PL値ベース。
    has_12_months = bool(
        employees_detail
        and any(e.get('monthly_wages_full') for e in employees_detail)
    )
    # 12ヶ月明細から月並びを推定（事業年度開始月始まり）。employees_detail 各人で
    # 同じ並びになる前提（_build_employees_detail_from_ledger が一括生成）。
    if has_12_months:
        for _e in employees_detail:
            _labels = _e.get('month_labels_full')
            if _labels and len(_labels) == 12:
                month_labels_full = list(_labels)
                break
        else:
            month_labels_full = [f'{i + 1}月' for i in range(12)]
    else:
        month_labels_full = [f'{i + 1}月' for i in range(12)]

    # Sheet 2 に合計行が出力されるか（金額が皆無なら合計行は出ない）。
    # ledger_total_cell の事前計算と Sheet 2 の実出力で**同じ条件**を使い、
    # 「Sheet 1 が存在しない合計行を参照して0表示」になる参照ズレを防ぐ。
    has_any_amount = bool(employees_detail) and any(
        (e.get('monthly_wages_full') and any(e['monthly_wages_full']))
        or e.get('m1') or e.get('m2') or e.get('m3')
        for e in (employees_detail or [])
    )

    # Sheet 2 の「12ヶ月在籍のみ合計」セル位置を事前計算（Sheet 1 から参照するため）。
    # Sheet 2 レイアウト（後段で実装）:
    #   row 4 ヘッダー / row 5〜(4+N) データ / row (5+N) 合計（全員） / row (6+N) 合計（12ヶ月在籍）
    # 12ヶ月合計列は Q列（FIRST_MONTH_COL=E=5 から12列で P=16、その次 Q=17）。
    ledger_total_cell: str | None = None
    if has_12_months and employees_detail and has_any_amount:
        _sheet2_target_row = 6 + len(employees_detail)
        ledger_total_cell = f"'従業員別明細'!Q{_sheet2_target_row}"

    # クリーンモード判定: 賃金台帳ベース運用が成立しているか
    # True: 決算書PL値由来セクション（損益計算書データ／役員報酬の控除／PL値参考）を全削除し、
    #       賃金台帳ベース1本でレイアウトを単純化（ユーザー指示 2026-05-19）。
    #       テンプレ転記用セクションは決算書PL値を直書きする（上のセル参照ではない）。
    # False: 賃金台帳由来データなし → 従来の決算書PL値ベース＋参考値併記レイアウト。
    clean_mode = bool(has_12_months and ledger_total_cell)

    # FTE 計算（R215「FTE換算従業員数」用）
    # 公募要領（IT2026 通常枠 p.9-10）原文:
    #   「基準年度および算出対象年度に全月分の給与等の支給を受けた従業員のみ算定対象。
    #    パート・非常勤含む。中途入退社者は除外。」
    # → 中途者は **除外**（按分ではない）。docs/補助金_実務知識ベース.md:51-52, 203 参照
    # - 12ヶ月在籍正社員: 1.0
    # - 12ヶ月在籍パート: 月間労働時間/標準月間労働時間
    # - 中途入退社者（full_year=False）: 完全除外
    fte_seishain = 0.0
    fte_part = 0.0
    excluded_midyear = 0       # 中途者として除外した人数（凡例表示用）
    excluded_zero = 0          # 給与支給0円で除外した人数（凡例表示用）
    excluded_partial_zero = 0  # 0円支給月の混在で除外した人数（凡例表示用、v0.2.69）
    excluded_bonus_only = 0    # 賞与のみ受給で除外した人数（凡例表示用、v0.2.70）
    if employees_detail:
        for e in employees_detail:
            if not e.get('full_year', True):
                # 除外理由を凡例で出し分ける（除外扱いはどれも同じ）:
                #   0円明記月と支給月の混在 / 全月0円明記（賞与あり・なし） /
                #   None ベースの従来中途者
                if has_partial_zero_months_detail(e):
                    excluded_partial_zero += 1
                elif is_all_zero_explicit_detail(e):
                    if (e.get('annual_bonus') or 0) > 0:
                        excluded_bonus_only += 1
                    else:
                        excluded_zero += 1
                else:
                    excluded_midyear += 1
                continue
            # 給与支給0円（年計≤0）は R215 算定対象外（公募要領「全月分の給与等の
            # 支給を受けた従業員」に非該当。2026-06-10 MTG決定）
            if is_zero_wage_detail(e):
                excluded_zero += 1
                continue
            if is_full_time_employment(e.get('type')):
                fte_seishain += 1.0
            else:
                monthly_h = e.get('monthly_hours', 0)
                if monthly_h <= 0:
                    continue
                fte_part += monthly_h / standard_monthly
        fte_adjusted = fte_seishain + fte_part
    else:
        # employees_detail が無い場合（賃金台帳・賃金状況報告シートともに取得失敗）の
        # フォールバック: 雇用区分の単純集計（中途者特定不可、警告表示）
        fte_seishain = float(seishain_count)
        fte_adjusted = seishain_count + fte_part

    # ===== Sheet 1: 給与支給総額計算 =====
    ws1 = wb.active
    ws1.title = '給与支給総額計算'

    _cell(ws1, 2, 2, '給与支給総額計算書', TITLE_FONT, border=None)
    # 社名はそのまま表示する（「株式会社」固定の接頭辞は個人事業主・
    # 有限会社等で誤表示になるため付けない）
    _cell(ws1, 3, 2, company_name, Font(name='游ゴシック', size=12), border=None)
    # 事業年度ラベルは PL 由来（AI抽出）or 賃金台帳期間からの導出。誤読リスクのあるセル
    _cell(ws1, 4, 2, f'事業年度: {fiscal_year_label}', NORMAL_FONT,
          fill=FILL_AI_EXTRACTED, border=None)
    _cell(ws1, 4, 4, f'出所: {pl_source}', SMALL_FONT, border=None)

    # 凡例（AI抽出セルの視認説明）— 計算結果より前に置いてユーザーに認識させる
    r = 5
    _cell(ws1, r, 2,
          '※薄オレンジ色のセル＝AI（Claude）が決算書PDFから抽出した値です。'
          '誤読の可能性があるため、決算書原本と必ず照合してください。',
          SMALL_FONT, border=None)
    ws1.cell(r, 2).fill = FILL_AI_EXTRACTED

    # P/Lデータ
    # 販管費・原価部の合算が実施されたかでセクションヘッダーを切替（建設業・製造業対応）
    breakdown = getattr(financial, 'breakdown', {}) or {}
    has_cost_section_merge = any(
        (breakdown.get(k, {}) or {}).get('cost_section', 0) > 0
        for k in ('salary', 'misc_wages', 'bonus', 'legal_welfare', 'welfare')
    )
    pl_section_header = (
        '【損益計算書データ（販管費＋製造原価報告書 合算）】'
        if has_cost_section_merge else '【損益計算書データ（販管費）】'
    )

    # 各勘定科目に対応する販管費側ラベル・原価部側ラベル
    # （cost_section_label には「給料手当/賃金/給料/労務費」など複数勘定が含まれるため
    #  代表ラベルとして "賃金等" のような総称を採用）
    # ※ clean_mode（賃金台帳ベース運用）でもテンプレ転記用の出所表記で使うため、
    #   if not clean_mode: ブロックより前で定義しておく。
    PL_ITEM_LABELS = {
        'salary':        ('給料手当', '賃金等（賃金/給料/労務費）'),
        'misc_wages':    ('雑給', '雑給'),
        'bonus':         ('賞与', '賞与'),
        'legal_welfare': ('法定福利費', '法定福利費'),
        'welfare':       ('福利厚生費', '福利厚生費'),
    }

    def _ai_source_tag(key: str) -> str:
        """AI抽出マーカー + ページ番号の表示文字列を生成（「決算書PDF」の重複を回避）。

        - ページ番号取得済み: 'AI抽出：決算書PDF p.3,5'
        - PDF テキストに値なし: 'AI抽出：決算書PDF（⚠PDFに値なし／AI誤読の可能性）'
        - ページ番号未取得（キー未登録）: 'AI抽出：決算書PDF'
        """
        if key not in pl_value_pages:
            return 'AI抽出：決算書PDF'
        pages = pl_value_pages[key]
        if not pages:
            return 'AI抽出：決算書PDF（⚠PDFに値なし／AI誤読の可能性）'
        return f'AI抽出：決算書PDF p.{",".join(map(str, pages))}'

    def _component_label(label: str, value: int, expected: str, ver: dict, side: str) -> str:
        """内訳1成分のラベル文を生成（PDF機械検証結果を反映）。

        Args:
            label: 表示ラベル（例: '販管費「給料手当」'）
            value: 金額
            expected: 期待されるセクション（'pl' or 'cost'）
            ver: 該当 key の検証結果 dict（空なら PDF照合なし）
            side: 'pl_section' or 'cost_section'

        セクション照合の振る舞い:
            - expected と一致: '販管費「X」200,000円 (PDF p.3で確認)'
            - 反対側に出現: '⚠ AI判定「X」200,000円(本来販管費だがPDF上は製造原価ページ p.5)'
            - 両セクションに同値: '販管費「X」200,000円 (⚠両セクションのページに同値あり p.3,5)'
            - PDFに存在せず: '販管費「X」200,000円 (⚠PDFテキストに該当数値なし／AI誤読の可能性)'
            - セクション未判定ページ: '販管費「X」200,000円 (PDF p.7に出現／セクション特定不可)'
            - 検証情報なし: '販管費「X」200,000円'（AI判定のみ）
        """
        cls = ver.get(f'{side}_class', 'none') if ver else 'none'
        pages = ver.get(f'{side}_pages', []) if ver else []
        page_str = ','.join(map(str, pages)) if pages else ''
        if not ver or cls == 'none':
            # 検証情報なし（PDFテキスト層なし or キー未対応）
            return f'{label}{value:,}円'
        if cls == expected:
            return f'{label}{value:,}円 (PDF p.{page_str}で確認)'
        if cls == 'both':
            return f'{label}{value:,}円 (⚠両セクションのページに同値あり p.{page_str})'
        if cls == 'absent':
            return f'{label}{value:,}円 (⚠PDFテキストに該当数値なし／AI誤読の可能性)'
        if cls == 'unknown':
            return f'{label}{value:,}円 (PDF p.{page_str}に出現／セクション特定不可)'
        # 反対側セクションに出現（AI誤分類の可能性）
        opposite_jp = '製造原価' if expected == 'pl' else '販管費'
        expected_jp = '販管費' if expected == 'pl' else '製造原価'
        return (
            f'⚠ AI判定「{label}」{value:,}円 '
            f'(本来{expected_jp}だがPDF上は{opposite_jp}ページ p.{page_str} に出現)'
        )

    def _verified_source_tag(key: str) -> str:
        """合算後の値そのものに対する出所ラベル。

        breakdown が機械検証済みなら「PDF照合済」、PDFテキスト層が無ければ
        「AI判定のみ（PDF照合不可）」と明示する。
        """
        if pl_breakdown_verification.get(key):
            return 'PDFテキストで機械照合済'
        if key in pl_value_pages and pl_value_pages[key]:
            return f'AI抽出：決算書PDF p.{",".join(map(str, pl_value_pages[key]))}'
        if key in pl_value_pages:
            return 'AI抽出：決算書PDF（⚠PDF照合できず）'
        return 'AI抽出：決算書PDF'

    def _build_pl_note(key: str, default_pl_note: str = '', excluded: bool = False) -> str:
        """販管費＋原価部の内訳を含む備考文を生成（機械検証ラベル付き）。

        verification info あり（PDF テキスト層から検証済）:
            "内訳: 販管費「給料手当」200,000円 (PDF p.3で確認) ＋
                  製造原価「賃金等」13,870,373円 (PDF p.5で確認)"
        verification なし（画像PDF・テスト等）:
            "内訳: 販管費「給料手当」200,000円 ＋ 製造原価「賃金等」13,870,373円
             (AI抽出：決算書PDF p.3,5)"
        breakdown 情報なし: 旧来の default_pl_note にフォールバック（テスト互換）
        """
        bd = (breakdown.get(key) or {}) if isinstance(breakdown, dict) else {}
        pl_v = int(bd.get('pl_section') or 0)
        cost_v = int(bd.get('cost_section') or 0)
        excluded_tag = '｜※給与支給総額から除外' if excluded else ''
        ai_tag = _ai_source_tag(key)
        pl_label, cost_label = PL_ITEM_LABELS.get(key, (key, key))
        ver = pl_breakdown_verification.get(key, {})
        has_verification = bool(ver) and any(
            ver.get(f'{s}_class', 'none') != 'none'
            for s in ('pl_section', 'cost_section')
        )
        # 検証なし時の合算後出所タグ（成分ごとのエビデンスを出せない代わり）
        agg_tag = f'（{ai_tag}）' if not has_verification else ''

        if pl_v > 0 and cost_v > 0:
            pl_part = _component_label(
                f'販管費「{pl_label}」', pl_v, 'pl', ver, 'pl_section')
            cost_part = _component_label(
                f'製造原価「{cost_label}」', cost_v, 'cost', ver, 'cost_section')
            return f'内訳: {pl_part} ＋ {cost_part}{excluded_tag}{agg_tag}'
        if pl_v > 0 and cost_v == 0:
            pl_part = _component_label(
                f'販管費「{pl_label}」', pl_v, 'pl', ver, 'pl_section')
            return f'{pl_part}{excluded_tag}{agg_tag}'
        if pl_v == 0 and cost_v > 0:
            cost_part = _component_label(
                f'製造原価「{cost_label}」', cost_v, 'cost', ver, 'cost_section')
            return f'{cost_part}{excluded_tag}{agg_tag}'
        # breakdown 情報なし → 旧来表示にフォールバック（決算書構造の固定マッピング表現）
        if default_pl_note:
            return f'{default_pl_note}（{ai_tag}）'
        return f'販管費「{pl_label}」より{excluded_tag}（{ai_tag}）'

    if not clean_mode:
        # 決算書PL値由来セクション（賃金台帳ベース運用が成立しない場合のフォールバック表示）。
        # clean_mode=True のときはユーザー指示（2026-05-19）により以下を全て省略し、
        # 賃金台帳ベース1本のレイアウトに統一する:
        #   - 損益計算書データ（販管費）セクション
        #   - 役員報酬の控除セクション
        #   - 給与支給総額算定の【参考】2行
        #   - 1人当たり給与支給総額の【参考】式3つ
        # 決算書PL値はテンプレ転記用セクションのみで直書きで提示する。
        r = 7
        _cell(ws1, r, 2, pl_section_header, HEADER_FONT, border=None)
        _cell(ws1, r, 3, f'出所: {pl_source}', SMALL_FONT, border=None)
        r += 1
        for i, h in enumerate(['科目', '金額（円）', '備考']):
            _cell(ws1, r, 2 + i, h, HEADER_FONT_WHITE, fill=FILL_HEADER)

        items = [
            ('給料手当', financial.salary,
                _build_pl_note('salary', default_pl_note='販管費より｜正社員給与')),
            ('雑給', financial.misc_wages,
                _build_pl_note('misc_wages', default_pl_note='販管費より｜パート・アルバイト給与')),
            ('賞与', financial.bonus,
                _build_pl_note('bonus', default_pl_note='販管費より')),
            ('法定福利費', financial.legal_welfare,
                _build_pl_note('legal_welfare', excluded=True)),
            ('福利厚生費', financial.welfare,
                _build_pl_note('welfare', excluded=True)),
        ]
        # 各科目のセル位置を後段の Excel 計算式で参照できるよう記録
        item_rows: dict[str, int] = {}
        item_keys = ['salary', 'misc_wages', 'bonus', 'legal_welfare', 'welfare']
        for (name, amount, note), key in zip(items, item_keys):
            r += 1
            _cell(ws1, r, 2, name)
            # AI抽出値はオレンジで塗る（除外行は灰色優先）
            is_excluded = '除外' in note
            cell_fill = FILL_GRAY if is_excluded else FILL_AI_EXTRACTED
            _cell(ws1, r, 3, amount, fmt=NUMBER_FMT, fill=cell_fill)
            _cell(ws1, r, 4, note, SMALL_FONT, fill=cell_fill if is_excluded else None)
            if is_excluded:
                ws1.cell(r, 2).fill = FILL_GRAY
            item_rows[key] = r

        r += 1
        a_row = r  # (A) の行を覚えておく → 後段で参照
        # 給与関連合計 (A) を Excel 式で表現（給料手当+雑給+賞与）。
        # ユーザーが Excel 上で値を編集しても合計が自動更新される。
        a_formula = (
            f'=C{item_rows["salary"]}+C{item_rows["misc_wages"]}+C{item_rows["bonus"]}'
        )
        _cell(ws1, r, 2, '給与関連合計（A）', BOLD_FONT, fill=FILL_BLUE)
        _cell(ws1, r, 3, a_formula, BOLD_FONT, NUMBER_FMT, FILL_BLUE)
        _cell(ws1, r, 4,
              f'機械計算: C{item_rows["salary"]}+C{item_rows["misc_wages"]}+C{item_rows["bonus"]}（給料手当+雑給+賞与）',
              SMALL_FONT, fill=FILL_BLUE)

        # 役員報酬
        r += 2
        _cell(
            ws1, r, 2,
            '【役員報酬の控除】※公募要領 p.10／応募申請の手引き p.24「役員報酬・役員人数は含めません」',
            HEADER_FONT, border=None,
        )
        # ソース判定: 賃金状況報告シート由来か、PL 由来（年額一致）か
        yakuin_source_is_pl = (
            financial.officer_compensation > 0
            and yakuin_hoshu_3m == int(financial.officer_compensation / 4)
        )
        if yakuin_source_is_pl:
            yakuin_source_label = f'出所: {pl_source}（決算書PDF）'
        elif wage_report_source:
            yakuin_source_label = f'出所: {wage_report_source}'
        else:
            yakuin_source_label = '出所: （不明）'
        _cell(ws1, r, 3, yakuin_source_label, SMALL_FONT, border=None)
        r += 1

        if yakuin_source_is_pl:
            # PL 由来: 決算書PDFに年額が直接記載されているため、年額1段で表示。
            # （3ヶ月合計の概念は決算書には無いので、÷4 → ×4 の逆算を出力に出さない）
            b_row = r
            _cell(ws1, r, 2, '役員報酬（年額）（B）', BOLD_FONT, fill=FILL_YELLOW)
            _cell(ws1, r, 3, int(financial.officer_compensation),
                  BOLD_FONT, NUMBER_FMT, FILL_AI_EXTRACTED)
            _cell(
                ws1, r, 4,
                f'決算書記載: 損益計算書「販売費及び一般管理費」内「役員報酬」（年額）'
                f'｜{_ai_source_tag("officer_compensation")}',
                SMALL_FONT, fill=FILL_AI_EXTRACTED,
            )
        else:
            # 賃金状況報告シート由来 or 不明: 3ヶ月合計 → 年額（×4） の2段表示。
            # 賃金状況報告シート（補助金提出書類のひとつ）は3ヶ月合計の様式なので、×4で年額化する。
            yakuin_3m_row = r
            _cell(ws1, r, 2, '役員報酬（3ヶ月合計）')
            yakuin_note = (
                '賃金状況報告シートより（3ヶ月合計）'
                if wage_report_source else '※出所不明（賃金状況報告シート未取込）'
            )
            _cell(ws1, r, 3, yakuin_hoshu_3m, fmt=NUMBER_FMT)
            _cell(ws1, r, 4, yakuin_note, SMALL_FONT)
            r += 1
            b_row = r
            _cell(ws1, r, 2, '役員報酬（年額）（B）', BOLD_FONT, fill=FILL_YELLOW)
            _cell(ws1, r, 3, f'=C{yakuin_3m_row}*4', BOLD_FONT, NUMBER_FMT, FILL_YELLOW)
            _cell(ws1, r, 4, f'機械計算: C{yakuin_3m_row}×4（3ヶ月合計 × 4）',
                  SMALL_FONT, fill=FILL_YELLOW)
    else:
        # クリーンモード: 決算書PL値由来セクションを全省略。
        # r は凡例直後の r=5 のまま → 後続 `r += 2` で r=7 となり、給与支給総額セクションが B7 から開始。
        # 後段で参照される変数はスタブ（このパスでは参照されない式は出さない）。
        a_row = None
        b_row = None
        item_rows = {}

    # 給与支給総額
    # 公募要領 p.10／応募申請の手引き p.24「役員報酬・役員人数は含めません」+
    # 実務知識ベース（docs/補助金_実務知識ベース.md）の R216 定義に従い、
    # 「給与支給総額」は **賃金台帳の12ヶ月課税給与合計（役員除外・12ヶ月在籍者のみ）**
    # を採用する。賃金台帳由来データが無い場合のみ決算書PL値ベースにフォールバック。
    # 決算書PL値は突合チェック用の参考値として併記する。
    r += 2
    _cell(
        ws1, r, 2,
        '【給与支給総額の算定】※公募要領 p.10／応募申請の手引き p.24：'
        '給与支給総額は賃金台帳の課税給与（役員報酬・通勤費非課税分・中途者除外）',
        HEADER_FONT, border=None,
    )
    r += 1
    if not clean_mode:
        # 参考値1: (A) の再掲（役員報酬込み・決算書PL値）
        _cell(ws1, r, 2, '【参考】給与関連合計（決算書PL値・役員報酬含む）')
        _cell(ws1, r, 3, f'=C{a_row}', fmt=NUMBER_FMT)
        _cell(ws1, r, 4,
              f'機械計算: =C{a_row}（=A の再掲。決算書PL値ベース・役員報酬込みの参考値）',
              SMALL_FONT)
        r += 1
        # 参考値2: (A-B) 役員報酬除外（決算書PL値ベース）— 突合用
        _cell(ws1, r, 2, '【参考】給与支給総額（決算書PL値・役員除外）')
        _cell(ws1, r, 3, f'=C{a_row}-C{b_row}', fmt=NUMBER_FMT)
        _cell(ws1, r, 4,
              f'機械計算: =C{a_row}-C{b_row}（決算書PL値ベース。賃金台帳ベース値との'
              f'突合チェックに使用。通勤費非課税分・中途者を含むため R216 母数とは差が出る）',
              SMALL_FONT)
        r += 1
    # メイン: 賃金台帳ベース給与支給総額（R216）
    total_excl_row = r  # 後段の賃上げ計画シート・1人当たり計算からシート間参照する
    if ledger_total_cell:
        # 賃金台帳由来の12ヶ月合計（12ヶ月在籍者のみ・役員除外済み）
        _cell(ws1, r, 2,
              '給与支給総額（賃金台帳ベース／R216）', BOLD_FONT, fill=FILL_GREEN)
        _cell(ws1, r, 3, f'={ledger_total_cell}',
              BOLD_FONT, NUMBER_FMT, FILL_GREEN)
        _cell(ws1, r, 4,
              f'機械計算: ={ledger_total_cell}（「従業員別明細」シートの12ヶ月在籍合計を参照。'
              f'出典: 賃金台帳「{ledger_source}」の各従業員の課税支給合計列を12ヶ月分集計し、'
              f'12ヶ月在籍者のみを抽出した値。'
              f'公募要領 p.10／実務知識ベース R216 定義：役員報酬・非課税通勤手当を除外）',
              SMALL_FONT, fill=FILL_GREEN)
    else:
        # フォールバック: 賃金台帳由来データなし → 決算書PL値 (A-B) を採用
        _cell(ws1, r, 2,
              '給与支給総額（決算書PL値ベース／役員報酬除外）', BOLD_FONT, fill=FILL_YELLOW)
        _cell(ws1, r, 3, f'=C{a_row}-C{b_row}', BOLD_FONT, NUMBER_FMT, FILL_YELLOW)
        _cell(ws1, r, 4,
              f'機械計算: =C{a_row}-C{b_row}（賃金台帳由来データが無いため決算書PL値で代用。'
              f'通勤費非課税分・中途者を含むため R216 厳密値より過大評価の可能性あり）',
              SMALL_FONT, fill=FILL_YELLOW)

    # 従業員数
    r += 2
    _cell(ws1, r, 2, '【従業員数と1人当たり給与支給総額】', HEADER_FONT, border=None)
    # 雇用区分の人数集計: 賃金状況報告シートがあればそちら、なければ賃金台帳から逆算
    headcount_source = wage_report_source or ledger_source
    _cell(ws1, r, 3, f'出所: {headcount_source}（役員数は{registry_source}）',
          SMALL_FONT, border=None)
    r += 1
    for i, h in enumerate(['項目', '人数/金額', '備考']):
        _cell(ws1, r, 2 + i, h, HEADER_FONT_WHITE, fill=FILL_HEADER)

    for name, val, note in [
        ('正規雇用従業員', f'{seishain_count}人', ''),
        ('契約社員', '0人', ''),
        ('パート・アルバイト', f'{part_count}人', ''),
        ('役員', f'{yakuin_count}人', '※従業員数に含まず'),
    ]:
        r += 1
        _cell(ws1, r, 2, name)
        _cell(ws1, r, 3, val)
        _cell(ws1, r, 4, note, SMALL_FONT)

    r += 1
    c_row = r  # 1人当たり計算で参照する用
    _cell(ws1, r, 2, '従業員合計（C）', BOLD_FONT, fill=FILL_BLUE)
    # 値は数値、表示は「3人」と維持（書式 '0"人"' で人付き表示）。
    # 数値で書くことで下段の (A)÷(C) 等の Excel 式から参照可能になる。
    _cell(ws1, r, 3, total_emp, BOLD_FONT, fmt='0"人"', fill=FILL_BLUE)
    _cell(ws1, r, 4, '', fill=FILL_BLUE)

    # FTE
    if employees_detail:
        r += 2
        _cell(ws1, r, 2, '【FTE換算（12ヶ月在籍者のみ／中途入退社は除外）】',
              HEADER_FONT, border=None)
        _cell(ws1, r, 3,
              f'出所: 賃金台帳「{ledger_source}」（雇用形態・月間労働時間列を集計）',
              SMALL_FONT, border=None)
        r += 1
        _cell(ws1, r, 2, '標準年間労働時間')
        _cell(ws1, r, 3, f'{STANDARD_ANNUAL_HOURS}時間')
        _cell(ws1, r, 4,
              f'40h/週 × 52週（IT2026 通常枠公募要領 p.9-10の標準値）',
              SMALL_FONT)
        r += 1
        seishain_fte_row = r
        _cell(ws1, r, 2, '正社員FTE合計（12ヶ月在籍のみ）')
        _cell(ws1, r, 3, round(fte_seishain, 2), fmt='0.00')
        _cell(ws1, r, 4,
              f'引用元: 賃金台帳「{ledger_source}」より、各従業員の「雇用形態」列が'
              f'『正社員』または『契約社員』を含み、かつ12ヶ月在籍している人数。'
              f'中途入退社者は公募要領 p.9-10「全月分の給与等の支給を受けた従業員」により除外',
              SMALL_FONT)
        r += 1
        part_fte_row = r
        _cell(ws1, r, 2, 'パートFTE換算合計（12ヶ月在籍のみ）')
        _cell(ws1, r, 3, round(fte_part, 2), fmt='0.00')
        _cell(ws1, r, 4,
              f'引用元: 賃金台帳「{ledger_source}」より、各パート・アルバイト従業員の'
              f'「月間平均労働時間」 ÷ ({STANDARD_ANNUAL_HOURS}/12 = {round(STANDARD_ANNUAL_HOURS/12, 1)}時間) で正社員換算し合算。'
              f'12ヶ月在籍者のみ',
              SMALL_FONT)
        if excluded_midyear:
            r += 1
            _cell(ws1, r, 2, '中途入退社で除外した人数')
            _cell(ws1, r, 3, excluded_midyear)
            _cell(ws1, r, 4,
                  '※R215 算定対象外（公募要領「全月分の給与等の支給を受けた従業員」）',
                  SMALL_FONT)
            ws1.cell(r, 2).fill = FILL_GRAY
            ws1.cell(r, 3).fill = FILL_GRAY
            ws1.cell(r, 4).fill = FILL_GRAY
        if excluded_partial_zero:
            r += 1
            _cell(ws1, r, 2, '0円支給月があり除外した人数')
            _cell(ws1, r, 3, excluded_partial_zero)
            _cell(ws1, r, 4,
                  '※R215 算定対象外（公募要領「全月分の給与等の支給を受けた従業員」に非該当。'
                  '中途入社・退職なら正しい除外。休職等で算入する場合は該当月を実支給額に修正、'
                  '0円月に賞与支給があった場合は該当月に賞与額を記入し年間賞与欄から同額を'
                  '差し引いて再実行）',
                  SMALL_FONT)
            ws1.cell(r, 2).fill = FILL_GRAY
            ws1.cell(r, 3).fill = FILL_GRAY
            ws1.cell(r, 4).fill = FILL_GRAY
        if excluded_bonus_only:
            r += 1
            _cell(ws1, r, 2, '賞与のみ受給で除外した人数')
            _cell(ws1, r, 3, excluded_bonus_only)
            _cell(ws1, r, 4,
                  '※R215 算定対象外（月次給与が全月0円のため「全月分の給与等の支給」に'
                  '非該当と整理＝2026-06-11ルール。年間賞与も給与支給総額に算入していません。'
                  '算入すべき場合は支給実態のある月に賞与額を記入し年間賞与欄から同額を'
                  '差し引いて再実行）',
                  SMALL_FONT)
            ws1.cell(r, 2).fill = FILL_GRAY
            ws1.cell(r, 3).fill = FILL_GRAY
            ws1.cell(r, 4).fill = FILL_GRAY
        if excluded_zero:
            r += 1
            _cell(ws1, r, 2, '給与支給0円で除外した人数')
            _cell(ws1, r, 3, excluded_zero)
            _cell(ws1, r, 4,
                  '※R215 算定対象外（公募要領「全月分の給与等の支給を受けた従業員」に非該当。'
                  '賃金台帳に0円と明記された従業員。退職済み行・無給の家族従業者等でないか要確認）',
                  SMALL_FONT)
            ws1.cell(r, 2).fill = FILL_GRAY
            ws1.cell(r, 3).fill = FILL_GRAY
            ws1.cell(r, 4).fill = FILL_GRAY
        r += 1
        d_row = r  # 1人当たり計算で参照する用
        # FTE換算後（D）を Excel 関数化（正社員FTE + パートFTE）
        _cell(ws1, r, 2, 'FTE換算後従業員数（D）', BOLD_FONT, fill=FILL_GREEN)
        _cell(ws1, r, 3, f'=C{seishain_fte_row}+C{part_fte_row}',
              BOLD_FONT, '0.00', FILL_GREEN)
        _cell(ws1, r, 4,
              f'機械計算: C{seishain_fte_row}+C{part_fte_row}（正社員FTE + パートFTE）',
              SMALL_FONT, fill=FILL_GREEN)
    else:
        d_row = None  # employees_detail なし時は (A)÷(D) 系は出さない

    # 1人当たり計算
    r += 2
    _cell(ws1, r, 2, '【1人当たり給与支給総額】', Font(name='游ゴシック', size=12, bold=True), border=None)
    if clean_mode:
        _cell(ws1, r, 3,
              f'出所: 賃金台帳「{ledger_source}」（分子）／ 賃金台帳より算出した FTE（分母）',
              SMALL_FONT, border=None)
    r += 1
    for i, h in enumerate(['算出方法', '金額', '備考']):
        _cell(ws1, r, 2 + i, h, HEADER_FONT_WHITE, fill=FILL_HEADER_DARK)

    # 各算出方法を Excel 式で。ユーザーが上流セルを編集すると即時再計算。
    # ゼロ除算は IF で防御。
    # メイン値（推奨）の母数（分子）は「賃金台帳ベース給与支給総額」セル C{total_excl_row} を使う。
    # clean_mode=True 時: メイン式のみ（決算書PL値由来の参考式は省略）
    # clean_mode=False 時: 参考式(A)÷(C)/(A-B)÷(C)/(A)÷(D)も併記
    main_label_suffix = '（賃金台帳ベース・推奨）' if ledger_total_cell else '（決算書PL値・参考）'
    if clean_mode:
        # クリーンモード: メイン式（賃金台帳÷FTE）のみ
        if employees_detail and fte_adjusted > 0 and d_row is not None:
            calc_methods = [
                (
                    f'1人当たり給与支給総額 = 給与支給総額 ÷ FTE{main_label_suffix}',
                    f'=IF(C{d_row}=0,0,ROUND(C{total_excl_row}/C{d_row},0))',
                    f'分子=C{total_excl_row}（賃金台帳12ヶ月合計／R216）／'
                    f'分母=C{d_row}（賃金台帳由来のFTE：正社員FTE+パートFTE）',
                ),
            ]
        else:
            calc_methods = [
                (
                    f'1人当たり給与支給総額 = 給与支給総額 ÷ 従業員数{main_label_suffix}',
                    f'=IF(C{c_row}=0,0,ROUND(C{total_excl_row}/C{c_row},0))',
                    f'分子=C{total_excl_row}（賃金台帳12ヶ月合計）／'
                    f'分母=C{c_row}（従業員合計）',
                ),
            ]
    else:
        # 従来モード: 決算書PL値由来の参考式も併記
        calc_methods = [
            ('【参考】(A)÷(C) 決算書PL値・頭数割り',
             f'=IF(C{c_row}=0,0,ROUND(C{a_row}/C{c_row},0))', ''),
            ('【参考】(A-B)÷(C) 決算書PL値・役員除外・頭数',
             f'=IF(C{c_row}=0,0,ROUND((C{a_row}-C{b_row})/C{c_row},0))', ''),
        ]
        if employees_detail and fte_adjusted > 0 and d_row is not None:
            calc_methods.extend([
                ('【参考】(A)÷(D) 決算書PL値・FTE換算',
                 f'=IF(C{d_row}=0,0,ROUND(C{a_row}/C{d_row},0))', ''),
                (f'1人当たり給与支給総額 = 給与支給総額 ÷ FTE{main_label_suffix}',
                 f'=IF(C{d_row}=0,0,ROUND(C{total_excl_row}/C{d_row},0))', ''),
            ])
        else:
            calc_methods.append(
                (f'1人当たり給与支給総額 = 給与支給総額 ÷ 従業員数{main_label_suffix}',
                 f'=IF(C{c_row}=0,0,ROUND(C{total_excl_row}/C{c_row},0))', ''),
            )

    for i, (label, formula, note) in enumerate(calc_methods):
        r += 1
        is_last = (i == len(calc_methods) - 1)
        _cell(ws1, r, 2, label, BOLD_FONT if is_last else NORMAL_FONT,
              fill=FILL_GREEN if is_last else None)
        _cell(ws1, r, 3, formula, RESULT_FONT if is_last else NORMAL_FONT,
              NUMBER_FMT, FILL_GREEN if is_last else None)
        if note:
            _cell(ws1, r, 4, note, SMALL_FONT,
                  fill=FILL_GREEN if is_last else None)
            ws1.cell(r, 4).alignment = Alignment(wrap_text=True, vertical='top')

    # テンプレート転記用（全項目 AI 抽出：決算書PDF由来）
    r += 2
    _cell(ws1, r, 2, '【2026テンプレート転記用】', HEADER_FONT, border=None)
    _cell(ws1, r, 3, f'出所: {pl_source}', SMALL_FONT, border=None)
    r += 1
    _cell(ws1, r, 2,
          '※下記すべてAI抽出値（決算書PDF由来）。テンプレ転記前に決算書原本と照合してください。',
          SMALL_FONT, border=None)
    ws1.cell(r, 2).fill = FILL_AI_EXTRACTED
    # 決算書のどの表に載っている値かをセル右側の備考に明示する
    # （会計指針で標準化されているため固定マッピングで 100% 正確）
    # 加えて PDF テキストから逆引きしたページ番号も付記する
    # 出所表記は AI に生成させず Python の固定マッピングで詳しく書く。
    # 決算書の標準フォーマット（会計指針）に基づく固定対応なので 100% 正確。
    # 業種別の呼称揺れや計算式も併記し、人間が決算書原本と照合するときの目印にする。
    PL_LOCATIONS = {
        'salary': (
            '損益計算書「販売費及び一般管理費」内「給料手当」'
            '（製造業は製造原価報告書「賃金」「給料」「労務費」も合算対象）'
        ),
        'misc_wages': '損益計算書「販売費及び一般管理費」内「雑給」',
        'bonus': (
            '損益計算書「販売費及び一般管理費」内「賞与」'
            '「賞与引当金繰入額」'
        ),
        'revenue': (
            '損益計算書 第1段「売上高」'
            '（建設業＝「完成工事高」、不動産業＝「営業収益」等の呼称あり）'
        ),
        'gross_profit': '損益計算書 第3段「売上総利益」（＝売上高 − 売上原価）',
        'operating_profit': (
            '損益計算書 第5段「営業利益」'
            '（＝売上総利益 − 販売費及び一般管理費）'
        ),
        'ordinary_profit': (
            '損益計算書 第7段「経常利益」'
            '（＝営業利益 ＋ 営業外収益 − 営業外費用）'
        ),
        'depreciation': (
            '損益計算書「販売費及び一般管理費」内「減価償却費」'
            '（製造業は製造原価報告書「減価償却費」と合算。複数箇所に分散記載される場合あり）'
        ),
    }
    # テンプレ転記用の値の出し方:
    # - clean_mode=True: 上の損益計算書データセクションが存在しないため、給料手当・雑給・賞与も
    #   決算書PL値（financial.*）を直書きする。
    # - clean_mode=False: 上のセクションがあるので、給料手当・雑給・賞与は上のセルを Excel 式で参照
    #   する（ユーザーが上で誤読を訂正すると転記用も自動追従）。
    if clean_mode:
        template_items = [
            ('給料手当', financial.salary, 'salary', True),
            ('雑給', financial.misc_wages, 'misc_wages', True),
            ('賞与手当', financial.bonus, 'bonus', True),
            ('売上高', financial.revenue, 'revenue', False),
            ('粗利益', financial.gross_profit, 'gross_profit', False),
            ('営業利益', financial.operating_profit, 'operating_profit', False),
            ('経常利益', financial.ordinary_profit, 'ordinary_profit', False),
            ('減価償却費', financial.depreciation, 'depreciation', False),
        ]
    else:
        template_items = [
            ('給料手当（販管費E5）', f'=C{item_rows["salary"]}', 'salary', True),
            ('雑給（販管費E6）', f'=C{item_rows["misc_wages"]}', 'misc_wages', True),
            ('賞与手当（販管費E7）', f'=C{item_rows["bonus"]}', 'bonus', True),
            ('売上高（B10）', financial.revenue, 'revenue', False),
            ('粗利益（B11）', financial.gross_profit, 'gross_profit', False),
            ('営業利益（B12）', financial.operating_profit, 'operating_profit', False),
            ('経常利益（B13）', financial.ordinary_profit, 'ordinary_profit', False),
            ('減価償却費（B14）', financial.depreciation, 'depreciation', False),
        ]
    for name, val, page_key, is_ref in template_items:
        r += 1
        _cell(ws1, r, 2, name)
        _cell(ws1, r, 3, val, fmt=NUMBER_FMT, fill=FILL_AI_EXTRACTED)
        location = PL_LOCATIONS[page_key]
        if is_ref:
            # 給料手当/雑給/賞与は販管費＋製造原価の合算が入りうるため、
            # ヘルパー _build_pl_note で内訳を生成して併記する。
            breakdown_note = _build_pl_note(
                page_key, default_pl_note=f'販管費「{name.split("（")[0]}」より'
            )
            if clean_mode:
                note = (
                    f'{breakdown_note}\n'
                    f'決算書記載: {location}'
                )
            else:
                note = (
                    f'{breakdown_note}\n'
                    f'上の販管費セル C{item_rows[page_key]} を参照／'
                    f'決算書記載: {location}'
                )
        else:
            # 売上高・粗利益・営業利益・経常利益・減価償却費は単一値で内訳なし。
            # 既存の決算書セクション説明 + AI抽出マーカー + ページ番号タグで出所を明示。
            note = (
                f'決算書記載: {location}'
                f'（{_ai_source_tag(page_key)}）'
                f'{_page_tag(page_key)}'
            )
        _cell(ws1, r, 4, note, SMALL_FONT)
        # 長い出所表記が読みやすいよう折り返し表示
        ws1.cell(r, 4).alignment = Alignment(wrap_text=True, vertical='top')

    ws1.column_dimensions['A'].width = 2
    # B列は項目名（例: '1人当たり給与支給総額 = 給与支給総額 ÷ FTE（賃金台帳ベース・推奨）'）
    # まで入れるため余裕を持たせる
    ws1.column_dimensions['B'].width = 44
    ws1.column_dimensions['C'].width = 20
    # D列は転記用の詳細な出所表記を入れるため広め。wrap_text=True と併用して折り返し表示
    ws1.column_dimensions['D'].width = 58

    # 備考列（B/C/D）に wrap_text を一括適用 + 長文行の行高を見積もり調整。
    # 個別セルで wrap_text を設定し忘れても、表示崩れ（長文がセル境界を越える／隣セルに被る）
    # を防ぐ。行高は openpyxl が「自動」を計算してくれないため、文字数から見積もって明示設定する。
    _apply_readable_layout(ws1, text_columns=(2, 3, 4), max_widths={2: 44, 3: 20, 4: 58})

    # ===== Sheet 2: 従業員別明細 =====
    # 12ヶ月モード（賃金台帳由来）: 月12列＋12ヶ月合計＋月間平均
    # 3ヶ月モード（賃金状況報告シート由来 or データ欠落）: 従来の3ヶ月レイアウト
    # 後段の Sheet 1 (賃金台帳ベース給与支給総額) と Sheet 3 (賃上げ計画) で
    # 合計列の位置を参照するので、各セル位置を変数に保持する。
    ledger_total_ref: str | None = None  # Sheet 1 から参照する「12ヶ月在籍合計」セルアドレス
    if employees_detail:
        ws2 = wb.create_sheet('従業員別明細')
        title_suffix = '（12ヶ月）' if has_12_months else '（直近3ヶ月）'
        _cell(ws2, 2, 2, f'従業員別給与明細{title_suffix}', TITLE_FONT, border=None)
        _cell(ws2, 2, 4, f'出所: {ledger_source}', SMALL_FONT, border=None)
        _cell(ws2, 3, 2,
              '※氏名・雇用形態・月給・時給はAIが賃金台帳から読み取った値です。'
              '誤読の可能性があるため賃金台帳原本と照合してください。',
              SMALL_FONT, border=None)
        ws2.cell(3, 2).fill = FILL_AI_EXTRACTED

        # ヘッダー構成
        if has_12_months:
            month_headers = [f'{lbl}支給額' for lbl in month_labels_full]
            headers = (
                ['No', '氏名', '雇用形態']
                + month_headers
                + ['年間給与計(賞与込)', '月間平均', '時給', '月間平均時間', 'FTE', '最低賃金判定', '備考']
            )
            # 列位置インデックス（B=2 ベース）
            FIRST_MONTH_COL = 5            # E列が事業年度開始月
            LAST_MONTH_COL = FIRST_MONTH_COL + 11  # P列が事業年度末月
            ANNUAL_TOTAL_COL = LAST_MONTH_COL + 1  # Q列 = 12ヶ月合計
            MONTHLY_AVG_COL = ANNUAL_TOTAL_COL + 1  # R列 = 月間平均
            HR_COL = MONTHLY_AVG_COL + 1            # S列
            HOURS_COL = HR_COL + 1                   # T列
            FTE_COL = HOURS_COL + 1                  # U列
            JUDGE_COL = FTE_COL + 1                  # V列
            NOTE_COL = JUDGE_COL + 1                 # W列
        else:
            headers = ['No', '氏名', '雇用形態', '1月基本給', '2月基本給', '3月基本給',
                       '3ヶ月平均', '時給', '月間平均時間', 'FTE', '最低賃金判定', '備考']
            FIRST_MONTH_COL = 5
            LAST_MONTH_COL = 7
            ANNUAL_TOTAL_COL = 8  # 3ヶ月平均（互換のため変数名は ANNUAL だが意味は 3M平均）
            MONTHLY_AVG_COL = None
            HR_COL = 9
            HOURS_COL = 10
            FTE_COL = 11
            JUDGE_COL = 12
            NOTE_COL = 13

        r = 4
        for i, h in enumerate(headers):
            _cell(ws2, r, 2 + i, h, HEADER_FONT_WHITE, fill=FILL_HEADER)
            ws2.cell(r, 2 + i).alignment = Alignment(horizontal='center', wrap_text=True)

        # 中途入退社社員のチェック視認性向上のため、行全体を灰色塗りする
        FILL_INCOMPLETE = PatternFill(start_color='DDDDDD', end_color='DDDDDD', fill_type='solid')
        # データ欠落月（中途者の非在籍月）専用の薄塗り
        FILL_MONTH_NODATA = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        first_data_row = 5
        for e in employees_detail:
            r += 1

            is_seishain = is_full_time_employment(e.get('type'))
            full_year = e.get('full_year', True)
            tenure_months = e.get('tenure_months', 12)
            # 給与支給0円（年計≤0）は R215/R216 算定対象外。FTE は 0 で表示する
            zero_wage = full_year and is_zero_wage_detail(e)
            # 在籍月数を反映した FTE（中途入退社は分母12を按分）
            tenure_factor = min(tenure_months, 12) / 12 if tenure_months > 0 else 0
            if zero_wage:
                fte = 0.0
            elif is_seishain:
                fte = 1.0 * tenure_factor
            else:
                monthly_h = e.get('monthly_hours', 0)
                fte = (monthly_h / standard_monthly) * tenure_factor if standard_monthly else 0

            # 行全体の塗り（優先: 算定対象外（中途/0円） > 非正規 > 通常）
            if not full_year or zero_wage:
                row_fill = FILL_INCOMPLETE
            elif not is_seishain:
                row_fill = FILL_GRAY
            else:
                row_fill = None

            # No / 氏名 / 雇用形態
            _cell(ws2, r, 2, e['no'], fill=row_fill)
            _cell(ws2, r, 3, e['name'], fill=row_fill)
            _cell(ws2, r, 4, e['type'], fill=row_fill)

            if has_12_months:
                wages_full = e.get('monthly_wages_full') or [0.0] * 12
                mask_full = e.get('month_data_mask') or [True] * 12
                in_service = [v for v, has in zip(wages_full, mask_full) if has and v > 0]
                avg_per_month = sum(in_service) / len(in_service) if in_service else 0

                # 月別12列を書き込み（非在籍月は薄塗りで視認性UP）
                for m_idx in range(12):
                    col = FIRST_MONTH_COL + m_idx
                    val = wages_full[m_idx]
                    has_data = mask_full[m_idx] if m_idx < len(mask_full) else True
                    # 在籍月なのに0円ということもあり得るので、マスクと0判定は分離
                    fill_to_use = row_fill
                    if not has_data:
                        # 行塗りより明確に「データなし月」を示す
                        fill_to_use = FILL_MONTH_NODATA if row_fill is None else row_fill
                    _cell(ws2, r, col, val if has_data else '',
                          fmt=NUMBER_FMT, fill=fill_to_use)

                # 年間給与計（賞与込）= 12ヶ月の月次課税給与合計 ＋ 年間賞与。
                # 月次セル（賞与抜き）と年間賞与を分離保持し、R216 はここで合算する
                # （公募要領 p.10: 給与支給総額に賞与を含む。月次に混ぜると最低賃金判定が歪む）。
                col_first = get_column_letter(FIRST_MONTH_COL)
                col_last = get_column_letter(LAST_MONTH_COL)
                _bonus = int(round(float(e.get('annual_bonus', 0) or 0)))
                _sum_expr = f'SUM({col_first}{r}:{col_last}{r})'
                _total_formula = f'={_sum_expr}+{_bonus}' if _bonus > 0 else f'={_sum_expr}'
                _cell(ws2, r, ANNUAL_TOTAL_COL, _total_formula,
                      BOLD_FONT, fmt=NUMBER_FMT, fill=row_fill)
                # 月間平均（在籍月のみ）
                _cell(ws2, r, MONTHLY_AVG_COL, round(avg_per_month), fmt=NUMBER_FMT, fill=row_fill)
            else:
                # 3ヶ月モード（賃金状況報告シート由来）
                m_vals = [e.get('m1', 0), e.get('m2', 0), e.get('m3', 0)]
                in_service = [v for v in m_vals if v > 0]
                avg3 = sum(in_service) / len(in_service) if in_service else 0
                for i, v in enumerate(m_vals):
                    _cell(ws2, r, FIRST_MONTH_COL + i, v, fmt=NUMBER_FMT, fill=row_fill)
                _cell(ws2, r, ANNUAL_TOTAL_COL, round(avg3), fmt=NUMBER_FMT, fill=row_fill)

            # 時給 / 月間平均時間 / FTE / 最低賃金判定 / 備考
            _cell(ws2, r, HR_COL, e.get('hr', 0), fmt=NUMBER_FMT, fill=row_fill)
            _cell(ws2, r, HOURS_COL, round(e.get('monthly_hours', 0), 1),
                  fmt='0.0', fill=row_fill)
            _cell(ws2, r, FTE_COL, round(fte, 2), fmt='0.00', fill=row_fill)
            judge_val = e.get('judge') or '-'
            _cell(ws2, r, JUDGE_COL, judge_val, fill=row_fill)

            note_parts = []
            if not full_year:
                # 0円明記月持ちは「在籍」と書くと休職者に対して誤りになるため
                # 「支給Nヶ月」表記で区別する（tenure_months は支給>0 の月数）
                if has_partial_zero_months_detail(e):
                    note_parts.append(
                        f'0円支給月あり（支給{tenure_months}ヶ月／R215・R216算定対象外）'
                    )
                elif is_all_zero_explicit_detail(e):
                    if (e.get('annual_bonus') or 0) > 0:
                        note_parts.append(
                            '賞与のみ受給（月次給与が全月0円／R215・R216算定対象外）'
                        )
                    else:
                        note_parts.append('給与支給0円（R215/R216 算定対象外）')
                else:
                    note_parts.append(f'中途入退社（在籍{tenure_months}ヶ月）')
                labels = [l for l in e.get('last_three_labels', []) if l]
                if labels:
                    note_parts.append(f'実体: {"/".join(labels)}')
            if zero_wage:
                note_parts.append('給与支給0円（R215/R216 算定対象外）')
            _bonus_note = float(e.get('annual_bonus', 0) or 0)
            if _bonus_note > 0:
                note_parts.append(f'年間給与計に賞与{_bonus_note:,.0f}円を算入（月次列には含めない）')
            _cell(ws2, r, NOTE_COL, ' '.join(note_parts), SMALL_FONT, fill=row_fill)

        # ── 合計行（全員 / 12ヶ月在籍のみの2段）────────────────────────
        # 12ヶ月在籍のみ合計は、Sheet 1 の「賃金台帳ベース給与支給総額（R216）」と
        # Sheet 3 の「賃上げ計画」基準値からシート間参照される。
        # has_any_amount は関数冒頭（ledger_total_cell の事前計算）で同一条件で算出済み。
        # ここで条件を変えると Sheet 1 の参照セルと実出力がズレるため再計算しない。
        if has_any_amount:
            last_data_row = r
            r += 1
            FILL_SUBTOTAL_ALL = PatternFill(start_color='B4C7E7', end_color='B4C7E7', fill_type='solid')
            FILL_SUBTOTAL_TARGET = PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid')

            # 合計（全員）
            all_row = r
            _cell(ws2, r, 2, '', BOLD_FONT, fill=FILL_SUBTOTAL_ALL)
            _cell(ws2, r, 3, '合計（全員）', BOLD_FONT, fill=FILL_SUBTOTAL_ALL)
            _cell(ws2, r, 4, '', fill=FILL_SUBTOTAL_ALL)
            # 月別 + 12ヶ月合計（or 3ヶ月平均）列を SUM
            sum_cols = list(range(FIRST_MONTH_COL, ANNUAL_TOTAL_COL + 1))
            if MONTHLY_AVG_COL is not None:
                sum_cols.append(MONTHLY_AVG_COL)
            for col_idx in sum_cols:
                col_letter = get_column_letter(col_idx)
                _cell(
                    ws2, r, col_idx,
                    f'=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})',
                    BOLD_FONT, fmt=NUMBER_FMT, fill=FILL_SUBTOTAL_ALL,
                )
            # 残り列はブランク（合計対象外）
            for col_idx in range(HR_COL, NOTE_COL + 1):
                _cell(ws2, r, col_idx, '', fill=FILL_SUBTOTAL_ALL)

            # 合計（12ヶ月在籍のみ）— R216 母数（給与支給0円の人は算定対象外）
            target_rows = [
                first_data_row + i for i, e in enumerate(employees_detail)
                if e.get('full_year', True) and not is_zero_wage_detail(e)
            ]
            r += 1
            target_row = r
            _cell(ws2, r, 2, '', BOLD_FONT, fill=FILL_SUBTOTAL_TARGET)
            _cell(ws2, r, 3, '合計（12ヶ月在籍のみ）', BOLD_FONT, fill=FILL_SUBTOTAL_TARGET)
            _cell(ws2, r, 4, '', fill=FILL_SUBTOTAL_TARGET)
            for col_idx in sum_cols:
                col_letter = get_column_letter(col_idx)
                if target_rows:
                    parts = '+'.join(f'{col_letter}{tr}' for tr in target_rows)
                    formula = f'={parts}'
                else:
                    formula = 0
                _cell(
                    ws2, r, col_idx, formula,
                    BOLD_FONT, fmt=NUMBER_FMT, fill=FILL_SUBTOTAL_TARGET,
                )
            for col_idx in range(HR_COL, NOTE_COL + 1):
                _cell(ws2, r, col_idx, '', fill=FILL_SUBTOTAL_TARGET)

            # Sheet 1 から参照する「R216 母数」のセルアドレス
            # 12ヶ月モードでは 12ヶ月合計列（Q列）、3ヶ月モードは「3ヶ月平均」列で
            # 厳密には R216 母数にはならないので Sheet 1 側でフォールバック分岐する。
            if has_12_months:
                ledger_total_ref = (
                    f"'従業員別明細'!{get_column_letter(ANNUAL_TOTAL_COL)}{target_row}"
                )

        # 凡例
        r += 2
        _cell(ws2, r, 2,
              '※灰色（濃）行＝直近事業年度に12ヶ月在籍していない社員（中途入社・退職含む）、'
              '0円支給月がある社員、給与支給0円または賞与のみ受給の社員（備考欄参照）。'
              'R215/R216の母数（12ヶ月在籍のみ合計）からは除外されます。',
              SMALL_FONT, border=None)
        if has_12_months:
            r += 1
            _cell(ws2, r, 2,
                  '　月列は事業年度開始月から時系列で12ヶ月分。データ欠落月は薄灰色で表示。',
                  SMALL_FONT, border=None)
        else:
            r += 1
            _cell(ws2, r, 2,
                  '　1月/2月/3月の列見出しは便宜表示で、中途者の列は実在籍月の時系列順（備考の「実体」欄を参照）。',
                  SMALL_FONT, border=None)
        r += 1
        _cell(ws2, r, 2,
              '※灰色（薄）行＝非正規雇用（パート・アルバイト）。',
              SMALL_FONT, border=None)
        r += 1
        _cell(ws2, r, 2,
              '※最低賃金判定の「-」＝このシートでは判定なし（加点判定タスクで都道府県を指定した場合に判定されます）。',
              SMALL_FONT, border=None)

        # 列幅
        if has_12_months:
            # No, 氏名, 雇用形態, 月12列, 12ヶ月合計, 月間平均, 時給, 月間時間, FTE, 判定, 備考
            widths = [4, 5, 14, 12] + [10] * 12 + [13, 11, 8, 12, 8, 12, 30]
        else:
            widths = [4, 5, 14, 12, 12, 12, 12, 12, 8, 13, 8, 12, 30]
        for i, w in enumerate(widths):
            ws2.column_dimensions[get_column_letter(i + 1)].width = w

    # ===== Sheet 3: 賃上げ計画 =====
    # 機械計算セルは全部 Excel 関数化（ユーザーが C5 を編集すると D5-F5 が自動再計算）
    ws3 = wb.create_sheet('賃上げ計画')
    _cell(ws3, 2, 2, '賃上げ計画シミュレーション', TITLE_FONT, border=None)

    r = 4
    for i, h in enumerate(['', '直近決算期\n(実績値)', '1年目計画', '2年目計画', '3年目計画']):
        _cell(ws3, r, 2 + i, h, HEADER_FONT_WHITE, fill=FILL_HEADER)
        ws3.cell(r, 2 + i).alignment = Alignment(horizontal='center', wrap_text=True)

    r += 1
    basis_row = r  # 給与支給総額 行（C列 = 直近実績、D-F 列 = 計画値）
    _cell(ws3, r, 2, '給与支給総額', BOLD_FONT)
    # C5: 直近実績（給与支給総額計算シートの「給与支給総額（公募要領定義／役員報酬除外）」と連動）
    # 公募要領 p.10／応募申請の手引き p.24「役員報酬・役員人数は含めません」を根拠に、
    # 賃上げ計画の母数は (A-B) = total_excl_row を参照する（旧実装は a_row だった）。
    _cell(ws3, r, 3, f"='給与支給総額計算'!C{total_excl_row}", fmt=NUMBER_FMT)
    # D5, E5, F5: 前年 × 1.03（年率3%）。ROUND で円単位に丸め
    prev_col = 'C'
    for i in range(1, 4):
        col_letter = get_column_letter(3 + i)  # D, E, F
        _cell(ws3, r, 3 + i, f'=ROUND({prev_col}{r}*1.03, 0)', fmt=NUMBER_FMT)
        prev_col = col_letter

    r += 1
    _cell(ws3, r, 2, '増加率（対基準年）')
    _cell(ws3, r, 3, '-')
    # D6, E6, F6: (D5 - C5) / C5 → 累積増加率
    for i in range(1, 4):
        col_letter = get_column_letter(3 + i)
        _cell(ws3, r, 3 + i,
              f'=({col_letter}{basis_row}-$C${basis_row})/$C${basis_row}',
              fmt=PCT_FMT)

    ws3.column_dimensions['B'].width = 28
    for col in ['C', 'D', 'E', 'F']:
        ws3.column_dimensions[col].width = 18

    # 保存
    wb.save(str(output_path))
    wb.close()
    logger.info(f'給与支給総額計算 保存: {output_path}')
    return output_path
