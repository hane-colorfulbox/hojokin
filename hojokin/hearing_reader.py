# -*- coding: utf-8 -*-
"""ヒアリングシートExcel読み取り"""
from __future__ import annotations

import logging
from pathlib import Path
import openpyxl

logger = logging.getLogger(__name__)

# 全角→半角変換テーブル
_ZEN2HAN = str.maketrans(
    '０１２３４５６７８９'
    'ＡＢＣＤＥＦＧＨＩＪＫＬＭＮＯＰＱＲＳＴＵＶＷＸＹＺ'
    'ａｂｃｄｅｆｇｈｉｊｋｌｍｎｏｐｑｒｓｔｕｖｗｘｙｚ'
    '－（）　＠．／：，',
    '0123456789'
    'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    'abcdefghijklmnopqrstuvwxyz'
    '-() @./：,',
)


def normalize_value(value):
    """全角数字・英字を半角に変換。数値化できるものは数値にする。"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).translate(_ZEN2HAN).strip()
    # 先頭0の数字列は電話番号等なので文字列のまま返す
    s_clean = s.replace(',', '').replace('円', '').replace('人', '').replace('時間', '')
    if s_clean.startswith('0') and s_clean.isdigit() and len(s_clean) >= 2:
        return s_clean
    try:
        if '.' in s_clean:
            return float(s_clean)
        return int(s_clean)
    except ValueError:
        return s


def normalize_phone(value) -> str | None:
    """数値の電話番号を先頭0補完した文字列に変換"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        s = str(int(value))
        if len(s) == 9:
            return '0' + s
        if len(s) == 10 and not s.startswith('0'):
            return '0' + s
        return s
    return str(value)


def read_hearing_sheet(path: Path) -> dict[int, any]:
    """
    ヒアリングシートを読み取り、{行番号: 値} の辞書を返す。
    シート名は「基本情報」を使用。
    """
    wb = openpyxl.load_workbook(path, data_only=True)

    # シート名を探す（「基本情報」優先、なければ最初のシート）
    sheet_name = None
    for name in wb.sheetnames:
        if '基本情報' in name:
            sheet_name = name
            break
    if sheet_name is None:
        sheet_name = wb.sheetnames[0]
        logger.warning(f'「基本情報」シートが見つからず。{sheet_name} を使用')

    ws = wb[sheet_name]
    data = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        row_num = row[0].row
        label = row[1].value if len(row) > 1 else None  # B列
        value = row[2].value if len(row) > 2 else None   # C列

        if label is not None:
            data[row_num] = {
                'label': str(label).strip(),
                'value': normalize_value(value),
            }

    wb.close()
    logger.info(f'ヒアリングシート読込: {path.name}, {len(data)}行')
    return data


def transfer_hearing_to_tenki(hearing_data: dict, ws_tenki, mapping: list[tuple[int, int, bool]]) -> int:
    """
    ヒアリングデータを転記シートに転記。
    mapping: [(ヒアリング行, 転記行, 電話番号変換フラグ), ...]
    転記した件数を返す。
    """
    from .template_filler import _safe_write_cell

    count = 0
    for h_row, t_row, is_phone in mapping:
        if h_row not in hearing_data:
            continue
        value = hearing_data[h_row]['value']
        if value is None:
            continue

        if is_phone:
            value = normalize_phone(value)

        _safe_write_cell(ws_tenki, t_row, 2, value)
        count += 1
        logger.debug(f'転記: ヒアリング行{h_row} → 転記行{t_row}: {value!r}')

    return count
