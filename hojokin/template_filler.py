# -*- coding: utf-8 -*-
"""申請書テンプレートへのデータ転記"""
from __future__ import annotations

import logging
import shutil
from pathlib import Path
import openpyxl

from openpyxl.cell.cell import MergedCell

from .models import ExtractionResult
from .config import TemplateMapping, get_min_wage

logger = logging.getLogger(__name__)


def _safe_write_cell(ws, row: int, col: int, value):
    """結合セルに対応した安全な書き込み。

    openpyxl では結合範囲の左上以外のセルに書き込むと
    ``'MergedCell' object attribute 'value' is read-only`` が出る。
    そのため、対象セルが MergedCell の場合は結合範囲の左上セルに書き込む。
    """
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                ws.cell(
                    row=merged_range.min_row,
                    column=merged_range.min_col,
                ).value = value
                return
    cell.value = value


def clear_manual_cells(wb: openpyxl.Workbook, mapping: TemplateMapping) -> int:
    """テンプレートの手入力セルをクリア（数式は残す）"""
    cleared = 0

    def is_formula(v):
        return isinstance(v, str) and v.startswith('=')

    # 転記シート: B列のテキスト項目範囲
    if '転記' in wb.sheetnames:
        ws_t = wb['転記']
        start, end = mapping.tenki_text_range
        for r in range(start, end):
            cell = ws_t.cell(row=r, column=2)
            if cell.value is not None and not is_formula(cell.value):
                cell.value = None
                cleared += 1

    # 申請内容シート: C列
    if '申請内容' in wb.sheetnames:
        ws_s = wb['申請内容']
        start, end = mapping.shinsei_clear_range
        preserve = set(mapping.preserve_rows)
        for row in ws_s.iter_rows(min_row=start, max_row=end):
            cell_c = row[2] if len(row) > 2 else None
            if cell_c is None or cell_c.value is None:
                continue
            if cell_c.row in preserve:
                continue
            if is_formula(cell_c.value):
                continue
            cell_c.value = None
            cleared += 1

    # 給与計算シート: マッピング対象セル
    if mapping.kyuyo_sheet_name in wb.sheetnames:
        ws_k = wb[mapping.kyuyo_sheet_name]
        for field_name, (row, col) in mapping.kyuyo.items():
            cell = ws_k.cell(row=row, column=col)
            if cell.value is not None and not is_formula(str(cell.value)):
                cell.value = None
                cleared += 1

    logger.info(f'{cleared}セル クリア完了')
    return cleared


def fill_shinsei_sheet(ws, mapping: TemplateMapping, data: ExtractionResult) -> list[str]:
    """申請内容シートにデータを転記。転記した項目のログリストを返す。"""
    writes = []
    m = mapping.shinsei
    co = data.company
    fi = data.financial
    ai = data.ai_judgment

    def write(field: str, value, label: str = ''):
        if field not in m:
            return
        if value is None:
            return
        # Excelが数式と誤認する文字列を防止
        if isinstance(value, str) and value.startswith('='):
            value = ' ' + value
        _safe_write_cell(ws, m[field], 3, value)
        writes.append(f'行{m[field]:3d} [{label or field}]: {str(value)[:50]}')

    # ── 履歴事項全部証明書 or 本人確認資料 ──
    if mapping.is_kojin:
        # 個人事業主は履歴事項がないため固定値で埋める。
        # 現住所・氏名・生年月日は担当者が本人確認資料から手書き記入する前提。
        write('headquarters_address', co.address, '現在住所')
        write('established_date', co.established_date, '事業開始年月日')
        write('capital', 0, '資本金')
        write('fin_capital', 0, '資本金(財務)')
        write('fiscal_month', '12月', '決算月')
        write('officer_count_prev', 1, '役員数(前期)')
        write('rep_name', co.representative_name, '代表者氏名')
        write('rep_kana', co.representative_kana, '代表者氏名(フリガナ)')
    else:
        write('headquarters_address', co.address, '本店所在地')
        write('established_date', co.established_date, '設立年月日')
        write('capital', co.capital, '資本金')
        write('fiscal_month', fi.fiscal_month, '決算月')

        # 代表者（法人）
        officer_count = 1 + len(co.officers)
        write('officer_count', officer_count, '役員数(申請時)')
        write('officer_count_prev', officer_count, '役員数(前期)')
        write('rep_title', co.representative_title, '代表者役職')
        write('rep_name', co.representative_name, '代表者氏名')
        write('rep_kana', co.representative_kana, '代表者フリガナ')

        # 役員 (最大10名)
        for i, officer in enumerate(co.officers[:10]):
            idx = i + 1
            write(f'officer_{idx}_title', officer.get('title'), f'役員({idx})役職')
            write(f'officer_{idx}_name', officer.get('name'), f'役員({idx})氏名')
            write(f'officer_{idx}_kana', officer.get('kana'), f'役員({idx})フリガナ')

    # ── 認定・補助金系 ──
    write('past_subsidies', 'なし', '過年度交付決定')
    write('eruboshi', '認定なし', 'えるぼし')
    write('kurumin', '認定なし', 'くるみん')

    # ── AI判断項目 ──
    write('industry_code', ai.industry_code, '業種コード')
    write('industry_text', ai.industry_text, '業種分類')
    write('business_description', ai.business_description, '事業内容')
    write('management_intent', ai.management_intent, '経営意欲')
    write('future_goals', ai.future_goals, '将来目標')
    write('security_status', ai.security_status, 'セキュリティ')
    write('business_types', ai.business_types, '行っている事業')
    write('it_investment_status', ai.it_investment_status, 'IT投資状況')
    write('it_utilization_status', ai.it_utilization_status, 'IT活用状況')
    # ── プルダウン選択型項目（番号付き選択肢）──
    write('weakness', ai.weakness, '弱み')
    write('it_investment_process', ai.it_investment_process, 'IT投資プロセス')
    write('improvement_process', ai.improvement_process, '改善プロセス')
    write('expected_effect_dept', ai.expected_effect_dept, '強化部門')
    write('expected_effect', ai.expected_effect, '期待効果')

    # ── インボイス枠特有の項目 ──
    write('it_utilization_scope', ai.it_utilization_scope, 'IT電子化範囲')
    write('invoice_related_work', ai.invoice_related_work, 'インボイス対応業務')

    # ── 最低賃金 ──
    min_wage = get_min_wage(co.address)
    if min_wage:
        write('min_wage', f'{min_wage[0]}/{min_wage[1]}円', '地域別最低賃金')
    elif ai.min_wage_text:
        write('min_wage', ai.min_wage_text, '地域別最低賃金')

    # ── 賃上げ関連（デフォルト値） ──
    write('wage_raise_declaration', '■はい\n□いいえ', '賃上げ表明')
    write('wage_raise_amount', '＋50円', '賃上げ幅')
    write('wage_raise_method',
          '□社内掲示板などへの掲載によって\n■朝礼時、会議、面談時など口頭によって\n□書面、電子メールによって\n□その他',
          '表明方法')

    # ── ツール名 ──
    if data.estimate.tool_name:
        write('tool_name', data.estimate.tool_name, 'ツール名')

    # ── 財務情報（数式参照を直接値で上書き）──
    # Phase 2: 低信頼項目は空欄+警告（write 関数を信頼度対応に拡張）
    fi_conf = getattr(fi, 'confidence', None) or {}

    def write_fin(field: str, value, label: str, conf_key: str):
        """財務系の write: confidence['xxx'].level == 'low' なら空欄+警告"""
        if field not in mapping.shinsei:
            return
        c = fi_conf.get(conf_key)
        if c and getattr(c, 'level', 'high') == 'low':
            # mapping.shinsei は dict[str, int] (行番号のみ、C列固定)。
            # mapping.kyuyo の dict[str, tuple] と混同しないこと。
            row = mapping.shinsei[field]
            writes.append(
                f'⚠ 申請内容 行{row:3d} C列 [{label}]: 低信頼のため空欄 '
                f'(理由: {getattr(c, "reason", "")})'
            )
            return
        write(field, value, label)

    write_fin('fin_revenue', fi.revenue, '売上高', 'revenue')
    write_fin('fin_gross_profit', fi.gross_profit, '粗利益', 'gross_profit')
    write_fin('fin_operating_profit', fi.operating_profit, '営業利益', 'operating_profit')
    write_fin('fin_ordinary_profit', fi.ordinary_profit, '経常利益', 'ordinary_profit')
    write_fin('fin_depreciation', fi.depreciation, '減価償却費', 'depreciation')
    # 人件費合計の信頼度: salary/misc_wages/bonus/travel いずれかが low なら全体 low
    personnel = (fi.salary or 0) + (fi.misc_wages or 0) + (fi.bonus or 0) + (fi.travel_expense or 0)
    personnel_low = any(
        fi_conf.get(k) and getattr(fi_conf[k], 'level', 'high') == 'low'
        for k in ('salary', 'misc_wages', 'bonus', 'travel_expense')
    )
    if personnel_low and 'fin_personnel' in mapping.shinsei:
        # mapping.shinsei は int（行番号のみ）、C列固定
        row = mapping.shinsei['fin_personnel']
        writes.append(
            f'⚠ 申請内容 行{row:3d} C列 [人件費]: 低信頼項目を含むため空欄'
        )
    else:
        write('fin_personnel', personnel, '人件費')
    # fin_capital は個人事業主の場合、上部で0固定済み。法人のみ co.capital を転記
    if not mapping.is_kojin:
        write('fin_capital', co.capital, '資本金(財務)')

    # ── 1人当たり給与支給総額の計画値（賃金台帳から算出時のみ）──
    # wage_plan は fill_template() から渡される場合のみ有効
    # （この関数のスコープ外で処理される）

    return writes


def fill_kyuyo_sheet(ws, mapping: TemplateMapping, data: ExtractionResult) -> list[str]:
    """給与計算シートに財務データを転記。

    Phase 2: financial.confidence を見て、低信頼項目は空欄+警告マーカー扱い。
    申請書側の処理タスクの empty_cells に追加されるよう writes に '⚠' マーカー付き行を返す。
    """
    writes = []
    fi = data.financial
    conf = getattr(fi, 'confidence', None) or {}
    m = mapping.kyuyo

    def write(field: str, value, label: str, conf_key: str = ''):
        if field not in m:
            return
        # 信頼度チェック: low なら空欄+警告マーカー（書込スキップ）
        c = conf.get(conf_key) if conf_key else None
        if c and getattr(c, 'level', 'high') == 'low':
            row, col = m[field]
            col_letter = chr(64 + col)
            writes.append(
                f'⚠ 給与計算 行{row:3d} {col_letter}列 [{label}]: 低信頼のため空欄 '
                f'(理由: {getattr(c, "reason", "")})'
            )
            return
        row, col = m[field]
        _safe_write_cell(ws, row, col, value)
        col_letter = chr(64 + col)
        writes.append(f'給与計算 行{row:3d} {col_letter}列 [{label}]: {value:,}')

    write('revenue', fi.revenue, '売上高', 'revenue')
    write('gross_profit', fi.gross_profit, '粗利益', 'gross_profit')
    write('operating_profit', fi.operating_profit, '営業利益', 'operating_profit')
    write('ordinary_profit', fi.ordinary_profit, '経常利益', 'ordinary_profit')
    write('depreciation', fi.depreciation, '減価償却費', 'depreciation')
    write('salary', fi.salary, '給料手当', 'salary')
    write('misc_wages', fi.misc_wages, '雑給', 'misc_wages')
    write('bonus', fi.bonus, '賞与手当', 'bonus')
    write('officer_comp', fi.officer_compensation, '役員報酬', 'officer_compensation')
    write('travel_expense', fi.travel_expense, '旅費交通費', 'travel_expense')

    return writes


def check_empty_cells(wb: openpyxl.Workbook) -> list[str]:
    """申請内容シートで空のままのセルを一覧表示"""
    ws = wb['申請内容']
    empty = []

    skip_keywords = {
        # 操作手順・ボタン
        '次へ', 'クリック', '宣誓', 'ファイル添付', 'アンケート',
        '計画数値入力', '書類添付', '交付申請情報', '申請要件確認',
        '事務局へ提出', '提出完了', '認証コード', '最終確認',
        '内容確認', '注意！',
        # セクションヘッダ・ラベル
        '項目', '添付資料', 'チェック項目', 'オレンジ',
        '財務情報', '経営状況', '賃金情報',
        '基本情報入力', '申請類型選択', '支援事業者入力',
        '申請要件に関する確認', '⇩必要に応じて',
        # gBizID自動取得項目（手入力不要）
        '法人番号', '事業者名', '事業者名フリガナ', '郵便番号',
        # 転記シートから手動コピーする項目
        '店舗事業所数', '事業者URL', '主な事業内容',
        '強み', '時間がかかっている', '月間何時間', 'どの機能',
        '何％', '浮いた時間', '売上目標', '属性の取引先',
        '担当部署', '担当者氏名', '担当者メールアドレス',
        '担当者電話番号', '担当者携帯番号', '代表電話番号',
        # 外部サイト確認項目
        'SECURITY ACTION照合', 'SECURITY ACTION自己宣言',
        'IT戦略ナビ', '省力化ナビ',
        # 別添資料（ファイル添付）
        '履歴事項全部証明書', '納税証明書', '決算書', 'その他資料',
        # 個人事業主の別添資料
        '身分証明書', '確定申告書', '収支内訳書', '青色申告',
        # 給与計画（賃金台帳がある場合に自動入力、なければスキップ）
        '給与支給総額', '従業員数（全期間', '賃上げを行いますか',
        '事業計画期間における', '計画数値',
        # 賃金状況関連（手動確認）
        '賃金状況', '最低賃金近傍', '最低賃金未満',
        '事業実施年度内', '交付申請の直近月',
        # 従業員がいない場合の項目
        '従業員がいない場合', '従業員を雇用する場合',
        # ここまで入力確認
        'ここまで入力',
        # プロンプト
        'プロンプト',
        # 補助事業者登録（手動確認項目）
        '補助事業者登録',
        # 代表者フリガナ・代表電話番号（転記シートから）
        '代表者氏名（フリガナ）',
    }

    # 使われていない役員枠を除外（役員(N)で値がないもの）
    def is_empty_officer_slot(label_str, row_num):
        """役員(N)のラベルだが値が空の場合True"""
        import re
        return bool(re.match(r'役員（[0-9０-９]+）', label_str))

    # 「⬇︎従業員がいない場合」セクション配下は空セルを報告しない
    # （従業員がいる場合は上部の賃上げ項目を埋めれば十分）
    in_no_employee_section = False

    for row in ws.iter_rows(min_row=35, max_row=250):
        row_num = row[0].row
        label = row[1].value if len(row) > 1 else None
        value = row[2].value if len(row) > 2 else None

        if label is not None:
            label_str_raw = str(label).strip()
            if '従業員がいない場合' in label_str_raw:
                in_no_employee_section = True
            elif '賃金状況' in label_str_raw or '最低賃金近傍' in label_str_raw:
                in_no_employee_section = False

        if label is None or value is not None:
            continue

        label_str = str(label).strip()
        if any(kw in label_str for kw in skip_keywords):
            continue

        if in_no_employee_section:
            continue

        # 使われていない役員枠はスキップ
        if is_empty_officer_slot(label_str, row_num):
            continue

        # 数値だけのラベル（テンプレート上で「数値=ラベル」になってしまっているセル）はスキップ
        # 例: 行149 のラベルが「67146344」（粗利の数値そのまま）になっているケース
        # これは check 対象としては誤検出（実際は計算式参照セルで、表示用の数値）
        ascii_only = label_str.replace(',', '').replace('-', '').replace('.', '').replace(' ', '').replace('　', '')
        if ascii_only.isdigit():
            continue

        empty.append(f'行{row_num:3d} [{label_str[:60]}]')

    return empty


def fill_template(
    template_path: Path,
    output_path: Path,
    mapping: TemplateMapping,
    hearing_data: dict,
    extraction: ExtractionResult,
    tenki_texts: dict[int, str] | None = None,
    wage_plan: dict[str, float] | None = None,
) -> list[str]:
    """
    テンプレートをコピーし、全データを転記して保存。
    空セルのリストを返す。

    wage_plan: 1人当たり給与支給総額の計画値
        {'year_0': 基準年, 'year_1': 1年目, 'year_2': 2年目, 'year_3': 3年目}
    """
    from .hearing_reader import transfer_hearing_to_tenki

    # テンプレートコピー
    shutil.copy2(template_path, output_path)
    logger.info(f'テンプレートコピー: {template_path.name} → {output_path.name}')

    wb = openpyxl.load_workbook(output_path)

    # STEP 1: サンプルデータクリア
    cleared = clear_manual_cells(wb, mapping)
    logger.info(f'STEP 1: {cleared}セル クリア')

    # STEP 2: ヒアリング → 転記
    count = 0
    if '転記' in wb.sheetnames and hearing_data:
        count = transfer_hearing_to_tenki(hearing_data, wb['転記'], mapping.hearing_to_tenki)
    logger.info(f'STEP 2: ヒアリング → {count}件転記')

    # テキスト項目（転記シートの行17-25等）
    if tenki_texts and '転記' in wb.sheetnames:
        ws_t = wb['転記']
        for row, text in tenki_texts.items():
            _safe_write_cell(ws_t, row, 2, text)

    # STEP 3: PDF → 申請内容 + 給与計算
    if '申請内容' in wb.sheetnames:
        shinsei_writes = fill_shinsei_sheet(wb['申請内容'], mapping, extraction)
        for w in shinsei_writes:
            logger.info(f'STEP 3: {w}')

    if mapping.kyuyo_sheet_name in wb.sheetnames:
        kyuyo_writes = fill_kyuyo_sheet(wb[mapping.kyuyo_sheet_name], mapping, extraction)
        for w in kyuyo_writes:
            logger.info(f'STEP 3: {w}')

    # STEP 3.5: 給与支給総額の計画値を申請内容シートに転記
    if wage_plan and '申請内容' in wb.sheetnames:
        ws_shinsei = wb['申請内容']
        m = mapping.shinsei
        # 従業員数（FTE換算）
        if 'employee_count_fte' in m and 'employee_count_fte' in wage_plan:
            fte = wage_plan['employee_count_fte']
            _safe_write_cell(ws_shinsei, m['employee_count_fte'], 3, round(fte, 1))
            logger.info(f'STEP 3.5: 行{m["employee_count_fte"]:3d} [従業員数FTE]: {fte:.1f}人')
        # 給与支給総額（基準年 + 3年計画）
        plan_fields = [
            ('wage_total_base', 'wage_total_base', '給与支給総額(基準年)'),
            ('wage_total_y1', 'wage_total_y1', '給与支給総額(1年目)'),
            ('wage_total_y2', 'wage_total_y2', '給与支給総額(2年目)'),
            ('wage_total_y3', 'wage_total_y3', '給与支給総額(3年目)'),
        ]
        for field, plan_key, label in plan_fields:
            if field in m and plan_key in wage_plan:
                val = round(wage_plan[plan_key])
                _safe_write_cell(ws_shinsei, m[field], 3, val)
                logger.info(f'STEP 3.5: 行{m[field]:3d} [{label}]: {val:,}円')

    # STEP 3.6: 生産性指標シートの総労働時間(B40)を賃金台帳実績で上書き（通常枠）
    # 通常枠の「事業者あたりの総労働時間」は B38*B39（人数×平均時間）の式だが、
    # 賃金台帳の実績がある場合は基準年だけ直接値で上書きする。C40-E40(計画年次)は
    # 既存の `=C38*C39` 等の式を保持する。
    if (
        wage_plan
        and 'total_annual_hours' in wage_plan
        and '生産性指標給与支給総額計算' in wb.sheetnames
    ):
        ws_prod = wb['生産性指標給与支給総額計算']
        hours = wage_plan['total_annual_hours']
        _safe_write_cell(ws_prod, 40, 2, hours)
        logger.info(f'STEP 3.6: 生産性指標 B40 [総労働時間]: {hours:,.1f}時間')

    # STEP 4: 空セル確認
    empty = check_empty_cells(wb)
    logger.info(f'STEP 4: 空セル {len(empty)}件')

    # 保存
    wb.save(output_path)
    wb.close()
    logger.info(f'保存完了: {output_path}')

    return empty


# ============================================================
# データソースシート（申請書作成タスク用）
# ============================================================
# 申請書作成タスクの出力 Excel に「データソース」シートを追加し、
# 各抽出値の出所（ファイル名 / PDFページ番号 / 抽出経路 / 信頼度）を
# 一覧化する。1次振り返り MTG（2026-05-14）の要望
# 「抽出した数値の根拠となるデータソースを記録し、出力資料に含める」
# への対応。
#
# 設計方針:
#   - 既存テンプレートには触らず、新規シートを末尾に追加するだけ
#   - 値の取得元別にセクション分け（履歴事項/PL/賃金台帳/AI判断…）
#   - PDF はテキスト層が取れればページ番号、取れなければ「ファイル全体」
#   - 信頼度は high/medium/low をそのまま表示（financial.confidence 由来）

_DATA_SOURCE_SHEET_NAME = 'データソース'

# (label, financial 属性名, confidence キー) の対応表
_PL_FIELDS_FOR_SOURCE = [
    ('売上高',         'revenue',              'revenue'),
    ('売上原価',       'cost_of_sales',        'cost_of_sales'),
    ('売上総利益',     'gross_profit',         'gross_profit'),
    ('営業利益',       'operating_profit',     'operating_profit'),
    ('経常利益',       'ordinary_profit',      'ordinary_profit'),
    ('当期純利益',     'net_profit',           'net_profit'),
    ('給料手当',       'salary',               'salary'),
    ('雑給',           'misc_wages',           'misc_wages'),
    ('賞与',           'bonus',                'bonus'),
    ('役員報酬',       'officer_compensation', 'officer_compensation'),
    ('法定福利費',     'legal_welfare',        'legal_welfare'),
    ('福利厚生費',     'welfare',              'welfare'),
    ('減価償却費',     'depreciation',         'depreciation'),
]


def _fmt_pages(pages: list[int] | None) -> str:
    """[1,3,5] → 'p.1,3,5'  / 空 → '（PDF全体）'"""
    if not pages:
        return '（PDF全体）'
    return 'p.' + ','.join(str(p) for p in pages)


def _fmt_pl_source(
    pages_entry: dict | None,
    pl_path: 'Path | None',
    cost_report_path: 'Path | None',
) -> tuple[str, str]:
    """PL値の出所表示を組み立てる（製造原価報告書とのマージケースに対応）。

    pages_entry: _compute_pl_value_pages の戻り値の1エントリ
                 {'pl': [pages], 'cost': [pages]} 形式 / None
    Returns:
        (出所ファイル名表示, ページ番号表示)
        - PL のみで見つかった: ('決算書.pdf', 'p.3,5')
        - 製造原価のみで見つかった: ('原価.pdf', 'p.2')
        - 両方で見つかった: ('決算書.pdf + 原価.pdf', 'PL p.3 / 原価 p.2')
        - 見つからなかった: (pl_path.name or '-', '（PDF全体）')
    """
    pl_name = pl_path.name if pl_path else '-'
    cost_name = cost_report_path.name if cost_report_path else ''
    if not pages_entry:
        return pl_name, '（PDF全体）'
    pl_pages = pages_entry.get('pl') or []
    cost_pages = pages_entry.get('cost') or []
    if pl_pages and cost_pages:
        return (
            f'{pl_name} + {cost_name}' if cost_name else pl_name,
            f'PL p.{",".join(map(str, pl_pages))} / 原価 p.{",".join(map(str, cost_pages))}',
        )
    if pl_pages:
        return pl_name, 'p.' + ','.join(map(str, pl_pages))
    if cost_pages:
        return cost_name or pl_name, 'p.' + ','.join(map(str, cost_pages))
    # どちらも空（テキスト層無し or 値見つからず）
    return pl_name, '（PDF全体）'


def _confidence_label(conf_obj) -> str:
    """FieldConfidence → '高'/'中'/'低'/'-' に変換"""
    if conf_obj is None:
        return '-'
    level = getattr(conf_obj, 'level', 'high')
    return {'high': '高', 'medium': '中', 'low': '低'}.get(level, level)


def add_data_source_sheet(
    output_path: Path,
    *,
    hearing_path: Path | None,
    registry_path: Path | None,
    pl_path: Path | None,
    cost_report_path: Path | None,
    tax_path: Path | None,
    estimate_path: Path | None,
    wage_ledger_paths: list[Path],
    extraction,  # ExtractionResult
    pl_value_pages: dict[str, list[int]],
    wage_plan: dict | None,
    wage_extraction_method: str,
) -> None:
    """申請書出力 Excel に「データソース」シートを追加して保存し直す。

    既存シートには一切手を触れず、末尾に新規シートを追加するだけ。
    シートが既に存在する場合は中身を作り直す（再実行時の上書き対応）。
    """
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    if not output_path.exists():
        logger.warning(f'add_data_source_sheet: 出力ファイル不在 {output_path}')
        return

    try:
        wb = openpyxl.load_workbook(output_path)
        # 既存シートがあれば削除して作り直し
        if _DATA_SOURCE_SHEET_NAME in wb.sheetnames:
            del wb[_DATA_SOURCE_SHEET_NAME]
        ws = wb.create_sheet(_DATA_SOURCE_SHEET_NAME)

        # スタイル
        TITLE = Font(name='游ゴシック', size=14, bold=True)
        SECTION = Font(name='游ゴシック', size=11, bold=True, color='FFFFFF')
        HEADER = Font(name='游ゴシック', size=10, bold=True)
        NORMAL = Font(name='游ゴシック', size=10)
        SMALL = Font(name='游ゴシック', size=9, color='666666')
        SECTION_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        HEADER_FILL = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ALT_FILL = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        THIN = Side(style='thin', color='BFBFBF')
        BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

        # タイトル
        ws.cell(1, 1, '申請書 — データソース一覧').font = TITLE
        ws.cell(2, 1, (
            '各抽出値が「どのファイル／どのページ」から来ているかの一覧。'
            '提出前の人間チェック時にこのシートを見て原本と突合してください。'
            'PDFのページ番号は機械的に逆引きしているため、複数ページに同値があれば全て列挙されます。'
        )).font = SMALL
        ws.cell(3, 1, (
            '※ ページ番号「（PDF全体）」= PDFのテキスト層が取れなかった（画像PDFなど）ため'
            '機械的にページ特定できなかったケース。原本を目視で確認してください。'
        )).font = SMALL

        # ヘッダー行（行5）
        headers = ['カテゴリ', '項目', '抽出値', '出所ファイル', 'ページ/位置', '抽出経路', '信頼度']
        for col, h in enumerate(headers, 1):
            c = ws.cell(5, col, h)
            c.font = HEADER
            c.fill = HEADER_FILL
            c.border = BORDER
            c.alignment = Alignment(horizontal='center', vertical='center')

        row = 6

        def _section(title: str):
            """セクション見出し行を1行書く"""
            nonlocal row
            c = ws.cell(row, 1, title)
            c.font = SECTION
            c.fill = SECTION_FILL
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
            row += 1

        def _row(category: str, label: str, value, source_file: str,
                 pages: str, method: str, confidence: str = '-'):
            """データ行を1行書く"""
            nonlocal row
            fill = ALT_FILL if row % 2 == 0 else None
            cells = [category, label, value, source_file, pages, method, confidence]
            for col, v in enumerate(cells, 1):
                c = ws.cell(row, col, v)
                c.font = NORMAL
                c.border = BORDER
                if fill:
                    c.fill = fill
                # 値は右寄せ（数値が多いため）、それ以外は左寄せ
                if col == 3 and isinstance(v, (int, float)):
                    c.alignment = Alignment(horizontal='right', vertical='top')
                    c.number_format = '#,##0'
                else:
                    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            row += 1

        # ── ヒアリングシート ──
        _section('① ヒアリングシート（Excel直接読取）')
        if hearing_path:
            _row('ヒアリング', '（全項目）', '転記シートに反映', hearing_path.name,
                 '（Excel全体）', '直接読取（API不要）', '高')
        else:
            _row('ヒアリング', '-', '（未提供）', '-', '-', '-', '-')

        # ── 履歴事項全部証明書 ──
        # 注意: ExtractionResult はデフォルトで CompanyInfo() を持つので、
        # `extraction.company` のtruthinessだけ見ると API 残高切れで実抽出
        # スキップされたケースでも True になる。実値（name 等）が入って
        # いるかで判定する。
        _section('② 履歴事項全部証明書（AI抽出 — Claude Sonnet 4.6）')
        _company = getattr(extraction, 'company', None)
        _company_extracted = bool(_company and (getattr(_company, 'name', '') or
                                                getattr(_company, 'representative_name', '')))
        if registry_path and _company_extracted:
            co = extraction.company
            _row('履歴事項', '会社名',           getattr(co, 'name', '') or '-',
                 registry_path.name, '（PDF全体）', 'AI抽出', '高')
            _row('履歴事項', '代表者氏名',       getattr(co, 'representative_name', '') or '-',
                 registry_path.name, '（PDF全体）', 'AI抽出', '高')
            _row('履歴事項', '代表者役職',       getattr(co, 'representative_title', '') or '-',
                 registry_path.name, '（PDF全体）', 'AI抽出', '高')
            _row('履歴事項', '本店所在地',       getattr(co, 'address', '') or '-',
                 registry_path.name, '（PDF全体）', 'AI抽出', '高')
            _row('履歴事項', '設立年月日',       getattr(co, 'established_date', '') or '-',
                 registry_path.name, '（PDF全体）', 'AI抽出', '高')
            _row('履歴事項', '資本金',           getattr(co, 'capital', 0) or 0,
                 registry_path.name, '（PDF全体）', 'AI抽出', '高')
            officers = getattr(co, 'officers', None) or []
            _row('履歴事項', '役員数（代表者除く）', len(officers),
                 registry_path.name, '（PDF全体）', 'AI抽出',
                 '高' if officers else '低')
        elif registry_path:
            _row('履歴事項', '-', '（API残高切れ等で AI 抽出スキップ）',
                 registry_path.name, '-', '-', '低')
        else:
            _row('履歴事項', '-', '（未提供）', '-', '-', '-', '-')

        # ── 損益計算書 ──
        # 同様にデフォルト FinancialData() のtruthiness判定を避け、
        # 実数値（revenue 等の主要項目）が入っているかで判定する。
        _section('③ 損益計算書（AI抽出 — 値→ページ逆引き）')
        _fin = getattr(extraction, 'financial', None)
        _fin_extracted = bool(_fin and (
            getattr(_fin, 'revenue', 0) or getattr(_fin, 'cost_of_sales', 0) or
            getattr(_fin, 'gross_profit', 0) or getattr(_fin, 'operating_profit', 0) or
            getattr(_fin, 'salary', 0) or getattr(_fin, 'fiscal_year_end', '')
        ))
        if pl_path and _fin_extracted:
            fin = extraction.financial
            confidence_map = getattr(fin, 'confidence', {}) or {}
            # 事業年度
            _row('決算書', '事業年度開始日', getattr(fin, 'fiscal_year_start', '') or '-',
                 pl_path.name, '（PDF全体）', 'AI抽出',
                 _confidence_label(confidence_map.get('fiscal_year_start')))
            _row('決算書', '事業年度終了日', getattr(fin, 'fiscal_year_end', '') or '-',
                 pl_path.name, '（PDF全体）', 'AI抽出',
                 _confidence_label(confidence_map.get('fiscal_year_end')))
            # PL各項目（値→ページ逆引きで PL.pdf / 製造原価.pdf 両方を探索）
            for label, attr, conf_key in _PL_FIELDS_FOR_SOURCE:
                val = getattr(fin, attr, 0) or 0
                if val == 0:
                    # 0円科目は決算書に該当行が無いことが多いのでスキップ
                    continue
                src_file, pages_str = _fmt_pl_source(
                    pl_value_pages.get(attr), pl_path, cost_report_path,
                )
                _row('決算書', label, val, src_file, pages_str, 'AI抽出',
                     _confidence_label(confidence_map.get(conf_key)))
            # 製造原価報告書がマージされている場合の注記（値はPLとマージ済）
            if cost_report_path:
                _row('決算書', '製造原価報告書', '（人件費等をPLとマージ済）',
                     cost_report_path.name, '（PDF全体）', 'AI抽出（マージ）', '高')
        elif pl_path:
            _row('決算書', '-', '（API残高切れ等で AI 抽出スキップ）',
                 pl_path.name, '-', '-', '低')
        else:
            _row('決算書', '-', '（未提供）', '-', '-', '-', '-')

        # ── 納税証明書 ──
        # 同様: TaxCertificate() のデフォルト値は tax_type='' / tax_amount=0
        _section('④ 納税証明書（AI抽出）')
        _tx = getattr(extraction, 'tax', None)
        _tx_extracted = bool(_tx and (getattr(_tx, 'tax_type', '') or
                                      getattr(_tx, 'tax_amount', 0)))
        if tax_path and _tx_extracted:
            tx = extraction.tax
            _row('納税証明', '税目', getattr(tx, 'tax_type', '') or '-',
                 tax_path.name, '（PDF全体）', 'AI抽出', '高')
            _row('納税証明', '税額', getattr(tx, 'tax_amount', 0) or 0,
                 tax_path.name, '（PDF全体）', 'AI抽出', '高')
        elif tax_path:
            _row('納税証明', '-', '（API残高切れ等で AI 抽出スキップ）',
                 tax_path.name, '-', '-', '低')
        else:
            _row('納税証明', '-', '（未提供）', '-', '-', '-', '-')

        # ── 見積書 ──
        # EstimateData() のデフォルトは tool_name=''。Excel直読は API 不要なので
        # 残高切れの影響を受けない（Phase 1 で処理済み）。
        _section('⑤ 見積書')
        _est = getattr(extraction, 'estimate', None)
        _est_extracted = bool(_est and getattr(_est, 'tool_name', ''))
        if estimate_path and _est_extracted:
            est = extraction.estimate
            method = 'AI抽出' if estimate_path.suffix.lower() == '.pdf' else '直接読取（API不要）'
            _row('見積書', 'ツール名', getattr(est, 'tool_name', '') or '-',
                 estimate_path.name, '（全体）', method, '高')
        elif estimate_path:
            _row('見積書', '-', '（抽出失敗 — ファイル名や記載項目を確認）',
                 estimate_path.name, '-', '-', '低')
        else:
            _row('見積書', '-', '（未提供）', '-', '-', '-', '-')

        # ── 賃金台帳 ──
        _section('⑥ 賃金台帳（決定論パーサー直読 — AI不使用）')
        if wage_ledger_paths and wage_plan:
            ledger_summary = (
                wage_ledger_paths[0].name if len(wage_ledger_paths) == 1
                else f'{wage_ledger_paths[0].name} 他 {len(wage_ledger_paths) - 1} 件'
            )
            _row('賃金台帳', '一人当たり給与支給総額(基準年)',
                 int(round(wage_plan.get('wage_total_base', 0) /
                           max(wage_plan.get('employee_count_fte', 1), 1))),
                 ledger_summary, '（Excel全体）', wage_extraction_method, '高')
            _row('賃金台帳', '給与支給総額(基準年)', int(round(wage_plan.get('wage_total_base', 0))),
                 ledger_summary, '（Excel全体）', wage_extraction_method, '高')
            _row('賃金台帳', '給与支給総額(1年目計画)', int(round(wage_plan.get('wage_total_y1', 0))),
                 '（基準年×1.03 自動計算）', '-', '機械計算', '高')
            _row('賃金台帳', '給与支給総額(2年目計画)', int(round(wage_plan.get('wage_total_y2', 0))),
                 '（基準年×1.03² 自動計算）', '-', '機械計算', '高')
            _row('賃金台帳', '給与支給総額(3年目計画)', int(round(wage_plan.get('wage_total_y3', 0))),
                 '（基準年×1.03³ 自動計算）', '-', '機械計算', '高')
            _row('賃金台帳', '従業員数（FTE換算）',
                 round(wage_plan.get('employee_count_fte', 0), 1),
                 ledger_summary, '（Excel全体）', wage_extraction_method, '高')
            if 'total_annual_hours' in wage_plan:
                _row('賃金台帳', '年間総労働時間（役員除く）',
                     round(wage_plan['total_annual_hours'], 1),
                     ledger_summary, '（Excel全体）', wage_extraction_method, '高')
        elif wage_ledger_paths:
            _row('賃金台帳', '-', '（読取失敗 — 「賃金台帳の作成」タスクで再整形を推奨）',
                 wage_ledger_paths[0].name, '-', wage_extraction_method, '低')
        else:
            _row('賃金台帳', '-', '（未提供）', '-', '-', '-', '-')

        # ── AI判断 ──
        # AIJudgment() のデフォルトは industry_code='' / business_description=''
        _section('⑦ AI判断（履歴事項 + ヒアリング + 見積書から Claude が生成）')
        _aj = getattr(extraction, 'ai_judgment', None)
        _aj_extracted = bool(_aj and (getattr(_aj, 'industry_code', '') or
                                       getattr(_aj, 'business_description', '')))
        if _aj_extracted:
            aj = extraction.ai_judgment
            biz = getattr(aj, 'business_description', '') or ''
            _row('AI判断', '業種コード（4桁）', getattr(aj, 'industry_code', '') or '-',
                 'AI生成（履歴事項+ヒアリング+見積参照）', '-', 'AI生成', '中')
            _row('AI判断', f'事業内容（{len(biz)}文字）',
                 biz if len(biz) <= 80 else biz[:80] + '…',
                 'AI生成（履歴事項+ヒアリング+見積参照）', '-', 'AI生成', '中')
        else:
            _row('AI判断', '-', '（API残高切れ等で生成スキップ／失敗）',
                 '-', '-', '-', '低')

        # 列幅
        widths = {1: 12, 2: 26, 3: 28, 4: 32, 5: 14, 6: 22, 7: 8}
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = w

        # 行高（タイトル・説明）
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 30
        ws.row_dimensions[3].height = 22

        wb.save(output_path)
        wb.close()
        logger.info(f'データソースシート追加完了: {output_path.name}')
    except Exception as e:
        # データソースシートは補助情報なので、書込失敗しても申請書本体は維持
        logger.warning(f'データソースシート追加に失敗（申請書本体は保存済み）: {e}', exc_info=True)


# 「一人当たり給与支給総額」タスク向けに、給与支給総額計算シートから
# 決算書PDF由来のセクション（【2026テンプレート転記用】＋見出し注記＋給料手当/雑給/賞与手当/
# 売上高/粗利益/営業利益/経常利益/減価償却費）を機械的に削除する。
# 賃金台帳のみを参照源としてシートを完成させる仕様。
_PL_SECTION_HEADER_MARKER = '【2026テンプレート転記用】'
_PL_SECTION_END_LABELS = (
    '給料手当', '雑給', '賞与手当',
    '売上高', '粗利益', '営業利益', '経常利益', '減価償却費',
)


def strip_pl_section_from_wage_sheet(output_path: Path) -> int:
    """給与支給総額計算シートから決算書PDF由来セクションを削除する。

    削除範囲: 「【2026テンプレート転記用】」見出し行から、その下の
    決算書由来項目（給料手当〜減価償却費）の最後の行まで。
    途中の空行・注記行（『※下記すべてAI抽出値…』等）も連帯して削除。

    Returns:
        削除した行数（0 ならセクションが見つからなかった）
    """
    if not output_path.exists():
        logger.warning(f'strip_pl_section: 出力ファイル不在 {output_path}')
        return 0

    wb = openpyxl.load_workbook(output_path)
    try:
        target_sheet = '給与支給総額計算'
        if target_sheet not in wb.sheetnames:
            logger.warning(
                f'strip_pl_section: シート「{target_sheet}」が見つからずスキップ'
            )
            return 0
        ws = wb[target_sheet]

        # 見出し行を B列で検索
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
            cell_b = row[1] if len(row) > 1 else None
            if cell_b is not None and isinstance(cell_b.value, str) \
                    and _PL_SECTION_HEADER_MARKER in cell_b.value:
                header_row = cell_b.row
                break

        if header_row is None:
            logger.info(
                f'strip_pl_section: セクション見出し未検出のためスキップ '
                f'(既に削除済みの可能性)'
            )
            return 0

        # 見出し行以降で、決算書由来ラベル群の最終出現行を探す
        last_target_row = header_row
        for row in ws.iter_rows(
            min_row=header_row, max_row=ws.max_row, max_col=2,
        ):
            cell_b = row[1] if len(row) > 1 else None
            if cell_b is None or not isinstance(cell_b.value, str):
                continue
            label = cell_b.value.strip()
            if any(label == end_label for end_label in _PL_SECTION_END_LABELS):
                last_target_row = max(last_target_row, cell_b.row)

        amount = last_target_row - header_row + 1
        ws.delete_rows(header_row, amount)
        logger.info(
            f'strip_pl_section: 行{header_row}〜{last_target_row} ({amount}行) を削除'
        )

        wb.save(output_path)
        return amount
    finally:
        wb.close()
