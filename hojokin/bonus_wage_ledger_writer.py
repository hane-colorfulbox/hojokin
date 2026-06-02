# -*- coding: utf-8 -*-
"""加点判定用賃金台帳テンプレートへの書き出し。

read_bonus_wage_ledger（wage_reader.py）と対になる writer。レイアウト定数は
wage_reader.BWL_* を単一の真実として共有する（列がズレると判定が壊れるため）。
入力は BonusWageEmployee（(年,月)→基本給 を保持）。月列は暦月固定。
"""
from __future__ import annotations

import datetime as _dt
import logging
from pathlib import Path

import openpyxl

from .wage_reader import (
    BonusWageEmployee,
    BONUS1_WINDOW,
    BWL_APPYM_CELL,
    BWL_COL_EMPTYPE,
    BWL_COL_HOURS,
    BWL_COL_LATEST,
    BWL_COL_NAME,
    BWL_COL_NO,
    BWL_COL_WINDOW_START,
    BWL_DATA_START_ROW,
    BWL_PREF_CELL,
    BWL_SHEET_NAME,
    prev_month,
)

logger = logging.getLogger(__name__)


def write_bonus_wage_ledger(
    employees: list[BonusWageEmployee],
    template_path: Path,
    output_path: Path,
    *,
    prefecture: str = '',
    application_ym: tuple[int, int] | None = None,
) -> Path:
    """BonusWageEmployee リストを加点判定用賃金台帳テンプレに書き出す。

    C2=都道府県、C3=交付申請月（日付）。各従業員行に No/氏名/雇用形態/月間所定労働時間/
    令和6年10月〜令和7年9月の基本給（F〜Q）/交付申請直近月の基本給（R）を書く。
    """
    wb = openpyxl.load_workbook(str(template_path))
    ws = wb[BWL_SHEET_NAME] if BWL_SHEET_NAME in wb.sheetnames else wb[wb.sheetnames[0]]

    if prefecture:
        ws.cell(*BWL_PREF_CELL, value=prefecture)
    if application_ym:
        cell = ws.cell(BWL_APPYM_CELL[0], BWL_APPYM_CELL[1],
                       _dt.datetime(application_ym[0], application_ym[1], 1))
        cell.number_format = 'yyyy/mm'
    latest_ym = prev_month(application_ym) if application_ym else None

    for i, emp in enumerate(employees):
        row = BWL_DATA_START_ROW + i
        ws.cell(row, BWL_COL_NO, emp.no if emp.no else i + 1)
        ws.cell(row, BWL_COL_NAME, emp.name)
        ws.cell(row, BWL_COL_EMPTYPE, emp.employment_type or '')
        if emp.scheduled_hours is not None:
            ws.cell(row, BWL_COL_HOURS, round(emp.scheduled_hours, 1))
        for j, ym in enumerate(BONUS1_WINDOW):
            base = emp.monthly_base.get(ym)
            if base is not None:
                ws.cell(row, BWL_COL_WINDOW_START + j, round(base))
        if latest_ym is not None:
            base = emp.monthly_base.get(latest_ym)
            if base is not None:
                ws.cell(row, BWL_COL_LATEST, round(base))

    wb.save(str(output_path))
    wb.close()
    logger.info(f'加点判定用賃金台帳保存: {output_path}（{len(employees)}名）')
    return output_path
