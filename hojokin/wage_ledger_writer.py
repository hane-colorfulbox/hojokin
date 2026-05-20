# -*- coding: utf-8 -*-
"""
賃金台帳テンプレート書き込みモジュール

入力 PDF/Excel/CSV から抽出した WageEmployee リストを、
ツール側が再読込できる Excel テンプレート（ツール/賃金台帳テンプレート.xlsx 規格）
に転記する。

対象テンプレート: `ツール/賃金台帳テンプレート.xlsx`
  - シート「従業員別明細」
  - ヘッダー行 5: B5=No, C5=氏名, D5=雇用形態,
                  E5=月間平均時間, F5=時給,
                  G5〜R5=1月〜12月, S5=年間通勤手当
  - データ行: B6 以降（テンプレ既定 10 行、超える分は追記）

設計方針:
  - monthly_wages[i] は **カレンダー月固定**（Index 0=1月, 11=12月）。
    決算月情報を使った並び替えはここでは行わない（ツール側 `_fiscal_month_order`
    が読込時に事業年度順に並び替える既存設計を踏襲）。
  - 雇用形態の正規化はテンプレ種別（法人/個人事業主）で分岐。
  - 抽出経路ラベル（Document AI / Sonnet 画像 等）と手書きPDF 検出結果は
    出力 xlsx の B1〜B3 に明示し、目視チェックを促す。
"""
from __future__ import annotations

import logging
import re
import shutil
import unicodedata
from collections.abc import Iterable
from dataclasses import dataclass
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .wage_reader import MONTH_NAMES, WageEmployee

logger = logging.getLogger(__name__)


# テンプレートのセル位置（マニュアル §1 と一致）
HEADER_ROW = 5
FIRST_DATA_ROW = 6
DEFAULT_LAST_DATA_ROW = 15  # テンプレ既定 10 行（B6〜B15）
SHEET_NAME = '従業員別明細'
MEMO_SHEET_NAME = '変換メモ'  # 新規追加するメタ情報シート

COL_NO = 2          # B
COL_NAME = 3        # C
COL_EMP_TYPE = 4    # D
COL_AVG_HOURS = 5   # E
COL_HOURLY = 6      # F
COL_MONTH_START = 7  # G = 1月
COL_MONTH_END = 18   # R = 12月
COL_TRANSPORT = 19   # S

# 警告メタ書込先（B1〜B3）。B4 はテンプレ既定の注釈を残す
META_ROW_TITLE = 1
META_ROW_EXTRACTION = 2
META_ROW_VERIFICATION = 3

# 手書き判定の閾値
HANDWRITTEN_TEXT_PER_PAGE_THRESHOLD = 100  # 文字/ページ
HANDWRITTEN_MIN_PAGES_FOR_JUDGEMENT = 1     # 最低この枚数で判定

# ルールベース役員疑い検出の閾値（B 改善）
# 月別給与の変動係数 (std/mean) がこの値を下回り、かつ
# 月平均が SUSPECT_HIGH_WAGE_YEN 以上 もしくは 全社員平均の SUSPECT_HIGH_WAGE_MULTIPLIER 倍以上
# のとき「役員疑い」フラグを立てる（雇用形態の上書きはしない）。
SUSPECT_CV_THRESHOLD = 0.05                 # 変動係数 < 0.05（月額がほぼ一定）
SUSPECT_HIGH_WAGE_YEN = 700_000             # 月平均が 70万円以上
SUSPECT_HIGH_WAGE_MULTIPLIER = 2.0          # 全社員平均の 2倍以上
SUSPECT_MIN_FULL_MONTHS = 6                 # 判定に最低必要な「給与あり」月数


# ─────────────────────────────────────────────────────────────
# 雇用形態の正規化
# ─────────────────────────────────────────────────────────────

# 既に正規化済みのターゲット表記
_NORMALIZED_TARGETS = {'役員', '正社員', 'パート・アルバイト'}

# 法人で「役員」とすべき肩書（部分一致で判定）
_OFFICER_KEYWORDS_CORP = (
    '代表取締役', '取締役', '監査役', '監事',
    '理事長', '理事',
    '執行役員', '会長', '社長', '副社長', '専務', '常務',
)

# 個人事業主の「事業主本人」表記。代表者本人は R216 算定対象外（給与でなく利益帰属）。
# 賃金台帳に通常は載らないが、誤って載った場合のセーフティネットとして役員扱い（R216除外）にする。
_KOJIN_OWNER_KEYWORDS = ('事業主', '個人事業主')

# 個人事業主の「専従者」表記（青色事業専従者を含む）。R216 算定対象内。
# 「専従者」は『事業主』『事業専従者』を含むため、_KOJIN_OWNER_KEYWORDS より先に判定する。
_KOJIN_SENJUSHA_KEYWORDS = ('専従者',)

# パート系キーワード
_PART_KEYWORDS = ('パート', 'アルバイト', '非常勤', 'ＰＴ', '時給')


def _norm_text(s: str) -> str:
    """NFKC 正規化 + 空白除去（小文字化はしない、漢字氏名想定）"""
    if not s:
        return ''
    n = unicodedata.normalize('NFKC', str(s))
    return re.sub(r'[\s　]+', '', n)


def normalize_employment_type(raw: str, is_kojin: bool = False) -> str:
    """雇用形態を「役員 / 正社員 / パート・アルバイト」の3値に正規化する。

    is_kojin=True（個人事業主テンプレ）時:
      - 「事業主」「専従者」は『正社員』扱い（公募要領上、個人事業主に役員概念なし）
      - その他は法人と同じルール
    """
    raw_clean = (raw or '').strip()
    if not raw_clean:
        # provenance を残す形で空判定を上位に伝えたいが、ここは「正社員」に丸める。
        # 「(推定)」付きの値は呼び出し元で温存される設計（後段 wage_validator が拾う）。
        return '正社員'

    # 既に「(推定)」等の provenance 付きならそのまま保持
    if '(推定)' in raw_clean or '（推定）' in raw_clean:
        return raw_clean

    nval = _norm_text(raw_clean)

    # 既に正規化済みターゲットならそのまま
    if raw_clean in _NORMALIZED_TARGETS:
        return raw_clean
    # 「役員」を含む（プロンプトが正規化を試みた痕跡）→ 役員確定
    if '役員' in nval:
        return '役員'

    # 個人事業主モード
    if is_kojin:
        # 専従者（青色事業専従者を含む）は R216 算定対象内 → 正社員扱い
        # 注意: 「事業専従者」も「事業主」キーワードを含むため、専従者判定を先に行う
        if any(k in nval for k in _KOJIN_SENJUSHA_KEYWORDS):
            return '正社員'
        # 事業主本人は R216 算定対象外 → 役員扱い（セーフティネット）
        if any(k in nval for k in _KOJIN_OWNER_KEYWORDS):
            return '役員'

    # 法人: 役員相当キーワード
    if not is_kojin and any(k in nval for k in _OFFICER_KEYWORDS_CORP):
        return '役員'

    # パート系
    if any(k in nval for k in _PART_KEYWORDS):
        return 'パート・アルバイト'

    # それ以外（職位名「現場代理人」「主任」「係長」「営業」「契約社員」等）
    # → 雇用形態列の3値表現としては「正社員」に丸める
    return '正社員'


# ─────────────────────────────────────────────────────────────
# 手書き PDF 検出
# ─────────────────────────────────────────────────────────────

def detect_handwritten_pdf(pdf_path: Path) -> tuple[bool, str]:
    """PDF が「手書きまたは品質の低い写真スキャン」か推定する。

    判定:
      - PyMuPDF でテキスト層を抽出 → 文字数/ページが
        HANDWRITTEN_TEXT_PER_PAGE_THRESHOLD 未満で
      - かつ Document AI を未実行（呼出側で別途検証）の場合に True

    本関数はテキスト層のみで「テキストPDFか印字スキャンPDFか」を判定する。
    OCR 後の精度は extract_wage_ledger 側のログ・実経路で別途追跡される。

    戻り値: (is_likely_handwritten, reason)
    """
    try:
        import fitz  # type: ignore
    except ImportError:
        logger.warning('PyMuPDF 未導入 — 手書き検出は無効化、False を返します')
        return False, 'PyMuPDF未導入'

    try:
        doc = fitz.open(str(pdf_path))
    except Exception as e:
        logger.warning(f'PDFオープン失敗: {pdf_path.name} ({e}) — 手書き判定スキップ')
        return False, f'PDFオープン失敗: {e}'

    try:
        if len(doc) < HANDWRITTEN_MIN_PAGES_FOR_JUDGEMENT:
            return False, 'ページ数不足'

        total_chars = 0
        for page in doc:
            text = page.get_text('text') or ''
            total_chars += len(text.strip())
        avg_chars = total_chars / len(doc)
        is_likely = avg_chars < HANDWRITTEN_TEXT_PER_PAGE_THRESHOLD
        reason = (
            f'テキスト層 平均{avg_chars:.0f}文字/ページ '
            f'(閾値{HANDWRITTEN_TEXT_PER_PAGE_THRESHOLD})'
        )
        return is_likely, reason
    finally:
        doc.close()


# ─────────────────────────────────────────────────────────────
# 名寄せ・役員照合（A 改善）
# ─────────────────────────────────────────────────────────────

# 名寄せキーは wage_reader._normalize_name_key と同等の処理を使う
# （NFKC + 空白除去 + OCR異体字置換）。重複定義を避けるため遅延 import で参照する。

def _name_key(name: str) -> str:
    """賃金台帳従業員名と履歴事項の役員氏名を突合するための正規化キー。

    wage_reader._normalize_name_key を流用（NFKC + 空白除去 + 異体字置換）。
    """
    from .wage_reader import _normalize_name_key
    return _normalize_name_key(name)


def match_officer_names_to_employees(
    employees: list[WageEmployee],
    officer_names: list[str],
) -> list[tuple[int, str, str]]:
    """賃金台帳の従業員と履歴事項の役員氏名を完全照合する（NFKC 名寄せキー基準）。

    Args:
        employees: 賃金台帳から抽出された WageEmployee リスト
        officer_names: 履歴事項PDFから取れた役員氏名のリスト
            （代表者 + 一般役員、重複排除済み）

    Returns:
        マッチした従業員のリスト。要素は (employee_index, 元の雇用形態, 役員氏名)。
        index は employees 内の位置。

    名寄せ方針:
        - 完全一致のみ（NFKC + 空白除去後）。部分一致は誤検出を生むため避ける。
        - 同姓同名は同一人物として扱う（中小企業規模で実用上問題なし）。
        - 元の雇用形態は呼出側で「変換メモ」シートに保存して根拠を残す。
    """
    if not employees or not officer_names:
        return []

    # 役員名の正規化キーセット（空キーは除外）
    officer_keys = {_name_key(n) for n in officer_names if n and n.strip()}
    officer_keys.discard('')
    if not officer_keys:
        return []

    matches: list[tuple[int, str, str]] = []
    for idx, emp in enumerate(employees):
        if not emp.name:
            continue
        key = _name_key(emp.name)
        if not key:
            continue
        if key in officer_keys:
            matches.append((idx, emp.employment_type or '', emp.name))
    return matches


# ─────────────────────────────────────────────────────────────
# ルールベース役員疑い検出（B 改善）
# ─────────────────────────────────────────────────────────────

@dataclass
class OfficerSuspect:
    """ルールベースで「役員疑い」と判定された従業員の情報。

    雇用形態は上書きしない。「変換メモ」シートで人間が確認する用。
    """
    name: str
    employment_type: str          # そのままの表記（上書きしないため元値）
    monthly_avg: float            # 月別給与の平均
    monthly_std: float            # 月別給与の標準偏差
    coefficient_of_variation: float  # 変動係数 = std / mean
    months_with_data: int
    reason: str                    # 判定理由（高額 or 倍率超え）


def _calc_variation(values: list[float]) -> tuple[float, float, float]:
    """月別給与の平均・標準偏差・変動係数を返す。

    値が空 or 1件以下なら (0, 0, 1.0) を返す（変動係数1.0で「疑いなし」扱い）。
    """
    n = len(values)
    if n < 2:
        return (values[0] if n == 1 else 0.0), 0.0, 1.0
    mean = sum(values) / n
    if mean <= 0:
        return mean, 0.0, 1.0
    variance = sum((v - mean) ** 2 for v in values) / n
    std = variance ** 0.5
    cv = std / mean if mean > 0 else 1.0
    return mean, std, cv


def detect_officer_suspects(
    employees: list[WageEmployee],
    *,
    excluded_indices: set[int] | None = None,
) -> list[tuple[int, OfficerSuspect]]:
    """ルールベースで「役員疑い」を検出する。

    雇用形態は変更しない（出力 xlsx の D列はそのまま）。
    「変換メモ」シートに『役員疑いN名』として一覧化して人間判断を仰ぐ。

    Args:
        employees: 全従業員リスト
        excluded_indices: 既に「役員」として確定済みの従業員 index（重複検出回避）

    Returns:
        (employee_index, OfficerSuspect) のリスト。月平均給与の降順でソート。

    判定:
        1. monthly_wages から None を除いた有効データが SUSPECT_MIN_FULL_MONTHS 以上
        2. 変動係数 (std/mean) < SUSPECT_CV_THRESHOLD（月額がほぼ一定）
        3. 月平均 ≥ SUSPECT_HIGH_WAGE_YEN OR 全社員平均の SUSPECT_HIGH_WAGE_MULTIPLIER 倍以上
        4. 既に「役員」確定済み（履歴事項照合済み）の人は除外
    """
    excluded = excluded_indices or set()

    # 各従業員の月平均給与を index で索引可能にしておく（後で「自分を除いた他社員平均」を出す）
    monthly_avg_by_idx: dict[int, float] = {}
    for idx, emp in enumerate(employees):
        if idx in excluded:
            continue
        valid = [w for w in emp.monthly_wages if w is not None and w > 0]
        if not valid:
            continue
        monthly_avg_by_idx[idx] = sum(valid) / len(valid)

    suspects: list[tuple[int, OfficerSuspect]] = []
    for idx, emp in enumerate(employees):
        if idx in excluded:
            continue
        valid = [w for w in emp.monthly_wages if w is not None and w > 0]
        if len(valid) < SUSPECT_MIN_FULL_MONTHS:
            continue
        mean, std, cv = _calc_variation(valid)
        if cv >= SUSPECT_CV_THRESHOLD:
            continue
        # 自分を除いた他社員の平均と比較（自分を含めると平均が引き上げられて倍率判定が成立しにくい）
        others = [v for j, v in monthly_avg_by_idx.items() if j != idx]
        peers_avg = sum(others) / len(others) if others else 0.0

        # 高額判定: 70万円以上 OR 他社員平均の2倍以上
        is_high_abs = mean >= SUSPECT_HIGH_WAGE_YEN
        is_high_rel = (
            peers_avg > 0 and mean >= peers_avg * SUSPECT_HIGH_WAGE_MULTIPLIER
        )
        if not (is_high_abs or is_high_rel):
            continue

        reasons = []
        if is_high_abs:
            reasons.append(f'月平均 {mean:,.0f}円 ≥ {SUSPECT_HIGH_WAGE_YEN:,}円')
        if is_high_rel:
            reasons.append(
                f'他社員平均({peers_avg:,.0f}円)の {mean/peers_avg:.1f}倍'
            )
        reason = f'変動係数{cv:.3f} かつ ' + ' / '.join(reasons)
        suspects.append((
            idx,
            OfficerSuspect(
                name=emp.name,
                employment_type=emp.employment_type or '',
                monthly_avg=mean,
                monthly_std=std,
                coefficient_of_variation=cv,
                months_with_data=len(valid),
                reason=reason,
            ),
        ))

    # 月平均の降順で並べ替え（疑いの強い順に上から表示）
    suspects.sort(key=lambda x: -x[1].monthly_avg)
    return suspects


# ─────────────────────────────────────────────────────────────
# パート時間欠落検知（C 改善）
# ─────────────────────────────────────────────────────────────

@dataclass
class PartTimeHoursMissing:
    """パート・アルバイトで月間労働時間が取れていない従業員。

    FTE換算 (R215) に必要なので、ユーザーに所定労働時間の手入力を促す。
    """
    name: str
    employment_type: str          # 正規化前の元値
    months_with_wage: int          # 給与は取れている月数（在籍期間の参考）


def detect_part_time_hours_missing(
    employees: list[WageEmployee],
    *,
    is_kojin: bool = False,
) -> list[tuple[int, PartTimeHoursMissing]]:
    """パート・アルバイトのうち、月間労働時間がすべて欠落している人を検出する。

    Args:
        employees: 賃金台帳の全従業員
        is_kojin: 雇用形態正規化のための個人事業主フラグ

    Returns:
        (employee_index, PartTimeHoursMissing) のリスト

    判定:
        1. 正規化後の雇用形態が「パート・アルバイト」
        2. monthly_hours が全て None
        3. monthly_avg_hours も 0 以下
        4. かつ給与（monthly_wages）が1件以上ある（完全に空の人は除外）
    """
    missing: list[tuple[int, PartTimeHoursMissing]] = []
    for idx, emp in enumerate(employees):
        norm = normalize_employment_type(emp.employment_type or '', is_kojin=is_kojin)
        if norm != 'パート・アルバイト':
            continue
        has_any_hours = (
            any(h is not None and h > 0 for h in emp.monthly_hours)
            or (emp.monthly_avg_hours or 0) > 0
        )
        if has_any_hours:
            continue
        months_with_wage = sum(1 for w in emp.monthly_wages if w is not None)
        if months_with_wage == 0:
            continue
        missing.append((
            idx,
            PartTimeHoursMissing(
                name=emp.name,
                employment_type=emp.employment_type or '',
                months_with_wage=months_with_wage,
            ),
        ))
    return missing


# ─────────────────────────────────────────────────────────────
# テンプレート書込
# ─────────────────────────────────────────────────────────────

@dataclass
class WriteResult:
    """書き込み結果のサマリ"""
    output_path: Path
    employee_count: int
    officer_count: int
    full_year_count: int
    midyear_count: int
    handwritten_files: list[str]
    extraction_path: str
    warnings: list[str]
    # 改善 A: 履歴事項照合で「役員」に上書きされた従業員（氏名, 元の雇用形態）
    officer_matches: list[tuple[str, str]] = None
    # 改善 B: ルールベースで「役員疑い」と判定された従業員
    officer_suspects: list[OfficerSuspect] = None
    # 改善 C: パート時間欠落
    part_time_missing: list[PartTimeHoursMissing] = None

    def __post_init__(self):
        if self.officer_matches is None:
            self.officer_matches = []
        if self.officer_suspects is None:
            self.officer_suspects = []
        if self.part_time_missing is None:
            self.part_time_missing = []


def _clear_data_rows(ws, last_row: int) -> None:
    """テンプレ既定のデータ行（B6〜B15 + 既定の D列ラベル）を全クリアする。

    11名以上の案件で「既定の正社員/パート・アルバイト」表記が残らないよう、
    書き込み開始前にデータ範囲を全消去する。

    注意: openpyxl の `ws.cell(row=r, column=c, value=None)` は
    value=None だと「value 引数未指定」と同じ扱いになり既存値を消さない。
    そのため value 属性に直接 None を代入する。
    """
    for r in range(FIRST_DATA_ROW, last_row + 1):
        for c in range(COL_NO, COL_TRANSPORT + 1):
            ws.cell(row=r, column=c).value = None


def _write_meta(
    ws,
    *,
    company_name: str,
    fiscal_month: int | None,
    extraction_path: str,
    employee_count: int,
    handwritten_files: list[str],
    warnings: list[str],
) -> None:
    """B1〜B3 に抽出経路・警告メタを書き込む。"""
    # タイトル行（既定 B2 のテンプレタイトルは残しつつ、B1 を新規に使う）
    title = f'{company_name}_賃金台帳一覧' if company_name else '賃金台帳一覧'
    if fiscal_month:
        title += f'（決算月: {fiscal_month}月）'
    c1 = ws.cell(row=META_ROW_TITLE, column=COL_NO, value=title)
    c1.font = Font(bold=True, size=12, color='1A1A2E')

    # 抽出経路（赤字で目立たせる）
    is_image_fallback = 'image' in extraction_path.lower() or 'fallback' in extraction_path.lower()
    is_docai = 'DocAI' in extraction_path or 'document_ai' in extraction_path.lower()
    if is_image_fallback:
        path_msg = (
            f'⚠ 抽出経路: {extraction_path} — '
            'Document AI で抽出できず Sonnet 画像経路にフォールバックしました。'
            '原本と全数値を必ず照合してください。'
        )
        color = 'C00000'
        fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
    elif is_docai:
        path_msg = f'抽出経路: {extraction_path}（Document AI + Sonnet 4.6）'
        color = '333333'
        fill = None
    else:
        path_msg = f'抽出経路: {extraction_path}'
        color = '333333'
        fill = None
    c2 = ws.cell(row=META_ROW_EXTRACTION, column=COL_NO, value=path_msg)
    c2.font = Font(bold=True, size=10, color=color)
    if fill is not None:
        c2.fill = fill

    # 検証指示（手書き警告込み）
    verify_msg = f'検出 {employee_count}名'
    if handwritten_files:
        verify_msg += (
            f' / 手書き判定 {len(handwritten_files)}件: '
            f'{", ".join(handwritten_files[:3])}'
            f'{"…" if len(handwritten_files) > 3 else ""}'
        )
        verify_msg += ' ⚠ 手書きPDFは精度が低い可能性があります、必ず原本照合してください'
    if warnings:
        verify_msg += ' / ' + ' / '.join(warnings[:3])
    c3 = ws.cell(row=META_ROW_VERIFICATION, column=COL_NO, value=verify_msg)
    c3.font = Font(
        bold=bool(handwritten_files or warnings),
        size=10,
        color='C00000' if (handwritten_files or warnings) else '333333',
    )


def _write_employee_row(ws, row: int, no: int, emp: WageEmployee, is_kojin: bool) -> None:
    """1名分のデータをテンプレ行に書き込む。"""
    ws.cell(row=row, column=COL_NO, value=no)
    ws.cell(row=row, column=COL_NAME, value=emp.name)
    norm_type = normalize_employment_type(emp.employment_type, is_kojin=is_kojin)
    ws.cell(row=row, column=COL_EMP_TYPE, value=norm_type)

    # 雇用形態が「正社員」または「役員」の場合、E列「月間平均時間」と F列「時給」は
    # 出力しない（月給制が前提のため、AI が誤って値を入れても xlsx には載せない）。
    # パート・アルバイトのみ時給情報を活用する設計（R215 FTE 換算）と整合させる。
    skip_hours_and_rate = norm_type in ('正社員', '役員')

    # E列: 月間平均時間（>0 かつ パートのみ書込、整数化）
    if not skip_hours_and_rate and emp.monthly_avg_hours and emp.monthly_avg_hours > 0:
        ws.cell(row=row, column=COL_AVG_HOURS, value=round(emp.monthly_avg_hours, 1))

    # F列: 時給（>0 かつ パートのみ書込、整数化）
    if not skip_hours_and_rate and emp.hourly_rate and emp.hourly_rate > 0:
        ws.cell(row=row, column=COL_HOURLY, value=int(round(emp.hourly_rate)))

    # G〜R列: monthly_wages[0..11]（カレンダー月そのまま）
    for i, val in enumerate(emp.monthly_wages[:12]):
        if val is None:
            continue
        cell = ws.cell(row=row, column=COL_MONTH_START + i, value=int(round(val)))
        cell.number_format = '#,##0'

    # S列: 年間通勤手当（AI が任意で抽出。> 0 のときのみ書込）
    # ツール側が読込時に在籍月数で均等割して各月から減算 → R216 を課税給与基準に補正
    atransport = getattr(emp, 'annual_transport_allowance', 0) or 0
    if atransport > 0:
        s_cell = ws.cell(row=row, column=COL_TRANSPORT, value=int(round(atransport)))
        s_cell.number_format = '#,##0'


def _write_memo_sheet(
    wb,
    *,
    company_name: str,
    fiscal_month: int | None,
    is_kojin: bool,
    extraction_path: str,
    handwritten_files: list[str],
    employees_total: int,
    officer_matches: list[tuple[str, str]],
    officer_suspects: list[OfficerSuspect],
    part_time_missing: list[PartTimeHoursMissing],
    data_source_files: list[str],
    additional_warnings: list[str],
) -> None:
    """「変換メモ」シートを新規追加して、人間チェック用のメタ情報を書き込む。

    このシートはツール側（read_wage_ledgers）が読まないため、自由に情報を書ける。
    """
    from datetime import datetime

    if MEMO_SHEET_NAME in wb.sheetnames:
        # 既存があれば一旦消す（書き直し）
        del wb[MEMO_SHEET_NAME]
    ms = wb.create_sheet(MEMO_SHEET_NAME)

    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    section_font = Font(bold=True, size=10, color='1A1A2E')
    note_font = Font(size=9, color='666666')

    def _set(row: int, col: int, value, font=None, fill=None, wrap=False):
        c = ms.cell(row=row, column=col, value=value)
        if font:
            c.font = font
        if fill:
            c.fill = fill
        if wrap:
            c.alignment = Alignment(wrap_text=True, vertical='top')

    r = 1
    title = f'{company_name}_賃金台帳変換メモ' if company_name else '賃金台帳変換メモ'
    _set(r, 1, title, font=Font(bold=True, size=14, color='1A1A2E'))
    r += 1
    _set(
        r, 1,
        '※ このシートはツール側が読みません。人間チェック・トラブルシュート用です。',
        font=note_font,
    )
    r += 2

    # 1. 抽出メタ
    _set(r, 1, '【抽出メタ】', font=header_font, fill=header_fill)
    r += 1
    _set(r, 1, '変換日時')
    _set(r, 2, datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    r += 1
    _set(r, 1, '抽出経路')
    _set(r, 2, extraction_path or '未指定')
    r += 1
    _set(r, 1, '決算月')
    _set(r, 2, f'{fiscal_month}月' if fiscal_month else '（指定なし）')
    r += 1
    _set(r, 1, 'テンプレ種別')
    _set(r, 2, '個人事業主' if is_kojin else '法人')
    r += 1
    _set(r, 1, '抽出従業員数')
    _set(r, 2, employees_total)
    r += 2

    # 2. 役員判定の根拠
    _set(r, 1, '【役員判定の根拠】', font=header_font, fill=header_fill)
    r += 1
    if officer_matches:
        _set(
            r, 1,
            f'履歴事項照合で「役員」に上書きした従業員: {len(officer_matches)}名',
            font=section_font,
        )
        r += 1
        _set(r, 1, '氏名', font=Font(bold=True))
        _set(r, 2, '元の雇用形態', font=Font(bold=True))
        _set(r, 3, '上書き後', font=Font(bold=True))
        r += 1
        for name, original in officer_matches:
            _set(r, 1, name)
            _set(r, 2, original or '（空）')
            _set(r, 3, '役員')
            r += 1
        r += 1
    else:
        _set(
            r, 1,
            '履歴事項照合: 該当者なし（履歴事項PDF未投入 or 賃金台帳との氏名一致なし）',
            font=note_font,
        )
        r += 2

    if officer_suspects:
        _set(
            r, 1,
            f'⚠ 役員疑い（ルールベース検出 / 雇用形態は上書きしていません）: {len(officer_suspects)}名',
            font=Font(bold=True, size=10, color='C00000'),
        )
        r += 1
        _set(
            r, 1,
            '※ 月別給与の変動が極めて小さく、かつ高水準のため役員の可能性があります。'
            'PDF原本で確認のうえ、該当者の雇用形態を「役員」に手動変更してください。',
            font=note_font, wrap=True,
        )
        r += 1
        _set(r, 1, '氏名', font=Font(bold=True))
        _set(r, 2, '元雇用形態', font=Font(bold=True))
        _set(r, 3, '月平均', font=Font(bold=True))
        _set(r, 4, '変動係数', font=Font(bold=True))
        _set(r, 5, '理由', font=Font(bold=True))
        r += 1
        for s in officer_suspects:
            _set(r, 1, s.name)
            _set(r, 2, s.employment_type or '（空）')
            cell_avg = ms.cell(row=r, column=3, value=s.monthly_avg)
            cell_avg.number_format = '#,##0'
            cell_cv = ms.cell(row=r, column=4, value=round(s.coefficient_of_variation, 4))
            cell_cv.number_format = '0.0000'
            _set(r, 5, s.reason)
            r += 1
        r += 1
    else:
        _set(r, 1, 'ルールベース役員疑い: 該当者なし', font=note_font)
        r += 2

    # 3. パート時間欠落
    _set(r, 1, '【FTE換算用の時間情報】', font=header_font, fill=header_fill)
    r += 1
    if part_time_missing:
        _set(
            r, 1,
            f'⚠ 月間労働時間が取れなかったパート従業員: {len(part_time_missing)}名',
            font=Font(bold=True, size=10, color='C00000'),
        )
        r += 1
        _set(
            r, 1,
            '※ R215（従業員数 FTE換算）の精度に影響します。'
            '所定労働時間を顧客に確認し、該当行の E列「月間平均時間」に手入力してください。',
            font=note_font, wrap=True,
        )
        r += 1
        _set(r, 1, '氏名', font=Font(bold=True))
        _set(r, 2, '元雇用形態', font=Font(bold=True))
        _set(r, 3, '給与あり月数', font=Font(bold=True))
        r += 1
        for m in part_time_missing:
            _set(r, 1, m.name)
            _set(r, 2, m.employment_type or '（空）')
            _set(r, 3, m.months_with_wage)
            r += 1
        r += 1
    else:
        _set(r, 1, '時間欠落のパート従業員: 該当者なし', font=note_font)
        r += 2

    # 4. 手書きPDF判定
    _set(r, 1, '【手書き/低品質PDF判定】', font=header_font, fill=header_fill)
    r += 1
    if handwritten_files:
        _set(
            r, 1,
            f'⚠ テキスト層が薄いため手書き/写真スキャンと判定された PDF: '
            f'{len(handwritten_files)}件',
            font=Font(bold=True, size=10, color='C00000'),
        )
        r += 1
        _set(
            r, 1,
            '※ OCR精度の低下が想定されます。原本と全数値を必ず照合してください。',
            font=note_font, wrap=True,
        )
        r += 1
        for f in handwritten_files:
            _set(r, 1, f, font=note_font)
            r += 1
        r += 1
    else:
        _set(r, 1, '手書きPDF判定: 該当ファイルなし', font=note_font)
        r += 2

    # 5. データソース
    _set(r, 1, '【データソース（入力ファイル）】', font=header_font, fill=header_fill)
    r += 1
    if data_source_files:
        for f in data_source_files:
            _set(r, 1, f, font=note_font)
            r += 1
    else:
        _set(r, 1, '（記録なし）', font=note_font)
        r += 1
    r += 1

    # 6. その他の警告
    if additional_warnings:
        _set(r, 1, '【追加警告】', font=header_font, fill=header_fill)
        r += 1
        for w in additional_warnings:
            _set(r, 1, w, font=note_font, wrap=True)
            r += 1

    # 列幅調整
    ms.column_dimensions['A'].width = 22
    ms.column_dimensions['B'].width = 22
    ms.column_dimensions['C'].width = 14
    ms.column_dimensions['D'].width = 12
    ms.column_dimensions['E'].width = 50


def write_wage_ledger_to_template(
    employees: Iterable[WageEmployee],
    template_path: Path,
    output_path: Path,
    *,
    company_name: str = '',
    fiscal_month: int | None = None,
    is_kojin: bool = False,
    extraction_path: str = '',
    handwritten_files: list[str] | None = None,
    additional_warnings: list[str] | None = None,
    officer_names: list[str] | None = None,
    data_source_files: list[str] | None = None,
    detect_suspects: bool = True,
) -> WriteResult:
    """`WageEmployee` リストを賃金台帳テンプレートに転記して保存する。

    Args:
        employees: 書き込む従業員リスト
        template_path: 賃金台帳テンプレート Excel のパス
        output_path: 出力 xlsx のパス
        company_name: 出力ファイル名やタイトルに使用
        fiscal_month: 決算月（1〜12）。タイトル表示用
        is_kojin: 個人事業主テンプレ選択時に True
        extraction_path: 抽出経路ラベル（B2 セル表示用）
        handwritten_files: 手書き判定された PDF のファイル名リスト
        additional_warnings: B3 セルに追記する警告メッセージ
        officer_names: 履歴事項PDFから取れた役員氏名（代表者 + 一般役員）。
            None または空なら履歴事項照合をスキップ。
        data_source_files: 入力ファイルの一覧（変換メモシートに記録）
        detect_suspects: True ならルールベース役員疑い検出を実行

    Returns:
        WriteResult: 書き込み結果サマリ
    """
    employees_list = list(employees)
    handwritten_files = handwritten_files or []
    additional_warnings = additional_warnings or []
    officer_names = officer_names or []
    data_source_files = data_source_files or []

    if not template_path.exists():
        raise FileNotFoundError(
            f'賃金台帳テンプレートが見つかりません: {template_path}'
        )

    # ── A: 履歴事項照合で役員上書き ──
    # 上書きは「賃金台帳に書き込む前」に内部リストの employment_type を変更する。
    # 元の雇用形態は officer_matches に保持して変換メモシートに残す。
    officer_matches: list[tuple[str, str]] = []
    if officer_names:
        raw_matches = match_officer_names_to_employees(employees_list, officer_names)
        for idx, original_type, name in raw_matches:
            officer_matches.append((name, original_type))
            # 上書き: 元表記が既に「役員」を含んでいたらそのまま、それ以外は「役員」に
            if '役員' not in (employees_list[idx].employment_type or ''):
                employees_list[idx].employment_type = '役員'
        if officer_matches:
            logger.info(
                f'履歴事項照合: {len(officer_matches)}名を役員に上書き '
                f'({[m[0] for m in officer_matches]})'
            )

    # ── B: ルールベース役員疑い検出（A で上書き済みは除外） ──
    excluded_indices: set[int] = set()
    # 既に「役員」表記の従業員（A で上書き済み or 元から役員）を除外
    for i, emp in enumerate(employees_list):
        if '役員' in (emp.employment_type or ''):
            excluded_indices.add(i)

    officer_suspects: list[OfficerSuspect] = []
    if detect_suspects:
        raw_suspects = detect_officer_suspects(
            employees_list, excluded_indices=excluded_indices,
        )
        officer_suspects = [s for _, s in raw_suspects]
        if officer_suspects:
            logger.info(
                f'役員疑い検出: {len(officer_suspects)}名 '
                f'({[s.name for s in officer_suspects]})'
            )

    # ── C: パート時間欠落検知 ──
    part_time_missing = [
        m for _, m in detect_part_time_hours_missing(employees_list, is_kojin=is_kojin)
    ]
    if part_time_missing:
        logger.info(
            f'パート時間欠落: {len(part_time_missing)}名 '
            f'({[m.name for m in part_time_missing]})'
        )

    # ── テンプレを物理コピーしてから openpyxl で開く（スタイル・既定文言を保持） ──
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copyfile(template_path, output_path)

    wb = openpyxl.load_workbook(str(output_path))
    if SHEET_NAME not in wb.sheetnames:
        wb.close()
        raise ValueError(
            f'テンプレートにシート「{SHEET_NAME}」がありません。テンプレート構造が変更された可能性があります'
        )
    ws = wb[SHEET_NAME]

    # 既定データ行を全クリア（11名以上案件で既定の正社員/パート表記が残らないように）
    needed_rows = max(len(employees_list), DEFAULT_LAST_DATA_ROW - FIRST_DATA_ROW + 1)
    _clear_data_rows(ws, FIRST_DATA_ROW + needed_rows - 1)

    # データ転記
    officer_count = 0
    full_year_count = 0
    midyear_count = 0
    for i, emp in enumerate(employees_list):
        row = FIRST_DATA_ROW + i
        _write_employee_row(ws, row, no=i + 1, emp=emp, is_kojin=is_kojin)
        norm_type = normalize_employment_type(emp.employment_type, is_kojin=is_kojin)
        if norm_type == '役員':
            officer_count += 1
        if emp.is_full_year:
            full_year_count += 1
        else:
            midyear_count += 1

    # B1〜B3 の警告メタを構築（役員疑い・パート欠落の件数も載せる）
    meta_warnings = list(additional_warnings)
    if officer_matches:
        meta_warnings.append(f'役員照合 {len(officer_matches)}名 → 上書き済')
    if officer_suspects:
        meta_warnings.append(f'役員疑い {len(officer_suspects)}名 → 変換メモシート参照')
    if part_time_missing:
        meta_warnings.append(f'パート時間欠落 {len(part_time_missing)}名 → 変換メモシート参照')

    _write_meta(
        ws,
        company_name=company_name,
        fiscal_month=fiscal_month,
        extraction_path=extraction_path or '未指定',
        employee_count=len(employees_list),
        handwritten_files=handwritten_files,
        warnings=meta_warnings,
    )

    # 変換メモシート生成
    _write_memo_sheet(
        wb,
        company_name=company_name,
        fiscal_month=fiscal_month,
        is_kojin=is_kojin,
        extraction_path=extraction_path or '未指定',
        handwritten_files=handwritten_files,
        employees_total=len(employees_list),
        officer_matches=officer_matches,
        officer_suspects=officer_suspects,
        part_time_missing=part_time_missing,
        data_source_files=data_source_files,
        additional_warnings=additional_warnings,
    )

    wb.save(str(output_path))
    wb.close()
    logger.info(
        f'賃金台帳テンプレート書込完了: {output_path.name} '
        f'({len(employees_list)}名 / 役員{officer_count}名 / '
        f'全月{full_year_count}名 / 中途{midyear_count}名 / '
        f'役員照合{len(officer_matches)}名 / 役員疑い{len(officer_suspects)}名 / '
        f'パート時間欠落{len(part_time_missing)}名)'
    )

    return WriteResult(
        output_path=output_path,
        employee_count=len(employees_list),
        officer_count=officer_count,
        full_year_count=full_year_count,
        midyear_count=midyear_count,
        handwritten_files=handwritten_files,
        extraction_path=extraction_path,
        warnings=meta_warnings,
        officer_matches=officer_matches,
        officer_suspects=officer_suspects,
        part_time_missing=part_time_missing,
    )
