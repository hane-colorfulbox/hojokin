#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""案件フォルダの必要書類チェック（決定論パート）— case-docs-check スキル同梱

役割:
- ファイル名・拡張子・Excel構造など「機械的に決まる」チェックだけを行い、
  FAIL（不備確定）/ WARN（要確認）/ INFO を列挙する。
- 日付・年度・書類種別など「中身の意味読み」は一切しない。PDF の証明日や
  納税証明書の対象年度をここで正規表現判定すると、和暦・様式の表記揺れで
  誤 PASS / 誤 FAIL を量産するため。意味読みは Claude Code 本体が Read で行う
  （SKILL.md §5 / checklists/中身チェック手順.md）。

依存: 標準ライブラリのみで動く。openpyxl があれば Excel 構造チェック
     （テンプレ規格・ヒアリング様式判定）も実施し、無ければその項目を
     「判定不能（要目視）」に落として続行する。hojokin パッケージ・
     ネットワーク・API は使わない（配布先の他人PCでも headless で動く）。

同期: PATTERNS / OUTPUT_FILE_MARKERS / ALLOWED_EXTS / EXCLUDED_SUBFOLDERS /
     REGISTRY_CONFIRMED_KEYWORD / REQUIRED_CATS_BY_TASK /
     REQUIRED_CATS_APPLICATION_KOJIN は hojokin/pipeline.py（FileDetector）と
     app.py からの独立コピー。scripts/check_docscheck_sync.py が AST 突合し、
     不一致なら配布 ZIP のビルドを中止する（原本を変えたらここも直す）。

使い方:
  python check_docs.py --local <案件フォルダ>
  python check_docs.py --manifest manifest.json [--files-dir <DL済みフォルダ>]
  共通: --task both|application|katen  --entity 法人|個人|不明
        --frame 通常枠|インボイス枠|不明  --json <出力パス>
終了コード: 0=FAILなし / 1=FAILあり / 2=スクリプトエラー
"""
import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

# ══════════════ 同期対象定数（原本: hojokin/pipeline.py FileDetector / app.py）══════════════

PATTERNS = {
    'hearing': ['ヒアリング'],
    'registry': ['履歴事項', '登記簿', '登記事項'],
    'identity': ['運転免許証', '運転経歴証明書', '住民票', '本人確認'],
    'tax': ['納税証明'],
    'pl': ['損益計算書', '決算報告書', '決算書', '収支内訳書', '青色申告'],
    'cost_report': ['製造原価報告書', '原価報告書'],
    'estimate': ['見積', 'お見積'],
    'wage_report': ['賃金状況報告'],
    'wage_ledger': ['賃金台帳', '給与台帳'],
    'wage_data': ['支給控除一覧', '給与データ'],
}

OUTPUT_FILE_MARKERS = (
    '_AI版', '_給与支給総額計算', '_一人当たり給与支給総額',
    '_賃金台帳_AI集計',
    '_加点①', '_加点②',
)

ALLOWED_EXTS = {
    'hearing':     {'.xlsx', '.xlsm'},
    'registry':    {'.pdf'},
    'identity':    {'.pdf'},
    'tax':         {'.pdf'},
    'pl':          {'.pdf'},
    'cost_report': {'.pdf'},
    'estimate':    {'.xlsx', '.xlsm', '.pdf'},
    'wage_report': {'.xlsx', '.xlsm'},
    'wage_ledger': {'.xlsx', '.xlsm', '.csv'},
    'wage_data':   {'.pdf'},
}

EXCLUDED_SUBFOLDERS = {'申請時使用'}

REGISTRY_CONFIRMED_KEYWORD = '履歴事項'

REQUIRED_CATS_BY_TASK = {
    'application':           {'hearing', 'registry', 'pl'},
    'wage':                  {'wage_ledger'},
    'per_employee_wage':     {'wage_ledger'},
    'bonus':                 {'wage_ledger'},
    'bonus_wage_ledger_creation': {'wage_ledger'},
    'wage_ledger_creation':  {'wage_ledger'},
    'all':                   {'hearing', 'registry', 'pl'},
}

REQUIRED_CATS_APPLICATION_KOJIN = {'hearing', 'pl'}

# ══════════════ スキル固有定数（同期対象外）══════════════

CAT_LABELS = {
    'hearing': 'ヒアリングシート', 'registry': '履歴事項全部証明書',
    'identity': '本人確認資料', 'tax': '納税証明書',
    'pl': '決算書（PL）', 'cost_report': '製造原価報告書',
    'estimate': '見積書', 'wage_report': '賃金状況報告シート',
    'wage_ledger': '賃金台帳', 'wage_data': '支給控除一覧',
}

# 申請書作成で「必須ではないが実質必要」なカテゴリ（欠落は WARN）
RECOMMENDED_CATS_APPLICATION = ('estimate', 'tax', 'wage_ledger')

# 未分類でも「書類らしい」拡張子（CC の目視レビュー対象に載せる）
DOC_LIKE_EXTS = {'.pdf', '.xlsx', '.xlsm', '.xls', '.csv', '.jpg', '.jpeg', '.png', '.heic'}

# 提出エビデンス（ツール入力ではないが回収状況を報告する。ファイル名ヒント）
SUBMISSION_HINTS = {
    'IT戦略ナビwith 実施結果': ('IT戦略', 'デジwith'),
    '成長加速マッチング 画面キャプチャ': ('成長加速',),
    '省力化ナビ 実施結果': ('省力化',),
}

GOOGLE_MIME_EXT = {
    'application/vnd.google-apps.spreadsheet': '.xlsx',
    'application/vnd.google-apps.document': '.docx',
}
MIME_FOLDER = 'application/vnd.google-apps.folder'
MIME_SHORTCUT = 'application/vnd.google-apps.shortcut'

# 標準賃金台帳テンプレ規格（原本: ツール/賃金台帳テンプレート.xlsx 実測）
WAGE_SHEET = '従業員別明細'
WAGE_HEADER_ROW = 5
WAGE_HEADERS = [  # B5〜T5（プレーン月名。注記混入はツールの月解析を壊す）
    'No', '氏名', '雇用形態', '月間平均時間', '時給',
    '1月', '2月', '3月', '4月', '5月', '6月',
    '7月', '8月', '9月', '10月', '11月', '12月',
    '年間通勤手当', '年間賞与',
]

# 加点判定用台帳テンプレ規格（原本: ツール/加点判定用賃金台帳テンプレート.xlsx 実測。
# 列定数の単一の真実は hojokin/wage_reader.py の BWL_*）
BWL_SHEET = '加点判定用明細'
BWL_HEADER_ROW = 6
BWL_HEADERS = [  # B6〜R6（暦月固定: 令和6年10月〜令和7年9月＋交付申請直近月）
    'No', '氏名', '雇用形態', '月間所定\n労働時間',
    '令和6年10月\n基本給', '令和6年11月\n基本給', '令和6年12月\n基本給',
    '令和7年1月\n基本給', '令和7年2月\n基本給', '令和7年3月\n基本給',
    '令和7年4月\n基本給', '令和7年5月\n基本給', '令和7年6月\n基本給',
    '令和7年7月\n基本給', '令和7年8月\n基本給', '令和7年9月\n基本給',
    '交付申請直近月\n基本給',
]
BWL_PREF_CELL = 'C2'    # 事業場所在地（都道府県）
BWL_APPYM_CELL = 'C3'   # 交付申請月（yyyy/mm）— 空欄だと加点②判定不能

# ヒアリングシート様式フィンガープリント（原本: ツール/ヒアリングシート2026_*.xlsx 実測
# 2026-08-10 に4様式で再実測。基本情報シートB列ラベルの行位置。2行以上一致で様式確定。
# 様式改訂時はここと required_rows を再実測すること）
#
# 🔴 指紋は「自分の様式で最高点・他様式では1点以下」になるよう選ぶこと。
#    判定は最高点が同点だと『判別不能』に落ちて後続チェックごと落ちる。
#    2つの個人様式（通常枠個人／インボイス個人）は基本情報の r8〜r26 が完全一致するため、
#    見分けは r62（過去交付の条件行）と r84（通常枠=自社の強み／インボイス=事業所内最低賃金時給）で行う。
#    2026-08-10 実測のスコア行列（行=実ファイル・列=様式定義。対角3・他1以下）:
#                   通常枠法人 インボイス法人 インボイス個人 通常枠個人
#      通常枠法人        3        1        0        0
#      インボイス法人      0        3        0        0
#      インボイス個人      0        1        3        1
#      通常枠個人        0        1        1        3
HEARING_SHEET = '基本情報'
HEARING_FORMS = {
    '通常枠法人': {
        'frame': '通常枠', 'entity': '法人',
        'fingerprint': {6: '企業名', 53: '申請ツール名', 71: '自社の強み'},
        'required_rows': [
            6, 8, 10, 12, 15, 16, 17, 18, 19, 20, 21, 22, 23, 26, 28, 30, 32,
            34, 36, 38, 40, 42, 44, 46, 48, 57, 60, 62, 65, 66, 67, 68, 71,
            72, 73, 74, 77, 82, 84, 85, 86, 89, 90,
        ],
        'sa_id_row': 62,
    },
    'インボイス法人': {
        'frame': 'インボイス枠', 'entity': '法人',
        'fingerprint': {6: 'GビズIDプライム', 8: '企業名', 83: 'インボイス制度対応状況'},
        'required_rows': [
            6, 8, 10, 12, 14, 17, 18, 19, 20, 21, 22, 23, 24, 28, 30, 32, 34,
            36, 38, 40, 42, 44, 46, 48, 50, 58, 61, 63, 66, 67, 68, 69, 72,
            76, 79, 80, 81, 84, 89, 90,
        ],
        'sa_id_row': 63,
    },
    'インボイス個人': {
        'frame': 'インボイス枠', 'entity': '個人',
        # r16 生年月日 / r22 事業開始年月日 は通常枠個人と完全一致するため指紋から外した
        'fingerprint': {8: '屋号・商号', 62: '過去にサービス等生産性向上IT導入支援事業',
                        84: '事業所内最低賃金時給'},
        'required_rows': [
            6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 29, 30, 31, 32, 33, 34,
            35, 36, 40, 42, 44, 46, 48, 50, 52, 54, 56, 58, 60, 62, 70, 73,
            75, 78, 79, 80, 81, 84, 88, 91, 92, 93, 96, 101, 102,
        ],
        'sa_id_row': 75,
    },
    '通常枠個人': {
        'frame': '通常枠', 'entity': '個人',
        'fingerprint': {8: '屋号・商号', 84: '自社の強み',
                        87: 'どのようなプロセスに対してIT投資を行いました'},
        'required_rows': [
            6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 29, 30, 31, 32, 33, 34,
            35, 36, 39, 41, 43, 45, 47, 49, 51, 53, 55, 57, 59, 61, 70, 73,
            75, 78, 79, 80, 81, 84, 85, 86, 87, 90, 95, 97, 98, 99, 102, 103,
        ],
        'sa_id_row': 75,
    },
}

# 決算書ファイル名の期情報トークン（無いと mtime ガチャで前期誤選択のリスク）
PL_PERIOD_RES = [
    re.compile(r'第\s*\d+\s*期'),
    re.compile(r'令和\s*\d+\s*年'),
    re.compile(r'[HRhr]\s*\d+[._年]'),
    re.compile(r'\d{4}\s*年'),
    re.compile(r'\d{4}[-_.]\d{1,2}'),
]


def norm(s) -> str:
    if s is None:
        return ''
    return unicodedata.normalize('NFC', str(s)).strip()


# ══════════════ 列挙（ローカル / manifest）══════════════

def iter_local(folder: Path, excluded_dirs: list) -> list:
    """FileDetector._iter_files と同じ規則（. / _ 始まりディレクトリは降りない）
    ＋ Drive 運用の除外サブフォルダ（申請時使用）をローカルでも適用する。"""
    entries = []

    def walk(d: Path):
        try:
            children = sorted(d.iterdir())
        except PermissionError:
            return
        for p in children:
            if p.is_dir():
                name_nfc = norm(p.name)
                if name_nfc.startswith(('.', '_')):
                    continue
                if name_nfc in EXCLUDED_SUBFOLDERS:
                    excluded_dirs.append(name_nfc)
                    continue
                walk(p)
            elif p.is_file():
                entries.append({
                    'name': p.name, 'ext': p.suffix.lower(),
                    'parent': str(p.parent.relative_to(folder)),
                    'local_path': str(p), 'google_native': False,
                })

    walk(folder)
    return entries


def load_manifest(path: Path, files_dir: Path | None, excluded_dirs: list,
                  shortcuts: list) -> list:
    """manifest.json（Drive 列挙結果）を entries に変換する。

    形式: {"files": [{"name": str, "mimeType": str?, "parent_path": str?,
                      "id": str?, "size": ...?}]}
    parent_path に除外サブフォルダ名を含むものは第二の安全網としてここでも落とす
    （一次的には列挙時に降りないのが正）。
    """
    data = json.loads(path.read_text(encoding='utf-8'))
    entries = []
    local_index = {}
    if files_dir and files_dir.is_dir():
        for p in files_dir.rglob('*'):
            if p.is_file():
                local_index[norm(p.name)] = str(p)
    for f in data.get('files', []):
        name = f.get('name') or ''
        mime = f.get('mimeType') or ''
        parent = f.get('parent_path') or ''
        if mime == MIME_FOLDER:
            continue
        if mime == MIME_SHORTCUT:
            shortcuts.append(name)
            continue
        parts = [norm(x) for x in re.split(r'[/\\]', parent) if x]
        if any(x in EXCLUDED_SUBFOLDERS for x in parts):
            excluded_dirs.append(parent)
            continue
        ext = Path(name).suffix.lower()
        google_native = False
        if not ext and mime in GOOGLE_MIME_EXT:
            ext = GOOGLE_MIME_EXT[mime]
            google_native = True
        entries.append({
            'name': name, 'ext': ext, 'parent': parent,
            'local_path': local_index.get(norm(name)),
            'google_native': google_native, 'id': f.get('id'),
        })
    return entries


# ══════════════ 分類（FileDetector._scan のミラー）══════════════

def classify(entries):
    detected = {k: [] for k in PATTERNS}
    skipped = []            # (category, name, reason)
    excluded_outputs = []
    unmatched = []
    for e in entries:
        name_nfc = norm(e['name'])
        if name_nfc.startswith('~$'):
            continue
        if any(m in name_nfc for m in OUTPUT_FILE_MARKERS):
            excluded_outputs.append(e['name'])
            continue
        hit = False
        for category, keywords in PATTERNS.items():
            if any(kw in name_nfc for kw in keywords):
                hit = True
                allowed = ALLOWED_EXTS.get(category)
                if allowed is not None and e['ext'] not in allowed:
                    skipped.append((category, e['name'],
                                    f'拡張子{e["ext"] or "なし"}は{CAT_LABELS[category]}では非対応'))
                else:
                    detected[category].append(e)
                break
        if not hit:
            unmatched.append(e)
    return detected, skipped, excluded_outputs, unmatched


# ══════════════ findings ══════════════

class Report:
    def __init__(self):
        self.findings = []

    def add(self, level, code, message, files=None):
        self.findings.append({
            'level': level, 'code': code, 'message': message,
            'files': files or [],
        })

    def fail(self, code, msg, files=None):
        self.add('FAIL', code, msg, files)

    def warn(self, code, msg, files=None):
        self.add('WARN', code, msg, files)

    def info(self, code, msg, files=None):
        self.add('INFO', code, msg, files)

    def count(self, level):
        return sum(1 for f in self.findings if f['level'] == level)


def _try_openpyxl():
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


def check_presence(rep: Report, detected, skipped, task: str, entity: str):
    """層1: 存在チェック（タスク×法人/個人）"""
    names = {c: [e['name'] for e in v] for c, v in detected.items()}

    if task in ('both', 'application'):
        if entity == '個人':
            required = REQUIRED_CATS_APPLICATION_KOJIN
        else:
            required = REQUIRED_CATS_BY_TASK['application']
        for cat in sorted(required):
            if cat == 'registry' and entity == '不明':
                continue  # 下で両様扱い
            if not names.get(cat):
                rep.fail('MISSING-REQUIRED',
                         f'{CAT_LABELS[cat]} が見つかりません（申請書作成タスクの必須入力）')
        if entity == '不明' and not names.get('registry'):
            rep.warn('MISSING-REQUIRED-COND',
                     '履歴事項全部証明書が見つかりません（法人なら必須／個人事業主なら不要。'
                     '法人か個人かを確定してください）')
        for cat in RECOMMENDED_CATS_APPLICATION:
            if not names.get(cat):
                extra = ''
                if cat == 'wage_ledger':
                    extra = '（R215/R216 算定に実質必須。直近事業年度12ヶ月分を回収）'
                elif cat == 'tax':
                    extra = '（提出必須書類。法人=法人税「その1/その2」直近分、個人=所得税）'
                elif cat == 'estimate':
                    extra = '（申請金額・ツール名の源泉）'
                rep.warn('MISSING-RECOMMENDED',
                         f'{CAT_LABELS[cat]} が見つかりません{extra}')
        if entity == '個人' and not names.get('identity'):
            rep.warn('MISSING-IDENTITY',
                     '本人確認資料（運転免許証/運転経歴証明書/住民票）が見つかりません'
                     '（個人事業主の提出必須書類。ツール入力ではないため回収状況を確認）')
        # 登記書類の名前が正準（履歴事項）でない → 種別要確認
        loose = [e['name'] for e in detected.get('registry', [])
                 if REGISTRY_CONFIRMED_KEYWORD not in norm(e['name'])]
        if loose:
            rep.warn('REGISTRY-NAME',
                     '登記書類の名前に「履歴事項」が含まれません。現在事項全部証明書等の'
                     '可能性があるため、中身で種別を確認してください（本補助金は履歴事項'
                     '全部証明書のみ）', loose)
        # 決算書の期数（社内運用は直近2期回収）
        pl_n = len(names.get('pl') or [])
        if pl_n == 1:
            rep.warn('PL-SINGLE',
                     '決算書らしきPDFが1つだけです。社内運用は直近2期分回収'
                     '（BS・PL・販管費明細/製造原価明細を含む）。中身で期数を確認してください',
                     names['pl'])
        if pl_n >= 2:
            no_period = [n for n in names['pl']
                         if not any(rx.search(norm(n)) for rx in PL_PERIOD_RES)]
            if no_period:
                rep.warn('PL-NO-PERIOD',
                         'ファイル名から年度・期数が読めない決算書があります。複数PDF混在時に'
                         '前期を誤選択するリスクがあるため「第N期」等をファイル名に入れることを推奨',
                         no_period)

    # 賃金台帳が PDF でしか無い（拡張子スキップに落ちたケース）
    wl_pdf = [n for (cat, n, _r) in skipped if cat == 'wage_ledger' and n.lower().endswith('.pdf')]
    if wl_pdf:
        if names.get('wage_ledger'):
            rep.info('WAGE-PDF-EXTRA',
                     '賃金台帳PDFがありますが、Excel/CSVの賃金台帳も存在するためそちらが使われます',
                     wl_pdf)
        else:
            rep.warn('WAGE-PDF-ONLY',
                     '賃金台帳がPDFしかありません。このままでは申請書作成タスクに使えません。'
                     '→ ツールの「賃金台帳の作成」タスク（PDF可）か wagebook-convert スキルで'
                     '規格Excelに変換してください（顧客への再依頼は不要な場合があります）', wl_pdf)

    # その他の拡張子スキップ
    other_skips = [(c, n, r) for (c, n, r) in skipped
                   if not (c == 'wage_ledger' and n.lower().endswith('.pdf'))]
    for cat, name, reason in other_skips:
        rep.warn('EXT-SKIP', f'{reason}（{CAT_LABELS[cat]}として使えません）', [name])


def check_submissions(rep: Report, entries):
    """提出エビデンス（デジwith等）の回収状況。不在は「回収状況の確認」どまり
    （メール添付→Drive自動格納が遅延する運用もあるため FAIL にしない）。"""
    all_names = [norm(e['name']) for e in entries]
    for label, hints in SUBMISSION_HINTS.items():
        if not any(any(h in n for h in hints) for n in all_names):
            rep.info('SUBMISSION-UNSEEN',
                     f'{label} らしきファイルが見当たりません（提出物。回収状況を確認）')


def _header_diffs(ws, header_row, expected, start_col=2):
    diffs = []
    for i, exp in enumerate(expected):
        got = ws.cell(row=header_row, column=start_col + i).value
        if norm(got) != norm(exp):
            col = ws.cell(row=header_row, column=start_col + i).column_letter
            diffs.append(f'{col}{header_row}: 期待「{exp}」/ 実際「{got}」')
    return diffs


def check_excel_structures(rep: Report, detected, task: str, frame: str, entity: str):
    """層2-3: Excel 構造チェック（ローカル取得済みの xlsx のみ。openpyxl 必要）"""
    xlsx_wage = [e for e in detected.get('wage_ledger', [])
                 if e['ext'] in ('.xlsx', '.xlsm')]
    hearings = [e for e in detected.get('hearing', [])]

    if not _try_openpyxl():
        if xlsx_wage or hearings:
            rep.warn('NO-OPENPYXL',
                     'openpyxl が無いため Excel 構造チェック（テンプレ規格・ヒアリング様式判定）'
                     'をスキップしました。中身は目視確認してください')
        return {'bonus_ledgers': [], 'hearing_form': None}

    from openpyxl import load_workbook

    bonus_ledgers = []
    std_ledgers = []
    for e in xlsx_wage:
        if not e.get('local_path'):
            rep.info('NOT-LOCAL',
                     f'{e["name"]} はローカル未取得のため構造チェックをスキップ'
                     '（ダウンロード後に再実行可）')
            continue
        try:
            wb = load_workbook(e['local_path'], data_only=True)
        except Exception as ex:
            rep.warn('XLSX-OPEN-ERROR', f'{e["name"]} を開けません: {ex}', [e['name']])
            continue
        sheet_names = [norm(s) for s in wb.sheetnames]
        if BWL_SHEET in sheet_names:
            bonus_ledgers.append(e)
            ws = wb[wb.sheetnames[sheet_names.index(BWL_SHEET)]]
            diffs = _header_diffs(ws, BWL_HEADER_ROW, BWL_HEADERS)
            if diffs:
                rep.fail('TPL-BONUS-HEADER',
                         f'{e["name"]}: 加点判定用台帳のヘッダーが暦月固定テンプレ'
                         f'（令和6年10月〜令和7年9月＋交付申請直近月）から改変されています: '
                         + ' / '.join(diffs[:5]), [e['name']])
            if not norm(ws[BWL_PREF_CELL].value):
                rep.warn('BONUS-PREF-EMPTY',
                         f'{e["name"]}: 事業場所在地（都道府県、{BWL_PREF_CELL}）が未記入です'
                         '（最低賃金判定に必要）', [e['name']])
            if not norm(ws[BWL_APPYM_CELL].value):
                rep.warn('BONUS-APPYM-EMPTY',
                         f'{e["name"]}: 交付申請月（{BWL_APPYM_CELL}、yyyy/mm）が未記入です'
                         '（空欄だと加点②＝直近月判定ができません）', [e['name']])
        elif WAGE_SHEET in sheet_names:
            std_ledgers.append(e)
            ws = wb[wb.sheetnames[sheet_names.index(WAGE_SHEET)]]
            diffs = _header_diffs(ws, WAGE_HEADER_ROW, WAGE_HEADERS)
            if diffs:
                rep.fail('TPL-WAGE-HEADER',
                         f'{e["name"]}: 賃金台帳テンプレのB5:T5ヘッダーが規格と不一致です'
                         '（月見出しへの注記混入はツールが0名と誤読する実事故あり）: '
                         + ' / '.join(diffs[:5]), [e['name']])
        else:
            rep.info('WAGE-RAW',
                     f'{e["name"]}: テンプレ規格外の賃金台帳（給与ソフト出力等）。'
                     '対象事業年度・全員12ヶ月分の中身チェックへ', [e['name']])

    if task in ('both', 'katen') and not bonus_ledgers:
        rep.warn('BONUS-LEDGER-MISSING',
                 '加点判定用賃金台帳（シート「加点判定用明細」を持つExcel）が見つかりません。'
                 '加点を狙う案件は「加点判定用賃金台帳の作成」タスクで先に生成してください'
                 '（対象期間: 令和6年10月〜令和7年9月＋交付申請直近月。決算月によっては'
                 '前期・今期の賃金台帳2冊が必要）')

    hearing_form = _check_hearing(rep, hearings, frame, entity)
    return {'bonus_ledgers': [e['name'] for e in bonus_ledgers],
            'hearing_form': hearing_form}


def _sa_id_row_by_label(ws):
    """B列ラベルから SECURITY ACTION の行を探す（様式が判別できないとき用）"""
    for r in range(1, 120):
        if 'SECURITY ACTION' in norm(ws.cell(row=r, column=2).value).upper():
            return r
    return None


def _check_sa_id(rep: Report, ws, row: int, name: str):
    """SECURITY ACTION 自己宣言IDが旧システムの「40」始まりでないか（FAIL 相当）"""
    digits = re.sub(r'\D', '', norm(ws.cell(row=row, column=3).value))
    if digits.startswith('40'):
        rep.fail('SA-ID-OLD',
                 f'{name}: SECURITY ACTION自己宣言IDが「40」始まりです。'
                 '旧システムのIDのため新システムでの再取得が必要です（即時発行可）',
                 [name])


def _check_hearing(rep: Report, hearings, frame: str, entity: str):
    """ヒアリングシートの様式判定（フィンガープリント）＋必須セル未記入チェック"""
    from openpyxl import load_workbook

    form_found = None
    for e in hearings:
        if not e.get('local_path'):
            rep.info('NOT-LOCAL',
                     f'{e["name"]} はローカル未取得のため様式判定をスキップ'
                     '（ダウンロード後に再実行可）')
            continue
        try:
            wb = load_workbook(e['local_path'], data_only=True)
        except Exception as ex:
            rep.warn('XLSX-OPEN-ERROR', f'{e["name"]} を開けません: {ex}', [e['name']])
            continue
        sheet_names = [norm(s) for s in wb.sheetnames]
        if HEARING_SHEET not in sheet_names:
            rep.warn('HEARING-NO-SHEET',
                     f'{e["name"]}: 「{HEARING_SHEET}」シートがありません（様式外）。'
                     '正規のヒアリングシート様式か確認してください', [e['name']])
            continue
        ws = wb[wb.sheetnames[sheet_names.index(HEARING_SHEET)]]

        scores = {}
        for form, spec in HEARING_FORMS.items():
            scores[form] = sum(
                1 for row, label in spec['fingerprint'].items()
                if label in norm(ws.cell(row=row, column=2).value)
            )
        top = sorted(scores.values(), reverse=True)
        if top[0] < 2 or (len(top) > 1 and top[0] == top[1]):
            rep.warn('HEARING-FORM-UNKNOWN',
                     f'{e["name"]}: 様式を判別できません（フィンガープリント一致 {scores}）。'
                     '独自改変か旧様式の可能性。目視確認してください', [e['name']])
            # 様式が分からなくても、SECURITY ACTION の旧ID（FAIL相当）だけは必ず見る。
            # ここで continue して全部飛ばすと「未記入なし・SA-ID問題なし」と誤読される
            sa_row = _sa_id_row_by_label(ws)
            if sa_row:
                _check_sa_id(rep, ws, sa_row, e['name'])
            else:
                rep.warn('SA-ID-UNCHECKED',
                         f'{e["name"]}: 様式不明かつSECURITY ACTION欄を特定できず、'
                         '旧ID（40始まり）の確認ができていません。目視で確認してください',
                         [e['name']])
            continue
        best = max(scores, key=lambda k: scores[k])
        spec = HEARING_FORMS[best]
        form_found = best
        rep.info('HEARING-FORM', f'{e["name"]}: 様式判定 = {best}', [e['name']])

        if frame != '不明' and spec['frame'] != frame:
            rep.fail('HEARING-FORM-MISMATCH',
                     f'{e["name"]}: ヒアリングシートが「{best}」様式ですが、案件は'
                     f'「{frame}」です。様式取り違えは転記の広域ズレ（実事故あり）に直結'
                     'します。正しい様式での再記入が必要です', [e['name']])
        if entity != '不明' and spec['entity'] != entity:
            rep.fail('HEARING-ENTITY-MISMATCH',
                     f'{e["name"]}: ヒアリングシートが{spec["entity"]}様式ですが、'
                     f'案件は{entity}です。様式を確認してください', [e['name']])

        empty = []
        for row in spec['required_rows']:
            if not norm(ws.cell(row=row, column=3).value):
                label = norm(ws.cell(row=row, column=2).value).split('\n')[0][:30]
                empty.append(f'行{row}: {label}')
        if empty:
            head = ' / '.join(empty[:8])
            more = f' …ほか{len(empty) - 8}件' if len(empty) > 8 else ''
            rep.warn('HEARING-EMPTY-REQUIRED',
                     f'{e["name"]}: 記入欄（黄色セル相当）に未記入が{len(empty)}件: '
                     f'{head}{more}', [e['name']])

        _check_sa_id(rep, ws, spec['sa_id_row'], e['name'])
    return form_found


def build_content_targets(detected):
    """CC が中身を Read で確認すべきファイルと観点（SKILL.md §5 が消費）"""
    plan = {
        'registry': '種別（履歴事項全部証明書か）／証明日が3ヶ月以内か／全ページ（証明文・登記官押印）',
        'tax': '種別（その1/その2・税目）／対象事業年度が直近決算期と一致するか／税務署発行か',
        'pl': '期末日（自〜至）と決算月の整合／直近期か／2期分あるか／BS・PL・販管費明細ページの有無',
        'estimate': 'ツール名がヒアリングの申請ツール名と一致するか／金額（税抜）／宛名',
        'identity': '種別／有効期限（免許証）または発行日3ヶ月以内（住民票）／裏面の要否',
        'wage_ledger': '対象12ヶ月が直近事業年度と一致するか／全員12ヶ月分あるか／賞与の分離'
                       '／S列（通勤手当）が月列に含まれる分だけか',
        'hearing': '確認事項シートの「確認済」欄／未記入項目の妥当性',
        'wage_report': '様式（加点①/②）と記入状況',
    }
    targets = []
    for cat, view in plan.items():
        for e in detected.get(cat, []):
            targets.append({'category': cat, 'label': CAT_LABELS[cat],
                            'name': e['name'], 'check': view,
                            'local_path': e.get('local_path'),
                            'id': e.get('id')})
    return targets


def main() -> int:
    ap = argparse.ArgumentParser(description='案件フォルダの必要書類チェック（決定論パート）')
    ap.add_argument('--local', help='ローカル案件フォルダ')
    ap.add_argument('--manifest', help='Drive 列挙結果 manifest.json')
    ap.add_argument('--files-dir', help='DL済みファイルの置き場（manifest とファイル名で突合）')
    ap.add_argument('--task', default='both', choices=['both', 'application', 'katen'])
    ap.add_argument('--entity', default='不明', choices=['法人', '個人', '不明'])
    ap.add_argument('--frame', default='不明', choices=['通常枠', 'インボイス枠', '不明'])
    ap.add_argument('--json', help='機械可読の結果 JSON を書き出すパス')
    args = ap.parse_args()

    if bool(args.local) == bool(args.manifest):
        print('ERROR: --local か --manifest のどちらか一方を指定してください', file=sys.stderr)
        return 2

    excluded_dirs, shortcuts = [], []
    if args.local:
        folder = Path(args.local)
        if not folder.is_dir():
            print(f'ERROR: フォルダがありません: {folder}', file=sys.stderr)
            return 2
        entries = iter_local(folder, excluded_dirs)
    else:
        mpath = Path(args.manifest)
        if not mpath.is_file():
            print(f'ERROR: manifest がありません: {mpath}', file=sys.stderr)
            return 2
        entries = load_manifest(
            mpath, Path(args.files_dir) if args.files_dir else None,
            excluded_dirs, shortcuts)

    detected, skipped, excluded_outputs, unmatched = classify(entries)

    rep = Report()
    check_presence(rep, detected, skipped, args.task, args.entity)
    check_submissions(rep, entries)
    extra = check_excel_structures(rep, detected, args.task, args.frame, args.entity)

    for name in shortcuts:
        rep.warn('SHORTCUT', 'Driveショートカットのため実体を解決できません。'
                 '実体ファイルの場所を確認してください', [name])
    g_native = [e['name'] for e in entries if e.get('google_native')]
    if g_native:
        rep.info('GOOGLE-NATIVE',
                 'Googleネイティブ形式（スプレッドシート等）です。ツール投入時は '
                 'xlsx へのエクスポートが必要', g_native)
    if excluded_dirs:
        rep.info('EXCLUDED-DIR',
                 f'除外サブフォルダをスキップ: {sorted(set(excluded_dirs))}'
                 '（税理士要約版等は入力に使わない運用）')
    if excluded_outputs:
        rep.info('OUTPUT-EXCLUDED',
                 f'過去のツール出力 {len(excluded_outputs)} 件を入力の母数から除外',
                 excluded_outputs)
    doc_unmatched = [e['name'] for e in unmatched if e['ext'] in DOC_LIKE_EXTS]
    if doc_unmatched:
        rep.info('UNMATCHED',
                 '未分類の書類らしきファイルがあります。必要書類が非標準名で紛れて'
                 'いないか名前と中身を確認してください', doc_unmatched)

    targets = build_content_targets(detected)

    # ---- 人間可読出力 ----
    print(f'対象: {args.local or args.manifest} ／ タスク: {args.task} ／ '
          f'前提: {args.frame}・{args.entity}')
    print(f'走査 {len(entries)} ファイル（除外フォルダ {len(set(excluded_dirs))}・'
          f'出力物 {len(excluded_outputs)} 件除外）\n')
    print('■ 検出結果')
    for cat in PATTERNS:
        names = [e['name'] for e in detected[cat]]
        if names:
            print(f'  [{CAT_LABELS[cat]}] ' + ' / '.join(names))
    print('\n■ 指摘')
    order = {'FAIL': 0, 'WARN': 1, 'INFO': 2}
    tags = {'FAIL': '[FAIL]', 'WARN': '[要確認]', 'INFO': '[INFO]'}
    for f in sorted(rep.findings, key=lambda x: order[x['level']]):
        files = f' ← {", ".join(f["files"])}' if f['files'] else ''
        print(f'  {tags[f["level"]]} {f["code"]}: {f["message"]}{files}')
    print(f'\n集計: FAIL {rep.count("FAIL")} / 要確認 {rep.count("WARN")} / '
          f'INFO {rep.count("INFO")}')
    print(f'中身チェック対象（CC が Read で確認）: {len(targets)} ファイル')

    if args.json:
        result = {
            'context': {'task': args.task, 'entity': args.entity, 'frame': args.frame,
                        'hearing_form': extra.get('hearing_form'),
                        'bonus_ledgers': extra.get('bonus_ledgers')},
            'detected': {c: [e['name'] for e in v] for c, v in detected.items() if v},
            'skipped': [list(s) for s in skipped],
            'excluded_outputs': excluded_outputs,
            'unmatched': doc_unmatched,
            'findings': rep.findings,
            'content_check_targets': targets,
            'summary': {'fail': rep.count('FAIL'), 'warn': rep.count('WARN'),
                        'info': rep.count('INFO')},
        }
        Path(args.json).write_text(
            json.dumps(result, ensure_ascii=False, indent=1), encoding='utf-8')
        print(f'JSON: {args.json}')

    return 1 if rep.count('FAIL') else 0


if __name__ == '__main__':
    try:
        sys.exit(main())
    except SystemExit:
        raise
    except Exception as ex:  # 沈黙しない（verify_wagebook と同じ自己防衛）
        print(f'ERROR: スクリプト内部エラー: {ex}', file=sys.stderr)
        sys.exit(2)
