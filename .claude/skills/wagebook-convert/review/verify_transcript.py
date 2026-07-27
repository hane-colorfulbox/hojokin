# -*- coding: utf-8 -*-
"""原本転記シート群（SKILL.md §4.1.2）の受け入れ検査。openpyxl のみで動く（API不使用）。

使い方:
    python .claude/skills/wagebook-convert/review/verify_transcript.py "{会社名}_賃金台帳一覧.xlsx"

verify_wagebook.py が「従業員別明細」を検算するのに対し、こちらは「原本転記*」シート群が
**原本の代わりに使える状態か**を検証する。転記シートの用途（確認者が Excel の関数で
「課税支給額 − 通勤手当」等を計算する／加点判定に数値を転用する）が成り立つ条件を機械化した。

  T1 内訳の完全性 : 課税支給額 ＝ 支給項目（基本給〜課税支給額直前）の和 が各列で成立。
                    成立すれば、通勤手当に限らずどの内訳項目でも関数で加減算できる。
                    非課税項目が内訳に混在する台帳では、部分集合一致で課税対象の項目群を特定して報告。
  T2 数値セル     : 課税支給額と内訳が数値型（文字列だと Excel 関数が計算できない）。
  T3 合計の一致   : 原本の「合計」列 ＝ 転記値の SUM が全行一致（読み取り誤りの検出）。
                    原本側の不整合で一致しない行は、合計セルに色を付けておけば別枠で報告される。
  T4 加点転用     : 各月次ページに 基本給 と 労働時間の手掛かり（日額単価／時給／出勤日数／労働時間）がある。
  T5 通勤手当     : 通勤手当が課税支給額に含まれているか（included / excluded）を2式で判定し、
                    従業員別明細の S列と突合する。控除漏れ（included なのに S列が空）と
                    二重控除（excluded なのに S列に値）を検知する（SKILL.md §3.1 の全額控除方式）。
                    どちらの式も成立しない場合は「判定不能」＝控除せず報告に明記させる。

1ページ内に複数ブロック（月次／賞与／総合計）が横並びの形式に対応する（項目名列で区切る）。
T5 は行=項目×列=従業員（月次ページ型）と行=従業員×列=項目（一覧型）の両レイアウトを扱う
（entity_maps で正規化。後者を無言で素通りさせない）。
終了コード: 0=全項目OK / 1=要確認あり / 2=転記シートなし・構造解析不可。
"""
import re
import sys
import unicodedata
from itertools import combinations
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook  # noqa: E402

if len(sys.argv) < 2:
    print('使い方: python verify_transcript.py <出力xlsx>', file=sys.stderr)
    raise SystemExit(2)
BOOK = Path(sys.argv[1])
if not BOOK.exists():
    print(f'❌ ファイルが見つかりません: {BOOK}', file=sys.stderr)
    raise SystemExit(2)
LABEL_HEADS = ('項目', '氏名', 'No.', 'No')
SKIP_HEADS = ('合計', '検算', '項目', '氏名')
TAXABLE = ('課税支給額', '総支給額(課税)', '課税分給与額')
BASE = ('基本給', '賞与額', '基本給区分')
COMMUTE = ('通勤手当', '通勤費', '交通費', '通勤')
NONTAX_COMMUTE = ('非課税通勤',)      # 「非課税通勤費」等＝通勤の非課税分が単独で印字される形式
TAXED_COMMUTE = ('課税通勤',)         # 「課税通勤費」等
# 従業員別明細（同じブック内）の S列＝年間通勤手当。T5 でこの値と転記シートの判定を突合する。
LEDGER_SHEET, LEDGER_HDR_ROW = '従業員別明細', 5
LEDGER_COL_NAME, LEDGER_COL_S = 3, 19
# 支給項目の和（T1・T5）から除外するラベル。**完全一致**で判定する
# （前方一致にすると「休日出勤」が「休日出勤手当」を巻き添えで除外して和が崩れる）。
# 単価行（金額でなく単価）も除外対象。時給・日額の単価が支給欄に並ぶ台帳がある。
NON_PAY = ('項目', '氏名', '扶養家族', '出勤日数', '休日出勤', '有給休暇', '普通残業', '休日残業',
           '欠勤日数', '遅刻早退', '期間', '所属', '日額単価', 'No', '支給日', '出勤時間数',
           '欠勤', '不就労', '平日普通', '平日深夜', '休日普通', '休日深夜', '法定休日普通',
           '法定休日深夜', '内60時間超過', '内45時間超過', '有休日数', '有休残日数', '性別',
           '生年月日', '入社日', '退社日', '月額表区分', '役職',
           '時給単価', '時間単価', '単価')
# 単価の行（加点判定の時間換算給与の分子）。時給者・日給者でラベルが違う
RATE_HINTS = ('日額単価', '時給', '時間単価', '単価')
# 加点判定に転用できるか（T4）＝分子（単価/基本給）と分母（時間）の手掛かりがあるか。前方一致。
TIME_HINTS = RATE_HINTS + ('出勤日数', '出勤時間数', '労働時間', '総労働時間', '実労働時間',
                           '所定労働時間', '就業時間', '稼働時間', '時間数', '平日普通')

results, notes = [], []


def check(rid, label, cond, detail=''):
    results.append((rid, cond))
    print(f"  {'OK  ' if cond else 'NG  '} [{rid}] {label} {detail}")


def norm(v):
    return unicodedata.normalize('NFKC', str(v or '')).replace(' ', '').replace('　', '')


def find_header_row(ws):
    """右に2つ以上の見出しが並ぶ行を、この転記シートのヘッダー行とみなす。"""
    best, best_n = None, 0
    for r in range(1, min(ws.max_row, 14) + 1):
        if norm(ws.cell(r, 1).value) not in LABEL_HEADS:
            continue
        n = sum(1 for c in range(2, ws.max_column + 1) if ws.cell(r, c).value not in (None, ''))
        if n > best_n:
            best, best_n = r, n
    return best


def blocks_of(ws, hdr):
    """項目名列（'項目'/'氏名'）で区切ってブロックに分ける。"""
    label_cols = [c for c in range(1, ws.max_column + 1)
                  if norm(ws.cell(hdr, c).value) in LABEL_HEADS]
    out = []
    for i, lc in enumerate(label_cols):
        end = label_cols[i + 1] - 1 if i + 1 < len(label_cols) else ws.max_column
        data, total = [], None
        for c in range(lc + 1, end + 1):
            h = norm(ws.cell(hdr, c).value)
            if not h:
                continue
            if h.startswith('合計') and total is None:
                total = c
            elif not h.startswith(SKIP_HEADS):
                data.append(c)
        rows = {}
        for r in range(hdr, ws.max_row + 1):
            lab = norm(ws.cell(r, lc).value)
            if lab and lab not in LABEL_HEADS:
                rows.setdefault(lab, r)
        if data:
            out.append({'label_col': lc, 'data': data, 'total': total, 'rows': rows,
                        'heads': [norm(ws.cell(hdr, c).value) for c in data]})
    return out


def is_transposed(blk):
    """行=従業員 × 列=項目 の形式か（一覧型ページ）。列見出しに項目名が並ぶ。"""
    heads = blk['heads']
    return (any(h.startswith(BASE) for h in heads)
            and any(h.startswith(TAXABLE) for h in heads))


def entity_maps(ws, p, blk):
    """ブロックを [(氏名, {項目ラベル: Cell})] に正規化する。

    形式①（行=項目・列=従業員）と形式②（行=従業員・列=項目）を同じ形にして、
    どちらでも同一ロジックで判定できるようにする（②を素通りさせない）。
    dict は挿入順＝原本の印字順を保つので、支給項目の範囲を順序で決められる。
    Cell を返すので、値だけでなくセル位置（coordinate）も報告に使える。
    """
    if is_transposed(blk):
        cols = [(norm(ws.cell(p['hdr'], c).value), c) for c in blk['data']]
        return [(name, {lab: ws.cell(r, c) for lab, c in cols})
                for name, r in blk['rows'].items()]
    out = []
    for c in blk['data']:
        name = norm(ws.cell(p['name_row'], c).value) or norm(ws.cell(p['hdr'], c).value)
        out.append((name, {lab: ws.cell(r, c) for lab, r in blk['rows'].items()}))
    return out


def numof(v):
    return v if isinstance(v, (int, float)) else None


def item_sum(cells):
    """支給項目（基本給〜課税支給額の直前）と課税支給額を返す。

    returns ([(ラベル, Cell)] または None, 課税支給額の値 または None)
    順序は原本の印字順（entity_maps が挿入順を保つ）。
    """
    keys = list(cells)
    i_base = next((i for i, k in enumerate(keys) if k.startswith(BASE)), None)
    i_tax = next((i for i, k in enumerate(keys) if k.startswith(TAXABLE)), None)
    tax = numof(cells[keys[i_tax]].value) if i_tax is not None else None
    if i_base is None or i_tax is None or i_tax <= i_base:
        return None, tax
    items = [(k, cells[k]) for k in keys[i_base:i_tax]
             if k not in NON_PAY and not k.startswith('日額単価')]
    return items, tax


def sum_verdict(cells):
    """T1 の判定。returns (verdict, detail)。

    ok      : 課税支給額 ＝ 支給項目の和
    partial : 非課税項目が内訳に混在し、部分集合が課税支給額に一致（課税対象を特定できた）
    ng      : どの部分集合とも一致しない＝転記漏れ・読み取り誤りの疑い
    noitem  : 課税支給額はあるが支給項目が無い（集計行のみの転記の疑い）
    skip    : 課税支給額が数値でない（この列/行は対象外）
    """
    items, tax = item_sum(cells)
    if tax is None:
        return 'skip', ''
    if items is None:
        return 'noitem', ''
    vals = [(k, c.value) for k, c in items if numof(c.value)]
    s = sum(v for _, v in vals)
    if abs(s - tax) < 0.5:
        return 'ok', ''
    if len(vals) <= 14:
        for k in range(len(vals), 0, -1):
            for comb in combinations(range(len(vals)), k):
                if abs(sum(vals[i][1] for i in comb) - tax) < 0.5:
                    excl = [vals[i][0] for i in range(len(vals)) if i not in comb]
                    return 'partial', f'課税対象外＝{excl}'
    return 'ng', f'内訳和{s:,.0f} vs 課税{tax:,.0f}'


def commute_state(cells):
    """通勤手当が課税支給額に含まれているかを2式で判定する（SKILL.md §3.1 の判定表）。

    included : 課税支給額 ＝ Σ支給項目（通勤手当を含む）→ S列に通勤手当の年額を入れて控除する
    excluded : 課税支給額 ＝ Σ支給項目 − 通勤手当        → 既に除外済み。S列は空が正
    unknown  : どちらの式も成立しない（内訳が読めない）  → 控除せず §9 に「未控除」を明記
    none     : 通勤手当の印字がない（＝支給なし）

    T1 の部分集合探索は複数解があり得るのでここでは使わない。この2式だけで決める。
    """
    keys = list(cells)
    tc = numof(next((cells[k].value for k in keys if k.startswith(TAXED_COMMUTE)), None))
    nc = numof(next((cells[k].value for k in keys if k.startswith(NONTAX_COMMUTE)), None))
    if tc or nc:
        # 課税通勤費／非課税通勤費が別行で印字される形式は、その値でそのまま決まる
        return ('included', tc) if tc else ('excluded', nc)

    com_k = next((k for k in keys if k.startswith(COMMUTE)
                  and not k.startswith(NONTAX_COMMUTE + TAXED_COMMUTE)), None)
    com = numof(cells[com_k].value) if com_k else None
    if not (com and com > 0):
        return 'none', 0.0

    items, tax = item_sum(cells)
    if items is None or tax is None:
        return 'unknown', com
    s = 0.0
    for _, c in items:
        v = numof(c.value)
        if v:
            s += v
    if abs(s - tax) < 0.5:
        return 'included', com
    if abs((s - com) - tax) < 0.5:
        return 'excluded', com
    return 'unknown', com


def ledger_transport(wb):
    """従業員別明細の S列（年間通勤手当）を氏名キーで読む。シートが無ければ None。"""
    if LEDGER_SHEET not in wb.sheetnames:
        return None
    ws = wb[LEDGER_SHEET]
    out = {}
    for r in range(LEDGER_HDR_ROW + 1, ws.max_row + 1):
        nm = norm(ws.cell(r, LEDGER_COL_NAME).value)
        if nm:
            v = ws.cell(r, LEDGER_COL_S).value
            out[nm] = v if isinstance(v, (int, float)) else None
    return out


wb = load_workbook(BOOK, data_only=False)
# 「原本転記」で始まる全シートを候補にする。表紙（ヘッダー行を持たないシート）は
# find_header_row が None を返すので自動的に対象外になる（旧仕様の1枚構成にも対応）。
sheets = [s for s in wb.sheetnames if s.startswith('原本転記')]
print(f'ブック: {BOOK.name}\n転記データシート: {len(sheets)}枚\n')

parsed = {}
for sn in sheets:
    ws = wb[sn]
    hdr = find_header_row(ws)
    if hdr:
        name_row = next((r for r in range(1, min(ws.max_row, 14) + 1)
                         if norm(ws.cell(r, 1).value).startswith('氏名')), None)
        parsed[sn] = {'ws': ws, 'hdr': hdr, 'name_row': name_row or hdr,
                      'blocks': blocks_of(ws, hdr)}
    else:
        notes.append(sn)   # 表紙シート等（データ行を持たない）
if not sheets:
    print('❌ 「原本転記*」シートが無い。§4.1.2 の転記シート群を作ってから実行する'
          '（既存JSON流用でPDF原本が無い場合は §9 に省略理由を明記）', file=sys.stderr)
    raise SystemExit(2)
if not parsed:
    print('❌ 転記シートの構造を解析できない（ヘッダー行に「項目」or「氏名」が必要）。'
          '§4.1.2 のレイアウトを確認する', file=sys.stderr)
    raise SystemExit(2)
nblocks = sum(len(v['blocks']) for v in parsed.values())
print(f'解析: {len(parsed)}シート / {nblocks}ブロック'
      + (f' / 未解析 {notes}' if notes else '') + '\n')

print('=== T1 内訳の完全性（課税支給額 ＝ 支給項目の和） ===')
ok1, ng1, partial, noitem = 0, [], [], []
for sn, p in parsed.items():
    for bi, blk in enumerate(p['blocks']):
        # 賞与欄・総合計欄は支給項目の内訳を持たないのが正常（集計のみでも違反にしない）
        is_summary_block = blk['heads'] and all(
            h.startswith(('賞与', '総合計')) for h in blk['heads'])
        bkey, seen = f'{sn}#b{bi}', 0
        for name, cells in entity_maps(p['ws'], p, blk):
            verdict, detail = sum_verdict(cells)
            if verdict == 'skip':
                continue
            seen += 1
            if verdict == 'ok':
                ok1 += 1
            elif verdict == 'partial':
                partial.append(f'{sn}/{name}: {detail}')
                ok1 += 1
            elif verdict == 'noitem':
                # 課税支給額はあるのに内訳が1つもない＝集計行だけの転記（§4.1.2 🔴 違反）
                if is_summary_block:
                    if bkey not in noitem:
                        noitem.append(bkey)
                else:
                    ng1.append(f'{sn}/{name}: 課税支給額はあるが支給項目（基本給・各手当）が'
                               '無い＝集計行のみの転記の疑い')
            else:
                ng1.append(f'{sn}/{name}: {detail}')
        if not seen and bkey not in noitem:
            noitem.append(bkey)
if partial:
    uniq = sorted({x.split(': ', 1)[1] for x in partial})
    print(f'    非課税項目が内訳に混在: {len(partial)}件（部分集合一致で課税対象を特定）→ {uniq[:3]}')
if noitem:
    print(f'    支給項目行を持たないブロック（総合計欄など）: {len(noitem)}件 {noitem[:6]}')
check('T1', f'課税支給額＝内訳の和が全列で成立（{ok1}列）', not ng1, '; '.join(ng1[:4]))

print('=== T2 関数計算の可能性（数値セルであること） ===')
bad2 = []
for sn, p in parsed.items():
    for blk in p['blocks']:
        for name, cells in entity_maps(p['ws'], p, blk):
            items, _tax = item_sum(cells)
            if items is None:
                continue
            tax_k = next((k for k in cells if k.startswith(TAXABLE)), None)
            targets = ([(tax_k, cells[tax_k])] if tax_k else []) + items
            for lab, c in targets:
                if c.value is not None and not isinstance(c.value, (int, float)):
                    bad2.append(f'{sn}!{c.coordinate}({lab})={c.value!r}')
check('T2', '課税支給額と内訳がすべて数値セル', not bad2, '; '.join(bad2[:4]))

print('=== T3 原本の網羅性（合計列 ＝ 転記値のSUM） ===')
ok3, ng3, ng3_marked, nototal = 0, [], [], []
for sn, p in parsed.items():
    ws = p['ws']
    for bi, blk in enumerate(p['blocks']):
        if not blk['total']:
            nototal.append(f'{sn}#b{bi}')
            continue
        for r in range(p['hdr'], ws.max_row + 1):
            printed = ws.cell(r, blk['total']).value
            if not isinstance(printed, (int, float)):
                continue
            s = sum(ws.cell(r, c).value for c in blk['data']
                    if isinstance(ws.cell(r, c).value, (int, float)))
            if abs(s - printed) < 0.5:
                ok3 += 1
            else:
                lab = norm(ws.cell(r, blk['label_col']).value)
                cellfill = ws.cell(r, blk['total']).fill
                marked = bool(cellfill and cellfill.fill_type and
                              (cellfill.fgColor.rgb or '') not in ('00000000', None))
                msg = f'{sn} r{r}({lab}): SUM{s:,.0f} vs 印字{printed:,.0f}'
                (ng3_marked if marked else ng3).append(msg)
if nototal:
    print(f'    合計列を持たないブロック: {len(nototal)}件 {nototal[:6]}')
if ng3_marked:
    print(f'    原本側の不整合として黄色マーク済み: {len(ng3_marked)}件 {ng3_marked[:3]}')
check('T3', f'原本の合計とSUMが全行一致（{ok3}行・マーク済みの原本不整合は除く）',
      not ng3, '; '.join(ng3[:4]))

print('=== T4 加点判定への転用（基本給＋労働時間の手掛かり） ===')
ng4, rate_pages = [], []
for sn, p in parsed.items():
    monthly = [blk for blk in p['blocks']
               if blk['heads'] and not all(h.startswith(('賞与', '総合計')) for h in blk['heads'])]
    if not monthly:
        continue          # 賞与/総合計だけのページは月次の勤怠を持たない
    for blk in monthly:
        # 形式②（行=従業員・列=項目）では項目名が列見出し側にある
        labels = blk['heads'] if is_transposed(blk) else list(blk['rows'])
        has_base = any(h.startswith(BASE) for h in labels)
        has_time = any(h.startswith(TIME_HINTS) for h in labels)
        if not (has_base and has_time):
            ng4.append(f'{sn}: 基本給={has_base} 時間手掛かり={has_time}')
    rr = next((v for blk in monthly if not is_transposed(blk)
               for k, v in blk['rows'].items() if k.startswith(RATE_HINTS)), None)
    if rr:
        n = sum(1 for blk in monthly for c in blk['data']
                if isinstance(p['ws'].cell(rr, c).value, (int, float)))
        rate_pages.append((sn, n))
check('T4', f'月次ページすべてで基本給＋労働時間の手掛かりあり', not ng4, '; '.join(ng4[:4]))
print(f'    単価（日額/時給）が転記されたページ: {len(rate_pages)}/{len(parsed)}'
      + (f' {rate_pages}' if rate_pages else '（この台帳形式には単価の印字なし）'))

print('=== T5 通勤手当のS列整合（控除漏れ・二重控除の検知。§3.1） ===')
states, amounts = {}, {}
anon = {'included': 0, 'excluded': 0, 'unknown': 0}
for sn, p in parsed.items():
    for blk in p['blocks']:
        for name, m in entity_maps(p['ws'], p, blk):
            st, amt = commute_state(m)
            if st == 'none':
                continue
            if not name or name in LABEL_HEADS:
                anon[st] += 1          # 氏名が取れないブロック（No.見出しだけ等）
                continue
            states.setdefault(name, set()).add(st)
            if amt:
                amounts.setdefault(name, []).append(amt)

s_col = ledger_transport(wb)
missing, doubled, unknown_names, mixed, ok5 = [], [], [], [], 0
for name, sts in sorted(states.items()):
    if 'included' in sts and 'excluded' in sts:
        mixed.append(name)
        continue
    st = 'included' if 'included' in sts else ('excluded' if 'excluded' in sts else 'unknown')
    if st == 'unknown':
        unknown_names.append(name)
        continue
    if s_col is None:
        ok5 += 1
        continue
    sv = s_col.get(name)
    if st == 'included' and not sv:
        missing.append((name, sum(amounts.get(name, []))))
    elif st == 'excluded' and sv:
        doubled.append((name, sv))
    else:
        ok5 += 1

n_inc = sum(1 for s in states.values() if s == {'included'})
n_exc = sum(1 for s in states.values() if s == {'excluded'})
if not states and not any(anon.values()):
    print('    通勤手当の印字なし＝支給なし（控除不要）。§9 には「対象なし（支給なし）」と書く')
else:
    print(f'    判定: 課税支給額に含まれる {n_inc}名 / 非課税処理済み {n_exc}名 / '
          f'判定不能 {len(unknown_names)}名 / 混在 {len(mixed)}名'
          + (f' / 氏名が取れないブロック {anon}' if any(anon.values()) else ''))
if s_col is None and states:
    print(f'    ⚠ 同じブックに「{LEDGER_SHEET}」シートが無いため S列との突合はスキップした')
if missing:
    print(f'    🔴 [控除漏れ] {len(missing)}名 — 課税支給額に通勤手当が含まれているのに S列が空:')
    for nm, tot in missing[:10]:
        print(f'      {nm}: 転記シート上の通勤手当合計 {tot:,.0f} 円')
    print('    → 2026-07-27 運用では課税扱いの通勤手当も控除する（§3.1）。転記シートは対象年度外の'
          'ページも含むので、S列には**対象12ヶ月分だけ**を集計した年額を入れる')
if doubled:
    print(f'    🔴 [二重控除] {len(doubled)}名 — 非課税処理済みなのに S列に値がある:')
    for nm, sv in doubled[:10]:
        print(f'      {nm}: S列={sv:,.0f} 円 → 空欄にする')
    print('    → 課税支給合計にはそもそも通勤手当が入っていない。S列を埋めると引かれるのは'
          '基本給・残業手当で、R216 が過小になる')
if mixed:
    print(f'    ⚠ [判定が混在] {mixed[:10]} — ページによって課税/非課税処理が違う。'
          '対象事業年度のページの処理に合わせ、根拠を §9 に書く')
if unknown_names:
    print(f'    ℹ [判定不能] {len(unknown_names)}名 {unknown_names[:8]} — 内訳から課税支給額との関係が'
          '決まらない（諸手当に合算・項目の取り違え等）。**控除せず** §9 に「未控除（内訳不明）」と'
          '理由・人数を書く（顧客に内訳を催促しない。§3.1.1）')
check('T5', f'S列が §3.1 の判定表と整合（整合 {ok5}名）',
      not (missing or doubled or mixed),
      f'控除漏れ{len(missing)}名 / 二重控除{len(doubled)}名 / 混在{len(mixed)}名')

nfail = sum(1 for _, ok in results if not ok)
print()
if nfail:
    print(f'=== 要確認 {nfail}件 ===（NGはPDF原本で確認し、§9 報告に判断を明記してから完了報告へ。'
          '原本側の不整合なら該当セルを黄色にして注記する）')
else:
    print('=== PASS ===（T1〜T5 すべて成立。ただし転記の「真」はPDF原本なので '
          '§5.5-3 の目視突合は別途必要）')
sys.exit(1 if nfail else 0)
