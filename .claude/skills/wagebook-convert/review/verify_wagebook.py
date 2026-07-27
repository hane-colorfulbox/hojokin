"""賃金台帳変換の出力 xlsx を機械検算する『必須レビューゲート』(SKILL.md §5.5)。

設計方針（重要）:
- **openpyxl だけ**に依存する。hojokin パッケージ・Anthropic API・Streamlit を一切呼ばない
  → 配布先の他人PC（まっさら環境）・headless でも動く。課金ゼロ。
- **R215/R216 を再計算しない**。ツール側 hojokin/wage_calculator.py の算定式をここで再実装すると
  式がドリフトして「誤った PASS」を出す（CLAUDE.md 同期負債）。本スクリプトは
  **不変量の観測と FLAG 列挙だけ**を行う。R216 の正値判定は、リポジトリ/Streamlit 側の
  本物の read_wage_ledger 往復（最終ゲート）に委ねる。
- 自己防衛: シート名・列レイアウトがテンプレと違えば沈黙せず非ゼロ終了する。

使い方:
    python .claude/skills/wagebook-convert/review/verify_wagebook.py "{会社名}_賃金台帳一覧.xlsx"

出力（=§9 完了報告にそのまま貼る）:
    各行の Σ(G:R)+T、月別縦計、雇用形態内訳、各 FLAG、末尾に PASS/要対応 件数。
終了コード: 0=要対応なし / 1=要対応あり（FAIL or 要確認が残っている） / 2=ファイル/構造エラー。
"""
import re
import sys
import unicodedata
from collections import Counter
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

SHEET_NAME = '従業員別明細'
HEADER_ROW = 5
DATA_START_ROW = 6
COL_NAME = 3        # C
COL_TYPE = 4        # D 雇用形態
COL_HOURS = 5       # E 月平均時間
COL_HOURLY = 6      # F 時給
COL_MONTH_FIRST = 7   # G (1月)
COL_MONTH_LAST = 18   # R (12月)
COL_TRANSPORT = 19    # S 年間通勤手当
COL_BONUS = 20        # T 年間賞与
MONTH_LABELS = [f'{m}月' for m in range(1, 13)]

# 役員ラベル未正規化の検知トリガ（職位名「現場代理人/主任/営業」等は含めない＝誤検知回避）
OFFICER_TRIGGERS = ('代表取締役', '取締役', '監査役', '監事', '理事長', '理事')
PART_TRIGGERS = ('パート', 'アルバイト', '非常勤')
BONUS_SPIKE_RATIO = 1.5  # その人の非空月の中央値×この倍率を超える月は賞与混入の疑い

# 出力ブックに存在してよい可視シート（§4.1.2）。これ以外の可視シートはツールが表として読む＝FAIL。
ALLOWED_SHEET_NAMES = (SHEET_NAME, '記入ルール')
ALLOWED_SHEET_PREFIXES = ('原本転記', '変換メモ')
# FLAG-R（§4.1.2 🔴 全項目転記）: 原本転記シート群に明細行の項目名がこの数以上見つからなければ
# 「課税支給額など集計行だけの転記」の疑い。手掛かり語は給与台帳の典型的な明細項目。
TRANSCRIPT_ITEM_HINTS = ('基本給', '通勤手当', '所得税', '住民税', '健康保険', '厚生年金',
                         '雇用保険', '社会保険', '控除', '残業', '時間外', '出勤', '労働時間', '差引')
TRANSCRIPT_MIN_HINTS = 3


def _num(v):
    """数値化できれば float、できなければ None。空文字・None は None。"""
    if v is None or v == '':
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _median(xs):
    s = sorted(xs)
    n = len(s)
    if n == 0:
        return None
    mid = n // 2
    return s[mid] if n % 2 else (s[mid - 1] + s[mid]) / 2


def _extract_fiscal_month(ws):
    """ヘッダー注記（B1〜B4 等）から「（N月決算）」を抽出。見つからなければ None。
    SKILL.md §4 で B1/B3 に『…（N月決算）』を必ず書く規約。FLAG-T の発火判定に使う。"""
    texts = []
    for r in range(1, HEADER_ROW):       # 1〜4 行目
        for c in range(2, 5):            # B, C, D 列
            v = ws.cell(r, c).value
            if v:
                texts.append(str(v))
    m = re.search(r'(\d{1,2})\s*月決算', ' '.join(texts))
    if m:
        mm = int(m.group(1))
        if 1 <= mm <= 12:
            return mm
    return None


def _header_is_clean_month(text) -> bool:
    """月見出しが『ツールが確実に読める規格どおりの形』か判定（FLAG-H 用）。
    許容: 'N月'(注記なし) / '令和N年M月' / '平成N年M月' / 'RN.M' / 'YYYY年M月' / 'YYYY/MM'。
    '1月(R7.1)' のような注記付きは False（末尾に余計な文字＝規格外）。
    ※ ツール側 parse_ym_header は注記を許容するよう堅牢化済みだが、本ゲートは
      『規格どおりプレーン/和暦か』を強制し、旧ツール・他リーダーでの 0名誤読を pre-ship で防ぐ。"""
    if text is None:
        return False
    s = unicodedata.normalize('NFKC', str(text))
    s = re.sub(r'[\s　]+', '', s)
    if not s:
        return False
    if re.fullmatch(r'令和\d+年\d{1,2}月', s):
        return True
    if re.fullmatch(r'平成\d+年\d{1,2}月', s):
        return True
    if re.fullmatch(r'[RH]\d+[./-]\d{1,2}月?', s, re.IGNORECASE):
        return True
    if re.fullmatch(r'\d{4}[年/\-.]\d{1,2}月?', s):
        return True
    if re.fullmatch(r'\d{1,2}月', s):
        return True
    return False


def main(argv):
    if len(argv) < 2:
        print('使い方: python verify_wagebook.py <出力xlsx>', file=sys.stderr)
        return 2
    xlsx = Path(argv[1])
    if not xlsx.exists():
        print(f'❌ ファイルが見つかりません: {xlsx}', file=sys.stderr)
        return 2
    try:
        from openpyxl import load_workbook
    except ImportError:
        print('❌ openpyxl が未導入。機械検算をスキップし、SKILL.md §5-A〜§5-M を'
              '全件人手で実施すること（このゲートは飛ばさない）。', file=sys.stderr)
        return 2

    try:
        wb = load_workbook(xlsx, data_only=True)
    except Exception as e:  # noqa: BLE001 沈黙させず原因を出す
        print(f'❌ xlsx を開けません: {e}', file=sys.stderr)
        return 2
    if SHEET_NAME not in wb.sheetnames:
        print(f'❌ シート「{SHEET_NAME}」が無い（シート: {wb.sheetnames}）。'
              'ツール出力 Sheet2 等を誤って渡していないか確認。', file=sys.stderr)
        return 2
    ws = wb[SHEET_NAME]
    header = str(ws.cell(HEADER_ROW, COL_NAME).value or '')
    if '氏名' not in header:
        print(f'❌ {HEADER_ROW}行目C列が「氏名」でない（"{header}"）。テンプレと列レイアウトが'
              '違う可能性。ヘッダー行/データ開始行を確認。', file=sys.stderr)
        return 2

    # 原本転記シート（SKILL.md §4.1.2・2026-07-21 決定）の有無は表示のみの注意喚起。
    # 旧スキル出力・ツール出力には無いため FAIL/要対応にはしない（advisory）。
    if not any(str(t or '').strip().startswith('原本転記') for t in wb.sheetnames):
        print('⚠ 原本転記シートなし（2026-07-21 以降のスキル出力では §4.1.2 で必須。'
              '旧スキル出力・ツール出力ならこのままで可。省略理由を §9 報告に明記）\n')

    rows = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        name = ws.cell(r, COL_NAME).value
        if name is None or str(name).strip() == '':
            continue
        months = [_num(ws.cell(r, c).value) for c in range(COL_MONTH_FIRST, COL_MONTH_LAST + 1)]
        rows.append({
            'row': r,
            'name': str(name).strip(),
            'type': str(ws.cell(r, COL_TYPE).value or '').strip(),
            'hours': _num(ws.cell(r, COL_HOURS).value),
            'hourly': _num(ws.cell(r, COL_HOURLY).value),
            'months': months,
            'S': _num(ws.cell(r, COL_TRANSPORT).value),
            'T': _num(ws.cell(r, COL_BONUS).value),
        })

    if not rows:
        print('❌ データ行（B6以降）が空。転記されていない。', file=sys.stderr)
        return 2

    print(f'=== レビューゲート機械検算: {xlsx.name}（{len(rows)}名） ===\n')

    # --- 各行の年間（課税月計＋賞与） ---
    print('[各行 Σ(G:R)+T]（PDF各社員の課税年間合計と人手照合する材料）')
    for e in rows:
        gr = sum(v for v in e['months'] if v is not None)
        t = e['T'] or 0
        print(f"  行{e['row']:>2} {e['name']}: Σ(G:R)={gr:,.0f} +T={t:,.0f} → 計={gr + t:,.0f}")
    print()

    fail = []       # 修正必須（PASSをブロック）
    confirm = []    # 要確認（PDFで素性確認し報告に解決を明記してから進む）

    # --- 規格外シート名（FAIL: ツールのリーダーは「原本転記*」以外の可視シートを表として読む） ---
    irregular_sheets = []
    for t in wb.sheetnames:
        title = str(t or '').strip()
        if wb[t].sheet_state != 'visible':
            continue
        if title in ALLOWED_SHEET_NAMES:
            continue
        if any(title.startswith(p) for p in ALLOWED_SHEET_PREFIXES):
            continue
        irregular_sheets.append(t)
    if irregular_sheets:
        print('[FAIL 規格外シート名]（ツールは「原本転記*」以外の可視シートを賃金表として読む。'
              '独自名の転記シートは R215/R216 を壊す）')
        for t in irregular_sheets:
            print(f'  「{t}」 → 転記シートなら「原本転記_...」へ改名（§4.1.2 🔴）。'
                  '作業用シートなら削除 or 非表示に')
        fail.append(f'規格外シート名 {len(irregular_sheets)}件')
        print()

    # --- FLAG-R 集計行のみ転記の疑い（§4.1.2 🔴 全項目転記） ---
    transcript_sheets = [t for t in wb.sheetnames
                         if str(t or '').strip().startswith('原本転記')]
    if transcript_sheets:
        found_hints = set()
        for t in transcript_sheets:
            for row_vals in wb[t].iter_rows(values_only=True):
                for v in row_vals:
                    if isinstance(v, str):
                        for h in TRANSCRIPT_ITEM_HINTS:
                            if h in v:
                                found_hints.add(h)
                if len(found_hints) >= TRANSCRIPT_MIN_HINTS:
                    break
            if len(found_hints) >= TRANSCRIPT_MIN_HINTS:
                break
        if len(found_hints) < TRANSCRIPT_MIN_HINTS:
            print(f'[FLAG-R 原本転記が集計行のみの疑い]（明細項目名の検出 {len(found_hints)}種'
                  f'（{sorted(found_hints)}）< {TRANSCRIPT_MIN_HINTS}種）')
            print('  → §4.1.2 🔴 は全項目転記（基本給・各手当（通勤手当）・控除・勤怠の明細行まで'
                  '原本どおり）。「課税支給額」等の集計行だけの転記は仕様違反。PDFの印字行と比べて'
                  '不足行を転記すること。原本に明細行が本当に無い台帳ならその旨を§9報告に明記。')
            confirm.append('FLAG-R 原本転記が集計行のみの疑い（全項目転記か確認）')
            print()

    # --- ① 通勤手当S列（2026-07-27 運用: 月列に含まれる通勤手当を控除する）---
    # 月列の素性（課税支給合計か通勤費込み合計か）と通勤手当の課税区分は、従業員別明細だけを
    # 見るこのスクリプトでは決められない。転記シートがあれば verify_transcript.py の T5 が
    # 機械判定する（§5.5-2b）。ここでは値の有無を提示して報告への明記を促す。
    s_rows = [e for e in rows if e['S'] is not None and e['S'] > 0]
    print(f'[FLAG-S 通勤手当S列]（値あり {len(s_rows)}行 / 空 {len(rows) - len(s_rows)}行）')
    for e in s_rows:
        print(f"  行{e['row']} {e['name']}: S={e['S']:,.0f}")
    print('  → S列に入れるのは「月列（G〜R）の値に**含まれている**通勤手当の年額」（§3.1 の判定表）:')
    print('     ・月列=課税支給合計 かつ 通勤手当が課税支給額に含まれる → S列に年額（控除する）')
    print('     ・月列=課税支給合計 かつ 通勤手当が非課税処理済み       → S列は空（埋めると基本給を削る）')
    print('     ・月列=「合計（通勤費込み）」                          → S列に通勤手当の全額')
    if s_rows:
        confirm.append(f'FLAG-S {len(s_rows)}件（S列の値が §3.1 判定表どおりか確認）')
    else:
        print('  → S列が全行空。**「支給なし」「未控除（内訳不明）」「判定して空が正」のどれなのかを'
              '§9 報告に1行書く**（無言の未控除を許さない。§3.1.1）')
        confirm.append('FLAG-S S列が全行空（支給なし/未控除/空が正 のどれかを報告に明記）')
    print()

    # --- ② 月配置ズレ（観測材料のみ。確定はPDF突合＝§5.5-3で人手必須） ---
    monthly_totals = [
        sum(e['months'][m] for e in rows if e['months'][m] is not None)
        for m in range(12)
    ]
    print('[月別縦計]（全員の各月合計。バグ②月配置ズレ検出の“材料”）')
    print('  ' + ' / '.join(f'{lab}={tot:,.0f}' for lab, tot in zip(MONTH_LABELS, monthly_totals)))
    gap_rows = [
        e for e in rows
        if any(e['months'][m] is None for m in range(12))
        and not all(e['months'][m] is None for m in range(12))
    ]
    if gap_rows:
        print(f'  ⚠ 途中に空欄月のある社員 {len(gap_rows)}名 → §5.5-3 でこの全員＋無作為3名を'
              'PDFと1セル突合（縦計が出た≠OK。左詰めズレは年間合計検算を素通りする §1.1.0）:')
        for e in gap_rows:
            present = [MONTH_LABELS[m] for m in range(12) if e['months'][m] is not None]
            print(f"    行{e['row']} {e['name']}: 在={','.join(present)}")
        confirm.append(f'月配置: 空欄月あり{len(gap_rows)}名＋無作為3名をPDF1セル突合')
    print()

    # --- ③ 賞与の月セル混入（要確認: 賞与か繁忙期/歩合か人手判断） ---
    print('[FLAG-B 賞与の月セル混入候補]（他月比1.5倍超の突出月）')
    spike_found = False
    for e in rows:
        vals = [v for v in e['months'] if v is not None]
        med = _median(vals)
        if med and med > 0:
            spikes = [(MONTH_LABELS[m], e['months'][m]) for m in range(12)
                      if e['months'][m] is not None and e['months'][m] > med * BONUS_SPIKE_RATIO]
            if spikes:
                spike_found = True
                detail = ', '.join(f'{lab}={v:,.0f}' for lab, v in spikes)
                print(f"  行{e['row']} {e['name']}: {detail}（中央値{med:,.0f}）")
    if spike_found:
        print('  → 賞与なら T列へ分離（§4.1.1）。繁忙期残業/歩合なら月セルのままで正。PDF確認。')
        confirm.append('賞与混入候補あり（T列分離 or 残業/歩合かをPDF確認）')
    else:
        print('  なし')
    print()

    # --- 賞与の年度帰属（非暦年決算ガード §4.1.1-a）FLAG-T ---
    bonus_rows = [e for e in rows if e['T'] is not None and e['T'] > 0]
    if bonus_rows:
        fiscal_month = _extract_fiscal_month(ws)
        print('[FLAG-T 賞与の年度帰属]（T列に年間賞与がある行。非暦年決算の窓内/窓外確定）')
        if fiscal_month is not None and fiscal_month != 12:
            print(f'  決算月={fiscal_month}月（非暦年）／T列あり {len(bonus_rows)}名。'
                  '台帳に賞与の支給月が無ければ回数→月を推測確定しないこと。')
            print('  → 暫定なら B2＋完了報告に「含めた回/除外した回」＋要確認（支給年月 or '
                  '勘定科目内訳明細書＝法人R216の正式ソース）を明記（§4.1.1-a）。')
            confirm.append(
                f'賞与年度帰属: 決算月{fiscal_month}月（非暦年）でT列あり'
                '→支給月で確定 or 暫定＋要顧客確認を明記（§4.1.1-a）'
            )
        elif fiscal_month is None:
            print(f'  ⚠ 決算月をヘッダー注記から検出できず（B1/B3 に「（N月決算）」が無い）／'
                  f'T列あり {len(bonus_rows)}名。')
            print('  → 決算月が12月以外なら賞与支給月ベースで窓内か要確認（§4.1.1-a）。'
                  'B1/B3 に「（N月決算）」を入れると本チェックが自動で効く（advisory・exit影響なし）。')
        else:  # fiscal_month == 12
            print('  決算月=12月（事業年度＝暦年）。令和N年の賞与＝当年度で帰属ズレなし（ガード対象外）。')
        print()

    # --- ④-a 役員ラベル未正規化（FAIL） ---
    officer_residue = [
        e for e in rows
        if any(t in e['type'] for t in OFFICER_TRIGGERS) and '役員' not in e['type']
    ]
    if officer_residue:
        print('[FAIL 役員ラベル未正規化]（R216に役員混入。§2で「役員」へ正規化必須）')
        for e in officer_residue:
            print(f"  行{e['row']} {e['name']}: 雇用形態=「{e['type']}」→「役員」へ")
        fail.append(f'役員ラベル未正規化 {len(officer_residue)}件')
        print()

    # --- ④-b パートでE列(月平均時間)空（FAIL: FTE=1.0サイレント昇格でR215過大） ---
    part_no_hours = [
        e for e in rows
        if any(t in e['type'] for t in PART_TRIGGERS) and e['hours'] is None
    ]
    if part_no_hours:
        print('[FAIL パートでE列(月平均時間)空]（FTE=1.0昇格でR215過大。§1 E列必須）')
        for e in part_no_hours:
            print(f"  行{e['row']} {e['name']}: 雇用形態=「{e['type']}」だがE列空")
        fail.append(f'パートでE列空 {len(part_no_hours)}件')
        print()

    # --- ④-c パートの時給（F列）: 要確認。FTE は E列のみに依存するので FAIL にはしない ---
    # F列は R215/R216 の計算には使わないが、最低賃金判定・加点①の材料になる。
    # 欠落と桁誤り（E列・F列・月列のどれかの取り違え）を検知する。
    rate_missing, rate_odd = [], []
    for e in rows:
        if not any(t in e['type'] for t in PART_TRIGGERS):
            continue
        paid = [v for v in e['months'] if v]
        if e['hourly'] is None:
            if paid:
                rate_missing.append(e)
        elif e['hours'] and paid and e['hourly'] > 0:
            est = _median(paid) / e['hours']
            if not 0.75 <= est / e['hourly'] <= 1.25:
                rate_odd.append((e, est))
    if rate_missing or rate_odd:
        print('[要確認 パートの時給(F列)]（最賃判定・加点①の材料。原本転記の時給/単価行から埋める）')
        for e in rate_missing:
            print(f"  行{e['row']} {e['name']}: F列(時給)が空")
        for e, est in rate_odd:
            print(f"  行{e['row']} {e['name']}: F列={e['hourly']:,.0f} だが 月額中位÷E列={est:,.0f}"
                  '（±25%超の乖離。E列・F列・月列のどれかが誤り）')
        confirm.append(f'F列(時給) {len(rate_missing) + len(rate_odd)}件（欠落 or 月額との乖離）')
        print()

    # --- FLAG-H 月ヘッダー規格外（注記付き等。ツールが月列を検出できず0名誤読する事故の pre-ship 検知）---
    bad_headers = []
    for m in range(12):
        col = COL_MONTH_FIRST + m
        htext = ws.cell(HEADER_ROW, col).value
        col_has_data = any(e['months'][m] is not None for e in rows)
        if htext is None or str(htext).strip() == '':
            if col_has_data:
                bad_headers.append((col, '(見出し空欄・データ有り)'))
        elif not _header_is_clean_month(htext):
            bad_headers.append((col, repr(str(htext))))
    if bad_headers:
        print('[FAIL FLAG-H 月ヘッダーが規格外]（注記付き等。ツールが月列を検出できず0名誤読の恐れ）')
        for col, why in bad_headers:
            print(f"  {chr(64 + col)}5: {why} → プレーン「N月」or 和暦「令和N年M月」へ。"
                  '事業年度マッピングは B2/B3 注記に書く（§0.0／SKILL.md §1🔴）')
        fail.append(f'月ヘッダー規格外 {len(bad_headers)}件（FLAG-H）')
        print()

    # --- 0混入（要確認: 空欄であるべき箇所の0） ---
    zero_rows = [(e, [MONTH_LABELS[m] for m in range(12) if e['months'][m] == 0])
                 for e in rows]
    zero_rows = [(e, zs) for e, zs in zero_rows if zs]
    if zero_rows:
        print('[FLAG-0 月セルに0]（中途/退社/休職は空欄が規約。0は誤判定の元）')
        for e, zs in zero_rows:
            print(f"  行{e['row']} {e['name']}: {','.join(zs)} が0")
        confirm.append(f'月セルに0が {len(zero_rows)}名（空欄にすべきか確認）')
        print()

    # --- 雇用形態内訳・全月在籍非役員数（観測） ---
    type_counts = Counter(e['type'] or '(空)' for e in rows)
    print('[雇用形態内訳]', dict(type_counts))
    has_officer = any('役員' in (e['type'] or '') for e in rows)
    if not has_officer:
        print('  ⚠ 「役員」が0件。代表者が賃金台帳に居るなら §2.1 で役員判定したか確認。')
    full_year_non_officer = sum(
        1 for e in rows
        if all(v is not None for v in e['months']) and '役員' not in (e['type'] or '')
    )
    print(f'[全月在籍の非役員] {full_year_non_officer}名')
    if full_year_non_officer == 0:
        print('  ❌ 0名 → R216/R215 が台帳だけでは確定不能（§0.2/§6-9 役員報酬ベース要否を報告）')
        fail.append('全月在籍非役員0名')
    print()

    # --- 集計 ---
    print('=== 集計 ===')
    if not fail and not confirm:
        print('=== PASS ===（FAIL・要確認なし。ただし §5.5-3 のPDF白紙突合は別途必須）')
        return 0
    if fail:
        print(f'■ 修正必須 FAIL {len(fail)}件: ' + ' / '.join(fail))
    if confirm:
        print(f'■ 要確認 {len(confirm)}件（PDFで素性確認し報告に解決を明記）: ' + ' / '.join(confirm))
    print('=== 要対応あり ===（FAILは修正、要確認はPDF確認の上で解決を§9報告に明記してから完了報告へ）')
    return 1


if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
