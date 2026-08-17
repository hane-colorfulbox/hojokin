# 🔴 引退（2026-08-17）: 通常枠×個人の原本は、原本管理担当者が Drive で作成した正本
#    （独自レイアウト）に切り替わり、リポジトリ原本は scripts/patch_genpon_tsujo_kojin_from_drive.py
#    がそのエクスポートから生成する。本スクリプト（素材連結方式）はもう実行しないこと
#    （実行すると OUT_NAME の本番原本を旧レイアウトで上書きしてしまう）。
#    8/10 生成版が必要なときは git 履歴（コミット 13f4e0e）から取得する。
#
# 通常枠×個人事業主の申請書原本を機械組み立てする（旧方式）。
# 素材:
#   E = ツール/【原本_法人】企業名_通常枠_法人2026_v2.xlsx   … 骨格（通常枠の全ブロック）
#   G = ツール/【原本_個人】企業名_インボイス枠_個人2026.xlsx … 個人事業主ブロックの前例
# 出力（🔴 本番ファイルを直接上書きする。「_下書き」サフィックスは付かない）:
#   ツール/【原本_個人】企業名_通常枠_個人2026.xlsx
# 方式: 行セグメント単位のコピー + 意味ベース行マップで数式/DV/CF/結合/行高を全付け替え。
import sys, re, shutil, unicodedata
from copy import copy, deepcopy
from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8')

# 引退ガード: 現行の原本（坂平さん版レイアウト）を旧レイアウトで上書きしないための停止。
# 歴史的な再現目的で本当に実行する場合のみ --force-legacy を付ける。
if '--force-legacy' not in sys.argv:
    print('このスクリプトは引退しました（2026-08-17）。現行の原本生成は')
    print('scripts/patch_genpon_tsujo_kojin_from_drive.py を使ってください。')
    print('（旧方式を意図して実行する場合のみ --force-legacy を付与）')
    raise SystemExit(1)

BASE = Path(r'C:\Users\user\projects\カラフルボックス\補助金')
OUT_NAME = 'ツール/【原本_個人】企業名_通常枠_個人2026.xlsx'
LOG = []


def log(msg):
    LOG.append(msg)


def resolve(rel):
    p = BASE / rel
    if p.exists():
        return p
    want = unicodedata.normalize('NFC', p.name)
    for cand in p.parent.iterdir():
        if unicodedata.normalize('NFC', cand.name) == want:
            return cand
    raise FileNotFoundError(rel)


E_PATH = resolve('ツール/【原本_法人】企業名_通常枠_法人2026_v2.xlsx')
G_PATH = resolve('ツール/【原本_個人】企業名_インボイス枠_個人2026.xlsx')
OUT_PATH = BASE / OUT_NAME

# ───────────────────────── 転記シートの新レイアウト ─────────────────────────
# (dst_row, src('E'|'G'), src_row)。坂平版ヒアリングシート基本情報の鏡像。
TENKI_SPEC = [
    (3, 'G', 3), (4, 'G', 4),
    (6, 'G', 6), (8, 'G', 8), (10, 'G', 10), (12, 'G', 12), (14, 'G', 14),
    (16, 'G', 16), (18, 'G', 18), (20, 'G', 20), (22, 'G', 22), (24, 'G', 24),
    (26, 'G', 26),
    (28, 'G', 28), (29, 'G', 29), (30, 'G', 30), (31, 'G', 31), (32, 'G', 32),
    (33, 'G', 33), (34, 'G', 34), (35, 'G', 35), (36, 'G', 36),
    (38, 'G', 38), (40, 'G', 40), (42, 'G', 42), (44, 'G', 44), (46, 'G', 46),
    (48, 'G', 48),
    (50, 'G', 50), (52, 'G', 52), (54, 'G', 54), (56, 'G', 56), (58, 'G', 58),
    (60, 'G', 60), (61, 'G', 61), (62, 'G', 62), (63, 'G', 63), (64, 'G', 64),
    (65, 'E', 54),                      # 申請ツール名（通常枠固有・坂平版ヒアにも有り）
    (67, 'G', 67), (68, 'G', 68), (70, 'G', 70), (71, 'G', 71), (73, 'G', 73),
    (75, 'G', 75), (76, 'G', 76), (77, 'G', 77), (78, 'G', 78), (79, 'G', 79),
    (81, 'E', 71), (82, 'E', 72), (83, 'E', 73), (84, 'E', 74), (85, 'E', 75),  # 経営状況（通常枠固有）
    (87, 'G', 81), (88, 'G', 82), (90, 'G', 84),
    (92, 'E', 82),                      # 賃上げ幅選択（通常枠の❶❷❸ラベル）
    (93, 'G', 87), (94, 'G', 88), (95, 'G', 89), (96, 'G', 90),
    (98, 'G', 97), (99, 'G', 98), (100, 'G', 99),
]
TENKI_MAX_ROW = 110

# 転記参照の付け替えマップ（参照元セルの出自ごとに使い分ける）
TENKI_MAP_G = {r: r for r in range(1, 80)}
TENKI_MAP_G.update({81: 87, 82: 88, 84: 90, 87: 93, 88: 94, 89: 95, 90: 96,
                    97: 98, 98: 99, 99: 100})
TENKI_MAP_E = {6: 6, 54: 65, 71: 81, 72: 82, 73: 83, 74: 84, 75: 85, 82: 92}

# ───────────────────────── 申請内容シートのセグメント ─────────────────────────
SEGMENTS = [
    ('E', 1, 8),      # 冒頭警告・シート作成/Wチェック・GビズID・添付ヘッダ
    ('G', 9, 12),     # 個人の必要書類: 身分証明書/所得税納税証明/確定申告書/IT戦略ナビ
    ('E', 12, 50),    # 賃金状況報告シート行 + チェック項目 + 申請類型/補助率(通常枠) + 基本情報入力
    ('G', 42, 63),    # 個人の事業者情報 + 個人向け質問セット
    ('E', 71, 71),    # ツール名（シート9マスタ連動DV）
    ('G', 65, 90),    # プロンプト/事業内容AI/決算月12月/代表者/担当者/従業員/過去交付/認定
    ('G', 91, 109),   # SECURITY ACTION / IT戦略ナビ / 申請要件
    ('G', 110, 117),  # 財務情報: 前期従業員・役員数(個人=1)・労働時間
    ('E', 156, 161),  # 売上高〜人件費（生産性指標シート参照 = 通常枠）
    ('G', 124, 128),  # 資本金0固定・次へ・経営状況ヘッダ
    ('E', 167, 184),  # 経営状況ブロック（通常枠フル: 強み/弱み/シート9連動/データ連携）
    ('E', 185, 190),  # アンケート注記・次へ・書類添付ヘッダ
    ('G', 142, 152),  # 添付5点（身分/所得税納税/確定申告/青色申告/その他）
    ('G', 153, 166),  # 交付申請情報確認〜宣誓
    ('G', 167, 189),  # 計画数値入力（暦年ラベル）〜賃上げ表明〜従業員なし
    ('E', 233, 245),  # 加点②ブロック + 労働生産性指標（通常枠）
    ('E', 246, 261),  # 最終確認〜事務局へ提出〜提出完了しました
]
# E の申請内容の最終非空行は 261（B261=『提出完了しました』）。
# ここを 260 で止めると 261 が wipe されず、新レイアウトの末尾のはるか下に孤児として残る。
SHINSEI_MAX_WIPE = 261

# 個人事業主向けに差し替える注記（E/G とも法人前提の文言が残っているため）。
# (セル, 置換前の部分文字列, 置換後) の並び。部分一致にして「適用済み」を取り違えないようにする。
KOJIN_NOTE_FIXES = [
    ('D20',
     '・製造業をされてる場合は「製造原価報告書」も必要\n'
     '　→無い場合はお客様へ確認(※子会社で製造や外注等で無い場合もあり)',
     '・製造原価がある場合は内訳の分かる資料も必要\n'
     '　→無い場合はお客様へ確認'),
    ('D20',
     '・直近分の貸借対照表と損益計算書は申請時に提出',
     '・直近分の所得税の青色申告決算書（白色申告の場合は収支内訳書）は申請時に提出'),
    ('D56',
     '「履歴事項全部証明書の目的の1番上の事業」',
     '「確定申告書・青色申告決算書に記載の主たる事業」'),
    ('D191',
     '都道府県名のみ、登記から抜粋',
     '都道府県名のみ、事業所所在地から'),
]

# 事業内容セルのフォールバック（キャッシュ値）に残るインボイス文言
TSUJO_TEXT_FIXES = [
    ('インボイス制度対応を含む', ''),
    ('インボイス対応の請求書作成', '見積・請求書の作成'),
]


def build_seg_maps():
    e_map, g_map, plan = {}, {}, []
    dst = 1
    for src, s, e in SEGMENTS:
        for r in range(s, e + 1):
            plan.append((dst, src, r))
            (e_map if src == 'E' else g_map)[r] = dst
            dst += 1
    return e_map, g_map, plan


SHINSEI_MAP_E, SHINSEI_MAP_G, SHINSEI_PLAN = build_seg_maps()
# 置換されたE行への外部/意味参照の付け替え（コピーはしないが参照だけ新居へ）
SHINSEI_MAP_E_EXTRA = {151: SHINSEI_MAP_G[113], 152: SHINSEI_MAP_G[114],
                       153: SHINSEI_MAP_G[115], 154: SHINSEI_MAP_G[116],
                       155: SHINSEI_MAP_G[117], 162: SHINSEI_MAP_G[124]}
SHINSEI_MAP_G_EXTRA = {64: SHINSEI_MAP_E[71]}   # プロンプトのツール名参照→E由来ツール名行

FULL_E = {**SHINSEI_MAP_E, **SHINSEI_MAP_E_EXTRA}
FULL_G = {**SHINSEI_MAP_G, **SHINSEI_MAP_G_EXTRA}

RE_TENKI = re.compile(r"('転記'!\$?B\$?)(\d+)")
RE_SELF = re.compile(r"(?<![A-Za-z0-9_!:])(\$?[A-E]\$?)(\d{1,4})(?![0-9])")


def rewrite_formula(text, self_map, tenki_map, where):
    out = text
    def tenki_sub(m):
        row = int(m.group(2))
        if row not in tenki_map:
            log(f'!! 未解決の転記参照 {where}: {m.group(0)}')
            return m.group(0)
        return f"{m.group(1)}{tenki_map[row]}"
    out = RE_TENKI.sub(tenki_sub, out)

    parts = re.split(r"('[^']+'!)", out)  # シート名付き参照の直後トークンは除外して同一シート参照だけ置換
    for i, part in enumerate(parts):
        if i > 0 and parts[i - 1].endswith("'!"):
            # シート参照直後のセル/レンジはそのまま（範囲2点目は ':' lookbehind で保護）
            m = re.match(r"(\$?[A-Z]{1,2}\$?\d+(:\$?[A-Z]{1,2}\$?\d+)?)(.*)", part, re.S)
            if m:
                head, _, rest = m.group(1), m.group(2), m.group(3)
                parts[i] = head + RE_SELF.sub(lambda mm: _self_sub(mm, self_map, where), rest)
                continue
        parts[i] = RE_SELF.sub(lambda mm: _self_sub(mm, self_map, where), part)
    return ''.join(parts)


def _self_sub(m, self_map, where):
    row = int(m.group(2))
    if row not in self_map:
        log(f'?? 同一シート参照が行マップ外 {where}: {m.group(0)} (据え置き)')
        return m.group(0)
    return f"{m.group(1)}{self_map[row]}"


def copy_cell(src_cell, dst_cell):
    dst_cell.value = src_cell.value
    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.protection = copy(src_cell.protection)
        dst_cell.number_format = src_cell.number_format
    if src_cell.hyperlink is not None:
        from openpyxl.worksheet.hyperlink import Hyperlink
        h = src_cell.hyperlink
        dst_cell.hyperlink = Hyperlink(ref=dst_cell.coordinate, target=h.target,
                                       tooltip=h.tooltip, location=h.location)
        log(f'hyperlink随伴 {src_cell.coordinate}->{dst_cell.coordinate}: {(h.target or "")[:50]}')


def wipe_rows(ws, max_row, max_col):
    for mc in [m for m in list(ws.merged_cells.ranges) if m.min_row <= max_row]:
        ws.unmerge_cells(str(mc))
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.value = None
            cell.style = 'Normal'
            if cell.hyperlink is not None:
                cell.hyperlink = None
    for r in list(ws.row_dimensions.keys()):
        if r <= max_row:
            del ws.row_dimensions[r]


def translate_ranges(sqref_str, row_map, col_ok=lambda c: True):
    """複数レンジ文字列を行マップで平行移動。移動不能なメンバーは捨てて報告。"""
    kept = []
    for token in str(sqref_str).split():
        m = re.match(r"^([A-Z]{1,2})(\d+)(?::([A-Z]{1,2})(\d+))?$", token)
        if not m:
            log(f'?? 変換不能レンジ {token} → 破棄')
            continue
        c1, r1, c2, r2 = m.group(1), int(m.group(2)), m.group(3), m.group(4)
        if c2 is None:
            if r1 in row_map:
                kept.append(f'{c1}{row_map[r1]}')
            else:
                log(f'   レンジ外メンバー破棄: {token}')
        else:
            r2 = int(r2)
            rows = list(range(r1, r2 + 1))
            if all(r in row_map for r in rows):
                mapped = [row_map[r] for r in rows]
                if mapped == list(range(mapped[0], mapped[0] + len(mapped))):
                    kept.append(f'{c1}{mapped[0]}:{c2}{mapped[-1]}')
                    continue
            ok = [r for r in rows if r in row_map]
            for r in ok:
                kept.append(f'{c1}{row_map[r]}' if c1 == c2 or c2 is None else f'{c1}{row_map[r]}:{c2}{row_map[r]}')
            if len(ok) < len(rows):
                log(f'   レンジ一部破棄: {token} (残 {len(ok)}/{len(rows)})')
    return kept


def main():
    shutil.copy2(E_PATH, OUT_PATH)
    wb = openpyxl.load_workbook(OUT_PATH)
    esrc = openpyxl.load_workbook(E_PATH)
    gsrc = openpyxl.load_workbook(G_PATH)

    if wb.defined_names:
        for name in list(wb.defined_names.keys()):
            log(f'defined_name: {name} = {wb.defined_names[name].value}')

    # ── 転記シート再構築 ──
    ws = wb['転記']
    eT, gT = esrc['転記'], gsrc['転記']
    wipe_rows(ws, TENKI_MAX_ROW, ws.max_column)
    for dst, src, srow in TENKI_SPEC:
        s_ws = eT if src == 'E' else gT
        for c in range(1, 6):
            copy_cell(s_ws.cell(srow, c), ws.cell(dst, c))
    for col in 'ABCDE':
        if col in gT.column_dimensions:
            ws.column_dimensions[col].width = gT.column_dimensions[col].width

    # ── 申請内容シート再構築 ──
    ws2 = wb['申請内容']
    eS, gS = esrc['申請内容'], gsrc['申請内容']
    old_e_heights = {r: d.height for r, d in eS.row_dimensions.items() if d.height}
    old_g_heights = {r: d.height for r, d in gS.row_dimensions.items() if d.height}
    old_e_merges = [str(m) for m in eS.merged_cells.ranges]
    old_g_merges = [str(m) for m in gS.merged_cells.ranges]

    wipe_rows(ws2, SHINSEI_MAX_WIPE, 26)
    ws2.data_validations.dataValidation = []
    # 条件付き書式は全消しして後で再構築
    for rng in list(ws2.conditional_formatting):
        del ws2.conditional_formatting[rng.sqref]

    for dst, src, srow in SHINSEI_PLAN:
        s_ws = eS if src == 'E' else gS
        for c in range(1, 27):
            copy_cell(s_ws.cell(srow, c), ws2.cell(dst, c))

    # 数式の付け替え
    for dst, src, srow in SHINSEI_PLAN:
        self_map = FULL_E if src == 'E' else FULL_G
        tenki_map = TENKI_MAP_E if src == 'E' else TENKI_MAP_G
        for c in range(1, 27):
            cell = ws2.cell(dst, c)
            v = cell.value
            if isinstance(v, str) and v.startswith('='):
                nv = rewrite_formula(v, self_map, tenki_map, f'{src}{srow}->r{dst}{get_column_letter(c)}')
                if nv != v:
                    log(f'式書換 r{dst}{get_column_letter(c)} ({src}{srow}): {v[:60]!r} -> {nv[:60]!r}')
                    cell.value = nv

    # プロンプトの枠名調整（G由来 → 通常枠へ）
    prow = SHINSEI_MAP_G[65]
    pv = ws2.cell(prow, 3).value
    assert isinstance(pv, str) and 'インボイス枠' in pv, 'プロンプト行の特定に失敗'
    n1 = pv.count('デジタル化・AI導入補助金（インボイス枠）')
    pv = pv.replace('デジタル化・AI導入補助金（インボイス枠）', 'デジタル化・AI導入補助金（通常枠）')
    n2 = pv.count('インボイス制度への対応遅れや、具体的な業務のボトルネックを')
    pv = pv.replace('インボイス制度への対応遅れや、具体的な業務のボトルネックを', '具体的な業務のボトルネックを')
    ws2.cell(prow, 3).value = pv
    log(f'プロンプト調整: 枠名置換 {n1} 箇所 / インボイス文言除去 {n2} 箇所 (r{prow})')

    # 結合セル
    for m_str, row_map, tag in [(m, SHINSEI_MAP_E, 'E') for m in old_e_merges] + \
                               [(m, SHINSEI_MAP_G, 'G') for m in old_g_merges]:
        rng = openpyxl.worksheet.cell_range.CellRange(m_str)
        rows = list(range(rng.min_row, rng.max_row + 1))
        if all(r in row_map for r in rows):
            mapped = [row_map[r] for r in rows]
            if mapped == list(range(mapped[0], mapped[0] + len(mapped))):
                new = f'{get_column_letter(rng.min_col)}{mapped[0]}:{get_column_letter(rng.max_col)}{mapped[-1]}'
                try:
                    ws2.merge_cells(new)
                    log(f'結合 {tag} {m_str} -> {new}')
                except Exception as ex:
                    log(f'!! 結合失敗 {tag} {m_str} -> {new}: {ex}')
                continue
        log(f'結合破棄 {tag} {m_str}（行が出力に無い/不連続）')

    # データバリデーション
    for src_ws, row_map, tenki_map, tag in [(eS, SHINSEI_MAP_E, TENKI_MAP_E, 'E'),
                                            (gS, SHINSEI_MAP_G, TENKI_MAP_G, 'G')]:
        for dv in src_ws.data_validations.dataValidation:
            new_ranges = translate_ranges(dv.sqref, row_map)
            if not new_ranges:
                log(f'DV破棄 {tag} {str(dv.sqref)[:50]}')
                continue
            ndv = DataValidation(type=dv.type, formula1=dv.formula1, formula2=dv.formula2,
                                 operator=dv.operator, allow_blank=dv.allow_blank,
                                 showDropDown=dv.showDropDown, showInputMessage=dv.showInputMessage,
                                 showErrorMessage=dv.showErrorMessage, errorTitle=dv.errorTitle,
                                 error=dv.error, promptTitle=dv.promptTitle, prompt=dv.prompt)
            if ndv.formula1 and isinstance(ndv.formula1, str) and not ndv.formula1.startswith('"') \
               and "'" not in ndv.formula1:
                nf = rewrite_formula('=' + ndv.formula1, row_map, tenki_map, f'DV({tag})')[1:]
                if nf != ndv.formula1:
                    log(f'DV式書換 {tag}: {ndv.formula1[:50]} -> {nf[:50]}')
                    ndv.formula1 = nf
            for rng in new_ranges:
                ndv.add(rng)
            ws2.add_data_validation(ndv)
            log(f'DV {tag} {str(dv.sqref)[:40]} -> {" ".join(new_ranges)[:40]}')

    # 条件付き書式（相対式はアンカー移動で再翻訳: 旧sqref最小セル→新sqref最小セル）
    from openpyxl.formula.translate import Translator
    from openpyxl.utils import column_index_from_string

    def range_key(tok):
        m = re.match(r'^([A-Z]{1,2})(\d+)', tok)
        return (int(m.group(2)), column_index_from_string(m.group(1)))

    for src_ws, row_map, tag in [(eS, SHINSEI_MAP_E, 'E'), (gS, SHINSEI_MAP_G, 'G')]:
        for cf in src_ws.conditional_formatting:
            new_ranges = translate_ranges(cf.sqref, row_map)
            if not new_ranges:
                log(f'CF破棄 {tag} {str(cf.sqref)[:60]}')
                continue
            old_tokens = sorted(str(cf.sqref).split(), key=range_key)
            new_ranges = sorted(new_ranges, key=range_key)
            old_origin = re.match(r'^([A-Z]{1,2}\d+)', old_tokens[0]).group(1)
            new_origin = re.match(r'^([A-Z]{1,2}\d+)', new_ranges[0]).group(1)
            for rule in cf.rules:
                nrule = deepcopy(rule)
                if getattr(nrule, 'formula', None) and old_origin != new_origin:
                    nrule.formula = [Translator('=' + f, origin=old_origin).translate_formula(new_origin)[1:]
                                     for f in nrule.formula]
                    log(f'CF式再翻訳 {tag} {old_origin}->{new_origin}: {rule.formula} -> {nrule.formula}')
                ws2.conditional_formatting.add(' '.join(new_ranges), nrule)
            log(f'CF {tag} {str(cf.sqref)[:50]} -> {" ".join(new_ranges)[:50]}')

    # 行高
    for r, h in old_e_heights.items():
        if r in SHINSEI_MAP_E:
            ws2.row_dimensions[SHINSEI_MAP_E[r]].height = h
    for r, h in old_g_heights.items():
        if r in SHINSEI_MAP_G:
            ws2.row_dimensions[SHINSEI_MAP_G[r]].height = h

    # 個人事業主向けの注記差し替え（法人前提の文言が E/G から持ち越されている）
    for coord, before, after in KOJIN_NOTE_FIXES:
        cur = ws2[coord].value
        if isinstance(cur, str) and before in cur:
            ws2[coord].value = cur.replace(before, after)
            log(f'注記差し替え {coord}: {before[:24]!r} -> {after[:24]!r}')
        elif isinstance(cur, str) and after in cur:
            log(f'注記差し替え {coord}: 適用済み（スキップ）')
        else:
            log(f'!! 注記差し替え失敗 {coord}: 想定の文言が見つからない {str(cur)[:60]!r}')

    # 事業内容セルのキャッシュ値に残るインボイス文言（通常枠テンプレなので落とす）
    biz_row = SHINSEI_MAP_G[66]
    bv = ws2.cell(biz_row, 3).value
    if isinstance(bv, str):
        nv = bv
        for before, after in TSUJO_TEXT_FIXES:
            nv = nv.replace(before, after)
        if nv != bv:
            ws2.cell(biz_row, 3).value = nv
            log(f'事業内容フォールバックのインボイス文言を除去 (r{biz_row})')

    # シート9 の平文ガイドを新レイアウトの行番号へ（E の法人行番号のまま持ち越されている）
    tool_row = SHINSEI_MAP_E[71]        # ツール名
    prompt_row = SHINSEI_MAP_G[65]      # プロンプト
    p1, p2, p3 = (SHINSEI_MAP_E[179], SHINSEI_MAP_E[180], SHINSEI_MAP_E[181])
    sheet9_map = {71: tool_row, 72: prompt_row, 179: p1, 180: p2, 181: p3}

    def sheet9_sub(m):
        n = int(m.group(1))
        return f'C{sheet9_map[n]}' if n in sheet9_map else m.group(0)

    ws9 = wb['シート9']
    n9 = 0
    for row in ws9.iter_rows(min_row=1, max_row=30, min_col=1, max_col=5):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and 'C' in v:
                nv = re.sub(r'C(\d{2,3})', sheet9_sub, v)
                if nv != v:
                    cell.value = nv
                    n9 += 1
                    log(f'シート9 案内文 {cell.coordinate}: {v[:40]!r} -> {nv[:40]!r}')
    log(f'シート9 案内文の行番号を更新: {n9} セル '
        f'(ツール名C{tool_row} / プロンプトC{prompt_row} / 3項目C{p1}〜C{p3})')

    # 生産性指標シートの申請内容参照付け替え
    ws3 = wb['生産性指標給与支給総額計算']
    for row in ws3.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "'申請内容'!" in v:
                def sub(mm):
                    r = int(mm.group(2))
                    if r not in FULL_E:
                        log(f'!! 生産性指標の未解決参照: {mm.group(0)}')
                        return mm.group(0)
                    return f'{mm.group(1)}{FULL_E[r]}'
                nv = re.sub(r"('申請内容'!\$?C\$?)(\d+)", sub, v)
                if nv != v:
                    log(f'生産性指標 {cell.coordinate}: {v} -> {nv}')
                    cell.value = nv

    wb.save(OUT_PATH)
    print(f'saved: {OUT_PATH}')
    print(f'log lines: {len(LOG)}')
    logfile = Path(__file__).parent / 'build_log.txt'
    logfile.write_text('\n'.join(LOG), encoding='utf-8')
    print(f'log -> {logfile}')

    # 警告を終了コードに反映する（ログを人が読む前提にしない）
    fatal = [m for m in LOG if m.startswith('!!')]
    warn = [m for m in LOG if m.startswith('??')]
    if warn:
        print(f'\n?? 要確認 {len(warn)} 件:')
        for m in warn[:10]:
            print('  ', m)
    if fatal:
        print(f'\n!! 未解決 {len(fatal)} 件（ビルド失敗扱い）:')
        for m in fatal:
            print('  ', m)
    return 1 if fatal else 0


if __name__ == '__main__':
    sys.exit(main())
