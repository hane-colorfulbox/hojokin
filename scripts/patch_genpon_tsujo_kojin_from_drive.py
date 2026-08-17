# -*- coding: utf-8 -*-
"""坂平さん版 通常枠×個人 申請フォーマット（Drive エクスポート）から、ツール配布用の修正版原本を生成する。

正本の位置づけ（2026-08-17〜）:
    通常枠×個人の申請書原本の正本は、原本管理担当者が Drive 上で管理する
    「【原本/法人】企業名_通常枠/個人2026」（Google スプレッドシート）。
    リポジトリ版はその xlsx エクスポートに、下記の機械的修正だけを当てたコピー。
    Drive 側原本にはこちらから一切書き込まない（2026-07-28 運用）。

    旧 scripts/build_genpon_tsujo_kojin.py（法人v2＋インボイス個人の素材連結方式）は
    この方式への切替に伴い引退（8/10 生成版は git 履歴にのみ残る）。

適用する修正（機械的修正のみ。内容判断が要るものは docs/Drive原本_申請テンプレの指摘一覧.md で報告）:
    P1: シート9 A2 案内文の古いセル番地（C21・C170〜C172）→ 実レイアウト（C75・C153〜C155）
    P2: シート9 r4 ヘッダーの番地（C170/C171/C172・C72）→（C153/C154/C155・C76）
    P3: 申請内容 C53（屋号・商号フリガナ）の参照 '転記'!B8 → '転記'!B10（商号と同一参照の誤り）
    P4: 申請内容 C60（事業所所在地）の未配線 → ='転記'!B20（郵便番号C59は配線済みなのに住所だけ欠け）
    P5: 申請内容 C77 AI() の IFERROR フォールバック（生成失敗文のキャッシュ）→ ""
    P6: 申請内容 D77 の作業メモ（※AI Worksにて作成…）を削除
    P7: 申請内容 C54 IMPORTXML の IFERROR フォールバック（"#REF!" キャッシュ）→ ""
    P8: 申請内容 D20 の決算書注記が法人前提（貸借対照表・損益計算書）→ 青色申告決算書/収支内訳書へ
    P9: 申請内容 D56 の業種コード注記が法人前提（履歴事項全部証明書）→ 確定申告書・青色申告決算書へ
    P10: 申請内容 D189 の最低賃金注記が法人前提（登記から抜粋）→ 事業所所在地から
    （P8〜P10 の文言は旧 build_genpon_tsujo_kojin.py の KOJIN_NOTE_FIXES と同一）

実行方法:
    python scripts/patch_genpon_tsujo_kojin_from_drive.py [--input <Drive版xlsx>] [--allow-new-input]

    入力の既定は _debug/_drive_tsujo_kojin_20260817/【原本_法人】企業名_通常枠_個人2026.xlsx。
    入力は SHA-256 で検証する（取得来歴は同フォルダの PROVENANCE.md）。Drive 側が更新されたら
    新エクスポートを取得し、--allow-new-input で実行 → 差分を確認のうえ本スクリプトの
    期待値定数（EXPECTED_INPUT_SHA256 と FROM 文字列）を更新すること。

保存後の自己検証:
    出力を開き直し、(1) パッチ7点が期待どおり (2) それ以外の全セル値が入力と同一
    (3) シート構成・DV数・条件付き書式数・結合範囲・定義名が入力と同一、を確認。
    1件でも崩れたら exit 1（出力は残すが採用しない）。
"""
import argparse
import hashlib
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT = ROOT / '_debug' / '_drive_tsujo_kojin_20260817' / '【原本_法人】企業名_通常枠_個人2026.xlsx'
OUTPUT = ROOT / 'ツール' / '【原本_個人】企業名_通常枠_個人2026.xlsx'

# 2026-08-17 取得エクスポートの SHA-256（PROVENANCE.md と対）
EXPECTED_INPUT_SHA256 = 'be79cf795275907427c1ec54eea91a2ae610db816682445dc13223a5e306325a'

SHINSEI = '申請内容'
SHEET9 = 'シート9'

# (シート, 行, 列, 修正前（None=空セル）, 修正後)。適用前に「修正前」を厳密照合する。
PATCHES = [
    # P1: シート9 案内文
    (SHEET9, 2, 1,
     'ツール名は「申請内容」シートのC21で選択してください。選択するとC170〜C172が自動入力されます。',
     'ツール名は「申請内容」シートのC75で選択してください。選択するとC153〜C155が自動入力されます。'),
    # P2: シート9 ヘッダー
    (SHEET9, 4, 2, '業務プロセス改善（申請内容C170）', '業務プロセス改善（申請内容C153）'),
    (SHEET9, 4, 3, '強化したい部門・業務（申請内容C171）', '強化したい部門・業務（申請内容C154）'),
    (SHEET9, 4, 4, '期待効果（申請内容C172）', '期待効果（申請内容C155）'),
    (SHEET9, 4, 5, '資料リンク（C72プロンプトに挿入）', '資料リンク（C76プロンプトに挿入）'),
    # P3: フリガナの参照誤り
    (SHINSEI, 53, 3, "='転記'!B8", "='転記'!B10"),
    # P4: 事業所所在地の未配線
    (SHINSEI, 60, 3, None, "='転記'!B20"),
    # P5: AI() のキャッシュ残骸（DUMMYFUNCTION ラッパーは温存＝Google シートに戻すと AI() が復元される）
    (SHINSEI, 77, 3,
     '=IFERROR(__xludf.DUMMYFUNCTION("AI(C76)"),"ヒアリング情報が不足しているため、文章を作成することができません。\n'
     '「業種・事業内容」「自社の強み」「現在の課題」「導入ツール」「具体的な改善点」などの詳細な情報をご提供いただけますでしょうか。'
     '情報をいただければ、ご指定の「250文字以上、255文字以内」の範囲で申請用の文章を作成いたします。")',
     '=IFERROR(__xludf.DUMMYFUNCTION("AI(C76)"),"")'),
    # P6: 作業メモ削除
    (SHINSEI, 77, 4, '※AI Worksにて作成\n参考：過去事業内容', None),
    # P7: IMPORTXML の "#REF!" キャッシュ
    (SHINSEI, 54, 3,
     '=IFERROR(__xludf.DUMMYFUNCTION("IMPORTXML(""https://api.excelapi.org/post/zipcode?address=""&ENCODEURL(C55),""."")"),"#REF!")',
     '=IFERROR(__xludf.DUMMYFUNCTION("IMPORTXML(""https://api.excelapi.org/post/zipcode?address=""&ENCODEURL(C55),""."")"),"")'),
    # P8: 決算書注記の個人化（法人: 貸借対照表・損益計算書 → 個人: 青色申告決算書/収支内訳書）
    (SHINSEI, 20, 4,
     '・製造業をされてる場合は「製造原価報告書」も必要\n'
     '　→無い場合はお客様へ確認(※子会社で製造や外注等で無い場合もあり)\n'
     '・直近分の貸借対照表と損益計算書は申請時に提出',
     '・製造原価がある場合は内訳の分かる資料も必要\n'
     '　→無い場合はお客様へ確認\n'
     '・直近分の所得税の青色申告決算書（白色申告の場合は収支内訳書）は申請時に提出'),
    # P9: 業種コード注記の個人化（履歴事項全部証明書 → 確定申告書・青色申告決算書）
    (SHINSEI, 56, 4,
     '複数の事業を行っている場合\n'
     '⇨「履歴事項全部証明書の目的の1番上の事業」もしくは「売上の高い事業」\n'
     '→「申請ツール」により整合制のある事業を選択する（中小企業定義に関係するので売上の立っている事業以外を選択する場合は相談）\n'
     '\n'
     '日本標準産業分類(令和５年[2023年]７月改定)\n'
     '細分類コードを入力（数字４桁）',
     '複数の事業を行っている場合\n'
     '⇨「確定申告書・青色申告決算書に記載の主たる事業」もしくは「売上の高い事業」\n'
     '→「申請ツール」により整合制のある事業を選択する（中小企業定義に関係するので売上の立っている事業以外を選択する場合は相談）\n'
     '\n'
     '日本標準産業分類(令和５年[2023年]７月改定)\n'
     '細分類コードを入力（数字４桁）'),
    # P10: 最低賃金注記の個人化（登記から抜粋 → 事業所所在地から）
    (SHINSEI, 189, 4,
     '都道府県名のみ、登記から抜粋\n令和7年度地域別最低賃金額',
     '都道府県名のみ、事業所所在地から\n令和7年度地域別最低賃金額'),
]

PATCHED_ADDRS = {(sheet, r, c) for sheet, r, c, _, _ in PATCHES}


def sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, 'rb') as f:
        for chunk in iter(lambda: f.read(1 << 20), b''):
            h.update(chunk)
    return h.hexdigest()


def _norm(v):
    """比較用: 改行コードの揺れだけ吸収（openpyxl は LF で返す想定だが念のため）"""
    if isinstance(v, str):
        return unicodedata.normalize('NFC', v.replace('\r\n', '\n'))
    return v


def snapshot(path: Path):
    wb = openpyxl.load_workbook(path, data_only=False)
    cells = {}
    meta = {}
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is not None:
                    cells[(ws.title, c.row, c.column)] = _norm(c.value)
        meta[ws.title] = {
            'dv': len(ws.data_validations.dataValidation),
            'cf': len(ws.conditional_formatting),
            'merged': sorted(str(r) for r in ws.merged_cells.ranges),
            'state': ws.sheet_state,
        }
    names = sorted(wb.defined_names.keys())
    order = wb.sheetnames
    wb.close()
    return cells, meta, names, order


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--input', type=Path, default=DEFAULT_INPUT)
    ap.add_argument('--allow-new-input', action='store_true',
                    help='SHA-256 が期待値と違う入力を許可（Drive 更新後の再取込時。差分確認とFROM定数の更新を忘れないこと）')
    args = ap.parse_args()

    src = args.input
    if not src.exists():
        print(f'NG: 入力が見つかりません: {src}')
        return 1

    digest = sha256(src)
    if digest != EXPECTED_INPUT_SHA256:
        if args.allow_new_input:
            print(f'!! 入力 SHA-256 が期待値と異なります（--allow-new-input 指定で続行）: {digest}')
        else:
            print('NG: 入力 SHA-256 が期待値と異なります。Drive 側が更新された可能性があります。')
            print(f'    実際:   {digest}')
            print(f'    期待値: {EXPECTED_INPUT_SHA256}')
            print('    新エクスポートを意図して使う場合は差分を確認のうえ --allow-new-input を付け、')
            print('    本スクリプトの期待値定数も更新してください。')
            return 1

    wb = openpyxl.load_workbook(src, data_only=False)
    ng = 0
    for sheet, r, c, before, after in PATCHES:
        cell = wb[sheet].cell(row=r, column=c)
        actual = _norm(cell.value)
        if actual != _norm(before):
            print(f'NG: {sheet}!{cell.coordinate} の現状が期待と不一致（Drive 側が変わった可能性）')
            print(f'    実際: {str(actual)[:120]!r}')
            print(f'    期待: {str(before)[:120]!r}')
            ng += 1
            continue
        cell.value = after
        if after is None and cell.hyperlink is not None:
            # D77 の作業メモは「過去事業内容」参考シートへのハイパーリンク付き。値だけ消すと
            # リンク文字列が残るため一緒に除去する
            cell.hyperlink = None
        print(f'OK: {sheet}!{cell.coordinate} を修正')
    if ng:
        print(f'\nNG {ng} 件のため保存しません')
        return 1

    wb.save(OUTPUT)
    wb.close()
    print(f'\n保存: {OUTPUT}')

    # ── 自己検証: 入力とのフルセル diff（パッチ7点以外は不変であること） ──
    src_cells, src_meta, src_names, src_order = snapshot(src)
    out_cells, out_meta, out_names, out_order = snapshot(OUTPUT)

    problems = []
    if src_order != out_order:
        problems.append(f'シート構成が変化: {src_order} -> {out_order}')

    for key in src_cells.keys() | out_cells.keys():
        sv, ov = src_cells.get(key), out_cells.get(key)
        if key in PATCHED_ADDRS:
            continue
        if sv != ov:
            sheet, r, c = key
            problems.append(f'意図外の変化: {sheet}!r{r}c{c}: {str(sv)[:60]!r} -> {str(ov)[:60]!r}')

    for sheet, r, c, _, after in PATCHES:
        ov = out_cells.get((sheet, r, c))
        if _norm(ov) != _norm(after):
            problems.append(f'パッチ未反映: {sheet}!r{r}c{c}: {str(ov)[:60]!r}')

    for title, sm in src_meta.items():
        om = out_meta.get(title)
        if om is None:
            continue
        for k in ('dv', 'cf', 'state'):
            if sm[k] != om[k]:
                problems.append(f'{title}.{k} が変化: {sm[k]} -> {om[k]}')
        if sm['merged'] != om['merged']:
            problems.append(f'{title} の結合範囲が変化')
    if src_names != out_names:
        problems.append(f'定義名が変化: {src_names} -> {out_names}')

    if problems:
        print('\n自己検証 NG:')
        for p in problems[:30]:
            print(' ', p)
        return 1

    print(f'自己検証 OK: パッチ {len(PATCHES)} 点以外の全セル・DV・条件付き書式・結合・定義名は入力と同一')
    print(f'出力 SHA-256: {sha256(OUTPUT)}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
