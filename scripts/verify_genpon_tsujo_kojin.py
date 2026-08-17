# -*- coding: utf-8 -*-
"""通常枠×個人 申請書原本（坂平さん版レイアウト・修正版）の構造検証。

対象: ツール/【原本_個人】企業名_通常枠_個人2026.xlsx
     （scripts/patch_genpon_tsujo_kojin_from_drive.py の出力。2026-08-17 に
       素材連結方式の旧原本から坂平さん版 Drive 正本ベースへ切り替えた）

検証観点（NG が 1 件でもあれば exit 1）:
    V1  シート構成（10シート・欠落なし）
    V2  エラー値（#REF! 等）の非残存（数式文字列内のキャッシュ含む）
    V3  法人前提キーワードの非残存（履歴事項/法人税/貸借対照表/損益計算書/登記/インボイス）
    V4  シート9: 案内文とヘッダーが実レイアウトの番地（C75選択・C153〜C155・C76 プロンプト）
        を指し、旧番地（C21・C170〜C172・C72）が残っていないこと。ツール名 DV = A5:A12
    V5  申請内容の要所配線: C53=転記B10 / C60=転記B20 / C54 IMPORTXML(ENCODEURL(C55)) /
        C77 AI(C76) / E77 LEN(C77) / C153〜C155 = VLOOKUP($C$75, シート9)
    V6  config.MAPPING_2026_TSUJO_KOJIN との整合: shinsei 全フィールドの行に B列ラベルがあり、
        ツール名・事業内容・決算月など要所のラベル一致 / preserve_rows が非数式セルを指す /
        hearing_to_tenki の転記行に A列ラベルがある
    V7  転記シート: 自由記述ゾーン（B28〜B37）が tenki_text_range と対応
    V8  DV 数（申請内容=14）と決算月 DV（C78 ← プルダウン用!$A$2:$A$13）
    V9  最終非空行 = 239 以内かつ『提出完了しました』の行が存在
    V10 ヒアリングシート（ツール/ヒアリングシート2026_通常枠個人.xlsx）との対応:
        hearing_to_tenki のヒアリング行に基本情報 B列ラベルがあること

実行方法:
    python scripts/verify_genpon_tsujo_kojin.py
"""
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
GENPON = ROOT / 'ツール' / '【原本_個人】企業名_通常枠_個人2026.xlsx'
HEARING = ROOT / 'ツール' / 'ヒアリングシート2026_通常枠個人.xlsx'

sys.path.insert(0, str(ROOT))
from hojokin.config import MAPPING_2026_TSUJO_KOJIN as MAPPING  # noqa: E402

EXPECTED_SHEETS = ['転記', 'プルダウン用', '申請内容', '生産性指標給与支給総額計算', '申請金額',
                   '小規模事業者確認', 'シート9', '分類表', '商品マスタ', '費用内訳マスタ']
ERROR_VALUES = ('#REF!', '#N/A', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!')
HOJIN_KEYWORDS = ('履歴事項全部証明書', '法人税', '貸借対照表', '損益計算書', '登記', 'インボイス')
# 法人語の許容箇所（内容を確認済みの残存。増えたら NG にして中身を見る）
HOJIN_ALLOW: set[tuple[str, str]] = set()

ANCHOR_LABELS = {
    'tool_name': 'ツール名',
    'business_description': '事業内容（255文字以内）',
    'fiscal_month': '決算月',
    'rep_name': '代表者氏名',
    'security_status': 'セキュリティの状況',
    'min_wage_hourly': '事業所内最低賃金時給',
    'wage_raise_amount': '賃上げ幅',
}


def norm(s) -> str:
    if s is None:
        return ''
    return re.sub(r'\s+', '', unicodedata.normalize('NFKC', str(s)))


def main() -> int:
    ng = []
    ok = 0

    def check(cond: bool, label: str, detail: str = ''):
        nonlocal ok
        if cond:
            ok += 1
        else:
            ng.append(f'{label}' + (f': {detail}' if detail else ''))

    wb = openpyxl.load_workbook(GENPON, data_only=False)

    # V1 シート構成
    check(wb.sheetnames == EXPECTED_SHEETS, 'V1 シート構成', f'{wb.sheetnames}')

    # V2 エラー値 / V3 法人語
    hojin_hits = []
    err_hits = []
    for sh in wb.worksheets:
        for row in sh.iter_rows():
            for cell in row:
                v = cell.value
                if not isinstance(v, str):
                    continue
                if any(e in v for e in ERROR_VALUES):
                    err_hits.append(f'{sh.title}!{cell.coordinate}')
                for kw in HOJIN_KEYWORDS:
                    if kw in v and (sh.title, cell.coordinate) not in HOJIN_ALLOW:
                        hojin_hits.append(f'{sh.title}!{cell.coordinate}:{kw}')
    check(not err_hits, 'V2 エラー値残存', ', '.join(err_hits[:8]))
    check(not hojin_hits, 'V3 法人前提キーワード残存', ', '.join(hojin_hits[:8]))

    # V4 シート9
    s9 = wb['シート9']
    a2 = str(s9.cell(row=2, column=1).value or '')
    check('C75' in a2 and 'C153〜C155' in a2, 'V4 シート9 A2 案内文の番地', a2[:60])
    check('C21' not in a2 and 'C170' not in a2, 'V4 シート9 A2 旧番地の残存', a2[:60])
    for col, want in ((2, 'C153'), (3, 'C154'), (4, 'C155'), (5, 'C76')):
        h = str(s9.cell(row=4, column=col).value or '')
        check(want in h, f'V4 シート9 r4 c{col} ヘッダー番地', h[:40])
        check('C17' not in h and 'C72' not in h.replace('C76', ''), f'V4 シート9 r4 c{col} 旧番地の残存', h[:40])
    ws = wb['申請内容']
    dv_c75 = [dv for dv in ws.data_validations.dataValidation if 'C75' in str(dv.sqref)]
    check(bool(dv_c75) and "'シート9'!$A$5:$A$12" in str(dv_c75[0].formula1 if dv_c75 else ''),
          'V4 ツール名 DV (C75←シート9A5:A12)', str(dv_c75[0].formula1) if dv_c75 else 'DVなし')

    # V5 要所配線
    check(ws.cell(row=53, column=3).value == "='転記'!B10", 'V5 C53 フリガナ参照', str(ws.cell(row=53, column=3).value))
    check(ws.cell(row=60, column=3).value == "='転記'!B20", 'V5 C60 事業所所在地参照', str(ws.cell(row=60, column=3).value))
    c54 = str(ws.cell(row=54, column=3).value or '')
    check('IMPORTXML' in c54 and 'ENCODEURL(C55)' in c54, 'V5 C54 IMPORTXML', c54[:60])
    c77 = str(ws.cell(row=77, column=3).value or '')
    check('AI(C76)' in c77, 'V5 C77 AI(C76)', c77[:60])
    e77 = str(ws.cell(row=77, column=5).value or '')
    check(norm(e77).lower() == '=len(c77)', 'V5 E77 LEN(C77)', e77)
    for r, idx in ((153, 2), (154, 3), (155, 4)):
        f = str(ws.cell(row=r, column=3).value or '')
        check(f"VLOOKUP($C$75,'シート9'!$A$5:$D$50,{idx},0)" in f, f'V5 C{r} シート9連動', f[:60])
    check(ws.cell(row=77, column=4).value is None, 'V5 D77 作業メモ除去', str(ws.cell(row=77, column=4).value)[:40])

    # V6 config との整合
    for field, row in MAPPING.shinsei.items():
        label = ws.cell(row=row, column=2).value
        check(label is not None and str(label).strip() != '', f'V6 shinsei {field} r{row} にラベルあり', repr(label))
    for field, want in ANCHOR_LABELS.items():
        row = MAPPING.shinsei[field]
        got = norm(str(ws.cell(row=row, column=2).value or '').split('\n')[0])
        check(got == norm(want), f'V6 アンカー {field} r{row}', f'{got[:30]} != {want}')
    for r in MAPPING.preserve_rows:
        v = ws.cell(row=r, column=3).value
        check(v is not None, f'V6 preserve r{r} に値あり', '空')

    tk = wb['転記']
    for hr, tr, _ in MAPPING.hearing_to_tenki:
        a = tk.cell(row=tr, column=1).value
        check(a is not None, f'V6 転記 r{tr} にラベルあり (hearing r{hr})', '空')

    # V7 転記 自由記述ゾーン
    lo, hi = MAPPING.tenki_text_range
    check((lo, hi) == (28, 38), 'V7 tenki_text_range', f'{(lo, hi)}')
    check(norm(tk.cell(row=28, column=1).value) == norm('主な事業内容'), 'V7 転記 A28', str(tk.cell(row=28, column=1).value))
    check(norm(tk.cell(row=37, column=1).value).startswith(norm('どのような属性の取引先')), 'V7 転記 A37',
          str(tk.cell(row=37, column=1).value))

    # V8 DV
    dvs = ws.data_validations.dataValidation
    check(len(dvs) == 14, 'V8 申請内容 DV数=14', str(len(dvs)))
    dv_c78 = [dv for dv in dvs if 'C78' in str(dv.sqref)]
    check(bool(dv_c78) and 'プルダウン用' in str(dv_c78[0].formula1 if dv_c78 else ''),
          'V8 決算月 DV (C78)', str(dv_c78[0].formula1) if dv_c78 else 'DVなし')

    # V9 最終行
    last = 0
    submit_done = None
    for row in ws.iter_rows(min_col=2, max_col=3):
        for cell in row:
            if cell.value is not None:
                last = max(last, cell.row)
                if isinstance(cell.value, str) and '提出完了しました' in cell.value:
                    submit_done = cell.row
    check(submit_done is not None, 'V9 『提出完了しました』行', '見つからない')
    check(last <= MAPPING.shinsei_clear_range[1], 'V9 最終非空行がクリア範囲内', f'last={last}')

    wb.close()

    # V10 ヒアリングシート対応
    hb = openpyxl.load_workbook(HEARING, read_only=True)
    hs = hb['基本情報']
    for hr, tr, _ in MAPPING.hearing_to_tenki:
        v = hs.cell(row=hr, column=2).value
        check(v is not None and str(v).strip() != '', f'V10 ヒアリング r{hr} にラベルあり (転記 r{tr})', '空')
    hb.close()

    print(f'OK {ok} 項目 / NG {len(ng)} 項目')
    if ng:
        print('\nNG 明細:')
        for x in ng:
            print(' ', x)
        return 1
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
