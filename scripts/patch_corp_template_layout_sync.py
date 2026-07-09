# -*- coding: utf-8 -*-
"""法人テンプレ2種を Drive 版原本（スプレッドシート）と同期し、行高つぶれ等を根治する。

背景:
    v0.2.87 のプルダウン位置修正後、Drive「2026」フォルダのスプレッドシート版原本
    （2026-07-03 更新）と全面比較した結果、リポジトリ側にのみ残る欠陥が判明した。
    1) 行の高さがプルダウンと同じ「+9行ずれ」で旧行に据え置かれ、スペーサー高(6pt)が
       本文行に当たり出力の『次へ』バナー等がつぶれる（2026-07-09 スクショ指摘）
    2) 通常枠のプルダウン用!B2:B48（都道府県/最低賃金）が R6 値のまま（Drive は R7）
    3) 成長率の案内文言が「1.5%以上」のまま（公募要領 2026 は 3.0%/3.5%。
       docs/補助金_実務知識ベース.md「賃上げ要件との関連」参照。Drive は 3.0%）
    4) 案内文・転記式 5 セルが B:E 結合の巻き添えで消失（インボイス C113/C126/C187/C224、
       通常枠 C133/C240。Drive は非結合で値あり。v0.2.59/61/74 と同族の結合欠陥）
    Drive 版の DV 位置は v0.2.87 の修正結果と完全一致（独立裏付け）。

修正内容（ソース＝Drive 版 xlsx エクスポート）:
    - 申請内容シートの全行の高さ・非表示フラグを Drive 版からコピー
    - 上記 5 セルの結合解除＋Drive 値を復元
    - 成長率文言（通常枠 B220 / インボイス B205）を Drive 値に差し替え
    - 通常枠 プルダウン用!B2:B48 を Drive 値（R7 最賃）に更新
    - DV・その他のセル値・結合には触れない（ツールマスタ連動の採用は別タスク）

実行方法:
    python scripts/patch_corp_template_layout_sync.py <Drive版通常枠.xlsx> <Drive版インボイス.xlsx>

冪等性:
    再実行しても同じ最終状態に収束する（結合解除は無ければ素通り、値・高さは同値上書き）。
"""
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
SHEET = '申請内容'
PULLDOWN_SHEET = 'プルダウン用'

# (repoテンプレ, 結合解除+C列復元行, 文言差替えセル(B列), プルダウン用B列更新するか)
TARGETS = [
    {
        'repo': ROOT / 'ツール' / '【原本_法人】企業名_通常枠_法人2026_v2.xlsx',
        'restore_rows': [133, 240],
        'growth_cell': 'B220',
        'sync_minwage': True,
    },
    {
        'repo': ROOT / 'ツール' / '【原本_法人】企業名_インボイス枠_法人2026_v2.xlsx',
        'restore_rows': [113, 126, 187, 224],
        'growth_cell': 'B205',
        'sync_minwage': False,
    },
]
GROWTH_EXPECT_OLD = '1.5％以上'
GROWTH_EXPECT_NEW = '3.0％以上'


def dv_inventory(ws):
    return sorted((str(dv.sqref), str(dv.formula1)) for dv in ws.data_validations.dataValidation)


def snapshot_bc(ws, max_row):
    return {(r, c): ws.cell(r, c).value for r in range(1, max_row + 1) for c in (2, 3)}


def sync_file(drive_path: Path, cfg: dict) -> bool:
    repo_path = cfg['repo']
    print(f'\n===== {repo_path.name} =====')
    wb_d = openpyxl.load_workbook(drive_path, data_only=False)
    wb_r = openpyxl.load_workbook(repo_path, data_only=False)
    ws_d, ws_r = wb_d[SHEET], wb_r[SHEET]
    n_rows = max(ws_d.max_row, ws_r.max_row)

    dv_before = dv_inventory(ws_r)
    bc_before = snapshot_bc(ws_r, n_rows)

    # 1) 行高・非表示フラグを全行コピー
    n_h = 0
    for r in range(1, n_rows + 1):
        hd = ws_d.row_dimensions[r].height if r in ws_d.row_dimensions else None
        hidden_d = ws_d.row_dimensions[r].hidden if r in ws_d.row_dimensions else False
        rd = ws_r.row_dimensions[r]
        if rd.height != hd or rd.hidden != hidden_d:
            rd.height = hd
            rd.hidden = hidden_d
            n_h += 1
    print(f'  行高/表示 同期: {n_h}行')

    # 2) 結合解除＋C列復元（Drive値をそのまま移植）
    for row in cfg['restore_rows']:
        rng = f'B{row}:E{row}'
        for mr in list(ws_r.merged_cells.ranges):
            if str(mr) == rng:
                ws_r.unmerge_cells(rng)
        val = ws_d.cell(row, 3).value
        ws_r.cell(row, 3).value = val
        vs = str(val).replace('\n', '⏎')[:40]
        print(f'  復元: C{row} = {vs!r}（{rng} 結合解除）')

    # 3) 成長率文言
    gc = cfg['growth_cell']
    gv_d = ws_d[gc].value
    if GROWTH_EXPECT_NEW not in str(gv_d):
        print(f'  !! Drive {gc} に {GROWTH_EXPECT_NEW} が無い: {gv_d!r}。中止')
        return False
    ws_r[gc].value = gv_d
    print(f'  文言: {gc} を 3.0% 版に差替え')

    # 4) 最賃リスト（通常枠のみ）
    if cfg['sync_minwage']:
        pd_, pr_ = wb_d[PULLDOWN_SHEET], wb_r[PULLDOWN_SHEET]
        n_m = 0
        for r in range(2, 49):
            v = pd_.cell(r, 2).value
            if pr_.cell(r, 2).value != v:
                pr_.cell(r, 2).value = v
                n_m += 1
        print(f'  最賃リスト更新: {n_m}件（R7値）')

    wb_r.save(repo_path)
    wb_r.close()

    # ---- 事後検証 ----
    wb2 = openpyxl.load_workbook(repo_path, data_only=False)
    ws2 = wb2[SHEET]
    ok = True

    if dv_inventory(ws2) != dv_before:
        print('  NG: DV一覧が変化した')
        ok = False

    for r in range(1, n_rows + 1):
        hd = ws_d.row_dimensions[r].height if r in ws_d.row_dimensions else None
        hidden_d = ws_d.row_dimensions[r].hidden if r in ws_d.row_dimensions else False
        if ws2.row_dimensions[r].height != hd or ws2.row_dimensions[r].hidden != hidden_d:
            print(f'  NG: r{r} 行高/表示が Drive と不一致')
            ok = False

    allowed = {(row, 3) for row in cfg['restore_rows']}
    gc_row = int(cfg['growth_cell'][1:])
    allowed.add((gc_row, 2))
    bc_after = snapshot_bc(ws2, n_rows)
    for key, v_old in bc_before.items():
        if bc_after.get(key) != v_old and key not in allowed:
            print(f'  NG: 想定外のセル変化 {key}: {v_old!r} → {bc_after.get(key)!r}')
            ok = False
    for row in cfg['restore_rows']:
        if bc_after[(row, 3)] != ws_d.cell(row, 3).value:
            print(f'  NG: C{row} の復元値不一致')
            ok = False
        if f'B{row}:E{row}' in {str(m) for m in ws2.merged_cells.ranges}:
            print(f'  NG: B{row}:E{row} の結合が残存')
            ok = False
    if GROWTH_EXPECT_NEW not in str(ws2[cfg['growth_cell']].value):
        print(f'  NG: {cfg["growth_cell"]} 文言未反映')
        ok = False
    if cfg['sync_minwage']:
        pr2 = wb2[PULLDOWN_SHEET]
        pd_ = wb_d[PULLDOWN_SHEET]
        for r in range(2, 49):
            if pr2.cell(r, 2).value != pd_.cell(r, 2).value:
                print(f'  NG: プルダウン用 B{r} 不一致')
                ok = False
                break
    wb2.close()
    wb_d.close()
    print(f'  検証 {"OK" if ok else "NG"}')
    return ok


def main() -> None:
    if len(sys.argv) != 3:
        print('usage: python scripts/patch_corp_template_layout_sync.py '
              '<Drive版通常枠.xlsx> <Drive版インボイス.xlsx>')
        sys.exit(2)
    sources = [Path(sys.argv[1]), Path(sys.argv[2])]
    all_ok = True
    for drive_path, cfg in zip(sources, TARGETS):
        if not drive_path.exists():
            print(f'ソースが無い: {drive_path}')
            sys.exit(2)
        all_ok = sync_file(drive_path, cfg) and all_ok
    print('\n=== 結果:', 'OK' if all_ok else 'NG（要確認）', '===')
    if not all_ok:
        sys.exit(1)


if __name__ == '__main__':
    main()
