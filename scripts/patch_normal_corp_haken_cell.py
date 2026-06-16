# -*- coding: utf-8 -*-
"""通常枠・法人テンプレの「従業員数：派遣社員」(r120) 欠陥を修正する。

背景:
    申請内容シートの r120(従業員数：派遣社員) が B120:E120 で結合され、かつ転記シート
    参照式が欠落していた（C120=None）ため、派遣社員数が空欄表示になっていた。
    （派遣社員数は転記シート B45 に書き込まれるのに、申請内容が引けていない）
    正常な兄弟 r117/118/119/121(正規/契約/パート/その他) は C="='転記'!B39/41/43/47" を
    持ち表示できている。配置の規則どおり派遣社員は B45（hearing_to_tenki: 44→45）。
    v0.2.59/v0.2.61 がインボイス法人の同種欠陥を直したが、通常枠の派遣社員行は網から漏れていた。
    2026-06-16 社内フィードバックMTGで村上さんが「従業員数の派遣のところだけセルが結合されて
    数字が入らない」と指摘した事象に一致。

修正内容:
    - B120:E120 の結合を解除
    - C120="='転記'!B45" を復元
    - D120="申請時の人数"（兄弟と同じ注記）を復元
    - 正常な兄弟 r121(その他) から C/D/E の書式と行高を複製（値は複製しない）

実行方法:
    python scripts/patch_normal_corp_haken_cell.py

冪等性:
    既に修正済みなら結合解除・式設定はそのまま通り、再実行しても結果は同じ。
"""
import sys
from copy import copy
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
NORMAL_CORP = ROOT / 'ツール' / '【原本_法人】企業名_通常枠_法人2026_v2.xlsx'

SHEET = '申請内容'
TARGET_ROW = 120           # 従業員数：派遣社員
SIBLING_ROW = 121          # 従業員数：その他（正常な兄弟）
FORMULA = "='転記'!B45"
NOTE = '申請時の人数'
STYLE_COLS = (3, 4, 5)     # C / D / E（B見出しセルは触らない）


def _copy_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


def main() -> None:
    wb = openpyxl.load_workbook(NORMAL_CORP, data_only=False)
    ws = wb[SHEET]

    dv_before = len(ws.data_validations.dataValidation) if ws.data_validations else 0
    sibling_height = ws.row_dimensions[SIBLING_ROW].height

    rng = f'B{TARGET_ROW}:E{TARGET_ROW}'
    for mr in list(ws.merged_cells.ranges):
        if str(mr) == rng:
            ws.unmerge_cells(rng)

    for col in STYLE_COLS:
        _copy_style(ws.cell(SIBLING_ROW, col), ws.cell(TARGET_ROW, col))
    if sibling_height is not None:
        ws.row_dimensions[TARGET_ROW].height = sibling_height

    ws.cell(TARGET_ROW, 3).value = FORMULA
    ws.cell(TARGET_ROW, 4).value = NOTE

    wb.save(NORMAL_CORP)

    # 事後検証
    wb2 = openpyxl.load_workbook(NORMAL_CORP, data_only=False)
    ws2 = wb2[SHEET]
    merged = {str(m) for m in ws2.merged_cells.ranges}
    dv_after = len(ws2.data_validations.dataValidation) if ws2.data_validations else 0
    b = ws2.cell(TARGET_ROW, 2).value
    c = ws2.cell(TARGET_ROW, 3).value
    d = ws2.cell(TARGET_ROW, 4).value
    is_unmerged = rng not in merged
    print('=== 修正後の検証 ===')
    print(f'  r{TARGET_ROW}: B={b!r} C={c!r} D={d!r} 非結合={is_unmerged}')
    print(f'  兄弟 r119 C={ws2.cell(119, 3).value!r}（不変であること）')
    print(f'  兄弟 r121 C={ws2.cell(121, 3).value!r}（不変であること）')
    print(f'  申請内容 data_validations: {dv_before} -> {dv_after}（維持されること）')
    ok = (
        is_unmerged
        and c == FORMULA
        and d == NOTE
        and b == '従業員数：派遣社員'
        and dv_after == dv_before
    )
    wb2.close()
    print('=== 結果:', 'OK' if ok else 'NG（要確認）', '===')
    if not ok:
        sys.exit(1)


if __name__ == '__main__':
    main()
