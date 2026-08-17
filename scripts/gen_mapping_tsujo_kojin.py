# -*- coding: utf-8 -*-
"""config.MAPPING_2026_TSUJO_KOJIN を「ラベル引き当て」で機械生成する（2026-08-17 全面改版）。

方式の変更経緯:
    旧版は scripts/build_genpon_tsujo_kojin.py（素材連結方式）の行写像から平行移動で導出していた。
    2026-08-13 に原本管理担当者が独自レイアウトの正本（Drive）を作成し、リポジトリ原本は
    scripts/patch_genpon_tsujo_kojin_from_drive.py がそのエクスポートから生成する方式に変わったため、
    「新原本の B列/A列ラベルから行番号を引き当てる」方式に書き換えた。
    原本・ヒアリングシートの行が変わったら、本スクリプトを再実行して差分を確認し、
    --emit の出力で config.py の定数を差し替える。

入力（リポジトリ内の現物。ファイル名固定）:
    ツール/【原本_個人】企業名_通常枠_個人2026.xlsx   （修正版原本 = patch スクリプトの出力）
    ツール/ヒアリングシート2026_通常枠個人.xlsx        （Drive 原本バイトの写し）

実行方法:
    python scripts/gen_mapping_tsujo_kojin.py           # 検証＋サマリのみ（書き込みなし）
    python scripts/gen_mapping_tsujo_kojin.py --emit 出力先.py   # config 用の定数ブロックを書き出す

検証で 1 件でも NG があれば生成中止（exit 1）。
"""
import argparse
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
GENPON = ROOT / 'ツール' / '【原本_個人】企業名_通常枠_個人2026.xlsx'
HEARING = ROOT / 'ツール' / 'ヒアリングシート2026_通常枠個人.xlsx'

HEARING_SCAN = (5, 104)   # ヒアリング基本情報のラベル走査範囲
TENKI_SCAN = (5, 100)     # 転記シートのラベル走査範囲
SHINSEI_SCAN = (5, 240)   # 申請内容シートのラベル走査範囲

# ── shinsei: フィールド → 申請内容 B列ラベル ──
# 引き当て規則: 正規化（NFKC＋全空白除去）後、まず完全一致、無ければ前方一致（一意のときだけ）。
# 同一ラベルが複数行に出る場合は (ラベル, 出現順) で指定する。
SHINSEI_FIELD_LABELS = [
    ('headquarters_address', '現在住所', 0),
    ('industry_code', '業種コード（数字４桁）', 0),
    ('industry_text', '業種_大分類/中分類/小分類/細分類', 0),
    ('capital', '資本金', 0),            # 基本情報側（1つ目）
    ('tool_name', 'ツール名', 0),
    ('business_description', '事業内容（255文字以内）', 0),
    ('fiscal_month', '決算月', 0),
    ('rep_name', '代表者氏名', 0),
    ('rep_kana', '代表者氏名（フリガナ）', 0),
    ('past_subsidies', '過去年度交付決定', 0),
    ('eruboshi', 'えるぼし認定', 0),
    ('kurumin', 'くるみん認定', 0),
    ('business_types', '行っている事業に該当するものすべて選択', 0),
    ('officer_count_prev', '代表者・役員数', 0),
    ('fin_revenue', '売上高', 0),
    ('fin_gross_profit', '粗利益', 0),
    ('fin_operating_profit', '営業利益', 0),
    ('fin_ordinary_profit', '経常利益', 0),
    ('fin_depreciation', '減価償却費', 0),
    ('fin_personnel', '人件費', 0),
    ('fin_capital', '資本金', 1),        # 財務側（2つ目）
    ('management_intent', '経営意欲', 0),
    ('strength', '強み', 0),             # 完全一致で「強み（転記）」とは区別される
    ('weakness', '弱み', 0),
    ('it_investment_amount', 'これまでのIT投資の年間金額', 0),
    ('it_investment_process', 'どのようなプロセスに対してＩＴ投資を行ったか', 0),
    ('security_status', 'セキュリティの状況', 0),
    ('future_goals', '事業をどのように変えていきますか？（将来目標）', 0),
    ('min_wage', '主たる事業場の所在地/地域別最低賃金', 0),
    ('min_wage_hourly', '事業所内最低賃金時給', 0),
    ('employee_count_fte', '従業員数（全期間在籍していない従業員は除外', 0),
    ('wage_total_base', '事業計画期間における給与支給総額', 0),
    ('wage_total_y1', '給与支給総額', 0),
    ('wage_total_y2', '給与支給総額', 1),
    ('wage_total_y3', '給与支給総額', 2),
    ('wage_raise_declaration', '上記の賃上げ計画を従業員へ表明しましたか？', 0),
    # 🔴 新原本は「賃上げ幅」が2箇所ある（r199=表明済み計画の幅 / r208=計画期間内に雇用する場合の幅）。
    #    ツールが書くのは表明ブロック側＝「賃上げ幅」（完全一致）。「賃上げ幅を選択」(r208) ではない。
    ('wage_raise_amount', '賃上げ幅', 0),
    ('wage_raise_method', '表明を行った方法', 0),
    ('wage_raise_date', '表明を行った日付', 0),
]

# ヒアリング⇔転記のラベルゆれ（正規化しても一致しないペアの明示対応）
TENKI_LABEL_ALIASES = {
    'GビズIDプライム': 'GビズID',   # ヒアリング側ラベル → 転記側ラベル
}

# 転記先が無いことが確認済みのヒアリング行（増減したら生成中止して原因を見る）
#   r103/r104（賃金状況の2問）: 坂平さん版の転記シートに着地行が無く、申請内容側にも
#   転記参照が無い（加点確認は申請内容の手動チェック欄で運用する設計）。回答はヒアリング
#   シート上で確認する。
EXPECTED_UNPAIRED_HEARING = {103, 104}

# 見出し・注記行（転記対象にしない）の判定
HEADING_PREFIXES = ('◼', '■', '【', '※', '⇨', '↓')


def norm(s) -> str:
    if s is None:
        return ''
    s = unicodedata.normalize('NFKC', str(s))
    return re.sub(r'\s+', '', s)


def first_line(s) -> str:
    return str(s).split('\n')[0] if s is not None else ''


def load_labels(path: Path, sheet: str, col: int, lo: int, hi: int):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    ws = wb[sheet]
    labels = []   # (row, raw_label)
    for r in range(lo, hi + 1):
        v = ws.cell(row=r, column=col).value
        if v is not None and str(v).strip():
            labels.append((r, str(v)))
    wb.close()
    return labels


def is_heading(label: str) -> bool:
    t = str(label).strip()
    return t.startswith(HEADING_PREFIXES)


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument('--emit', type=Path, default=None)
    args = ap.parse_args()

    ng = []

    # ══ 1. shinsei: 申請内容 B列ラベル → 行番号 ══
    shinsei_labels = load_labels(GENPON, '申請内容', 2, *SHINSEI_SCAN)
    by_exact = {}
    for r, raw in shinsei_labels:
        by_exact.setdefault(norm(first_line(raw)), []).append(r)
    full_norm = [(r, norm(raw)) for r, raw in shinsei_labels]

    shinsei = {}
    shinsei_comment = {}
    for field, label, nth in SHINSEI_FIELD_LABELS:
        key = norm(label)
        rows = by_exact.get(key, [])
        if not rows:
            # 前方一致（申請内容側のラベルが注記付きで長い場合）
            rows = [r for r, nl in full_norm if nl.startswith(key)]
        if len(rows) <= nth:
            ng.append(f'shinsei: {field} のラベル「{label}」(出現{nth + 1}番目) が見つからない (候補={rows})')
            continue
        shinsei[field] = rows[nth]
        raw = next(raw for r, raw in shinsei_labels if r == rows[nth])
        shinsei_comment[field] = first_line(raw)[:30]

    # 引き当ての重複チェック（別フィールドが同じ行を指していないか）
    seen_rows = {}
    for f, r in shinsei.items():
        if r in seen_rows:
            ng.append(f'shinsei: {f} と {seen_rows[r]} が同じ行 {r} を指している')
        seen_rows[r] = f

    # ══ 2. hearing_to_tenki: ラベル一致で再構成 ══
    hearing_labels = [(r, raw) for r, raw in load_labels(HEARING, '基本情報', 2, *HEARING_SCAN)]
    tenki_labels = [(r, raw) for r, raw in load_labels(GENPON, '転記', 1, *TENKI_SCAN)]

    tenki_by_label = {}
    tenki_by_first_line = {}
    for r, raw in tenki_labels:
        if is_heading(raw):
            continue
        tenki_by_label.setdefault(norm(raw), []).append(r)
        tenki_by_first_line.setdefault(norm(first_line(raw)), []).append(r)

    pairs = []          # (hearing_row, tenki_row, phone_flag, label)
    unpaired_hearing = []
    used_rows = set()   # 消費済みの転記行（同一ラベル複数行は行番号昇順に消費）
    for hr, raw in hearing_labels:
        if is_heading(raw):
            continue
        raw_key = str(raw).strip()
        if raw_key in TENKI_LABEL_ALIASES:
            key = norm(TENKI_LABEL_ALIASES[raw_key])
        else:
            key = norm(raw)
        rows = [r for r in tenki_by_label.get(key, []) if r not in used_rows]
        if not rows:
            # ラベルに選択肢や注記が続いて全文一致しないケース（例: 賃上げ幅選択の❶❷❸）は
            # 先頭行どうしで引き当てる
            rows = [r for r in tenki_by_first_line.get(norm(first_line(raw)), []) if r not in used_rows]
        if not rows:
            unpaired_hearing.append((hr, first_line(raw)[:40]))
            continue
        tr = rows[0]
        used_rows.add(tr)
        phone = ('電話番号' in key) or ('携帯番号' in key)
        pairs.append((hr, tr, phone, first_line(raw)[:30]))

    if {r for r, _ in unpaired_hearing} != EXPECTED_UNPAIRED_HEARING:
        ng.append(f'転記先の無いヒアリング行が想定 {sorted(EXPECTED_UNPAIRED_HEARING)} と異なる: '
                  f'{sorted(r for r, _ in unpaired_hearing)}')

    # 転記側で未使用の入力行（ラベルがあるのにヒアリングから来ない行）→ 想定外なら生成中止
    unused_tenki = [(r, first_line(raw)[:40]) for r, raw in tenki_labels
                    if not is_heading(raw) and r not in used_rows]
    if unused_tenki:
        ng.append(f'ヒアリングから値が来ない転記行がある（対応の判断が必要）: {sorted(r for r, _ in unused_tenki)}')

    # ══ 3. 申請内容の '転記'!B 参照が転記対象行を指しているかの整合 ══
    wb = openpyxl.load_workbook(GENPON, data_only=False)
    ws = wb['申請内容']
    tenki_target_rows = {t for _, t, _, _ in pairs}
    text_lo, text_hi = 28, 37   # 自由記述ゾーン（ヒアリング29-38 → 転記28-37）
    for row in ws.iter_rows(min_col=3, max_col=3):
        cell = row[0]
        v = cell.value
        if isinstance(v, str) and v.startswith('=') and '転記' in v:
            for m in re.finditer(r"転記'!B(\d+)", v):
                t = int(m.group(1))
                if t not in tenki_target_rows and not (text_lo <= t <= text_hi):
                    blabel = first_line(ws.cell(row=cell.row, column=2).value or '')[:30]
                    ng.append(f"申請内容 C{cell.row}「{blabel}」が転記B{t} を参照するが、転記対象行でない")

    # ══ 4. 給与計算シートの kyuyo 座標検証（変更なしを確認） ══
    KYUYO_EXPECT = [
        ('revenue', 10, 2, '売上高', 1), ('gross_profit', 11, 2, '粗利益', 1),
        ('operating_profit', 12, 2, '営業利益', 1), ('ordinary_profit', 13, 2, '経常利益', 1),
        ('depreciation', 21, 5, '減価償却費', 4), ('salary', 5, 5, '給料手当', 4),
        ('misc_wages', 6, 5, '雑給', 4), ('bonus', 7, 5, '賞与手当', 4),
        ('travel_expense', 9, 5, '旅費交通費', 4),
    ]
    ky = wb['生産性指標給与支給総額計算']
    for name, r, c, label, label_col in KYUYO_EXPECT:
        actual = norm(ky.cell(row=r, column=label_col).value)
        if actual != norm(label):
            ng.append(f'kyuyo: {name} のラベル位置 r{r}c{label_col} が「{label}」でない（実際:「{actual[:20]}」）')

    # ══ 5. hearing_wage_raise_row ══
    wage_raise_row = None
    for hr, raw in hearing_labels:
        if norm(raw).startswith(norm('賃金引上げ計画における地域別最低賃金に対する賃上げ幅')):
            wage_raise_row = hr
            break
    if wage_raise_row is None:
        ng.append('hearing_wage_raise_row: 賃上げ幅選択の設問がヒアリングに見つからない')
    wb.close()

    # ══ レポート ══
    print(f'shinsei: {len(shinsei)}/{len(SHINSEI_FIELD_LABELS)} フィールド引き当て')
    print(f'hearing_to_tenki: {len(pairs)} ペア')
    if unpaired_hearing:
        print('\nヒアリング側で転記先が無い行（見出し以外）:')
        for r, l in unpaired_hearing:
            print(f'  hearing r{r}: {l}')
    if unused_tenki:
        print('\n転記側でヒアリングから来ない行:')
        for r, l in sorted(unused_tenki):
            print(f'  転記 r{r}: {l}')
    if ng:
        print(f'\nNG {len(ng)} 件 → 生成中止:')
        for x in ng:
            print(' ', x)
        return 1

    # ══ 出力ブロック生成 ══
    lines = []
    lines.append('MAPPING_2026_TSUJO_KOJIN = TemplateMapping(')
    lines.append('    hearing_to_tenki=[')
    lines.append('        # (ヒアリング行, 転記行, 電話番号変換)')
    for hr, tr, phone, label in pairs:
        lines.append(f'        ({hr}, {tr}, {phone}),   # {label}')
    lines.append('    ],')
    lines.append('    shinsei={')
    for field, _, _ in SHINSEI_FIELD_LABELS:
        if field in shinsei:
            lines.append(f"        '{field}': {shinsei[field]},   # {shinsei_comment[field]}")
    lines.append('    },')
    lines.append("    kyuyo_sheet_name='生産性指標給与支給総額計算',")
    lines.append("    kyuyo={'revenue': (10, 2), 'gross_profit': (11, 2), 'operating_profit': (12, 2), "
                 "'ordinary_profit': (13, 2), 'depreciation': (21, 5), 'salary': (5, 5), "
                 "'misc_wages': (6, 5), 'bonus': (7, 5), 'travel_expense': (9, 5)},")
    lines.append('    shinsei_clear_range=(5, 245),')
    lines.append('    tenki_text_range=(28, 38),')
    lines.append('    is_kojin=True,')
    lines.append('    preserve_rows=[147, 149, 153, 154, 155],')
    lines.append(f'    hearing_wage_raise_row={wage_raise_row},')
    lines.append(')')
    block = '\n'.join(lines)

    if args.emit:
        args.emit.write_text(block + '\n', encoding='utf-8')
        print(f'\n--emit: {args.emit} に書き出しました')
    else:
        print('\n（--emit 未指定のためブロックは表示のみ）\n')
        print(block)
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
