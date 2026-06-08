# -*- coding: utf-8 -*-
"""インボイス枠・法人テンプレの「従業員代表者」「事業所内最低賃金者」欠陥を修正する。

背景:
    申請内容シートの r211(従業員代表者) / r213(事業所内最低賃金者) が B:E 結合され、
    かつ転記シート参照式が欠落していたため両欄が空欄表示になっていた。
    （氏名は転記シート B76/B78 に書き込まれているのに、申請内容が引けていない）
    正常な兄弟 r212(給与担当者) は C212="='転記'!B77" を持ち表示できている。
    v0.2.59 が粗利益(r149)・セキュリティ(r161)の同種欠陥を直した際、式で引く設計の
    この2セルは config.shinsei のマップ先に無く、監査スキャンの網から漏れていた。

修正内容:
    - B211:E211 / B213:E213 の結合を解除
    - C211="='転記'!B76" / C213="='転記'!B78" を復元
    - 正常な兄弟 r212 から C/D/E の書式と行高を複製（値は複製しない）
    - D211/D213 の注記は通常枠の正規文を移植（給与担当者向け注記の流用を避ける）

実行方法:
    python scripts/patch_invoice_corp_rep_cells.py

冪等性:
    既に修正済みなら結合解除・式設定はそのまま通り、再実行しても結果は同じ。
"""
import sys
from copy import copy
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
INV_CORP = ROOT / 'ツール' / '【原本_法人】企業名_インボイス枠_法人2026_v2.xlsx'
NORMAL_CORP = ROOT / 'ツール' / '【原本_法人】企業名_通常枠_法人2026_v2.xlsx'

SHEET = '申請内容'
SIBLING_ROW = 212          # 給与担当者（正常な兄弟）

# (対象行, 復元する転記参照式, 通常枠の正規注記の取得元行)
FIXES = [
    (211, "='転記'!B76", 226),   # 従業員代表者   ← 通常枠 D226
    (213, "='転記'!B78", 228),   # 事業所内最低賃金者 ← 通常枠 D228
]
STYLE_COLS = (3, 4, 5)         # C / D / E（B見出しセルは触らない）


def _copy_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


def main() -> None:
    # 通常枠から正規の注記テキストを読む
    wb_n = openpyxl.load_workbook(NORMAL_CORP, data_only=False)
    ws_n = wb_n[SHEET]
    notes = {}
    for target_row, _, note_src_row in FIXES:
        notes[target_row] = ws_n.cell(note_src_row, 4).value
    wb_n.close()

    wb = openpyxl.load_workbook(INV_CORP, data_only=False)
    ws = wb[SHEET]

    sibling_height = ws.row_dimensions[SIBLING_ROW].height

    for target_row, formula, _ in FIXES:
        rng = f'B{target_row}:E{target_row}'
        # 結合解除（既に解除済みでも例外を握りつぶす）
        for mr in list(ws.merged_cells.ranges):
            if str(mr) == rng:
                ws.unmerge_cells(rng)

        # C/D/E の書式を兄弟 r212 から複製（値は入れない）
        for col in STYLE_COLS:
            _copy_style(ws.cell(SIBLING_ROW, col), ws.cell(target_row, col))

        # 行高を兄弟に合わせる
        if sibling_height is not None:
            ws.row_dimensions[target_row].height = sibling_height

        # 値の復元：C列＝転記参照式、D列＝通常枠の正規注記
        ws.cell(target_row, 3).value = formula
        ws.cell(target_row, 4).value = notes[target_row]

    wb.save(INV_CORP)

    # 事後検証
    wb2 = openpyxl.load_workbook(INV_CORP, data_only=False)
    ws2 = wb2[SHEET]
    merged = {str(m) for m in ws2.merged_cells.ranges}
    print('=== 修正後の検証 ===')
    ok = True
    for target_row, formula, _ in FIXES:
        c = ws2.cell(target_row, 3).value
        b = ws2.cell(target_row, 2).value
        rng = f'B{target_row}:E{target_row}'
        is_unmerged = rng not in merged
        is_formula = c == formula
        label_kept = isinstance(b, str) and b in ('従業員代表者', '事業所内最低賃金者')
        print(f'  r{target_row}: B={b!r} C={c!r} 非結合={is_unmerged} 式一致={is_formula} 見出し維持={label_kept}')
        ok = ok and is_unmerged and is_formula and label_kept
    # 既存の正常セルが壊れていないか
    print(f'  r212 給与担当者 C={ws2.cell(212, 3).value!r}（不変であること）')
    # 粗利益/セキュリティが不変か（v0.2.59修正の非回帰）
    print(f'  r149 粗利益 B={ws2.cell(149, 2).value!r}（"粗利益"のままであること）')
    print(f'  r161 セキュリティ B={ws2.cell(161, 2).value!r}（"セキュリティの状況"のままであること）')
    wb2.close()
    print('=== 結果:', 'OK' if ok else 'NG（要確認）', '===')
    if not ok:
        sys.exit(1)


if __name__ == '__main__':
    main()
