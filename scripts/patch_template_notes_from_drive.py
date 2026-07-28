# -*- coding: utf-8 -*-
"""Drive 版原本の「記入ガイド文・注記」だけをリポジトリ側テンプレへ取り込む。

背景:
    compare_drive_template.py で 2026-07-28 に実測したところ、Drive 原本には
    リポジトリ側に無い記入ガイドが増えていた（前期決算期末の人数、申請MTG時の
    選択指示、セキュリティ設問の選択肢、経理担当不在時の代替 など）。
    通常枠 B1 には「セカンドオピニオンの面談は完了していますか？」という
    提出前の警告も追加されていた。

    ただし Drive 版を丸ごと採用はできない（行番号固定のマッピングが崩れる／
    Drive 側に #REF! とプルダウン欠落が残る／リポジトリ側だけが持つ修正が
    巻き戻る）。そこで本スクリプトは **ツールの振る舞いを変えない差分だけ** を
    許可リストで明示して取り込む。

取り込むもの（許可リスト = NOTE_CELLS）:
    - 記入ガイド・注記セル（D列）と、通常枠 B1 の提出前警告
    - リポジトリ側のタイポ修正（「プルダウンを選択選択」→「選択」）

取り込まないもの（意図的に除外。理由は EXCLUDED を参照）:
    - 費用内訳マスタ / シート9 / ツールマスタ の新シートとその VLOOKUP 配線
      → ツール側（config のマッピング）の対応が要るため別タスク
    - リポジトリ側にだけある URL セル（Drive 側は空。取り込むと情報が消える）
    - 役員（8）→（８）の全角化（リポジトリの半角統一を維持）
    - インボイス法人 D164/D165（リポジトリのツール別記入例のほうが実務的）
    - 通常枠 D240（Drive はラベル、リポジトリは別セルに実URL。構造が違う）

実行方法:
    python scripts/patch_template_notes_from_drive.py <Drive版通常枠.xlsx> <Drive版インボイス法人.xlsx>

    引数は順不同。ファイル名で通常枠／インボイス法人を判別する。
    インボイス個人は差分ゼロのため対象外。

冪等性:
    再実行しても同じ最終状態に収束する（同値上書き）。差分が無ければ書き込まない。
"""
import shutil
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
TOOL_DIR = ROOT / 'ツール'
SHEET = '申請内容'

# 取り込む対象。テンプレ種別 -> (リポジトリ版ファイル名, 対象セル)
# セルは Drive 版の値をそのままコピーする（値はここに直書きしない＝原本が単一の真実）。
NOTE_CELLS = {
    '通常枠': (
        '【原本_法人】企業名_通常枠_法人2026_v2.xlsx',
        # B1=提出前警告 / D151-153=人数の注記 / D168・D182=選択肢の明示
        # D178=セキュリティ設問の選択肢 / D179・D181=タイポ修正 / D180・D183・D184=選択指示
        # D225=日付の注記 / D227=担当者欄の代替指示
        ('B1', 'D151', 'D152', 'D153', 'D168', 'D178', 'D179', 'D180',
         'D181', 'D182', 'D183', 'D184', 'D225', 'D227'),
    ),
    'インボイス法人': (
        '【原本_法人】企業名_インボイス枠_法人2026_v2.xlsx',
        # D143-145=人数の注記 / D187=画面共有の手順注記
        ('D143', 'D144', 'D145', 'D187'),
    ),
}

# 除外理由の記録（レビュー時に「なぜ入れていないか」を追えるようにする）
EXCLUDED = {
    '費用内訳マスタ・シート9・ツールマスタと配線': 'ツール側マッピングの対応が要る（別タスク）',
    'URL セル（通常枠 D114/D115/D204/D231・インボイス法人 D189/D217/D228）': 'Drive 側が空。取り込むと実URLが消える',
    '役員（8）→（８）の全角化': 'リポジトリの半角統一を維持（表示のみ・実害なし）',
    'インボイス法人 D164/D165': 'リポジトリのツール別記入例のほうが実務的',
    '通常枠 D240': 'Drive はラベル、リポジトリは別セルに実URL。構造が異なる',
}

# Drive 版ファイルの判別キーワード（エクスポート名に含まれる語）
DRIVE_KEYWORDS = {'通常枠': ('通常枠',), 'インボイス法人': ('インボイス枠', '法人')}


def _nfc(s: str) -> str:
    return unicodedata.normalize('NFC', s)


def resolve_in_dir(base: Path, name: str) -> Path | None:
    target = _nfc(name)
    for p in base.iterdir():
        if _nfc(p.name) == target:
            return p
    return None


def _merged_ranges_by_cell(ws) -> dict[str, object]:
    """座標 -> その座標を含む結合範囲。結合されていない座標は持たない。"""
    out = {}
    for rng in ws.merged_cells.ranges:
        for row in ws[str(rng)]:
            for c in row:
                out[c.coordinate] = rng
    return out


def classify(drive_path: Path) -> str | None:
    name = _nfc(drive_path.name)
    for kind, words in DRIVE_KEYWORDS.items():
        if all(_nfc(w) in name for w in words):
            return kind
    return None


def apply_one(kind: str, drive_path: Path) -> int:
    repo_name, cells = NOTE_CELLS[kind]
    repo_path = resolve_in_dir(TOOL_DIR, repo_name)
    if repo_path is None:
        print(f'❌ リポジトリ版が見つかりません: {repo_name}', file=sys.stderr)
        return -1

    ws_drive = openpyxl.load_workbook(drive_path, data_only=False)[SHEET]
    wb_repo = openpyxl.load_workbook(repo_path)
    ws_repo = wb_repo[SHEET]
    merged_at = _merged_ranges_by_cell(ws_repo)

    changes, skipped = [], []
    for coord in cells:
        new = ws_drive[coord].value
        old = ws_repo[coord].value
        rng = merged_at.get(coord)
        # 結合範囲の左上以外は書き込めない（openpyxl の MergedCell は読み取り専用）。
        # 結合を解くとレイアウトが変わるため、ここでは触らず一覧に出して判断に回す。
        if rng is not None and str(rng).split(':')[0] != coord:
            skipped.append((coord, str(rng), new))
            continue
        if old == new:
            continue
        ws_repo[coord] = new
        changes.append((coord, old, new))

    print(f'\n=== {kind} ({repo_path.name})')
    if skipped:
        print(f'  結合のため見送り: {len(skipped)} セル（解除するとレイアウトが変わるため別判断）')
        for coord, rng, new in skipped:
            print(f'    {coord} (結合 {rng}) に入るはずの Drive 値: {repr(new)[:70]}')
    if not changes:
        print('  取り込む差分なし（すでに反映済み）')
        wb_repo.close()
        return 0

    backup = repo_path.with_suffix('.xlsx.bak')
    shutil.copy2(repo_path, backup)
    wb_repo.save(repo_path)
    wb_repo.close()

    print(f'  バックアップ: {backup.name}')
    for coord, old, new in changes:
        print(f'  {coord}: {repr(old)[:60]} -> {repr(new)[:90]}')
    print(f'  {len(changes)} セルを更新')
    return len(changes)


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 1

    targets: dict[str, Path] = {}
    for raw in argv[1:]:
        p = Path(raw)
        if not p.exists():
            found = resolve_in_dir(p.parent, p.name) if p.parent.exists() else None
            if found is None:
                print(f'❌ 見つかりません: {p}', file=sys.stderr)
                return 1
            p = found
        kind = classify(p)
        if kind is None:
            print(f'⚠ 種別を判別できずスキップ: {p.name}', file=sys.stderr)
            continue
        targets[kind] = p

    if not targets:
        print('❌ 対象が0件です', file=sys.stderr)
        return 1

    total = 0
    for kind, path in targets.items():
        n = apply_one(kind, path)
        if n < 0:
            return 1
        total += n

    print(f'\n合計 {total} セルを取り込みました。')
    print('\n意図的に取り込まなかったもの:')
    for what, why in EXCLUDED.items():
        print(f'  - {what}: {why}')
    return 0


if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
