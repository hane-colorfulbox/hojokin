# -*- coding: utf-8 -*-
"""法人テンプレ2種の申請内容シートのプルダウン（データ検証）位置ずれを根治する。

背景:
    2026-04-06 の役員10名対応(v2)で申請内容シートに9行挿入した際、セルのラベルと
    config.py の行マッピングは+9行下げたが、データ検証（プルダウン）の適用範囲
    (sqref) だけ旧行のまま据え置かれた。結果、役員挿入点より下の全プルダウンが
    ラベルに対してきっかり9行上にずれて出力される（TODO T03。2026-07-09 に
    顧客案件のインボイス枠実出力で指摘を受け、実測で全数確認済み）。
    個人テンプレは正常（ずれ無し）で、修正後の整合状態の基準。

修正内容:
    - ずれている DV の sqref を +9 行の正位置へ付け替える（削除・追加はしない）
    - 通常枠 C135 の DV は formula1 が #REF! に壊れているため、インボイス法人と
      同一の業種リスト参照（'プルダウン用'!$C$2:$C$20。通常枠にも同リスト現存）に復旧
    - 役員挿入点より上の正常な DV（ツール名/決算月/担当者/設立年月日アンカー）は不変

実行方法:
    python scripts/patch_corp_template_dv_shift.py

冪等性:
    sqref が既に正位置なら skip。再実行しても結果は同じ。
"""
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils.cell import range_boundaries

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
SHEET = '申請内容'
BUSINESS_LIST_FORMULA = "'プルダウン用'!$C$2:$C$20"

# 各修正: (現sqref, 正sqref, formula1修復値 or None, {検証: 正位置セル→B列ラベル断片})
FIXES = {
    ROOT / 'ツール' / '【原本_法人】企業名_インボイス枠_法人2026_v2.xlsx': [
        ('C103', 'C112', None, {112: '過去年度交付決定'}),
        ('C107', 'C116', None, {116: 'えるぼし認定'}),
        ('C108', 'C117', None, {117: 'くるみん認定'}),
        ('C109 C121', 'C118 C130', None, {118: '省力化ナビ', 130: 'IT戦略ナビwithの実施有無'}),
        ('C128', 'C137', None, {137: '行っている事業'}),
        ('C189', 'C198', None, {198: '主たる事業場の所在地'}),
        ('C48 C201', 'C48 C210', None, {210: '表明を行った日付'}),
        ('C199 C208', 'C208 C217', None, {208: '賃上げ幅を選択', 217: '賃上げ幅を選択'}),
    ],
    ROOT / 'ツール' / '【原本_法人】企業名_通常枠_法人2026_v2.xlsx': [
        ('C113', 'C122', None, {122: '過去年度交付決定'}),
        ('C114', 'C123', None, {123: 'えるぼし認定'}),
        ('C115', 'C124', None, {124: 'くるみん認定'}),
        ('C116 C128', 'C125 C137', None, {125: '省力化ナビ', 137: 'IT戦略ナビwithの実施有無'}),
        ('C135', 'C144', BUSINESS_LIST_FORMULA, {144: '行っている事業'}),
        ('C161', 'C170', None, {170: '強み'}),
        ('C163', 'C172', None, {172: '弱み'}),
        ('C168', 'C177', None, {177: 'どのようなプロセスに対してＩＴ投資'}),
        ('C170', 'C179', None, {179: '改善したい業務プロセス'}),
        ('C171', 'C180', None, {180: '強化したい部門'}),
        ('C172', 'C181', None, {181: 'どんな効果を期待します'}),
        ('C204', 'C213', None, {213: '主たる事業場の所在地'}),
        ('C56:C57 C216', 'C56:C57 C225', None, {225: '表明を行った日付'}),
        ('C214 C223', 'C223 C232', None, {223: '賃上げ幅', 232: '賃上げ幅を選択'}),
    ],
}

# 触ってはいけない正常DV（修正後も存続していることを検証）
KEEP_SQREFS = {
    ROOT / 'ツール' / '【原本_法人】企業名_インボイス枠_法人2026_v2.xlsx': ['C61', 'C64', 'C2:C3'],
    ROOT / 'ツール' / '【原本_法人】企業名_通常枠_法人2026_v2.xlsx': ['C71', 'C74', 'C2:C3'],
}


def _cells_of(sqref: str) -> set[str]:
    cells = set()
    for token in str(sqref).split():
        c0, r0, c1, r1 = range_boundaries(token)
        for r in range(r0, r1 + 1):
            for c in range(c0, c1 + 1):
                cells.add(f'{openpyxl.utils.get_column_letter(c)}{r}')
    return cells


def patch_file(path: Path, fixes: list, keep: list[str]) -> bool:
    print(f'\n===== {path.name} =====')
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb[SHEET]
    dvs = list(ws.data_validations.dataValidation)
    n_before = len(dvs)
    merged_before = {str(m) for m in ws.merged_cells.ranges}
    sheets_before = list(wb.sheetnames)

    # 連鎖移動（例: C161→C170 と C170→C179）があるため、fix単位でなくファイル単位で
    # 冪等判定する。全fixの旧sqrefが揃っていれば未適用、全fixの新sqrefが揃っていれば
    # 適用済み。混在はテンプレが想定と違うので中止。
    by_sqref = {str(dv.sqref): dv for dv in dvs}
    sqrefs_now = set(by_sqref)
    if all(new in sqrefs_now for _o, new, _f, _l in fixes):
        print(f'  済み: 全{len(fixes)}件が既に正位置（変更なし）')
        wb.close()
        moved = 0
    elif all(old in sqrefs_now for old, _n, _f, _l in fixes):
        moved = 0
        for old, new, formula_fix, _labels in fixes:
            dv = by_sqref[old]
            dv.sqref = new
            if formula_fix is not None:
                dv.formula1 = formula_fix
            moved += 1
            note = f' formula1→{formula_fix}' if formula_fix else ''
            print(f'  移動: {old} → {new}{note}')
        wb.save(path)
        wb.close()
    else:
        missing = [old for old, _n, _f, _l in fixes if old not in sqrefs_now]
        print(f'  !! 適用済み/未適用が混在（不在の旧sqref: {missing}）。テンプレが想定と違うため中止')
        wb.close()
        return False

    # ---- 事後検証 ----
    wb2 = openpyxl.load_workbook(path, data_only=False)
    ws2 = wb2[SHEET]
    dvs2 = list(ws2.data_validations.dataValidation)
    sqrefs2 = {str(dv.sqref) for dv in dvs2}
    ok = True

    if len(dvs2) != n_before:
        print(f'  NG: DV総数 {n_before} → {len(dvs2)}')
        ok = False
    if {str(m) for m in ws2.merged_cells.ranges} != merged_before:
        print('  NG: 結合範囲が変化')
        ok = False
    if list(wb2.sheetnames) != sheets_before:
        print('  NG: シート構成が変化')
        ok = False

    old_cells = set()
    new_cells = set()
    for old, new, _f, labels in fixes:
        old_cells |= _cells_of(old)
        new_cells |= _cells_of(new)
        if new not in sqrefs2:
            print(f'  NG: 正位置 {new} にDVが無い')
            ok = False
        for row, frag in labels.items():
            label = ws2.cell(row, 2).value
            if not (isinstance(label, str) and frag in label):
                print(f'  NG: B{row} ラベル不一致（{frag!r} を含まない: {label!r}）')
                ok = False
    # 旧位置に迷子DVが残っていないか（新位置と重なるセルは除外して判定）
    stray = (old_cells - new_cells) & set().union(*(_cells_of(s) for s in sqrefs2))
    if stray:
        print(f'  NG: 旧位置にDVが残存: {sorted(stray)}')
        ok = False
    for s in keep:
        if s not in sqrefs2:
            print(f'  NG: 正常DV {s} が消失')
            ok = False
    if not any(dv.formula1 == BUSINESS_LIST_FORMULA for dv in dvs2) and \
            any(f == BUSINESS_LIST_FORMULA for _o, _n, f, _l in fixes if f):
        print('  NG: 業種リスト formula1 修復が反映されていない')
        ok = False
    if any('#REF!' in str(dv.formula1) for dv in dvs2):
        print('  NG: #REF! のDVが残存')
        ok = False

    wb2.close()
    print(f'  移動{moved}件 / DV総数 {n_before}→{len(dvs2)} / 検証 {"OK" if ok else "NG"}')
    return ok


def main() -> None:
    all_ok = True
    for path, fixes in FIXES.items():
        all_ok = patch_file(path, fixes, KEEP_SQREFS[path]) and all_ok
    print('\n=== 結果:', 'OK' if all_ok else 'NG（要確認）', '===')
    if not all_ok:
        sys.exit(1)


if __name__ == '__main__':
    main()
