# -*- coding: utf-8 -*-
"""
賃金台帳「総額提出用」テンプレート生成

賃金台帳（個人別の給与明細）の提出をためらう顧客向けに、会社全体の集計値だけを
提出してもらう Excel を生成する。顧客は「計算用（社内控え・非提出）」シートに個人別
データを入力し、シートが補助金ルール（デジタル化・AI導入補助金2026）に沿って自動で
対象者を絞り、R215（従業員数FTE）/ R216（給与支給総額）を算出する。顧客は「提出用
シート」の集計値だけをカラフルボックスへ渡す（個人明細は社内保管・非提出）。

算定ロジックは hojokin/wage_calculator.py と一致させている:
- R216 = Σ(対象者の[Σ月次総支給 − 通勤手当年額 + 課税賞与年額])
  ※通勤手当は課税・非課税を問わず控除する（2026-07-27 社内運用）
- R215 = 正社員・契約社員は各1.0、パートは 年間労働時間 ÷ 2080（STANDARD_ANNUAL_HOURS）
- 対象者 = 役員でない かつ 12ヶ月すべて支給>0（役員・中途・退職・休職・0円月は除外）
- 加点①②は個人×暦月データが必須のため本様式では扱わない（別テンプレ）

ツールとの意図的な差（2026-07-21 差動テスト後の設計判断）:
- 入力不備（雇用形態未選択/リスト外・パートの労働時間未入力・課税支給年額≤0）は、
  ツールのようにフォールバック（正社員扱い・FTE=1.0昇格）せず「要確認」として
  集計から外し、警告を提出用シートまで伝搬する。提出物なので無言の数字を出さない。
- 通勤手当の月割減算で月が0円化する超レアケースはツール＝丸ごと除外（W-WAGE-004 で警告）
  ／本様式＝年額差引のまま算入（公募要領は月割配賦を規定していないため様式側を正とする）。

使い方:
    python scripts/build_wage_summary_template.py            # 空テンプレを ツール/ に出力
    python scripts/build_wage_summary_template.py --sample   # 記入例入り（目視確認用）を出力
"""
import sys

sys.stdout.reconfigure(encoding='utf-8')

import argparse
import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.protection import SheetProtection

# ── 出力先 ──
OUTPUT_DIR = Path(__file__).resolve().parent.parent / 'ツール'
OUTPUT_NAME = '賃金台帳（総額提出用）テンプレート.xlsx'

# ── 算定パラメータ（wage_calculator と一致させる）──
STANDARD_ANNUAL_HOURS = 2080  # FTE換算の分母（config.STANDARD_ANNUAL_HOURS と一致）

# ── レイアウト ──
DATA_ROWS = 30
PERIOD_ROW = 3
PERIOD_WARN_ROW = 4
HEADER_ROW = 6
FIRST_DATA_ROW = 7
LAST_DATA_ROW = FIRST_DATA_ROW + DATA_ROWS - 1  # 36

# 集計ブロック（明細の下）
SUM_TITLE_ROW = LAST_DATA_ROW + 2         # 38
R216_ROW = SUM_TITLE_ROW + 1              # 39 給与支給総額
R215_ROW = SUM_TITLE_ROW + 2              # 40 従業員数FTE
FT_COUNT_ROW = SUM_TITLE_ROW + 3          # 41 フルタイム人数
PART_COUNT_ROW = SUM_TITLE_ROW + 4        # 42 パート人数
PART_HOURS_ROW = SUM_TITLE_ROW + 5        # 43 パート年間労働時間合計
TARGET_COUNT_ROW = SUM_TITLE_ROW + 6      # 44 対象者合計
PER_CAPITA_ROW = SUM_TITLE_ROW + 7        # 45 1人当たり
TOTAL_HOURS_ROW = SUM_TITLE_ROW + 8       # 46 対象者の年間総労働時間（参考）
CHECK_ROW = SUM_TITLE_ROW + 9             # 47 要確認・警告件数
DEN_ROW = SUM_TITLE_ROW + 10              # 48 分母（2080）

# ── 列（計算用シート）──
COL_NO = 1          # A
COL_NAME = 2        # B
COL_EMP = 3         # C
COL_M1 = 4          # D 1ヶ月目
COL_M12 = 15        # O 12ヶ月目
COL_BONUS = 16      # P 年間賞与
COL_TRANSPORT = 17  # Q 年間通勤手当
COL_HOURS = 18      # R 年間労働時間
COL_JUDGE = 19      # S 判定
COL_TAXABLE = 20    # T 課税支給年額
COL_FTE = 21        # U FTE
COL_WARN = 22       # V 警告
COL_LAST = COL_WARN

# 列文字（数式生成用）
L_NAME = get_column_letter(COL_NAME)        # B
L_EMP = get_column_letter(COL_EMP)          # C
L_M1 = get_column_letter(COL_M1)            # D
L_M12 = get_column_letter(COL_M12)          # O
L_BONUS = get_column_letter(COL_BONUS)      # P
L_TRANSPORT = get_column_letter(COL_TRANSPORT)  # Q
L_HOURS = get_column_letter(COL_HOURS)      # R
L_JUDGE = get_column_letter(COL_JUDGE)      # S
L_TAXABLE = get_column_letter(COL_TAXABLE)  # T
L_FTE = get_column_letter(COL_FTE)          # U
L_WARN = get_column_letter(COL_WARN)        # V
L_PERIOD_START = get_column_letter(COL_M1)          # D3
L_PERIOD_END = get_column_letter(COL_M1 + 2)        # F3

# ── 雇用形態 ──
EMP_TYPES = ['正社員', '契約社員', 'パート・アルバイト', '役員']
FULLTIME_TYPES = ['正社員', '契約社員']
PART_TYPE = 'パート・アルバイト'
OFFICER_TYPE = '役員'

JUDGE_TARGET = '集計対象'
JUDGE_OFFICER = '対象外（役員）'
JUDGE_NOTFULL = '対象外（全月支給なし）'
JUDGE_NEED_TYPE = '要確認（雇用形態未選択）'
JUDGE_NEED_HOURS = '要確認（労働時間未入力）'
JUDGE_NEED_AMOUNT = '要確認（金額の確認）'

WARN_NEED_TYPE = '⚠雇用形態をプルダウンから選択してください'
WARN_NEED_HOURS = '⚠パート・アルバイトは年間労働時間の入力が必要です'
WARN_NEED_AMOUNT = '⚠年間通勤手当が支給合計を上回っています。金額をご確認ください'
WARN_HOURS_OVER = '⚠年間労働時間が正社員の年間所定労働時間を超えています。桁をご確認ください'
WARN_PERIOD_EMPTY = '⚠対象期間が未入力です（決算期の開始日と終了日をご記入ください）'
WARN_PERIOD_NOT12 = '⚠対象期間が12ヶ月ちょうどになっていません（例：2025/7/1 〜 2026/6/30）'
WARN_PERIOD_BADDATE = '⚠対象期間の日付が読み取れません（日付形式でご入力ください）'

# ── 色 ──
C_HEADER = '1F4E78'   # 紺（ヘッダー）
C_INPUT = 'FFFF00'    # 黄（入力欄）
C_CALC = 'E2EFDA'     # 薄緑（自動計算）
C_SAMPLE = 'F2F2F2'   # 薄グレー（記入例）
C_NOTE = 'FFF2CC'     # 薄オレンジ（注意）

# ── フォント／罫線 ──
FONT = '游ゴシック'
F_TITLE = Font(name=FONT, size=14, bold=True)
F_SECTION = Font(name=FONT, size=11, bold=True, color=C_HEADER)
F_HEADER = Font(name=FONT, size=9, bold=True, color='FFFFFF')
F_NORMAL = Font(name=FONT, size=10)
F_SMALL = Font(name=FONT, size=9)
F_SMALL_GRAY = Font(name=FONT, size=9, color='808080')
F_BOLD = Font(name=FONT, size=10, bold=True)
F_RESULT = Font(name=FONT, size=12, bold=True, color='C00000')
F_NOTE = Font(name=FONT, size=10, color='974706')
F_ALERT = Font(name=FONT, size=10, bold=True, color='C00000')

THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

NUMFMT_YEN = '#,##0'
NUMFMT_HOUR = '#,##0'
NUMFMT_FTE = '0.0'
NUMFMT_DATE = 'yyyy/m/d'  # 月次列(幅10)に収まる表示。###### 化を避ける

FILL_HEADER = PatternFill(start_color=C_HEADER, end_color=C_HEADER, fill_type='solid')
FILL_INPUT = PatternFill(start_color=C_INPUT, end_color=C_INPUT, fill_type='solid')
FILL_CALC = PatternFill(start_color=C_CALC, end_color=C_CALC, fill_type='solid')
FILL_SAMPLE = PatternFill(start_color=C_SAMPLE, end_color=C_SAMPLE, fill_type='solid')
FILL_NOTE = PatternFill(start_color=C_NOTE, end_color=C_NOTE, fill_type='solid')

UNLOCKED = Protection(locked=False)
LOCKED = Protection(locked=True)

CALC_SHEET = '計算用（社内控え・非提出）'
SUBMIT_SHEET = '提出用シート（総額のみ）'
HELP_SHEET = '記入方法'

# 誤操作防止の保護（パスワードなし。列幅・行高の調整は顧客に開放する）
def sheet_protection() -> SheetProtection:
    return SheetProtection(
        sheet=True, selectLockedCells=False, selectUnlockedCells=False,
        formatCells=False, formatColumns=False, formatRows=False,
    )


# ── 数式 ──
def f_judge(r: int) -> str:
    """判定。入力不備は「要確認」で集計から外す（無言のフォールバックはしない）。"""
    type_ok = (
        f'OR({L_EMP}{r}="{FULLTIME_TYPES[0]}",{L_EMP}{r}="{FULLTIME_TYPES[1]}",'
        f'{L_EMP}{r}="{PART_TYPE}")'
    )
    full_year = (
        f'AND(COUNT({L_M1}{r}:{L_M12}{r})=12,'
        f'COUNTIF({L_M1}{r}:{L_M12}{r},">0")=12)'
    )
    taxable = (
        f'SUM({L_M1}{r}:{L_M12}{r})-N({L_TRANSPORT}{r})+N({L_BONUS}{r})'
    )
    return (
        f'=IF({L_NAME}{r}="","",'
        f'IF({L_EMP}{r}="{OFFICER_TYPE}","{JUDGE_OFFICER}",'
        f'IF(NOT({type_ok}),"{JUDGE_NEED_TYPE}",'
        f'IF(NOT({full_year}),"{JUDGE_NOTFULL}",'
        f'IF(AND({L_EMP}{r}="{PART_TYPE}",N({L_HOURS}{r})<=0),"{JUDGE_NEED_HOURS}",'
        f'IF({taxable}<=0,"{JUDGE_NEED_AMOUNT}","{JUDGE_TARGET}"))))))'
    )


def f_taxable(r: int) -> str:
    return (
        f'=IF({L_JUDGE}{r}="{JUDGE_TARGET}",'
        f'SUM({L_M1}{r}:{L_M12}{r})-N({L_TRANSPORT}{r})+N({L_BONUS}{r}),"")'
    )


def f_fte(r: int) -> str:
    # 判定=集計対象 の時点で雇用形態は3種のいずれか・パートは時間>0 が保証される
    return (
        f'=IF({L_JUDGE}{r}<>"{JUDGE_TARGET}",0,'
        f'IF(OR({L_EMP}{r}="{FULLTIME_TYPES[0]}",{L_EMP}{r}="{FULLTIME_TYPES[1]}"),1,'
        f'N({L_HOURS}{r})/IF(N($C${DEN_ROW})>0,$C${DEN_ROW},{STANDARD_ANNUAL_HOURS})))'
    )


def f_warn(r: int) -> str:
    return (
        f'=IF({L_JUDGE}{r}="{JUDGE_NEED_TYPE}","{WARN_NEED_TYPE}",'
        f'IF({L_JUDGE}{r}="{JUDGE_NEED_HOURS}","{WARN_NEED_HOURS}",'
        f'IF({L_JUDGE}{r}="{JUDGE_NEED_AMOUNT}","{WARN_NEED_AMOUNT}",'
        f'IF(AND({L_JUDGE}{r}="{JUDGE_TARGET}",{L_EMP}{r}="{PART_TYPE}",'
        f'N({L_HOURS}{r})>IF(N($C${DEN_ROW})>0,$C${DEN_ROW},{STANDARD_ANNUAL_HOURS})),"{WARN_HOURS_OVER}",""))))'
    )


def f_period_warn() -> str:
    s = f'${L_PERIOD_START}${PERIOD_ROW}'
    e = f'${L_PERIOD_END}${PERIOD_ROW}'
    return (
        f'=IF(OR({s}="",{e}=""),"{WARN_PERIOD_EMPTY}",'
        f'IFERROR(IF(EDATE({s},12)-1={e},"","{WARN_PERIOD_NOT12}"),'
        f'"{WARN_PERIOD_BADDATE}"))'
    )


def f_check() -> str:
    """要確認・警告の総件数。V列の⚠ ＋ 対象期間の警告。提出用シートのバナーが参照。"""
    return (
        f'=COUNTIF({L_WARN}{FIRST_DATA_ROW}:{L_WARN}{LAST_DATA_ROW},"⚠*")'
        f'+IF($A${PERIOD_WARN_ROW}="",0,1)'
    )


def f_month_header(n: int) -> str:
    """「Nヶ月目」＋対象期間入力後は実年月を自動表示（暦年取り違え防止）。"""
    s = f'${L_PERIOD_START}${PERIOD_ROW}'
    return (
        f'=IF({s}="","{n}ヶ月目",'
        f'"{n}ヶ月目"&CHAR(10)&IFERROR(TEXT(EDATE({s},{n - 1}),"yyyy年m月"),""))'
    )


# ── セルスタイル適用ヘルパー ──
def put(ws, row, col, value=None, *, font=F_NORMAL, fill=None, align='center',
        fmt=None, border=True, locked=True, wrap=False):
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.font = font
    if fill is not None:
        cell.fill = fill
    cell.alignment = Alignment(
        horizontal=align, vertical='center', wrap_text=wrap
    )
    if border:
        cell.border = BORDER
    if fmt:
        cell.number_format = fmt
    cell.protection = UNLOCKED if not locked else LOCKED
    return cell


# ── 記入方法シート ──
def build_help_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 108

    lines = [
        ('賃金台帳（総額提出用）シート', F_TITLE, None),
        ('', None, None),
        ('本シートは、賃金台帳（個人別の給与明細）の提出が難しい場合に、会社全体の', F_NORMAL, None),
        (f'「集計値（総額・人数）」だけをご提出いただくための様式です（従業員{DATA_ROWS}名分まで）。', F_NORMAL, None),
        ('', None, None),
        ('【ご記入の流れ】', F_BOLD, None),
        ('① 「計算用（社内控え・非提出）」シートに、対象期間（前期決算期の12ヶ月）と、', F_NORMAL, None),
        ('　 従業員お一人ずつの毎月の支給額などをご入力ください。', F_NORMAL, None),
        ('② 入力に応じて、補助金の要件に沿った合計・人数が自動で計算されます。', F_NORMAL, None),
        ('　 ⚠マークの警告が残っている間は集計が確定していません。すべて解消してください。', F_NOTE, None),
        ('③ 「提出用シート（総額のみ）」に自動反映されるので、会社名・記入日を入れて', F_NORMAL, None),
        ('　 内容をご確認のうえ、そのシートだけをPDF化してご提出ください。', F_NORMAL, None),
        ('', None, None),
        ('【入力欄の色】', F_BOLD, None),
        ('　←この色のセルがご入力いただく欄です。', F_NORMAL, None, FILL_INPUT),
        ('　←この色のセルは自動計算です（入力不要）。', F_NORMAL, None, FILL_CALC),
        ('', None, None),
        ('【毎月の金額に入れるもの（重要）】', F_BOLD, None),
        ('・各月の「総支給額」（基本給＋残業・役職・家族・住宅などの各種手当＋通勤手当を含む、', F_NORMAL, None),
        ('　社会保険料や税金を差し引く前の支給合計）を数字でご入力ください。', F_NORMAL, None),
        ('・「課税支給合計（課税計）」ではなく「総支給額」でお願いします（通勤手当を含む金額）。', F_SMALL, None),
        ('　通勤手当は「年間通勤手当」欄にご記入いただいた金額でこちらが差し引きます。', F_SMALL, None),
        ('・「1ヶ月目」＝対象期間（決算期）の最初の月です。暦年の1月ではありません。', F_NORMAL, None),
        ('　対象期間を入力すると、各月の見出しに実際の年月が自動表示されます。', F_SMALL, None),
        ('', None, None),
        ('【賞与（ボーナス）】', F_BOLD, None),
        ('・毎月の欄には混ぜず、「年間賞与」欄にまとめてご入力ください。', F_NORMAL, None),
        ('・入れるのは対象期間（決算期）の決算に費用計上した賞与の合計です', F_NORMAL, None),
        ('　（経費として計上した日で判断。決算書の賞与計上額と一致させてください）。', F_SMALL, None),
        ('', None, None),
        ('【年間通勤手当】', F_BOLD, None),
        ('・毎月の欄に通勤手当を含めた場合は、通勤手当の年間合計を必ずご記入ください。', F_NORMAL, None),
        ('　課税・非課税の区別は不要です（どちらも集計から除きます）。', F_NORMAL, None),
        ('　（空欄のままだと総額が過大になり、賃上げ目標が実態より高く設定されてしまいます。', F_SMALL, None),
        ('　金額は給与ソフトの通勤手当欄・賃金台帳の通勤手当の行でご確認いただけます）', F_SMALL, None),
        ('・通勤手当を支給していない場合、または毎月の欄に通勤手当を含めていない場合は空欄で。', F_NORMAL, None),
        ('', None, None),
        ('【氏名について】', F_BOLD, None),
        ('・氏名が空欄の行は集計されません。実名を書きたくない場合は「社員A」「従業員1」等の', F_NORMAL, None),
        ('　仮名でも構いません（このシート自体が社外非提出のため、実名でも問題ありません）。', F_SMALL, None),
        ('', None, None),
        ('【集計の対象になる方・ならない方】', F_BOLD, None),
        ('・役員（取締役など）は、そもそもご記入いただかなくて結構です（補助金の集計対象外）。', F_NORMAL, None),
        ('　記入した場合も、雇用形態で「役員」を選べば自動で集計から外れます。', F_SMALL, None),
        ('・中途入社・退職・休職などで12ヶ月そろわない方は、自動で集計対象外になります', F_NORMAL, None),
        ('　（補助金は「その年度に全月分の給与を受けた方」だけを数える決まりのためです）。', F_SMALL, None),
        ('・賞与のみ支給の月（月給0円・賞与だけ支給）がある方も、給与の支給がない月がある方として', F_NORMAL, None),
        ('　集計対象外になります（12ヶ月すべて給与の支給がある方だけを数える決まりのためです）。', F_SMALL, None),
        ('・パート・アルバイトの方は「年間労働時間」を必ずご入力ください', F_NORMAL, None),
        ('　（人数を正社員の労働時間に換算するために使います。未入力の間は集計に入りません）。', F_SMALL, None),
        ('・週の労働時間が正社員より短い契約社員の方は「パート・アルバイト」を選び、', F_NORMAL, None),
        ('　年間労働時間をご入力ください（労働時間で人数換算する決まりのためです）。', F_SMALL, None),
        ('・パート・アルバイトの人数換算に使う「正社員の年間所定労働時間」は、計算用シート下部で', F_NORMAL, None),
        ('　ご確認・変更できます（初期値2080時間＝週40h×52週。自社の所定労働時間に合わせてください）。', F_SMALL, None),
        ('', None, None),
        ('【この様式が使えないケース（担当までご相談ください）】', F_BOLD, None),
        (f'・従業員が{DATA_ROWS + 1}名以上いる／従業員を雇用していない（役員のみの）法人／', F_NORMAL, None),
        ('　直近の決算期が12ヶ月に満たない（設立初年度・決算期変更など）', F_NORMAL, None),
        ('', None, None),
        ('【重要】ご提出は「提出用シート（総額のみ）」だけをお願いします。', F_NOTE, FILL_NOTE),
        ('「計算用（社内控え・非提出）」シートには個人ごとの給与が含まれるため、提出先へは共有せず、', F_NOTE, FILL_NOTE),
        ('御社内で保管してください。（誤ってExcelの「ブック全体」をPDF化しても、計算用シートの', F_NOTE, FILL_NOTE),
        ('個人明細は印刷されない設定にしてあります）', F_NOTE, FILL_NOTE),
    ]
    r = 2
    for spec in lines:
        text, font, fill = spec[0], spec[1], spec[2]
        swatch = spec[3] if len(spec) > 3 else None
        cell = ws.cell(row=r, column=2, value=text)
        if font:
            cell.font = font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if fill:
            cell.fill = fill
        if swatch is not None:
            chip = ws.cell(row=r, column=1)
            chip.fill = swatch
            chip.border = BORDER
        r += 1


# ── 計算用シート ──
def build_calc_sheet(ws, with_sample=False):
    ws.sheet_view.showGridLines = False

    # 列幅
    ws.column_dimensions['A'].width = 4      # No
    ws.column_dimensions['B'].width = 14     # 氏名
    ws.column_dimensions['C'].width = 15     # 雇用形態
    for c in range(COL_M1, COL_M12 + 1):     # 月次
        ws.column_dimensions[get_column_letter(c)].width = 10
    ws.column_dimensions[L_BONUS].width = 11
    ws.column_dimensions[L_TRANSPORT].width = 13
    ws.column_dimensions[L_HOURS].width = 12
    ws.column_dimensions[L_JUDGE].width = 20
    ws.column_dimensions[L_TAXABLE].width = 13
    ws.column_dimensions[L_FTE].width = 7
    ws.column_dimensions[L_WARN].width = 34

    # 行1 注意書き
    c = ws.cell(row=1, column=1,
                value='※このシートは社内保管用です。個人ごとの給与情報を含むため、提出先には共有しないでください。')
    c.font = F_NOTE
    c.fill = FILL_NOTE
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=COL_LAST)

    # 行3 対象期間
    ws.cell(row=PERIOD_ROW, column=1, value='対象期間（前期決算期・12ヶ月）：').font = F_BOLD
    ws.cell(row=PERIOD_ROW, column=1).alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=PERIOD_ROW, start_column=1, end_row=PERIOD_ROW, end_column=3)
    put(ws, PERIOD_ROW, COL_M1, None, fill=FILL_INPUT, fmt=NUMFMT_DATE, locked=False, align='center')
    ws.cell(row=PERIOD_ROW, column=COL_M1 + 1, value='〜').alignment = Alignment(horizontal='center', vertical='center')
    put(ws, PERIOD_ROW, COL_M1 + 2, None, fill=FILL_INPUT, fmt=NUMFMT_DATE, locked=False, align='center')
    ws.cell(row=PERIOD_ROW, column=COL_M1 + 3,
            value='（例：2025/7/1 〜 2026/6/30）').font = F_SMALL_GRAY
    ws.cell(row=PERIOD_ROW, column=COL_M1 + 3).alignment = Alignment(horizontal='left', vertical='center')

    # 行4 対象期間の警告（未入力／12ヶ月でない／日付でない）
    pc = ws.cell(row=PERIOD_WARN_ROW, column=1, value=f_period_warn())
    pc.font = F_ALERT
    pc.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=PERIOD_WARN_ROW, start_column=1,
                   end_row=PERIOD_WARN_ROW, end_column=COL_LAST)

    # 行5 セクション見出し
    c = ws.cell(row=5, column=1,
                value='■ 従業員ごとの入力（黄＝入力／緑＝自動計算）　毎月の欄＝総支給額（控除前・通勤手当込／賞与は除く）')
    c.font = F_SECTION
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=COL_LAST)

    # 行6 ヘッダー
    headers = {
        COL_NO: 'No', COL_NAME: '氏名', COL_EMP: '雇用形態',
        COL_BONUS: '年間賞与\n(課税分)', COL_TRANSPORT: '年間通勤手当\n(全額)',
        COL_HOURS: '年間労働時間\n(パート必須)', COL_JUDGE: '判定\n(自動)',
        COL_TAXABLE: '課税支給年額\n(自動)', COL_FTE: 'FTE\n(自動)', COL_WARN: '警告\n(自動)',
    }
    ws.row_dimensions[HEADER_ROW].height = 32
    for col in range(COL_NO, COL_LAST + 1):
        if COL_M1 <= col <= COL_M12:
            # 対象期間の入力に連動して実年月を表示（暦年取り違え防止）
            put(ws, HEADER_ROW, col, f_month_header(col - COL_M1 + 1),
                font=F_HEADER, fill=FILL_HEADER, wrap=True)
        else:
            put(ws, HEADER_ROW, col, headers.get(col, ''), font=F_HEADER,
                fill=FILL_HEADER, wrap=True)

    # データ行
    month_cols = list(range(COL_M1, COL_M12 + 1))
    input_cols = [COL_NAME, COL_EMP] + month_cols + [COL_BONUS, COL_TRANSPORT, COL_HOURS]
    for i in range(DATA_ROWS):
        r = FIRST_DATA_ROW + i
        put(ws, r, COL_NO, i + 1, font=F_SMALL_GRAY, locked=True)
        # 入力列
        for col in input_cols:
            fmt = NUMFMT_YEN if col in month_cols or col in (COL_BONUS, COL_TRANSPORT) else None
            if col == COL_HOURS:
                fmt = NUMFMT_HOUR
            put(ws, r, col, None, fill=FILL_INPUT, locked=False,
                fmt=fmt, align='center' if col != COL_NAME else 'left')
        # 自動計算列
        put(ws, r, COL_JUDGE, f_judge(r), fill=FILL_CALC, font=F_SMALL, align='center')
        put(ws, r, COL_TAXABLE, f_taxable(r), fill=FILL_CALC, fmt=NUMFMT_YEN, align='center')
        put(ws, r, COL_FTE, f_fte(r), fill=FILL_CALC, fmt=NUMFMT_FTE, align='center')
        put(ws, r, COL_WARN, f_warn(r), fill=FILL_CALC, font=F_SMALL, align='left')

    # プルダウン（雇用形態）
    dv_emp = DataValidation(
        type='list', formula1='"' + ','.join(EMP_TYPES) + '"', allow_blank=True,
        showErrorMessage=True,
    )
    dv_emp.error = '一覧から選択してください'
    dv_emp.prompt = '正社員／契約社員／パート・アルバイト／役員 から選択'
    ws.add_data_validation(dv_emp)
    dv_emp.add(f'{L_EMP}{FIRST_DATA_ROW}:{L_EMP}{LAST_DATA_ROW}')

    # ── 集計ブロック ──
    title = ws.cell(row=SUM_TITLE_ROW, column=1,
                    value='■ 会社全体の集計（自動計算 → 「提出用シート」に反映）')
    title.font = F_SECTION
    title.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=SUM_TITLE_ROW, start_column=1, end_row=SUM_TITLE_ROW, end_column=6)

    judge_rng = f'${L_JUDGE}${FIRST_DATA_ROW}:${L_JUDGE}${LAST_DATA_ROW}'
    emp_rng = f'${L_EMP}${FIRST_DATA_ROW}:${L_EMP}${LAST_DATA_ROW}'
    hours_rng = f'${L_HOURS}${FIRST_DATA_ROW}:${L_HOURS}${LAST_DATA_ROW}'
    taxable_rng = f'{L_TAXABLE}{FIRST_DATA_ROW}:{L_TAXABLE}{LAST_DATA_ROW}'
    fte_rng = f'{L_FTE}{FIRST_DATA_ROW}:{L_FTE}{LAST_DATA_ROW}'

    rows = [
        (R216_ROW, '給与支給総額（R216）', f'=SUM({taxable_rng})', NUMFMT_YEN, F_RESULT),
        (R215_ROW, '従業員数（FTE換算・R215）', f'=ROUND(SUM({fte_rng}),1)', NUMFMT_FTE, F_RESULT),
        (FT_COUNT_ROW, '　内訳：フルタイム人数（正社員＋契約社員）',
         f'=COUNTIFS({judge_rng},"{JUDGE_TARGET}",{emp_rng},"{FULLTIME_TYPES[0]}")'
         f'+COUNTIFS({judge_rng},"{JUDGE_TARGET}",{emp_rng},"{FULLTIME_TYPES[1]}")', '0', F_NORMAL),
        (PART_COUNT_ROW, '　内訳：パート・アルバイト人数',
         f'=COUNTIFS({judge_rng},"{JUDGE_TARGET}",{emp_rng},"{PART_TYPE}")', '0', F_NORMAL),
        (PART_HOURS_ROW, '　内訳：パート・アルバイト 年間労働時間 合計',
         f'=SUMIFS({hours_rng},{judge_rng},"{JUDGE_TARGET}",{emp_rng},"{PART_TYPE}")',
         NUMFMT_HOUR, F_NORMAL),
        (TARGET_COUNT_ROW, '　対象者 合計人数',
         f'=COUNTIF({judge_rng},"{JUDGE_TARGET}")', '0', F_NORMAL),
        (PER_CAPITA_ROW, '　参考：1人当たり給与支給総額（FTE換算1人当たり）',
         f'=IFERROR($C${R216_ROW}/$C${R215_ROW},"")', NUMFMT_YEN, F_NORMAL),
        (TOTAL_HOURS_ROW, '　参考：時間欄に入力があった対象者の年間労働時間 合計',
         f'=SUMIFS({hours_rng},{judge_rng},"{JUDGE_TARGET}")', NUMFMT_HOUR, F_SMALL_GRAY),
        (CHECK_ROW, '　入力チェック：要確認・警告の件数（0 であること）',
         f_check(), '0', F_BOLD),
    ]
    for row, label, formula, fmt, font in rows:
        lab = ws.cell(row=row, column=1, value=label)
        lab.font = font if font in (F_RESULT,) else F_BOLD if row in (R216_ROW, R215_ROW) else F_NORMAL
        lab.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        put(ws, row, 3, formula, fill=FILL_CALC, fmt=fmt, font=font, align='center')

    # 年間所定労働時間（FTE換算の分母・顧客入力／初期値2080）
    lab = ws.cell(row=DEN_ROW, column=1,
                  value='正社員の年間所定労働時間（FTE換算の分母）')
    lab.font = F_NORMAL
    lab.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=DEN_ROW, start_column=1, end_row=DEN_ROW, end_column=2)
    put(ws, DEN_ROW, 3, STANDARD_ANNUAL_HOURS, fill=FILL_INPUT, fmt='#,##0',
        font=F_NORMAL, locked=False)
    dv_hours = DataValidation(
        type='whole', operator='greaterThan', formula1='0',
        allow_blank=True, showErrorMessage=True,
    )
    dv_hours.error = '1以上の整数（例：2080）でご入力ください'
    dv_hours.prompt = '正社員がフルタイムで1年間に働く所定労働時間（例：2080）'
    ws.add_data_validation(dv_hours)
    dv_hours.add(f'C{DEN_ROW}')
    note = ws.cell(row=DEN_ROW, column=4,
                   value='※例：2080時間（週40h×52週）。自社の正社員フルタイムの所定労働時間に変更できます。パート人数の換算に使います。')
    note.font = F_SMALL_GRAY
    note.alignment = Alignment(horizontal='left', vertical='center')

    if with_sample:
        _fill_sample(ws)

    # 見出し・氏名を固定表示（横スクロールでの行取り違え防止）
    ws.freeze_panes = f'{L_M1}{FIRST_DATA_ROW}'

    # 誤って「ブック全体」をPDF化しても個人明細が出ないよう、印刷範囲を注意書きに限定
    ws.print_area = f'A1:{get_column_letter(COL_LAST)}1'

    # シート保護（自動計算セルはロック、入力欄は編集可）
    ws.protection = sheet_protection()


# 目視確認・検証で共有する記入例
# （氏名, 雇用形態, 月次12ヶ月の総支給, 年間賞与, 年間通勤手当, 年間労働時間）
SAMPLE_ROWS = [
    ('（例）山田 太郎', '正社員', [280000] * 12, 700000, 120000, None),
    ('（例）佐藤 次郎', '正社員', [250000] * 12, 500000, 60000, None),
    ('（例）鈴木 三郎', '契約社員', [220000] * 12, 300000, 0, None),
    ('（例）田中 花子', 'パート・アルバイト', [96000] * 12, 0, 0, 1040),
    ('（例）高橋 桜', 'パート・アルバイト', [80000] * 12, 0, 0, 900),
    ('（例）伊藤 一郎（社長）', '役員', [800000] * 12, 2000000, 0, None),
    ('（例）渡辺 新人（10月入社）', '正社員', [None] * 9 + [250000] * 3, 0, 0, None),
]
SAMPLE_PERIOD = (datetime.date(2025, 7, 1), datetime.date(2026, 6, 30))


def _fill_sample(ws):
    """目視確認用の記入例（薄グレー）。役員・中途を含め、除外挙動を確認できる構成。"""
    ws.cell(row=PERIOD_ROW, column=COL_M1, value=SAMPLE_PERIOD[0])
    ws.cell(row=PERIOD_ROW, column=COL_M1 + 2, value=SAMPLE_PERIOD[1])
    for i, (name, emp, months, bonus, transport, hours) in enumerate(SAMPLE_ROWS):
        r = FIRST_DATA_ROW + i
        ws.cell(row=r, column=COL_NAME, value=name)
        ws.cell(row=r, column=COL_EMP, value=emp)
        for j, v in enumerate(months):
            if v is not None:
                ws.cell(row=r, column=COL_M1 + j, value=v)
        if bonus:
            ws.cell(row=r, column=COL_BONUS, value=bonus)
        if transport:
            ws.cell(row=r, column=COL_TRANSPORT, value=transport)
        if hours is not None:
            ws.cell(row=r, column=COL_HOURS, value=hours)
        # 記入例の行を薄グレーで示す
        for col in [COL_NAME, COL_EMP] + list(range(COL_M1, COL_M12 + 1)) + \
                [COL_BONUS, COL_TRANSPORT, COL_HOURS]:
            ws.cell(row=r, column=col).fill = FILL_SAMPLE


# ── 提出用シート ──
SUBMIT_COMPANY_ROW = 2
SUBMIT_DATE_ROW = 3
SUBMIT_BANNER_ROW = 4
SUBMIT_ITEMS = [
    (6, '対象期間（前期決算期・12ヶ月）', None, None),          # 式は build 内で組む
    (8, '給与支給総額（R216）', f'C{R216_ROW}', NUMFMT_YEN),
    (10, '従業員数（FTE換算・R215）', f'C{R215_ROW}', NUMFMT_FTE),
    (12, '　内訳：フルタイム人数（正社員＋契約社員）', f'C{FT_COUNT_ROW}', '0'),
    (14, '　内訳：パート・アルバイト人数', f'C{PART_COUNT_ROW}', '0'),
    (16, '　内訳：パート・アルバイト 年間労働時間 合計', f'C{PART_HOURS_ROW}', NUMFMT_HOUR),
    (18, '　参考：1人当たり給与支給総額（FTE換算1人当たり）', f'C{PER_CAPITA_ROW}', NUMFMT_YEN),
]
SUBMIT_LAST_ROW = SUBMIT_ITEMS[-1][0] + 1


def build_submit_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 24
    ws.column_dimensions['D'].width = 16

    title = ws.cell(row=1, column=2,
                    value='【本シートのみをPDF等でご提出ください】計算用シートは提出不要です。')
    title.font = F_BOLD
    title.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)

    # 会社名・記入日（入力欄）
    for row, label, fmt in (
        (SUBMIT_COMPANY_ROW, '会社名', None),
        (SUBMIT_DATE_ROW, '記入日', NUMFMT_DATE),
    ):
        lab = ws.cell(row=row, column=2, value=label)
        lab.font = F_NORMAL
        lab.alignment = Alignment(horizontal='left', vertical='center')
        lab.border = BORDER
        put(ws, row, 3, None, fill=FILL_INPUT, locked=False,
            fmt=fmt, align='left' if fmt is None else 'center')
        ws.row_dimensions[row].height = 20

    # 警告バナー（計算用シートの要確認件数を伝搬）
    banner = ws.cell(
        row=SUBMIT_BANNER_ROW, column=2,
        value=(
            f"=IF('{CALC_SHEET}'!$C${CHECK_ROW}=0,\"\","
            f"\"⚠ 計算用シートに未解消の要確認・警告が \"&'{CALC_SHEET}'!$C${CHECK_ROW}"
            f"&\" 件あります。解消してからご提出ください。\")"
        ),
    )
    banner.font = F_ALERT
    banner.alignment = Alignment(horizontal='left', vertical='center')
    ws.merge_cells(start_row=SUBMIT_BANNER_ROW, start_column=2,
                   end_row=SUBMIT_BANNER_ROW, end_column=4)

    def ref(cell):
        return f"='{CALC_SHEET}'!{cell}"

    s = f"'{CALC_SHEET}'!{L_PERIOD_START}{PERIOD_ROW}"
    e = f"'{CALC_SHEET}'!{L_PERIOD_END}{PERIOD_ROW}"
    period = (
        f'=IF(OR({s}="",{e}=""),"（未入力）計算用シートに対象期間をご記入ください",'
        f'TEXT({s},"yyyy年m月d日")&" 〜 "&TEXT({e},"yyyy年m月d日"))'
    )
    for row, label, cell, fmt in SUBMIT_ITEMS:
        formula = period if cell is None else ref(cell)
        lab = ws.cell(row=row, column=2, value=label)
        big = label.startswith(('給与支給総額', '従業員数'))
        lab.font = F_BOLD if big else F_NORMAL
        lab.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        lab.border = BORDER
        val = ws.cell(row=row, column=3, value=formula)
        val.font = F_RESULT if big else F_NORMAL
        val.alignment = Alignment(horizontal='center', vertical='center')
        val.fill = FILL_CALC
        val.border = BORDER
        if fmt:
            val.number_format = fmt
        ws.row_dimensions[row].height = 22

    # 印刷範囲を提出用シートに限定
    ws.print_area = f'A1:D{SUBMIT_LAST_ROW}'

    # 集計値の誤上書き防止（会社名・記入日のみ入力可）
    ws.protection = sheet_protection()


def build(with_sample=False) -> Path:
    wb = Workbook()
    ws_help = wb.active
    ws_help.title = HELP_SHEET
    build_help_sheet(ws_help)

    ws_calc = wb.create_sheet(CALC_SHEET)
    build_calc_sheet(ws_calc, with_sample=with_sample)

    ws_submit = wb.create_sheet(SUBMIT_SHEET)
    build_submit_sheet(ws_submit)

    wb.active = wb.sheetnames.index(HELP_SHEET)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    name = OUTPUT_NAME if not with_sample else OUTPUT_NAME.replace('.xlsx', '_記入例.xlsx')
    out = OUTPUT_DIR / name
    wb.save(out)
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sample', action='store_true', help='記入例入り（目視確認用）を出力')
    args = ap.parse_args()
    out = build(with_sample=args.sample)
    print(f'生成: {out}')


if __name__ == '__main__':
    main()
