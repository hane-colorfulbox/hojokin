# -*- coding: utf-8 -*-
"""Drive 版の申請テンプレ原本（スプレッドシート → xlsx エクスポート）と、
リポジトリ側のテンプレを機械比較し、取り込むべき差分を一覧で出す。

背景:
    申請テンプレの「正」は Drive の原本（営業・事務メンバーが編集する生きた原本）で、
    リポジトリ側はそれをツールが使える形に整えたコピー。ツール実行時は原本を
    shutil.copy2 で丸ごとコピーして値だけ書き込むため、リポジトリ側テンプレが
    Drive から乖離すると、生成物が原本と食い違う。

    ただし Drive 版を丸ごと採用はできない。理由は3つ。
    1) config.TemplateMapping が行番号固定。Drive 側で行が増減すると全項目がズレる
    2) Drive 版にも欠陥が残る（#REF! ・プルダウン欠落を過去に確認）
    3) リポジトリ版だけが持つ修正がある（結合欠陥の修正・商品マスタ刷新など）

    そこで「丸ごと採用してよいか」を毎回判定できるようにするのが本スクリプト。
    行ズレがゼロで Drive 側にエラー値が無ければ丸ごと採用の候補、
    どちらかが崩れていれば差分同期（patch_corp_template_layout_sync.py 系）に回す。

実行方法:
    python scripts/compare_drive_template.py <Drive版.xlsx> [リポジトリ版.xlsx]

    第2引数を省略すると、Drive 版のファイル名から対応するリポジトリ版を推定する。

出力:
    シート単位のサマリと、種別ごとの差分明細（既定で各 MAX_DETAIL 件まで）。
    最後に「丸ごと採用の可否」を判定して表示する。

読み取り専用:
    どちらのファイルにも書き込まない。Drive 側にも一切触れない。
"""
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
TOOL_DIR = ROOT / 'ツール'

# 明細の表示上限（多すぎると読めないので種別ごとに打ち切る）
MAX_DETAIL = 25
# 行ズレ検知に使うラベル列（申請テンプレは B 列に項目名が入る）
LABEL_COL = 2
# Excel のエラー値。Drive 側に残っていると丸ごと採用できない
ERROR_VALUES = ('#REF!', '#N/A', '#VALUE!', '#DIV/0!', '#NAME?', '#NULL!', '#NUM!')
# Drive 版ファイル名（エクスポート時の名前）→ リポジトリ版ファイル名
# 先頭一致で返すため、個人版（キーワードが法人版を包含する）を必ず法人版より先に置く。
# 例: Drive タイトル「【原本/法人】企業名_通常枠/個人2026」の DL 名は
#     「【原本_法人】企業名_通常枠_個人2026.xlsx」で「通常枠」にも一致してしまう。
REPO_BY_KEYWORD = (
    ('通常枠_個人', '【原本_個人】企業名_通常枠_個人2026.xlsx'),
    ('通常枠', '【原本_法人】企業名_通常枠_法人2026_v2.xlsx'),
    ('インボイス枠_個人', '【原本_個人】企業名_インボイス枠_個人2026.xlsx'),
    ('インボイス枠_法人', '【原本_法人】企業名_インボイス枠_法人2026_v2.xlsx'),
)


def _nfc(s: str) -> str:
    return unicodedata.normalize('NFC', s)


def resolve_in_dir(base: Path, name: str) -> Path | None:
    """日本語ファイル名の NFC/NFD 差を吸収してディレクトリ内から探す。"""
    target = _nfc(name)
    for p in base.iterdir():
        if _nfc(p.name) == target:
            return p
    return None


def guess_repo_path(drive_path: Path) -> Path | None:
    name = _nfc(drive_path.name)
    for keyword, repo_name in REPO_BY_KEYWORD:
        if _nfc(keyword) in name:
            return resolve_in_dir(TOOL_DIR, repo_name)
    return None


def load(path: Path):
    """数式そのものを比較したいので data_only=False で開く。"""
    return openpyxl.load_workbook(path, data_only=False)


def cell_map(ws) -> dict[str, object]:
    """座標 -> 値。空セルは持たない。"""
    out = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is not None and c.value != '':
                out[c.coordinate] = c.value
    return out


def label_rows(ws) -> dict[int, str]:
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, LABEL_COL).value
        if isinstance(v, str) and v.strip():
            out[r] = v.strip()
    return out


def find_errors(cells: dict[str, object]) -> list[str]:
    hits = []
    for coord, v in cells.items():
        if isinstance(v, str) and any(e in v for e in ERROR_VALUES):
            hits.append(f'{coord}={v[:40]}')
    return sorted(hits)


def dv_set(ws) -> set[tuple[str, str]]:
    """(適用範囲, 参照元) の集合。プルダウンの位置ずれ・欠落の検知に使う。"""
    out = set()
    for dv in ws.data_validations.dataValidation:
        out.add((str(dv.sqref), str(dv.formula1)))
    return out


def compare_sheet(name: str, ws_repo, ws_drive) -> dict:
    repo_cells, drive_cells = cell_map(ws_repo), cell_map(ws_drive)
    repo_labels, drive_labels = label_rows(ws_repo), label_rows(ws_drive)

    shifted = sorted(
        r for r in set(repo_labels) | set(drive_labels)
        if repo_labels.get(r) != drive_labels.get(r)
    )
    only_repo = sorted(set(repo_cells) - set(drive_cells))
    only_drive = sorted(set(drive_cells) - set(repo_cells))
    changed = sorted(
        c for c in set(repo_cells) & set(drive_cells)
        if str(repo_cells[c]) != str(drive_cells[c])
    )
    repo_dv, drive_dv = dv_set(ws_repo), dv_set(ws_drive)
    return {
        'sheet': name,
        'label_shift': shifted,
        'only_repo': only_repo,
        'only_drive': only_drive,
        'changed': changed,
        'dv_only_repo': sorted(repo_dv - drive_dv),
        'dv_only_drive': sorted(drive_dv - repo_dv),
        'err_repo': find_errors(repo_cells),
        'err_drive': find_errors(drive_cells),
        'repo_cells': repo_cells,
        'drive_cells': drive_cells,
    }


def _print_detail(title: str, items: list, fmt) -> None:
    if not items:
        return
    print(f'    {title}: {len(items)} 件')
    for it in items[:MAX_DETAIL]:
        print(f'      {fmt(it)}')
    if len(items) > MAX_DETAIL:
        print(f'      … 他 {len(items) - MAX_DETAIL} 件（MAX_DETAIL={MAX_DETAIL} で打ち切り）')


def report_sheet(res: dict) -> None:
    print(f'\n--- シート「{res["sheet"]}」')
    _print_detail('B列ラベルの行ズレ（最優先）', res['label_shift'], lambda r: f'r{r}')
    _print_detail('Drive のみに値', res['only_drive'],
                  lambda c: f'{c} = {str(res["drive_cells"][c])[:50]}')
    _print_detail('リポジトリのみに値', res['only_repo'],
                  lambda c: f'{c} = {str(res["repo_cells"][c])[:50]}')
    _print_detail('両方にあり値が違う', res['changed'],
                  lambda c: f'{c}: repo[{str(res["repo_cells"][c])[:28]}] -> drive[{str(res["drive_cells"][c])[:28]}]')
    _print_detail('プルダウン: Drive のみ', res['dv_only_drive'], lambda d: f'{d[0]} <- {d[1][:40]}')
    _print_detail('プルダウン: リポジトリのみ', res['dv_only_repo'], lambda d: f'{d[0]} <- {d[1][:40]}')
    _print_detail('エラー値（Drive 側）', res['err_drive'], str)
    _print_detail('エラー値（リポジトリ側）', res['err_repo'], str)
    if not any(res[k] for k in
               ('label_shift', 'only_drive', 'only_repo', 'changed',
                'dv_only_drive', 'dv_only_repo', 'err_drive', 'err_repo')):
        print('    差分なし')


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 1

    drive_path = Path(argv[1])
    if not drive_path.exists():
        found = resolve_in_dir(drive_path.parent, drive_path.name) if drive_path.parent.exists() else None
        if found is None:
            print(f'❌ Drive 版が見つかりません: {drive_path}', file=sys.stderr)
            return 1
        drive_path = found

    repo_path = Path(argv[2]) if len(argv) > 2 else guess_repo_path(drive_path)
    if repo_path is None or not repo_path.exists():
        print(f'❌ リポジトリ版を特定できません（第2引数で明示してください）: {repo_path}', file=sys.stderr)
        return 1

    print('Drive 版      :', drive_path.name)
    print('リポジトリ版  :', repo_path.name)

    wb_drive, wb_repo = load(drive_path), load(repo_path)
    common = [s for s in wb_repo.sheetnames if s in wb_drive.sheetnames]
    print(f'\n=== シート構成 ===')
    print('  共通         :', ', '.join(common) or '(なし)')
    print('  Drive のみ   :', ', '.join(s for s in wb_drive.sheetnames if s not in wb_repo.sheetnames) or '(なし)')
    print('  リポジトリのみ:', ', '.join(s for s in wb_repo.sheetnames if s not in wb_drive.sheetnames) or '(なし)')

    print('\n=== シート別の差分 ===')
    results = [compare_sheet(s, wb_repo[s], wb_drive[s]) for s in common]
    for res in results:
        report_sheet(res)

    total_shift = sum(len(r['label_shift']) for r in results)
    total_err_drive = sum(len(r['err_drive']) for r in results)
    total_dv_gap = sum(len(r['dv_only_repo']) for r in results)

    print('\n=== 丸ごと採用の可否 ===')
    print(f'  行ズレ                : {total_shift} 件')
    print(f'  Drive 側のエラー値    : {total_err_drive} 件')
    print(f'  Drive で欠けたプルダウン: {total_dv_gap} 件')
    if total_shift == 0 and total_err_drive == 0 and total_dv_gap == 0:
        print('  → 丸ごと採用の候補（行番号マッピングは保たれ、Drive 側に欠陥なし）')
        print('     ※ リポジトリのみに値があるセルは巻き戻りになるため、上の明細を確認すること')
    else:
        print('  → 丸ごと採用は不可。差分同期に回す（patch_corp_template_layout_sync.py 系）')

    wb_drive.close()
    wb_repo.close()
    return 0


if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
