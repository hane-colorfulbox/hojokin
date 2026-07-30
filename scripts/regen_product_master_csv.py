# -*- coding: utf-8 -*-
"""Drive 版原本（xlsx エクスポート）の商品マスタから scripts/product_master_2026.csv を再生成する。

Drive の商品マスタが更新されたときの取り込み1段目。生成した CSV を
`patch_update_product_master.py` に食わせてテンプレ3種へ反映する（2段構え）。

- 3ファイル（通常枠/インボイス法人=『商品マスタ』、個人=『商品マスタ のコピー』※Drive 側の
  シート名整理待ち）の内容が同一であることを検証してから、通常枠版を正として書き出す。
- A列（ベンダー）は Drive では結合セル＝左上のみ値を持つため前方フィルで全行展開
  （リポジトリ流儀。VLOOKUP・プルダウンの動作には影響しない）。
- 品名は原本に忠実（★プレフィックス・表記ゆれもそのまま。直すのは Drive 側＝坂平さん経由）。

実行方法:
    python scripts/regen_product_master_csv.py <Drive版エクスポートのフォルダ>
    （フォルダ内のファイル名は 通常枠/インボイス枠_法人/インボイス枠_個人 を含むこと）
"""
import csv
import sys
import unicodedata
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / 'scripts' / 'product_master_2026.csv'

# (タグ, ファイル名判別キーワード, 商品マスタのシート名)
SOURCES = [
    ('通常枠', ('通常枠',), '商品マスタ'),
    ('インボ法人', ('インボイス枠', '法人'), '商品マスタ'),
    ('個人', ('インボイス枠', '個人'), '商品マスタ のコピー'),
]
DATA_START_ROW = 3


def _nfc(s: str) -> str:
    return unicodedata.normalize('NFC', s)


def _norm(v):
    if v is None:
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def find_source(folder: Path, keywords: tuple[str, ...]) -> Path | None:
    for p in sorted(folder.glob('*.xlsx')):
        name = _nfc(p.name)
        if all(_nfc(k) in name for k in keywords):
            return p
    return None


def read_master(path: Path, sheet: str) -> list[dict]:
    ws = openpyxl.load_workbook(path, data_only=False)[sheet]
    rows = []
    vendor = ''
    r = DATA_START_ROW
    while True:
        name = _norm(ws.cell(r, 2).value)
        if not name:
            break
        a = _norm(ws.cell(r, 1).value)
        if a:
            vendor = a  # 結合セルの左上 → 以降の行へ前方フィル
        rows.append({
            'vendor': vendor,
            'name': name,
            'dlcode': _norm(ws.cell(r, 3).value),
            'mgmt_code': _norm(ws.cell(r, 5).value),
            'lower': _norm(ws.cell(r, 6).value),
            'upper': _norm(ws.cell(r, 7).value),
        })
        r += 1
    return rows


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(__doc__)
        return 1
    folder = Path(argv[1])
    if not folder.is_dir():
        print(f'❌ フォルダが見つかりません: {folder}', file=sys.stderr)
        return 1

    all_rows: dict[str, list[dict]] = {}
    for tag, keywords, sheet in SOURCES:
        path = find_source(folder, keywords)
        if path is None:
            print(f'❌ {tag} の Drive 版が見つかりません（キーワード: {keywords}）', file=sys.stderr)
            return 1
        all_rows[tag] = read_master(path, sheet)
        print(f'{tag}: {len(all_rows[tag])}件 ({path.name} / {sheet})')

    base = all_rows['通常枠']
    same = True
    for tag in ('インボ法人', '個人'):
        if all_rows[tag] != base:
            same = False
            print(f'⚠ {tag} が通常枠と不一致:')
            for i, (a, b) in enumerate(zip(base, all_rows[tag])):
                if a != b:
                    print(f'   行{DATA_START_ROW + i}: 通常枠={a} / {tag}={b}')
            if len(base) != len(all_rows[tag]):
                print(f'   件数差: 通常枠={len(base)} {tag}={len(all_rows[tag])}')
    if not same:
        print('❌ 3ファイルの商品マスタが同一でないため中止（正を決めてから再実行）')
        return 1

    with open(CSV_PATH, 'w', encoding='utf-8', newline='') as f:
        w = csv.DictWriter(f, fieldnames=['vendor', 'name', 'dlcode', 'mgmt_code', 'lower', 'upper'])
        w.writeheader()
        w.writerows(base)
    print(f'✅ CSV 書き出し: {CSV_PATH} ({len(base)}件)')

    # 空欄フィールドの点検（dlcode/mgmt_code/lower/upper が欠ける行を可視化）
    for i, m in enumerate(base):
        missing = [k for k in ('dlcode', 'mgmt_code', 'lower', 'upper') if not m[k]]
        if missing:
            print(f'  ⚠ B{DATA_START_ROW + i} [{m["name"][:40]}] 空欄: {missing}')
    print('\n次の手順: python scripts/patch_update_product_master.py でテンプレ3種へ反映')
    return 0


if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
