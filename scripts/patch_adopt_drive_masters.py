# -*- coding: utf-8 -*-
"""Drive 原本（2026-07-30 更新）のツールマスタ連動・費用内訳マスタを
リポジトリ版テンプレへ採用する。

採用するもの:
    通常枠:
      - 新シート「シート9」（ツール名→申請内容3項目の引当マスタ＋選択肢一覧）
      - 新シート「費用内訳マスタ」（申請金額シートの内訳 VLOOKUP 用）
      - 申請内容 C71 の DV をインライン → 'シート9'!$A$5:$A$12 参照へ
      - 申請内容 C179/C180/C181 を VLOOKUP 式化（既存の プルダウン用!I/J/K DV は除去）
        ※ config.MAPPING_2026_TSUJO 側も同時にマッピング除外＋preserve 化が必要
      - 申請金額 M〜AB 列の内訳 VLOOKUP（103セル）
      - B108-113 役員ラベルの全角化（Drive 準拠）
      - C71 の既定値 'scale' を撤去（Drive 準拠＝プルダウンで選ぶ）
    インボイス法人:
      - 新シート「ツールマスタ」「費用内訳マスタ」
      - 申請内容 C61 の DV → 'ツールマスタ'!$A$4:$A$11
      - 申請内容 C164/C165 を VLOOKUP 式化（config は preserve 済み＝変更不要）
      - 申請内容 C161 のセキュリティ記入例（実行時は AI 値で上書き＝無害）
      - 申請金額 M〜T 列の内訳 VLOOKUP（40セル）
      - B98-103 役員ラベルの全角化

是正レイヤ（Drive 側の既知欠陥を直して取り込む。Drive 原本には書き戻さない）:
    - シート9 A6 「Al Works」→「AI Works」（VLOOKUP キーと商品マスタ表記の整合）
    - シート9 内の案内・見出しの行番号 C21→C71 / C170〜C172→C179〜C181（実配線に一致）
    - ツールマスタ A1 の C155・C156→C164・C165（同上）

採用しないもの（リポジトリ側が正。事後検証で存続を確認）:
    - 給与支給総額計算 H11 は =H10/(B5+B6+B7) を維持（Drive の =H1/ は欠陥）
    - 通常枠 C144 の DV は 'プルダウン用'!$C$2:$C$20 を維持（Drive は #REF!）
    - 実URL セル（D231 等）・生産性指標 C35（3.5%行）

実行方法:
    python scripts/patch_adopt_drive_masters.py [Drive版フォルダ]
    （省略時 _debug/_drive_export_20260730/）

冪等性:
    新シートは削除→再作成、DV は除去→再追加、セルは同値上書き。再実行しても同じ。
"""
import re
import shutil
import sys
import unicodedata
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
TOOL_DIR = ROOT / 'ツール'
DEFAULT_DRIVE_DIR = ROOT / '_debug' / '_drive_export_20260730'
SHINSEI = '申請内容'
AMOUNT = '申請金額'
# 申請金額の内訳 VLOOKUP が入る範囲（Drive 実測: 通常枠 M3:AB9 / インボイス M3:T7）
AMOUNT_SCAN = dict(min_row=1, max_row=20, min_col=13, max_col=28)  # M〜AB

# 行番号参照の是正（正規表現 → 置換。数字の続きを壊さないよう (?!\d) ガード）
FIX_TSUJO = [
    (re.compile(r'C21(?!\d)'), 'C71'),
    (re.compile(r'C170(?!\d)'), 'C179'),
    (re.compile(r'C171(?!\d)'), 'C180'),
    (re.compile(r'C172(?!\d)'), 'C181'),
    (re.compile(r'Al Works'), 'AI Works'),
]
FIX_INVOICE = [
    (re.compile(r'C155(?!\d)'), 'C164'),
    (re.compile(r'C156(?!\d)'), 'C165'),
]

TARGETS = {
    '通常枠': dict(
        repo='【原本_法人】企業名_通常枠_法人2026_v2.xlsx',
        drive='【原本_法人】企業名_通常枠_法人2026_drive.xlsx',
        new_sheets=['シート9', '費用内訳マスタ'],
        fixes={'シート9': FIX_TSUJO},
        tool_cell='C71', tool_dv="'シート9'!$A$5:$A$12", tool_value=None,
        vlookup_cells=('C179', 'C180', 'C181'),   # 式は Drive からコピー
        drop_dv_cells=('C179', 'C180', 'C181'),   # 既存 DV（プルダウン用!I/J/K）を除去
        copy_cells=[f'B{r}' for r in range(108, 114)],  # 役員ラベル全角化
        invariants={  # (シート, セル) -> 期待値の部分文字列
            (SHINSEI, 'D231'): 'it-shien.smrj.go.jp',
            ('生産性指標給与支給総額計算', 'C35'): '=B35*1.035',
        },
        invariant_dv={(SHINSEI, 'C144'): "'プルダウン用'!$C$2:$C$20"},
    ),
    'インボイス法人': dict(
        repo='【原本_法人】企業名_インボイス枠_法人2026_v2.xlsx',
        drive='【原本_法人】企業名_インボイス枠_法人2026_drive.xlsx',
        new_sheets=['ツールマスタ', '費用内訳マスタ'],
        fixes={'ツールマスタ': FIX_INVOICE},
        tool_cell='C61', tool_dv="'ツールマスタ'!$A$4:$A$11", tool_value=None,
        vlookup_cells=('C164', 'C165'),
        drop_dv_cells=(),
        copy_cells=[f'B{r}' for r in range(98, 104)] + ['C161'],
        invariants={
            (SHINSEI, 'D189'): 'mhlw.go.jp',
            ('給与支給総額計算', 'H11'): '=H10/(B5+B6+B7)',
        },
        invariant_dv={},
    ),
}


def _nfc(s: str) -> str:
    return unicodedata.normalize('NFC', s)


def resolve_in_dir(base: Path, name: str) -> Path | None:
    target = _nfc(name)
    for p in base.iterdir():
        if _nfc(p.name) == target:
            return p
    return None


def _apply_fixes(value, patterns):
    if not isinstance(value, str):
        return value
    for pat, rep in patterns:
        value = pat.sub(rep, value)
    return value


def _copy_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


def _nonempty(ws):
    return [c for row in ws.iter_rows() for c in row if c.value not in (None, '')]


def copy_sheet(wb_repo, wb_drive, name, fixes):
    """Drive のシートをリポジトリ版へ複製（値・式・書式・DV・結合・列幅・リンク）。"""
    if name in wb_repo.sheetnames:
        del wb_repo[name]
    # 挿入位置: Drive の並びで直前にある「リポジトリにも存在するシート」の直後
    order = wb_drive.sheetnames
    idx = order.index(name)
    pos = len(wb_repo.sheetnames)
    for prev in reversed(order[:idx]):
        if prev in wb_repo.sheetnames:
            pos = wb_repo.sheetnames.index(prev) + 1
            break
    ws_d = wb_drive[name]
    ws_r = wb_repo.create_sheet(name, pos)

    for row in ws_d.iter_rows():
        for c in row:
            if c.value is None and not c.has_style and not c.hyperlink:
                continue
            dst = ws_r.cell(c.row, c.column)
            dst.value = _apply_fixes(c.value, fixes)
            if c.has_style:
                _copy_style(c, dst)
            if c.hyperlink:
                dst.hyperlink = c.hyperlink.target
    for rng in ws_d.merged_cells.ranges:
        ws_r.merge_cells(str(rng))
    for letter, dim in ws_d.column_dimensions.items():
        if dim.width:
            ws_r.column_dimensions[letter].width = dim.width
    for dv in ws_d.data_validations.dataValidation:
        new_dv = DataValidation(type=dv.type, formula1=dv.formula1, allow_blank=dv.allow_blank)
        new_dv.sqref = str(dv.sqref)
        ws_r.add_data_validation(new_dv)
    ws_r.sheet_state = ws_d.sheet_state
    return len(_nonempty(ws_d)), len(_nonempty(ws_r))


def _dv_covering(ws, coord):
    """coord に掛かる DV のリスト。"""
    col = re.match(r'([A-Z]+)(\d+)', coord)
    letter, row = col.group(1), int(col.group(2))
    hits = []
    for dv in ws.data_validations.dataValidation:
        for rng in dv.sqref.ranges:
            c0 = openpyxl.utils.get_column_letter(rng.min_col)
            c1 = openpyxl.utils.get_column_letter(rng.max_col)
            if c0 <= letter <= c1 and rng.min_row <= row <= rng.max_row:
                hits.append(dv)
                break
    return hits


def patch_one(tag, cfg, drive_dir):
    repo_path = resolve_in_dir(TOOL_DIR, cfg['repo'])
    drive_path = resolve_in_dir(drive_dir, cfg['drive'])
    if repo_path is None or drive_path is None:
        raise SystemExit(f'{tag}: ファイルが見つかりません repo={repo_path} drive={drive_path}')

    wb_r = openpyxl.load_workbook(repo_path, data_only=False)
    wb_d = openpyxl.load_workbook(drive_path, data_only=False)
    print(f'\n### {tag}: {repo_path.name}')

    backup = repo_path.with_suffix('.xlsx.adopt.bak')
    shutil.copy2(repo_path, backup)

    # 1) 新シートの複製（是正レイヤ適用）
    for name in cfg['new_sheets']:
        n_d, n_r = copy_sheet(wb_r, wb_d, name, cfg['fixes'].get(name, []))
        print(f'  シート「{name}」: Drive {n_d}セル -> リポジトリ {n_r}セル '
              f'(位置 {wb_r.sheetnames.index(name)})')

    sh_r, sh_d = wb_r[SHINSEI], wb_d[SHINSEI]

    # 2) ツール名セルの DV 差し替え＋既定値
    for dv in _dv_covering(sh_r, cfg['tool_cell']):
        sh_r.data_validations.dataValidation.remove(dv)
    new_dv = DataValidation(type='list', formula1=cfg['tool_dv'], allow_blank=True)
    new_dv.sqref = cfg['tool_cell']
    sh_r.add_data_validation(new_dv)
    sh_r[cfg['tool_cell']] = cfg['tool_value']
    print(f'  {cfg["tool_cell"]}: DV -> {cfg["tool_dv"]} / 値 -> {cfg["tool_value"]!r}')

    # 3) VLOOKUP 式（Drive からコピー）＋不要になった既存 DV の除去
    for coord in cfg['drop_dv_cells']:
        for dv in _dv_covering(sh_r, coord):
            sh_r.data_validations.dataValidation.remove(dv)
    for coord in cfg['vlookup_cells']:
        sh_r[coord] = sh_d[coord].value
        print(f'  {coord} = {sh_d[coord].value}')

    # 4) 個別セルのコピー（役員ラベル全角化・記入例）
    for coord in cfg['copy_cells']:
        old, new = sh_r[coord].value, sh_d[coord].value
        if old != new:
            sh_r[coord] = new
            print(f'  {coord}: {repr(old)[:36]} -> {repr(new)[:50]}')

    # 5) 申請金額の内訳 VLOOKUP
    amt_r, amt_d = wb_r[AMOUNT], wb_d[AMOUNT]
    n = 0
    for row in amt_d.iter_rows(**AMOUNT_SCAN):
        for c in row:
            if c.value in (None, ''):
                continue
            dst = amt_r.cell(c.row, c.column)
            dst.value = c.value
            if c.has_style:
                _copy_style(c, dst)
            n += 1
    print(f'  申請金額: 内訳 VLOOKUP {n}セル')

    wb_r.save(repo_path)
    wb_r.close()
    wb_d.close()
    print(f'  バックアップ: {backup.name}')
    return verify_one(tag, cfg, repo_path, drive_path)


def verify_one(tag, cfg, repo_path, drive_path):
    wb_r = openpyxl.load_workbook(repo_path, data_only=False)
    wb_d = openpyxl.load_workbook(drive_path, data_only=False)
    ok = True

    def check(label, cond):
        nonlocal ok
        print(f'  {"✅" if cond else "❌"} {label}')
        ok = ok and cond

    for name in cfg['new_sheets']:
        same_cells = len(_nonempty(wb_r[name])) == len(_nonempty(wb_d[name]))
        dv_same = (len(wb_r[name].data_validations.dataValidation)
                   == len(wb_d[name].data_validations.dataValidation))
        check(f'シート「{name}」セル数一致・DV数一致', same_cells and dv_same)

    for sheet_name, fixes in cfg['fixes'].items():
        ws = wb_r[sheet_name]
        text = '\n'.join(str(c.value) for c in _nonempty(ws))
        residue = [pat.pattern for pat, _ in fixes if pat.search(text)]
        check(f'「{sheet_name}」是正の取り残しなし', not residue)
        if residue:
            print(f'     取り残し: {residue}')

    sh = wb_r[SHINSEI]
    tool_dvs = [dv for dv in sh.data_validations.dataValidation
                if cfg['tool_cell'] in str(dv.sqref)]
    check(f'{cfg["tool_cell"]} の DV がマスタ参照',
          len(tool_dvs) == 1 and tool_dvs[0].formula1 == cfg['tool_dv'])
    for coord in cfg['vlookup_cells']:
        check(f'{coord} が Drive と同じ式', sh[coord].value == wb_d[SHINSEI][coord].value)
    for coord in cfg['drop_dv_cells']:
        check(f'{coord} の旧 DV 除去', not _dv_covering(sh, coord))

    amt_r, amt_d = wb_r[AMOUNT], wb_d[AMOUNT]
    cnt = lambda ws: sum(1 for row in ws.iter_rows(**AMOUNT_SCAN)
                         for c in row if c.value not in (None, ''))
    check('申請金額 内訳セル数一致', cnt(amt_r) == cnt(amt_d))
    prod_dvs = [dv for dv in amt_r.data_validations.dataValidation
                if '商品マスタ' in str(dv.formula1)]
    check('申請金額 商品DV（B3:B61）存続',
          len(prod_dvs) == 1 and '$B$61' in prod_dvs[0].formula1)

    for (sheet_name, coord), expect in cfg['invariants'].items():
        val = str(wb_r[sheet_name][coord].value or '')
        check(f'不変: {sheet_name}!{coord} に {expect[:30]}', expect in val)
    for (sheet_name, coord), expect in cfg['invariant_dv'].items():
        dvs = _dv_covering(wb_r[sheet_name], coord)
        check(f'不変: {sheet_name}!{coord} の DV = {expect}',
              len(dvs) == 1 and dvs[0].formula1 == expect)

    wb_r.close()
    wb_d.close()
    print(f'  === {tag}: {"OK" if ok else "NG（要確認）"} ===')
    return ok


def main(argv):
    drive_dir = Path(argv[1]) if len(argv) > 1 else DEFAULT_DRIVE_DIR
    if not drive_dir.exists():
        raise SystemExit(f'Drive 版フォルダが見つかりません: {drive_dir}')
    all_ok = all(patch_one(tag, cfg, drive_dir) for tag, cfg in TARGETS.items())
    print('\n=== 総合結果:', 'OK' if all_ok else 'NG（要確認）', '===')
    return 0 if all_ok else 1


if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
