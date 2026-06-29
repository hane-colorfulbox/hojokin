# -*- coding: utf-8 -*-
"""申請テンプレ3種の『商品マスタ』を最新の登録ITツール一覧へ刷新し、
『申請金額』シートのプルダウン（商品選択）を最新化する。

背景:
    申請ツールが生成するExcelの『申請金額』シートでツール（商品）を選ぶプルダウンが
    古い一覧を出していた（2026-06-26 補助金MTGで坂平さんが報告）。Drive上の原本
    （Googleシート）の『商品マスタ』は最新化済みだが、ツールはDriveを読まず同梱
    .xlsx を使うため反映されていなかった。

修正内容（同梱xlsxへの外科的データ移植。他シート・他セルは触らない）:
    - 『商品マスタ』B3:G... の行を最新一覧（scripts/product_master_2026.csv）へ差し替え。
      列対応: A=ベンダー/コンソーシアム(参考) / B=ITツール名 / C=No.(DLコード) /
      D=ITツールNo.(空) / E=管理コード / F=下限 / G=上限。
      ※『申請金額』のVLOOKUP（B:F の2・4列目、B:H の5列目）が参照するのは C/E/F。
    - 通常枠のみ『申請金額』のプルダウンが商品マスタを参照せず数式直書きだったため、
      インボイス枠と同じく `'商品マスタ'!$B$3:$B$<末尾>` 参照に作り直す（D3:D7）。
    - インボイス法人/個人は既存の商品マスタ参照DVの範囲を末尾行まで更新（$B$3:$B$28 → 末尾）。

データ元: scripts/product_master_2026.csv（Drive原本3マスターの商品マスタは同一内容と確認済み）。

実行方法:
    python scripts/patch_update_product_master.py

冪等性:
    商品マスタは毎回クリア→CSVから再書き込み。DVは対象を除去→再追加/再設定。
    再実行しても結果は同じ。
"""
import csv
import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
CSV_PATH = ROOT / 'scripts' / 'product_master_2026.csv'

MASTER_SHEET = '商品マスタ'
AMOUNT_SHEET = '申請金額'
DATA_START_ROW = 3                  # 商品マスタのデータ開始行（1=編集不可,2=ヘッダ）
CLEAR_UNTIL_ROW = 120               # クリア対象の下限（旧26件/新43件を十分カバー）
DROPDOWN_COLS_LETTER = 'D'          # 申請金額の商品選択列

# (テンプレ表示名, ファイル名, プルダウンsqref)
#   通常枠は D3:D7（A3..A7=1..5の5行）に合わせて単一DVへ作り直す。
#   インボイスは既存DV（D3:D6）の範囲のみ更新（sqrefは据え置き＝最小変更）。
TEMPLATES = [
    ('通常枠',     '【原本_法人】企業名_通常枠_法人2026_v2.xlsx',     'D3:D7', 'rebuild'),
    ('インボ法人', '【原本_法人】企業名_インボイス枠_法人2026_v2.xlsx', None,    'retarget'),
    ('個人',       '【原本_個人】企業名_インボイス枠_個人2026.xlsx',   None,    'retarget'),
]

# 非回帰チェック対象（触っていないのに変わっていないこと）
INVARIANT_FORMULAS = {  # 申請金額シートの主要数式
    'B3': "=VLOOKUP(D3,'商品マスタ'!B:F,2,FALSE)",
    'C3': "=VLOOKUP(D3,'商品マスタ'!B:F,4,FALSE)",
    'J3': "=VLOOKUP(D3,'商品マスタ'!B:H,5,FALSE)",
    'H8': '=SUM(H3:H7)',
    'H9': '=H8*0.1',
    'H11': '=SUM(H8:H10)',
}


def _to_int(v):
    v = (v or '').strip()
    return int(v) if v else None


def _load_master():
    with open(CSV_PATH, encoding='utf-8') as f:
        rows = list(csv.DictReader(f))
    if not rows:
        raise SystemExit('CSVが空です')
    return rows


def _copy_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


def _dv_touches_dropdown(dv):
    """DVのsqrefが D3:D7 のいずれかに掛かるか。"""
    for rng in dv.sqref.ranges:
        for col in range(rng.min_col, rng.max_col + 1):
            if openpyxl.utils.get_column_letter(col) == DROPDOWN_COLS_LETTER:
                if rng.min_row <= 7 and rng.max_row >= 3:
                    return True
    return False


def patch_one(tag, path, sqref, dv_mode, master):
    wb = openpyxl.load_workbook(path, data_only=False)
    sheetnames_before = list(wb.sheetnames)

    if MASTER_SHEET not in wb.sheetnames:
        raise SystemExit(f'{tag}: {MASTER_SHEET} シートがありません')
    ws = wb[MASTER_SHEET]
    amt = wb[AMOUNT_SHEET]

    # 申請金額シートの主要数式（非回帰用に編集前を記録）
    formulas_before = {addr: amt[addr].value for addr in INVARIANT_FORMULAS}
    # 全シートのデータ検証(プルダウン)件数を記録（申請金額以外は不変であるべき）
    dv_counts_before = {s: len(wb[s].data_validations.dataValidation) for s in wb.sheetnames}

    # --- 商品マスタの書式テンプレ（既存3行目セル）を参照退避（値クリアでは書式は消えない） ---
    style_cols = range(1, 8)  # A..G
    row3_cells = {c: ws.cell(DATA_START_ROW, c) for c in style_cols}

    # 想定外データ（新データ末尾より下に既存内容が無いか）を点検
    new_last = DATA_START_ROW + len(master) - 1
    stray = [r for r in range(new_last + 1, CLEAR_UNTIL_ROW + 1)
             if ws.cell(r, 2).value not in (None, '')]
    if stray:
        print(f'  ⚠ {tag}: 商品マスタ {new_last+1}行目以降に既存データ行 {stray} を検出（クリアされます）')

    # --- クリア（A..G, 3..CLEAR_UNTIL_ROW） ---
    for r in range(DATA_START_ROW, CLEAR_UNTIL_ROW + 1):
        for c in style_cols:
            ws.cell(r, c).value = None

    # --- 最新行を書き込み（D列=ITツールNo.は空のまま） ---
    for i, m in enumerate(master):
        r = DATA_START_ROW + i
        ws.cell(r, 1).value = m['vendor'] or None      # A ベンダー(参考)
        ws.cell(r, 2).value = m['name']                # B ITツール名
        ws.cell(r, 3).value = m['dlcode']              # C No.(DLコード)
        # D は空
        ws.cell(r, 5).value = m['mgmt_code']           # E 管理コード
        ws.cell(r, 6).value = _to_int(m['lower'])      # F 下限
        ws.cell(r, 7).value = _to_int(m['upper'])      # G 上限
        for c in style_cols:                            # 書式は既存3行目を踏襲
            _copy_style(row3_cells[c], ws.cell(r, c))

    # --- 申請金額シートのプルダウン ---
    ref = f"'{MASTER_SHEET}'!$B${DATA_START_ROW}:$B${new_last}"
    dv_list = amt.data_validations.dataValidation
    dv_before = len(dv_list)
    target_dvs = [dv for dv in dv_list if _dv_touches_dropdown(dv)]

    if dv_mode == 'rebuild':
        # 通常枠：商品選択の既存DV（インライン2件）を除去し、商品マスタ参照の単一DVを追加
        for dv in target_dvs:
            dv_list.remove(dv)
        new_dv = DataValidation(type='list', formula1=ref, allow_blank=True)
        new_dv.sqref = sqref
        amt.add_data_validation(new_dv)
    else:
        # インボイス：既存の商品マスタ参照DVの範囲だけ更新（sqrefは据え置き）
        if len(target_dvs) != 1:
            raise SystemExit(f'{tag}: 申請金額の商品DVが想定外 ({len(target_dvs)}件)')
        target_dvs[0].formula1 = ref

    wb.save(path)

    # ===== 事後検証 =====
    wb2 = openpyxl.load_workbook(path, data_only=False)
    ws2 = wb2[MASTER_SHEET]
    amt2 = wb2[AMOUNT_SHEET]
    ok = True

    # 1) 商品マスタの中身
    got = []
    r = DATA_START_ROW
    while ws2.cell(r, 2).value not in (None, ''):
        got.append(ws2.cell(r, 2).value)
        r += 1
    n_ok = len(got) == len(master)
    head_ok = got[0] == master[0]['name'] if got else False
    tail_ok = got[-1] == master[-1]['name'] if got else False
    below_blank = ws2.cell(new_last + 1, 2).value in (None, '')
    f3_ok = ws2.cell(DATA_START_ROW, 6).value == _to_int(master[0]['lower'])
    print(f'  商品マスタ: 件数={len(got)}(期待{len(master)}) 先頭={head_ok} 末尾={tail_ok} '
          f'直下空={below_blank} F3={ws2.cell(DATA_START_ROW,6).value!r}')
    ok = ok and n_ok and head_ok and tail_ok and below_blank and f3_ok

    # 2) DV
    dvs2 = amt2.data_validations.dataValidation
    prod_dvs = [dv for dv in dvs2 if _dv_touches_dropdown(dv)]
    dv_ok = (len(prod_dvs) == 1 and prod_dvs[0].formula1 == ref and prod_dvs[0].type == 'list')
    inline_gone = all(MASTER_SHEET in str(dv.formula1) for dv in prod_dvs)
    print(f'  申請金額DV: 総数 {dv_before}->{len(dvs2)} / 商品DV={len(prod_dvs)} '
          f'formula1={prod_dvs[0].formula1!r} sqref={[str(s) for s in prod_dvs[0].sqref.ranges]}')
    ok = ok and dv_ok and inline_gone

    # 3) 非回帰：シート一覧 / 申請金額の主要数式
    sheets_ok = list(wb2.sheetnames) == sheetnames_before
    formulas_now = {addr: amt2[addr].value for addr in INVARIANT_FORMULAS}
    formula_diffs = {a: (formulas_before[a], formulas_now[a])
                     for a in INVARIANT_FORMULAS if formulas_before[a] != formulas_now[a]}
    # 期待値とも一致するか（B3/C3/J3等）
    formula_expected_ok = all(formulas_now[a] == INVARIANT_FORMULAS[a] for a in INVARIANT_FORMULAS)
    print(f'  非回帰: シート一覧不変={sheets_ok} 申請金額数式変化={formula_diffs or "なし"} '
          f'数式期待一致={formula_expected_ok}')
    ok = ok and sheets_ok and not formula_diffs and formula_expected_ok

    # 申請金額以外のシートのプルダウン件数が不変であること（openpyxl保存での欠落検知）
    dv_counts_after = {s: len(amt2.parent[s].data_validations.dataValidation) for s in wb2.sheetnames}
    other_dv_changed = {s: (dv_counts_before.get(s), dv_counts_after.get(s))
                        for s in wb2.sheetnames
                        if s != AMOUNT_SHEET and dv_counts_before.get(s) != dv_counts_after.get(s)}
    print(f'  非回帰: 他シートのDV件数変化={other_dv_changed or "なし"} '
          f'(申請金額 {dv_counts_before.get(AMOUNT_SHEET)}->{dv_counts_after.get(AMOUNT_SHEET)})')
    ok = ok and not other_dv_changed

    wb2.close()
    print(f'  === {tag}: {"OK" if ok else "NG（要確認）"} ===\n')
    return ok, got


def main():
    master = _load_master()
    print(f'CSV読込: {len(master)}件 (末尾行=商品マスタ B{DATA_START_ROW + len(master) - 1})\n')
    all_ok = True
    written = None
    for tag, fname, sqref, mode in TEMPLATES:
        path = ROOT / 'ツール' / fname
        if not path.exists():
            raise SystemExit(f'テンプレが見つかりません: {path}')
        print(f'### {tag}: {fname}')
        ok, got = patch_one(tag, path, sqref, mode, master)
        all_ok = all_ok and ok
        written = got  # 全テンプレ同一の想定

    # 書き込んだ一覧を全件表示（人手クロスチェック用）
    print('--- 書き込んだ商品マスタ（全件） ---')
    for i, name in enumerate(written or []):
        print(f'  B{DATA_START_ROW + i}: {name}')

    print('\n=== 総合結果:', 'OK' if all_ok else 'NG（要確認）', '===')
    if not all_ok:
        sys.exit(1)


if __name__ == '__main__':
    main()
