# -*- coding: utf-8 -*-
"""加点判定用賃金台帳テンプレートに月別労働時間のオプション列（S〜AE）を追加する。

背景:
    テンプレは従業員1人=1行で労働時間欄が E列（月間所定労働時間）1つしかなく、
    時給制パート等の月ごとに変動する労働時間を保持できない。AI抽出は月別時間
    （monthly_hours_override）を取得しているのに、テンプレ往復（writer→reader）で
    失われ、全月が「基本給÷E列固定時間」で換算される。中途入社月など勤務時間の
    少ない月では時間換算給与が実時給の半分以下に歪み、法定最賃割れの見た目の値が
    公式の賃金状況報告シートに転記されるリスクがあった。

修正内容:
    - S6〜AD6: BONUS1_WINDOW 12ヶ月の月別労働時間ヘッダー（『令和6年10月\\n労働時間』形式）
    - AE6: 『交付申請直近月\\n労働時間』ヘッダー（書式は F6 から複製）
    - S5:AE5 結合バナー: 任意入力である旨の案内
    - B4 説明文に月別労働時間の追記＋結合を B4:R4 → B4:AE4 に拡張
    - データ行 7〜26 の S〜AE に F7 の書式（罫線等）を複製、列幅 S〜AE=11
    - 『記入ルール』シートに S〜AE列 の説明行を挿入

    列位置は hojokin.wage_reader の BWL_COL_HOURS_WINDOW_START / BWL_COL_HOURS_LATEST
    （単一の真実）を import して使用する。

実行方法:
    python scripts/patch_bonus_ledger_monthly_hours_cols.py

冪等性:
    ヘッダー・バナー・書式・列幅の設定は再実行同値。結合は既存範囲を確認してから
    unmerge/merge。『記入ルール』の行挿入は見出しの存在チェックでスキップ。
"""
import sys
from copy import copy
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

from hojokin.wage_reader import (  # noqa: E402
    BONUS1_WINDOW,
    BWL_COL_HOURS_LATEST,
    BWL_COL_HOURS_WINDOW_START,
    BWL_COL_WINDOW_START,
    BWL_DATA_START_ROW,
    BWL_HEADER_ROW,
    BWL_SHEET_NAME,
    ym_label,
)

TEMPLATE = ROOT / 'ツール' / '加点判定用賃金台帳テンプレート.xlsx'
RULE_SHEET = '記入ルール'

DATA_END_ROW = 26          # 罫線済みデータ行の末尾（テンプレ既存仕様）
BANNER_ROW = 5
NOTE_ROW = 4
COL_WIDTH = 11.0           # F〜R と同じ

BANNER_TEXT = '月別労働時間（任意）— 月ごとに変動する場合のみ入力。空欄の月はE列を使用'
B4_APPEND = ('時給制パート等で月により労働時間が変わる場合は S〜AE列に月別労働時間を入力'
             '（空欄の月はE列で換算）。')
RULE_TITLE = '月別労働時間（S〜AE列・任意）'
RULE_BODY = ('月ごとに所定労働時間が変わる場合のみ入力（入力した月はE列より優先、空欄の月は'
             'E列を使用）。時給制パートは未入力だと時間換算給与が歪み、最低賃金割れの'
             '誤表示につながるため必ず入力。')
LATEST_HEADER = '交付申請直近月\n労働時間'


def _copy_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.protection = copy(src.protection)
    dst.number_format = src.number_format


def _remerge(ws, old_range: str, new_range: str):
    for mr in list(ws.merged_cells.ranges):
        if str(mr) in (old_range, new_range):
            ws.unmerge_cells(str(mr))
    ws.merge_cells(new_range)


def main() -> None:
    wb = openpyxl.load_workbook(TEMPLATE)
    ws = wb[BWL_SHEET_NAME]

    first_letter = get_column_letter(BWL_COL_HOURS_WINDOW_START)   # S
    last_letter = get_column_letter(BWL_COL_HOURS_LATEST)          # AE

    # ヘッダー S6〜AD6 / AE6（書式は F6 から複製）
    src_header = ws.cell(BWL_HEADER_ROW, BWL_COL_WINDOW_START)
    for j, ym in enumerate(BONUS1_WINDOW):
        cell = ws.cell(BWL_HEADER_ROW, BWL_COL_HOURS_WINDOW_START + j)
        _copy_style(src_header, cell)
        cell.value = f'{ym_label(ym)}\n労働時間'
    cell = ws.cell(BWL_HEADER_ROW, BWL_COL_HOURS_LATEST)
    _copy_style(src_header, cell)
    cell.value = LATEST_HEADER

    # S5:AE5 案内バナー
    _remerge(ws, '', f'{first_letter}{BANNER_ROW}:{last_letter}{BANNER_ROW}')
    banner = ws.cell(BANNER_ROW, BWL_COL_HOURS_WINDOW_START)
    banner.value = BANNER_TEXT
    banner.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # B4 説明文の追記＋結合の張り直し（B4:R4 → B4:AE4）
    note = ws.cell(NOTE_ROW, 2)
    if B4_APPEND not in str(note.value or ''):
        note.value = f'{note.value or ""}{B4_APPEND}'
    _remerge(ws, f'B{NOTE_ROW}:R{NOTE_ROW}', f'B{NOTE_ROW}:{last_letter}{NOTE_ROW}')

    # データ行の書式複製（F7 → S〜AE 各行）と列幅
    for r in range(BWL_DATA_START_ROW, DATA_END_ROW + 1):
        src = ws.cell(r, BWL_COL_WINDOW_START)
        for c in range(BWL_COL_HOURS_WINDOW_START, BWL_COL_HOURS_LATEST + 1):
            _copy_style(src, ws.cell(r, c))
    for c in range(BWL_COL_HOURS_WINDOW_START, BWL_COL_HOURS_LATEST + 1):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTH

    # 記入ルールシートに説明行を挿入（時間換算給与の行の手前。冪等）
    rule = wb[RULE_SHEET]
    existing = {str(rule.cell(r, 1).value or '') for r in range(1, rule.max_row + 1)}
    if RULE_TITLE not in existing:
        insert_at = next(
            (r for r in range(1, rule.max_row + 1)
             if str(rule.cell(r, 1).value or '').startswith('時間換算給与')),
            rule.max_row + 1,
        )
        rule.insert_rows(insert_at)
        _copy_style(rule.cell(insert_at + 1, 1), rule.cell(insert_at, 1))
        _copy_style(rule.cell(insert_at + 1, 2), rule.cell(insert_at, 2))
        rule.cell(insert_at, 1, RULE_TITLE)
        rule.cell(insert_at, 2, RULE_BODY)

    wb.save(TEMPLATE)
    wb.close()

    # 事後検証
    wb2 = openpyxl.load_workbook(TEMPLATE)
    ws2 = wb2[BWL_SHEET_NAME]
    merged = {str(m) for m in ws2.merged_cells.ranges}
    dv = [str(d.sqref) for d in ws2.data_validations.dataValidation]
    first_hdr = ws2.cell(BWL_HEADER_ROW, BWL_COL_HOURS_WINDOW_START).value
    last_hdr = ws2.cell(BWL_HEADER_ROW, BWL_COL_HOURS_LATEST).value
    rule2 = wb2[RULE_SHEET]
    rule_rows = [str(rule2.cell(r, 1).value or '') for r in range(1, rule2.max_row + 1)]
    checks = {
        f'S6見出し={first_hdr!r}': first_hdr == f'{ym_label(BONUS1_WINDOW[0])}\n労働時間',
        f'AE6見出し={last_hdr!r}': last_hdr == LATEST_HEADER,
        'B4:AE4結合': f'B{NOTE_ROW}:{last_letter}{NOTE_ROW}' in merged,
        'S5:AE5結合': f'{first_letter}{BANNER_ROW}:{last_letter}{BANNER_ROW}' in merged,
        'C2プルダウン残存': 'C2' in dv,
        '記入ルール説明行': RULE_TITLE in rule_rows,
        '記入ルール重複なし': rule_rows.count(RULE_TITLE) == 1,
    }
    wb2.close()
    print('=== 修正後の検証 ===')
    for label, ok in checks.items():
        print(f'  {"OK" if ok else "NG"}: {label}')
    if all(checks.values()):
        print('=== 結果: OK ===')
    else:
        print('=== 結果: NG（要確認） ===')
        sys.exit(1)


if __name__ == '__main__':
    main()
