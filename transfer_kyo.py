# -*- coding: utf-8 -*-
"""
京のお肉処弘 — 補助金申請フォーマット自動転記スクリプト
原本テンプレート(2026) + ヒアリングシート(2025) + PDF読み取りデータ → 出力
"""
import sys
import shutil
import datetime
import openpyxl
from pathlib import Path

sys.stdout.reconfigure(encoding='utf-8')

# ── パス定義 ──
BASE = Path('c:/Users/user/projects/カラフルボックス/補助金')
RESOURCE = BASE / '京の食事処資料'


def find_file(directory, name_contains):
    """日本語パスの文字化け対策: iterdir で実ファイルを探す"""
    for p in directory.iterdir():
        if name_contains in p.name:
            return p
    return None


# ── 電話番号変換 ──
PHONE_DIGITS = {
    758022224: '0758022224',
    8057866988: '08057866988',
}


def normalize_phone(value):
    """数値の電話番号を先頭0補完した文字列に変換"""
    if isinstance(value, (int, float)):
        v = int(value)
        if v in PHONE_DIGITS:
            return PHONE_DIGITS[v]
        s = str(v)
        if len(s) == 9:
            return '0' + s
        if len(s) == 10 and not s.startswith('0'):
            return '0' + s
        return s
    return str(value) if value is not None else None


# ── ヒアリング行 → 転記行 マッピング ──
# (ヒアリング行, 転記行, 電話番号変換フラグ)
HEARING_TO_TENKI = [
    (6,  8,  False),   # 企業名
    (8,  10, False),   # 企業名（フリガナ）
    (10, 12, False),   # 店舗事業所数
    (12, 14, False),   # 事業者URL
    (14, 16, False),   # 事業内容 → 主な事業内容
    (20, 27, True),    # 代表電話番号
    (22, 29, False),   # 担当者氏名
    (24, 31, False),   # 担当者氏名フリガナ
    (26, 33, False),   # 担当者メールアドレス
    (28, 35, True),    # 担当者電話番号
    (30, 37, True),    # 担当者携帯番号
    (32, 39, False),   # 従業員数：正規雇用
    (34, 41, False),   # 従業員数：契約社員
    (36, 43, False),   # 従業員数：パートアルバイト
    (38, 45, False),   # 従業員数：派遣社員
    (40, 47, False),   # 従業員数：その他
    (42, 49, False),   # 過去に補助金
    (59, 58, False),   # えるぼし
    (62, 61, False),   # くるみん
    (64, 63, False),   # SECURITY ACTION自己宣言ID
    (69, 66, False),   # 従業員数正規雇用(前期)
    (70, 67, False),   # 従業員数契約社員(前期)
    (71, 68, False),   # 従業員数パート(前期)
    (72, 69, False),   # 年間平均労働時間
    (75, 72, False),   # 自社の強み
    (76, 73, False),   # 自社の弱み
    (77, 74, False),   # IT投資年間金額
    (78, 75, False),   # IT投資プロセス (None → skip)
    (81, 78, False),   # 事業所内最低賃金時給
    (83, 80, False),   # 賃金引上げ表明
    (85, 82, False),   # 賃金引上げ幅
    (87, 84, False),   # 従業員代表者
    (88, 85, False),   # 給与担当者
    (89, 86, False),   # 事業所内最低賃金者
]


# ── テキスト項目（転記シート行16-25） ──
TEXT_ITEMS = {
    17: '独自性ある食肉加工技術と営業力。8店舗展開による販売網と新商品開発力。',
    18: '在庫管理・業務管理がアナログで把握できておらず、社員の高齢化や退職が進み、'
        '人材が育たない状況。評価基準が不明確で従業員のモチベーション維持が困難。',
    19: '明確な評価制度がなく、従業員の目標設定や成果の可視化ができていない。'
        '若手の育成・定着が課題。',
    20: '人事評価の属人的運用により月20時間以上の管理工数が発生。'
        '離職に伴う採用・教育コストも増大。',
    21: '目標管理機能による全従業員の目標設定・進捗管理・評価の一元化。'
        'AI分析による評価の公平性担保。',
    22: '人事管理工数を月20時間から5時間に削減。'
        '評価制度確立により離職率を大幅改善。',
    23: '新メニュー・新商品の開発、販路拡大の営業活動、'
        '従業員のスキルアップ研修の実施。',
    24: '現在の約14.7億円から16億円（109%）を目指す。',
    25: '個人消費者、飲食店、スーパーマーケット、百貨店',
}


# ── 事業内容（255文字以内） ──
JIGYO_NAIYOU = (
    '京都を拠点に食肉の加工・販売および弁当・惣菜の製造販売を8店舗で展開する企業である。'
    '独自性ある商品力と強固な営業基盤が強みだが、業務管理のアナログ運用や人事評価基準の'
    '不明確さにより従業員の定着に課題を抱え、営業損失が発生している状況にある。'
    '本事業ではクラウド型人事評価ツール「cbox」を導入し、全従業員の目標設定・進捗管理・'
    '評価をデジタル化することで、公平な評価制度を確立し離職率を低減させる。'
    '業務効率化により生まれた時間を新メニュー開発や販路拡大の営業に充て、'
    '利益回復と年率1.5%以上の賃上げを実現する。'
)


# ── 業種分類 ──
GYOSHU_CODE = '0961'
GYOSHU_TEXT = (
    '大分類 E 製造業 / '
    '中分類 09 食料品製造業 / '
    '小分類 096 畜産食料品製造業 / '
    '細分類 0961 部分肉・冷凍肉製造業'
)


def step1_clear_sample(wb):
    """STEP 1: サンプルデータをクリア"""
    cleared = 0

    # 転記シート: B列 行16-25のテキスト項目をクリア
    ws_t = wb['転記']
    for r in range(16, 26):
        cell = ws_t.cell(row=r, column=2)
        if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith('=')):
            cell.value = None
            cleared += 1

    # 申請内容シート: C列の直接値をクリア（数式は残す）
    ws_s = wb['申請内容']
    for row in ws_s.iter_rows(min_row=5, max_row=ws_s.max_row):
        cell_c = row[2] if len(row) > 2 else None
        if cell_c is None:
            continue
        v = cell_c.value
        if v is not None and not (isinstance(v, str) and v.startswith('=')):
            cell_c.value = None
            cleared += 1

    # 生産性指標給与支給総額計算シート: 直接値をクリア
    ws_k = wb['生産性指標給与支給総額計算']
    direct_cells = [
        (10, 2), (11, 2), (12, 2), (13, 2),  # B列: 売上〜経常利益
        (5, 5), (6, 5), (7, 5), (8, 5), (9, 5),   # E列: 給与関連
        (5, 6), (6, 6), (7, 6), (8, 6), (9, 6),   # F列: 原価報告書
        (21, 5), (21, 6),                          # E21, F21: 減価償却費
    ]
    for r, c in direct_cells:
        cell = ws_k.cell(row=r, column=c)
        if cell.value is not None and not (isinstance(cell.value, str) and str(cell.value).startswith('=')):
            cell.value = None
            cleared += 1

    return cleared


def step2_hearing_to_tenki(ws_hearing, ws_tenki):
    """STEP 2: ヒアリングシートから転記シートへ（ラベルマッチングではなく行番号マッピング）"""
    count = 0
    skipped = []

    for h_row, t_row, is_phone in HEARING_TO_TENKI:
        value = ws_hearing.cell(row=h_row, column=3).value  # C列
        h_label = ws_hearing.cell(row=h_row, column=2).value  # B列（ログ用）
        t_label = ws_tenki.cell(row=t_row, column=1).value    # A列（ログ用）

        if value is None:
            skipped.append(f'  スキップ: ヒアリング行{h_row} [{h_label}] → 転記行{t_row} (値なし)')
            continue

        if is_phone:
            value = normalize_phone(value)

        ws_tenki.cell(row=t_row, column=2).value = value
        count += 1
        print(f'  転記: ヒアリング行{h_row:2d} [{h_label}] → 転記行{t_row:2d} [{t_label}]: {value!r}')

    # テキスト項目（行17-25）
    for t_row, text in TEXT_ITEMS.items():
        ws_tenki.cell(row=t_row, column=2).value = text
        t_label = ws_tenki.cell(row=t_row, column=1).value
        count += 1
        print(f'  テキスト: 転記行{t_row:2d} [{t_label}]: {text[:40]}...')

    for s in skipped:
        print(s)

    return count


def step3_pdf_to_sheets(ws_shinsei, ws_kyuyo):
    """STEP 3: PDF読み取りデータ → 申請内容シート + 給与計算シート"""
    writes = []

    def ws(row, val, label):
        ws_shinsei.cell(row=row, column=3).value = val
        writes.append(f'  申請内容 行{row:3d} [{label}]: {val!r}')

    def wk(row, col, val, label):
        ws_kyuyo.cell(row=row, column=col).value = val
        col_letter = chr(64 + col)
        writes.append(f'  給与計算 行{row:3d} {col_letter}列 [{label}]: {val}')

    # ── 履歴事項全部証明書 ──
    # 行52 事業者名フリガナ: 数式 ='転記'!B10 → そのまま
    ws(54, '京都市下京区四条大宮町２番地', '本店所在地')
    ws(55, GYOSHU_CODE, '業種コード')
    ws(56, GYOSHU_TEXT, '業種分類')
    ws(57, datetime.datetime(2024, 3, 13), '設立年月日')
    ws(58, 5_000_000, '資本金')
    ws(74, '3月', '決算月')
    ws(81, 4, '代表者・役員数(申請時)')
    ws(82, '代表取締役', '代表者役職')
    ws(83, '西田\u3000一弘', '代表者氏名')
    ws(84, 'ニシダ\u3000カズヒロ', '代表者氏名フリガナ')

    # 役員
    ws(87, '取締役', '役員(1)役職')
    ws(88, '西田\u3000賢弘', '役員(1)氏名')
    ws(89, 'ニシダ\u3000ケンヒロ', '役員(1)フリガナ')
    ws(90, '取締役', '役員(2)役職')
    ws(91, '飴田\u3000猛一', '役員(2)氏名')
    ws(92, 'アメダ\u3000タケカズ', '役員(2)フリガナ')
    ws(93, '取締役', '役員(3)役職')
    ws(94, '髙嶋\u3000滋夫', '役員(3)氏名')
    ws(95, 'タカシマ\u3000シゲオ', '役員(3)フリガナ')

    ws(144, 4, '代表者・役員数(前期)')

    # ── 認定系（ヒアリングの選択肢を変換） ──
    ws(113, 'なし', '過年度交付決定')
    ws(114, '認定なし', 'えるぼし認定')
    ws(115, '認定なし', 'くるみん認定')

    # ── 行っている事業 ──
    ws(134, 'E、製造業, I、卸売業、小売業, M、宿泊業、飲食サービス業', '行っている事業')

    # ── 経営状況（営業利益マイナス） ──
    ws(157, '□事業の拡大に積極的\n■事業の維持に注力\n□事業の売却・整備・廃業を考えている\n□特に意識したことは無い', '経営意欲')
    ws(171, '□事業の拡大\n■利益の確保\n□顧客の定着\n□人材の再配置による営業力や生産力の強化\n□従業員の人材育成、経営陣の経営能力の向上\n□顧客満足度・利便性やサービスの質の向上による満足度の向上\n□従業員の満足度の向上\n□その他（フリー記載）\n□わからない', '将来目標')

    # ── セキュリティの状況（SECURITY ACTION取得済み） ──
    ws(167, '□緊急時の対応マニュアルや手順を定め、定期的に訓練を行っている\n■パソコンやサーバなどには、IDやパスワードを設け情報セキュリティ管理を行っている\n□セキュリティ対策は講じていないため、対策を講じていく\n□セキュリティ対策を講じておらず、今後もその予定はない', 'セキュリティの状況')

    # ── 地域別最低賃金（京都府） ──
    ws(202, '京都府/1058円', '地域別最低賃金')

    # ── 賃上げ関連 ──
    ws(210, '■はい\n□いいえ', '賃上げ表明')
    ws(211, '＋50円', '賃上げ幅')
    ws(212, '□社内掲示板などへの掲載によって\n■朝礼時、会議、面談時など口頭によって\n□書面、電子メールによって\n□その他', '表明方法')
    ws(213, datetime.datetime(2026, 3, 1), '表明日付')

    # ── ツール名 ──
    ws(71, 'cbox', 'ツール名')

    # ── 事業内容（255文字以内） ──
    ws(73, JIGYO_NAIYOU, '事業内容')

    # ── IT投資関連 ──
    ws(162, '■今までIT投資を行っていなかった', 'IT投資状況')
    ws(164, '■ITツールを導入しておらず、今回が初めてである', 'IT活用状況')

    # ── 決算書 第2期（直近: R6.4.1〜R7.3.31） → 給与計算シート ──
    wk(10, 2, 1_476_107_055, '売上高')
    wk(11, 2, 623_244_955, '粗利益')
    wk(12, 2, -3_575_005, '営業利益')
    wk(13, 2, -2_034_572, '経常利益')

    # 減価償却費: E22 = E21 + F21 なので E21 に入れる
    wk(21, 5, 374_218, '減価償却費(E21)')

    # 給与関連（E列 = 販管費）
    wk(5, 5, 102_890_664, '給料手当')
    wk(6, 5, 79_487_112, '雑給')
    wk(7, 5, 11_501_000, '賞与手当')
    wk(8, 5, 0, '役員報酬')

    return writes


def step4_ai_text():
    """STEP 4: 事業内容テキストの文字数チェック"""
    length = len(JIGYO_NAIYOU)
    status = 'OK' if 250 <= length <= 255 else 'NG'
    return length, status


def step5_check_empty(wb):
    """STEP 5: 申請内容シートで空のままのセルを一覧表示"""
    ws = wb['申請内容']
    empty = []
    for row in ws.iter_rows(min_row=35, max_row=248):
        row_num = row[0].row
        b_val = row[1].value if len(row) > 1 else None
        c_val = row[2].value if len(row) > 2 else None
        if b_val is not None and c_val is None:
            empty.append(f'  行{row_num:3d} [{b_val}]')
    return empty


def main():
    template_path = find_file(BASE, '原本_法人')
    hearing_path = find_file(RESOURCE, 'ヒアリングシート')
    output_path = RESOURCE / '京のお肉処弘_通常枠_AI版.xlsx'

    if not template_path:
        print('エラー: 原本テンプレートが見つかりません')
        sys.exit(1)
    if not hearing_path:
        print('エラー: ヒアリングシートが見つかりません')
        sys.exit(1)

    print(f'原本テンプレート: {template_path.name}')
    print(f'ヒアリングシート: {hearing_path.name}')
    print(f'出力先: {output_path}')

    # テンプレートをコピー
    shutil.copy2(template_path, output_path)

    wb_h = openpyxl.load_workbook(hearing_path)
    wb_f = openpyxl.load_workbook(output_path)

    # ── STEP 1: サンプルデータクリア ──
    print('\n── STEP 1: サンプルデータをクリア ──')
    cleared = step1_clear_sample(wb_f)
    print(f'  {cleared}セル クリア完了')

    # ── STEP 2: ヒアリングシート → 転記シート ──
    print('\n── STEP 2: ヒアリングシート → 転記シート ──')
    count = step2_hearing_to_tenki(wb_h['基本情報'], wb_f['転記'])
    print(f'  計 {count}件 転記完了')

    # ── STEP 3: PDF読み取りデータ → 申請内容 + 給与計算 ──
    print('\n── STEP 3: PDF読み取りデータ → 申請内容 + 給与計算 ──')
    writes = step3_pdf_to_sheets(wb_f['申請内容'], wb_f['生産性指標給与支給総額計算'])
    for w in writes:
        print(w)

    # ── STEP 4: AI判断テキストチェック ──
    print('\n── STEP 4: 事業内容テキスト文字数チェック ──')
    length, status = step4_ai_text()
    print(f'  事業内容: {length}文字 [{status}]')
    if status == 'NG':
        print(f'  警告: 250〜255文字の範囲外です。調整が必要です。')

    # ── STEP 5: 空セル確認 ──
    print('\n── STEP 5: 申請内容シート 空セル一覧 ──')
    empty = step5_check_empty(wb_f)
    if empty:
        for e in empty:
            print(e)
        print(f'  計 {len(empty)}件 未入力')
    else:
        print('  空セルなし')

    # ── バリデーション ──
    print('\n── バリデーション ──')
    print('  履歴事項全部証明書: 取得日 2025/06/04 → 2025/09/02 までに申請要')
    print('  納税証明書: その1 → OK')
    print('  決算月: 3月 (R6.4.1〜R7.3.31)')
    print('  営業利益: マイナス(-3,575,005円) → 事業維持・利益確保を選択')

    # 保存
    wb_f.save(output_path)
    print(f'\n出力完了: {output_path}')


if __name__ == '__main__':
    main()
